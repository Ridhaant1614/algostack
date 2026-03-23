# -*- coding: utf-8 -*-
# AlgoStack Dashboard v10.9
from __future__ import annotations  # MUST be first executable line
# ── UTF-8 fix: prevents Windows Python 3.13 UnicodeEncodeError crash ──────
import sys as _sys, os as _os
_os.environ['PYTHONIOENCODING'] = 'utf-8'
_os.environ['PYTHONUTF8'] = '1'
if hasattr(_sys.stdout, 'reconfigure'):
    try: _sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception: pass
if hasattr(_sys.stderr, 'reconfigure'):
    try: _sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception: pass
print('[STARTUP] unified_dash_v3.py OK - Python', _sys.version_info.major, '.', _sys.version_info.minor, flush=True)
del _sys, _os

# ─────────────────────────────────────────────────────────────────────────

import json, logging, os, re, socket, subprocess, threading, time
import traceback, urllib.parse, urllib.request
from collections import deque
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pytz

try:
    from market_calendar import MarketCalendar
    _MC_AVAILABLE = True
except ImportError:
    _MC_AVAILABLE = False
    MarketCalendar = None

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [dash] %(levelname)s %(message)s")
log = logging.getLogger("unified_dash")

DASH_PORT   = int(os.getenv("UNIFIED_DASH_PORT", os.getenv("PORT", "8055")))
# v10.8: NGROK_TOKENS — only full authtokens work (cr_XXXX are API keys, NOT authtokens).
# We keep ONE valid authtoken. Tunnels tried in order: ngrok → cloudflared → SSH.
NGROK_TOKEN   = os.getenv("NGROK_AUTHTOKEN", "3BJFwkTTGUXId7wJVuxwYgvhzaR_4vBezPDWQKJVQJpS3M4vD")
NGROK_API_KEY = os.getenv("NGROK_API_KEY",   "3BJFjgQBSExlG3BmeIx9jDOIyHX_7tRFdhKruruAHCr1vkr58")
CLOUDFLARED   = os.getenv("CLOUDFLARED_PATH", "cloudflared")
DISABLE_CLOUDFLARE = os.getenv("DISABLE_CLOUDFLARE", "1").strip() in ("1", "true", "True")
DISABLE_PYNGROK    = os.getenv("DISABLE_PYNGROK", "1").strip() in ("1", "true", "True")
DISABLE_PUBLIC_TUNNEL = os.getenv("DISABLE_PUBLIC_TUNNEL", "0").strip() in ("1", "true", "True")
# v10.8: Keep process handles alive at module level so tunnels survive
_CF_PROC:    Optional[subprocess.Popen] = None
_NGROK_PROC: Optional[subprocess.Popen] = None  # ngrok CLI fallback handle
_LT_PROC:    Optional[subprocess.Popen] = None  # localtunnel process handle
_SSH_PROC:   Optional[subprocess.Popen] = None  # ssh-based tunnel handle (Pinggy)
_SSH_URL:    str = ""
_TUNNEL_LOCK = threading.Lock()
LEVELS_DIR  = "levels"
TRADE_DIR   = "trade_logs"
IST         = pytz.timezone("Asia/Kolkata")
CURRENT_X   = float(os.getenv("CURRENT_X_MULTIPLIER", "0.008575"))
TG_TOKEN    = os.getenv("TELEGRAM_BOT_TOKEN", "7587307352:AAG6RaiF4gO5I_ZFZ_4b8Gj7dnsu4GtPWFw")
TG_CHATS    = [c for c in [os.getenv("TELEGRAM_CHAT_ID","1376513391"),
                             os.getenv("TELEGRAM_CHAT_ID_2","793674804")] if c]
PUBLIC_LINK_PASSWORD = os.getenv("PUBLIC_LINK_PASSWORD", "Ridz@2004")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
GEMINI_API_KEY    = os.getenv("GEMINI_API_KEY", "AIzaSyB9fJ1geyar2gtgs-HYsVTOKEGwmNt_r08")
USDT_TO_INR = float(os.getenv("USDT_TO_INR", "84.0"))
# v10.9: live fetch wrapper — returns current INR rate
def _usdt_to_inr() -> float:
    global USDT_TO_INR
    try:
        import time as _tm2
        if _tm2.time() - getattr(_usdt_to_inr, '_ts', 0) < 3600:
            return USDT_TO_INR
        import urllib.request as _ur2, json as _j3
        with _ur2.urlopen("https://cdn.jsdelivr.net/npm/@fawazahmed0/currency-api@latest/v1/currencies/usd.json", timeout=4) as _rr:
            rate = float(_j3.loads(_rr.read())["usd"]["inr"])
            if 70 < rate < 120:
                USDT_TO_INR = rate; _usdt_to_inr._ts = _tm2.time()
    except Exception: pass
    return USDT_TO_INR
_usdt_to_inr._ts = 0.0
CRYPTO_X    = float(os.getenv("CRYPTO_X_MULTIPLIER", "0.008575"))
BESTX_FILE  = os.path.join("best_x_trades", "live_state.json")

# ── Startup URL dedup guard: send to Telegram ONLY ONCE per process ──────────
_STARTUP_URL_SENT: bool = False
_STARTUP_URL_LOCK  = threading.Lock()

# Mobile stability: tunnel churn can spam Telegram; throttle "tunnel restarted" updates.
_LAST_TUNNEL_TG: float = 0.0

SCANNER_DIRS = {
    1: os.path.join("sweep_results","scanner1_narrow_x0080_x0090"),
    2: os.path.join("sweep_results","scanner2_dual_x0010_x0160"),
    3: os.path.join("sweep_results","scanner3_widedual_x0010_x0320"),
}
SCANNER_DIRS_LEGACY = {
    2: os.path.join("sweep_results","scanner2_low_x0010_x0070"),
    3: os.path.join("sweep_results","scanner3_high_x0090_x0400"),
}
COMM_SCANNER_DIRS = {
    1: os.path.join("sweep_results","commodity_scanner1"),
    2: os.path.join("sweep_results","commodity_scanner2"),
    3: os.path.join("sweep_results","commodity_scanner3"),
}
CRYPTO_SCANNER_DIRS = {
    1: os.path.join("sweep_results","crypto_scanner1"),
    2: os.path.join("sweep_results","crypto_scanner2"),
    3: os.path.join("sweep_results","crypto_scanner3"),
}

# FIX 1: exact equity set — commodity/crypto symbols filtered out of equity levels table
EQUITY_STOCKS = frozenset({
    "NIFTY","BANKNIFTY","HDFCBANK","KOTAKBANK","SBIN","ICICIBANK","INDUSINDBK",
    "ADANIPORTS","ADANIENT","ASIANPAINT","BAJFINANCE","DRREDDY","SUNPHARMA",
    "INFY","TCS","TECHM","TITAN","TATAMOTORS","RELIANCE","INDIGO","JUBLFOOD",
    "BATAINDIA","PIDILITIND","ZEEL","BALKRISIND","VOLTAS","ITC","BPCL",
    "BRITANNIA","HEROMOTOCO","HINDUNILVR","UPL","SRF","TATACONSUM","BALRAMCHIN",
    "ABFRL","VEDL","COFORGE",
})
COMMODITY_SYMS = frozenset({"GOLD","SILVER","NATURALGAS","CRUDE","COPPER"})
CRYPTO_SYMS    = frozenset({"BTC","ETH","BNB","SOL","ADA"})

# ── Colours ───────────────────────────────────────────────────────────────────
BG="#0a0c10"; SB="#0d1117"; CARD="#161b22"; BORDER="#21262d"
ACCENT="#58a6ff"; GREEN="#3fb950"; RED="#f85149"
YELLOW="#e3b341"; AMBER="#d29922"; TEXT="#c9d1d9"; DIM="#6e7681"
ORANGE="#fb8f44"; PURPLE="#a371f7"; TEAL="#39d353"
FONT="-apple-system,BlinkMacSystemFont,'Segoe UI','JetBrains Mono',monospace"
EQ_COL="#58a6ff"; COMM_COL="#d29922"; CRYPTO_COL="#a371f7"

# Lightweight in-app chart cache (avoids external TradingView dependency).
_CHART_MAX_POINTS = 300
_CHART_HIST: Dict[str, deque] = {}


def _chart_push(sym: str, px: float) -> None:
    if px <= 0:
        return
    d = _CHART_HIST.get(sym)
    if d is None:
        d = deque(maxlen=_CHART_MAX_POINTS)
        _CHART_HIST[sym] = d
    d.append((datetime.now(IST).strftime("%H:%M:%S"), float(px)))


def _public_ipv4(timeout_s: float = 3.0) -> str:
    for u in ("https://api.ipify.org", "https://ipv4.icanhazip.com"):
        try:
            with urllib.request.urlopen(u, timeout=timeout_s) as r:
                ip = r.read().decode("utf-8", "ignore").strip()
                if re.match(r"^\d+\.\d+\.\d+\.\d+$", ip):
                    return ip
        except Exception:
            pass
    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  BACKGROUND DATA STORE
# ══════════════════════════════════════════════════════════════════════════════
class _DataStore:
    DATA_INTERVAL   = 0.75   # v10.9: 0.75s fast tier (was 0.5s) — reduces CPU by 33%
    XLSX_INTERVAL   = 45     # v10.9: 45s XLSX cache (was 30s)
    HIST_INTERVAL   = 90     # v10.9: 90s history (was 60s)
    SCAN_INTERVAL   = 2.0    # v10.9: 2s scanner JSON (was 1.5s)

    def __init__(self):
        self._lock = threading.Lock()
        self._d: Dict[str,Any] = {
            "live_prices":{}, "price_age":9999.0,
            "events":[], "levels":[], "levels_ts":0.0, "levels_mtime":0,
            "scanner":{1:None,2:None,3:None},
            "scanner_age":{1:9999.0,2:9999.0,3:9999.0},
            "optimizer_df":None, "opt_age":9999.0,
            "pub_url":None, "last_refresh":0.0, "error":None,
            "history":{}, "history_ts":0.0,
            "bestx_state":{}, "_date_str":"",
            # Commodity
            "commodity_prices":{}, "commodity_age":9999.0,
            "commodity_levels":{}, "commodity_events":[],
            "comm_scanner":{1:None,2:None,3:None},
            "comm_scanner_age":{1:9999.0,2:9999.0,3:9999.0},
            "comm_history":{}, "comm_history_ts":0.0,
            # Crypto
            "crypto_prices":{}, "crypto_age":9999.0,
            "crypto_levels":{}, "crypto_events":[],
            "crypto_anchor_time":"",
            "crypto_scanner":{1:None,2:None,3:None},
            "crypto_scanner_age":{1:9999.0,2:9999.0,3:9999.0},
            "crypto_history":{}, "crypto_history_ts":0.0,
            # v10.6: persistent price caches (survive engine restarts)
            "_comm_price_cache":{},   # last known commodity prices
            "_crypto_price_cache":{}, # last known crypto prices
            "_equity_price_cache":{}, # last known equity prices
            # v10.6: alert state
            "_alert_equity_stale": False,
            "_alert_comm_stale":   False,
            "_alert_crypto_stale": False,
            "_alert_tunnel_down":  False,
            "_last_alert_ts":      {},  # keyed by alert type → monotonic time
        }
        self._stop = threading.Event()
        self._xlsx_ts   = 0.0
        self._scan_ts   = 0.0
        self._last_mtime = 0.0  # v10.7 PERF: skip live_prices.json read if unchanged
        self._bestx_mtime = 0.0
        self._url_mtime = 0.0
        # Separate threads for fast/slow refresh
        threading.Thread(target=self._fast_loop, daemon=True, name="DataFast").start()
        threading.Thread(target=self._slow_loop, daemon=True, name="DataSlow").start()
        # v10.5: AlertMonitor standalone process handles all alerting.
        # _alert_loop removed from DataStore to prevent duplicate/equity-bot-only alerts.

    def get(self, key, default=None):
        with self._lock:
            return self._d.get(key, default)

    def _set(self, **kw):
        with self._lock:
            self._d.update(kw)
            self._d["last_refresh"] = time.monotonic()

    def _fast_loop(self):
        """Fast tier: live prices + events every 0.5s — no XLSX, no history."""
        while not self._stop.is_set():
            try: self._refresh_fast()
            except Exception as e:
                with self._lock: self._d["error"] = str(e)
            self._stop.wait(self.DATA_INTERVAL)

    def _slow_loop(self):
        """Slow tier: XLSX levels + history every 5s."""
        while not self._stop.is_set():
            try: self._refresh_slow()
            except Exception as e:
                pass
            self._stop.wait(15.0)

    def _alert_loop(self):
        """v10.5: Deprecated — all alerting moved to alert_monitor.py (correct bot routing).
        This stub exists for backward compatibility but does nothing."""
        while not self._stop.is_set():
            self._stop.wait(60)

    def _refresh_fast(self):
        """Read only cheap JSON/JSONL files — no blocking XLSX reads."""
        ds = datetime.now(IST).strftime("%Y%m%d")
        with self._lock:
            if ds != self._d.get("_date_str",""):
                self._d["levels_ts"]=0; self._d["levels_mtime"]=0
                self._d["_date_str"]=ds
        upd: Dict[str,Any] = {}

        # live_prices.json — equity + commodity + crypto
        # v10.7 PERF: only re-read if file changed (mtime check avoids unnecessary I/O)
        p = os.path.join(LEVELS_DIR,"live_prices.json")
        if os.path.exists(p):
            try:
                _mtime = os.path.getmtime(p)
                if _mtime == self._last_mtime:
                    pass  # file unchanged — skip read, reuse cached state
                else:
                    self._last_mtime = _mtime
                    age = round(time.time()-os.path.getmtime(p),1)
                    d   = _rj(p) or {}
                    eq_prices = d.get("equity_prices") or d.get("prices") or {}
                    COMM_SET = {"GOLD","SILVER","NATURALGAS","CRUDE","COPPER"}
                    CRYPTO_SET = {"BTC","ETH","BNB","SOL","ADA"}
                    eq_prices = {k:v for k,v in eq_prices.items() if k not in COMM_SET and k not in CRYPTO_SET}
                    comm_px = d.get("commodity_prices") or {}
                    cryp_px = d.get("crypto_prices") or {}
                    file_age = age
                    try:
                        from datetime import datetime as _dt
                        _now = time.time()
                        comm_ts = d.get("commodity_ts","")
                        if comm_ts and comm_px:
                            _ct = _dt.strptime(comm_ts, "%Y-%m-%d %H:%M:%S")
                            _ct = pytz.timezone("Asia/Kolkata").localize(_ct)
                            comm_age = round(_now - _ct.timestamp(), 1)
                        else:
                            comm_age = file_age if comm_px else 9999.0
                        cryp_ts = d.get("crypto_ts","")
                        if cryp_ts and cryp_px:
                            _ct = _dt.strptime(cryp_ts, "%Y-%m-%d %H:%M:%S")
                            _ct = pytz.timezone("Asia/Kolkata").localize(_ct)
                            cryp_age = round(_now - _ct.timestamp(), 1)
                        else:
                            cryp_age = file_age if cryp_px else 9999.0
                    except Exception:
                        comm_age = file_age if comm_px else 9999.0
                        cryp_age = file_age if cryp_px else 9999.0

                    # v10.7: update persistent caches when we have fresh data
                    with self._lock:
                        if eq_prices:   self._d["_equity_price_cache"].update(eq_prices)
                        if comm_px:     self._d["_comm_price_cache"].update(comm_px)
                        if cryp_px:     self._d["_crypto_price_cache"].update(cryp_px)

                    # fall back to cache if engine is down
                    if not comm_px:
                        with self._lock:
                            comm_px = dict(self._d["_comm_price_cache"])
                        if comm_px: comm_age = 9990.0
                    if not cryp_px:
                        with self._lock:
                            cryp_px = dict(self._d["_crypto_price_cache"])
                        if cryp_px: cryp_age = 9990.0
                    if not eq_prices:
                        with self._lock:
                            eq_prices = dict(self._d["_equity_price_cache"])
                        if eq_prices: age = 9990.0

                    upd.update(
                        live_prices      = eq_prices,
                        price_age        = age,
                        commodity_prices = comm_px,
                        commodity_age    = comm_age,
                        crypto_prices    = cryp_px,
                        crypto_age       = cryp_age,
                    )
            except Exception:
                pass

        # equity + commodity + crypto trade events (JSONL — cheap)
        upd["events"]          = _read_jsonl(os.path.join(TRADE_DIR,f"trade_events_{ds}.jsonl"))
        upd["commodity_events"]= _read_jsonl(os.path.join(TRADE_DIR,f"commodity_trade_events_{ds}.jsonl"))
        import datetime as _dt
        upd["crypto_events"]   = _read_jsonl(os.path.join(TRADE_DIR,f"crypto_trade_events_{datetime.now(IST).strftime('%Y%m%d')}.jsonl"))

        # scanners (every SCAN_INTERVAL seconds — lightweight JSON)
        now_m = time.monotonic()
        if now_m - self._scan_ts >= self.SCAN_INTERVAL:
            self._scan_ts = now_m
            # Equity scanners: sweep_results/scanner{N}_*/{date}/live_state.json
            for sid in (1,2,3):
                found=False
                for d_ in (SCANNER_DIRS.get(sid,""), SCANNER_DIRS_LEGACY.get(sid,"")):
                    if not d_: continue
                    path = os.path.join(d_, ds, "live_state.json")
                    if os.path.exists(path):
                        try:
                            age=round(time.time()-os.path.getmtime(path),1); st=_rj(path)
                            with self._lock:
                                self._d["scanner"][sid]=st; self._d["scanner_age"][sid]=age
                        except Exception: pass
                        found=True; break
                if not found:
                    # Fallback: latest subdir (no date subfolder variant)
                    for d_ in (SCANNER_DIRS.get(sid,""), SCANNER_DIRS_LEGACY.get(sid,"")):
                        if not d_ or not os.path.isdir(d_): continue
                        subs=sorted([f for f in os.listdir(d_) if os.path.isdir(os.path.join(d_,f))],reverse=True)
                        for sub in subs[:2]:
                            path=os.path.join(d_,sub,"live_state.json")
                            if os.path.exists(path):
                                try:
                                    age=round(time.time()-os.path.getmtime(path),1); st=_rj(path)
                                    with self._lock:
                                        self._d["scanner"][sid]=st; self._d["scanner_age"][sid]=age
                                except Exception: pass
                                break
            # Commodity scanners: same date-based structure
            for sid in (1,2,3):
                base=COMM_SCANNER_DIRS.get(sid,"")
                if not base: continue
                p=os.path.join(base,ds,"live_state.json")
                if not os.path.exists(p):
                    # Try without date subdir
                    p=os.path.join(base,"live_state.json")
                if os.path.exists(p):
                    try:
                        age=round(time.time()-os.path.getmtime(p),1); st=_rj(p)
                        with self._lock:
                            self._d["comm_scanner"][sid]=st; self._d["comm_scanner_age"][sid]=age
                    except Exception: pass
            # Crypto scanners: use 6h window-named dirs (not date dirs)
            # Pattern: sweep_results/crypto_scanner{N}/{window_id}/live_state.json
            # Also try date dir for compatibility
            for sid in (1,2,3):
                base=CRYPTO_SCANNER_DIRS.get(sid,"")
                if not base: continue
                found=False
                # First try date-based subdir
                p=os.path.join(base,ds,"live_state.json")
                if os.path.exists(p):
                    try:
                        age=round(time.time()-os.path.getmtime(p),1); st=_rj(p)
                        with self._lock:
                            self._d["crypto_scanner"][sid]=st; self._d["crypto_scanner_age"][sid]=age
                        found=True
                    except Exception: pass
                if not found and os.path.isdir(base):
                    # Try newest window dir (sorted by mtime desc)
                    try:
                        subs=sorted(
                            [f for f in os.listdir(base) if os.path.isdir(os.path.join(base,f))],
                            key=lambda f: os.path.getmtime(os.path.join(base,f)),
                            reverse=True)
                        for win in subs[:5]:
                            path=os.path.join(base,win,"live_state.json")
                            if os.path.exists(path):
                                try:
                                    age=round(time.time()-os.path.getmtime(path),1); st=_rj(path)
                                    with self._lock:
                                        self._d["crypto_scanner"][sid]=st
                                        self._d["crypto_scanner_age"][sid]=age
                                    found=True
                                except Exception: pass
                                break
                    except Exception: pass

        # public URL (mtime-cached)
        _url_path = os.path.join(LEVELS_DIR, "dashboard_url.json")
        try:
            if os.path.exists(_url_path):
                _mt = os.path.getmtime(_url_path)
                if _mt != self._url_mtime:
                    self._url_mtime = _mt
                    d = _rj(_url_path)
                    if d:
                        upd["pub_url"] = d.get("public_url")
        except Exception:
            pass

        # BestX state (mtime-cached)
        try:
            if os.path.exists(BESTX_FILE):
                _mt = os.path.getmtime(BESTX_FILE)
                if _mt != self._bestx_mtime:
                    self._bestx_mtime = _mt
                    bs = _rj(BESTX_FILE)
                    if bs:
                        upd["bestx_state"] = bs
        except Exception:
            pass

        self._set(**upd)

    def _refresh_slow(self):
        """Slow tier: XLSX parsing + history loading — won't block fast callbacks."""
        ds = datetime.now(IST).strftime("%Y%m%d")
        upd: Dict[str,Any] = {}
        now_m = time.monotonic()

        # equity levels XLSX (30s cache) — v10.8: search ALL xlsx, not just today's date.
        # Algofinal writes files with NEXT trading day in name (e.g. 20260323 written on 20260322 night).
        # Removing the date-gate that was blocking loads when filename date ≠ current date.
        with self._lock:
            cached_ts=self._d.get("levels_ts",0); cached_mt=self._d.get("levels_mtime",0)
        if now_m - cached_ts > self.XLSX_INTERVAL:
            try:
                # Search ALL equity xlsx files across levels/ and prevday/ — pick most recent by mtime
                all_dirs = [LEVELS_DIR,
                            os.path.join(LEVELS_DIR,"initial_levels_prevday"),
                            os.path.join(LEVELS_DIR,"initial_levels_930"),
                            os.path.join(LEVELS_DIR,"initial_levels_eod")]
                cands = []
                for d_ in all_dirs:
                    if os.path.isdir(d_):
                        cands += [os.path.join(d_,f) for f in os.listdir(d_)
                                  if f.endswith(".xlsx") and "initial_levels" in f
                                  and "commodity" not in f.lower()]
                if cands:
                    cands.sort(key=os.path.getmtime, reverse=True)
                    lmt = os.path.getmtime(cands[0])
                    if lmt != cached_mt or now_m - cached_ts > self.XLSX_INTERVAL:
                        loaded = _load_levels_xlsx(ds)
                        if loaded:  # only update if we actually got data
                            upd["levels"] = loaded
                            upd["levels_ts"] = now_m; upd["levels_mtime"] = lmt
                        elif not self._d.get("levels"):  # first boot, take whatever we can
                            upd["levels"] = loaded
                            upd["levels_ts"] = now_m
                else:
                    # No xlsx at all — load_levels_xlsx will return [] gracefully
                    if now_m - cached_ts > self.XLSX_INTERVAL * 4:
                        upd["levels"] = _load_levels_xlsx(ds)
                        upd["levels_ts"] = now_m
            except Exception: pass

        # commodity levels JSON — v10.8: fall back to prevday xlsx if JSON missing
        try:
            cp=os.path.join(LEVELS_DIR,f"commodity_initial_levels_{ds}.json")
            if os.path.exists(cp):
                d=_rj(cp) or {}
                if d.get("levels"): upd["commodity_levels"]=d["levels"]
            if not upd.get("commodity_levels") and not self._d.get("commodity_levels"):
                # Try any commodity JSON (any date)
                try:
                    jsons=sorted([f for f in os.listdir(LEVELS_DIR)
                                  if f.startswith("commodity_initial_levels_") and f.endswith(".json")
                                  and "latest" not in f],
                                 key=lambda f: os.path.getmtime(os.path.join(LEVELS_DIR,f)),
                                 reverse=True)
                    for jf in jsons[:3]:
                        d=_rj(os.path.join(LEVELS_DIR,jf)) or {}
                        if d.get("levels"):
                            upd["commodity_levels"]=d["levels"]; break
                except Exception: pass
            if not upd.get("commodity_levels") and not self._d.get("commodity_levels"):
                # Fall back to prevday commodity xlsx
                upd["commodity_levels"] = _load_commodity_levels_xlsx()
        except Exception: pass

        # crypto levels JSON
        try:
            crp=os.path.join(LEVELS_DIR,"crypto_initial_levels_latest.json")
            if os.path.exists(crp):
                d=_rj(crp) or {}
                upd["crypto_levels"]=d.get("levels",{}); upd["crypto_anchor_time"]=d.get("anchor_time","")
        except Exception: pass

        # history loads (every 60s)
        with self._lock: hist_ts=self._d.get("history_ts",0.0)
        if now_m-hist_ts>60:
            try: upd["history"]=_load_hist_generic(TRADE_DIR,r"^trade_events_(\d{8})\.jsonl$"); upd["history_ts"]=now_m
            except Exception: pass
        with self._lock: ch_ts=self._d.get("comm_history_ts",0.0)
        if now_m-ch_ts>60:
            try: upd["comm_history"]=_load_hist_generic(TRADE_DIR,r"^commodity_trade_events_(\d{8})\.jsonl$","net_pnl","gross_pnl"); upd["comm_history_ts"]=now_m
            except Exception: pass
        with self._lock: cr_ts=self._d.get("crypto_history_ts",0.0)
        if now_m-cr_ts>60:
            try: upd["crypto_history"]=_load_hist_generic(TRADE_DIR,r"^crypto_trade_events_(\d{8})\.jsonl$","net_pnl_inr","gross_pnl_inr"); upd["crypto_history_ts"]=now_m
            except Exception: pass

        # optimizer
        with _OPT_LOCK:
            if _OPT_DF is not None:
                upd["optimizer_df"]=_OPT_DF.copy(); upd["opt_age"]=time.monotonic()-_OPT_TS
        if upd.get("optimizer_df") is None:
            odf,cage=_load_opt_csv(ds)
            if odf is not None: upd["optimizer_df"]=odf; upd["opt_age"]=cage

        if upd:
            self._set(**upd)

    # Legacy _loop/_refresh kept for compatibility
    def _loop(self):
        pass
    def _refresh(self):
        pass  # Replaced by _refresh_fast + _refresh_slow in v10.1


_DS=_DataStore()
_OPT_DF=None; _OPT_LOCK=threading.Lock(); _OPT_TS=0.0

# v10.6: lightweight Telegram alert (used by DataStore alert_loop)
def _tg_alert(text: str) -> None:
    def _go():
        for cid in TG_CHATS:
            try:
                data=urllib.parse.urlencode({"chat_id":cid,"text":text}).encode()
                urllib.request.urlopen(f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage",
                                       data=data, timeout=10)
            except Exception: pass
    threading.Thread(target=_go, daemon=True, name="TgAlert").start()

def register_optimizer(df):
    global _OPT_DF,_OPT_TS
    try:
        if df is not None and not df.empty:
            with _OPT_LOCK: _OPT_DF=df.copy(); _OPT_TS=time.monotonic()
    except Exception: pass


# ── File helpers ──────────────────────────────────────────────────────────────
def _rj(path):
    try:
        with open(path,encoding="utf-8") as f: return json.load(f)
    except Exception: return None

def _read_jsonl(path):
    rows=[]
    try:
        with open(path,encoding="utf-8") as f:
            for line in f:
                line=line.strip()
                if line:
                    try: rows.append(json.loads(line))
                    except Exception: pass
    except Exception: pass
    return rows

def _load_levels_xlsx(ds):
    """Load equity levels. v10.8: searches ALL xlsx files, sorted by mtime.
    Skips COMMODITY_SYMS and CRYPTO_SYMS rows. Logs which file it loaded from.
    """
    try:
        import openpyxl
        cands=[]
        # Search all directories for any initial_levels XLSX (exclude commodity-specific files)
        # v10.9: search adjusted_levels FIRST (written at 09:30 re-anchor)
        for dirp in (os.path.join(LEVELS_DIR,"adjusted_levels"),
                     LEVELS_DIR,
                     os.path.join(LEVELS_DIR,"initial_levels_930"),
                     os.path.join(LEVELS_DIR,"initial_levels_prevday"),
                     os.path.join(LEVELS_DIR,"initial_levels_eod")):
            if os.path.isdir(dirp):
                cands+=[os.path.join(dirp,f) for f in os.listdir(dirp)
                        if f.endswith(".xlsx") and "initial_levels" in f
                        and "commodity" not in f.lower()
                        and "crypto" not in f.lower()]
        if not cands: return []
        cands = list(set(cands))
        cands.sort(key=os.path.getmtime, reverse=True)
        # Try newest-first, but skip files that don't yield a sane equity symbol count.
        # This prevents accidentally parsing crypto_initial_levels_*.xlsx (can yield ~2 rows).
        for best in cands[:12]:
            try:
                log.info("_load_levels_xlsx: reading %s", os.path.basename(best))
                wb=openpyxl.load_workbook(best,read_only=True,data_only=True)
                ws=wb.active; rows=list(ws.iter_rows(values_only=True)); wb.close()
                result=[]; i=0
                while i<len(rows):
                    r=rows[i]
                    if r and r[0] is not None and isinstance(r[0],str) and len(r[0])>=2:
                        sym=str(r[0]).upper().strip()
                        if sym in COMMODITY_SYMS or sym in CRYPTO_SYMS:
                            i+=7; continue
                        br=rows[i+1] if i+1<len(rows) else (None,)*9
                        sr=rows[i+3] if i+3<len(rows) else (None,)*9
                        def _v(row,idx,default=None):
                            try: return row[idx]
                            except Exception: return default
                        result.append({
                            "SYMBOL":sym,"PREV CLOSE":_v(br,0),
                            "BUY ABOVE":_v(br,3),"T1":_v(br,4),"T2":_v(br,5),"T3":_v(br,6),
                            "SELL BELOW":_v(sr,3),"ST1":_v(sr,4),"ST2":_v(sr,5),"ST3":_v(sr,6),
                            "SL":_v(br,0),
                        })
                        i+=7
                    else: i+=1
                if len(result) >= 10:
                    log.info("_load_levels_xlsx: loaded %d equity symbols", len(result))
                    return result
                log.warning("_load_levels_xlsx: ignoring %s (only %d equity rows)", os.path.basename(best), len(result))
            except Exception:
                pass
        return []
    except Exception as e:
        log.debug("_load_levels_xlsx: %s",e); return []



def _startup_reanchor_levels(levels: list, live_prices: dict, is_commodity: bool = False) -> list:
    """
    v10.9 Startup Re-anchor Fix:
    When the market is open and current price is OUTSIDE the current buy_above/sell_below
    range (or between them but levels are from prev session), shift levels so that:
      - sell_below = current_price * (1 - X)
      - buy_above  = current_price * (1 + X)
      - T1..T3 and ST1..ST3 computed from new buy_above/sell_below
    Uses X = 0.008575 (equity default) or symbol-specific X for commodities.
    Only applies if the current price is more than 0.5X away from existing levels.
    """
    if not levels or not live_prices:
        return levels

    from config import cfg as _cfg
    EQUITY_X   = _cfg.CURRENT_X_MULTIPLIER   # 0.008575
    COMM_X     = _cfg.COMM_X                  # dict

    adjusted = []
    for row in levels:
        sym = str(row.get("SYMBOL","")).upper().strip()
        px  = float(live_prices.get(sym, 0) or 0)
        if px <= 0:
            adjusted.append(row); continue

        buy_above  = float(row.get("BUY ABOVE", 0) or 0)
        sell_below = float(row.get("SELL BELOW", 0) or 0)

        if buy_above <= 0 or sell_below <= 0:
            adjusted.append(row); continue

        # Determine X for this symbol
        if is_commodity:
            x = COMM_X.get(sym, EQUITY_X)
        else:
            from sweep_core import SPECIAL_SYMBOLS, INDEX_SYMBOLS
            if sym in INDEX_SYMBOLS:
                x = 0.00343
            elif sym in SPECIAL_SYMBOLS:
                x = EQUITY_X * 0.6
            else:
                x = EQUITY_X

        x_val = x * (buy_above if buy_above > 0 else px)

        # Check if current price is outside the trading range
        price_in_range = sell_below <= px <= buy_above
        price_near_edge = (px < sell_below * 1.02) or (px > buy_above * 0.98)

        if not price_in_range or price_near_edge:
            # Re-anchor: place current price in the center of the range
            new_buy   = round(px + x_val, 2)
            new_sell  = round(px - x_val, 2)
            step      = x_val  # step = x for regular, 0.6x for special

            new_row = dict(row)
            new_row["BUY ABOVE"]  = new_buy
            new_row["SELL BELOW"] = new_sell
            new_row["T1"]  = round(new_buy + step,   2)
            new_row["T2"]  = round(new_buy + step*2, 2)
            new_row["T3"]  = round(new_buy + step*3, 2)
            new_row["ST1"] = round(new_sell - step,   2)
            new_row["ST2"] = round(new_sell - step*2, 2)
            new_row["ST3"] = round(new_sell - step*3, 2)
            new_row["SL"]  = round(new_sell, 2)  # buy SL = sell_below
            adjusted.append(new_row)
        else:
            adjusted.append(row)
    return adjusted

def _load_commodity_levels_xlsx() -> dict:
    """v10.8: Load commodity levels from prevday xlsx when no JSON available.
    Returns dict keyed by symbol e.g. {"GOLD": {"buy_above":..., "sell_below":..., ...}}
    """
    try:
        import openpyxl
        cands = []
        prevday = os.path.join(LEVELS_DIR, "initial_levels_prevday")
        for dirp in (LEVELS_DIR, prevday):
            if os.path.isdir(dirp):
                cands += [os.path.join(dirp,f) for f in os.listdir(dirp)
                          if f.endswith(".xlsx") and "commodity" in f.lower()]
        if not cands:
            # Also try the main equity xlsx — commodity rows may be embedded
            for dirp in (LEVELS_DIR, prevday):
                if os.path.isdir(dirp):
                    cands += [os.path.join(dirp,f) for f in os.listdir(dirp)
                              if f.endswith(".xlsx") and "initial_levels" in f
                              and "equity" not in f.lower()]
        if not cands: return {}
        cands.sort(key=os.path.getmtime, reverse=True)
        wb = openpyxl.load_workbook(cands[0], read_only=True, data_only=True)
        ws = wb.active; rows = list(ws.iter_rows(values_only=True)); wb.close()
        result = {}; i = 0
        while i < len(rows):
            r = rows[i]
            if r and r[0] is not None and isinstance(r[0], str):
                sym = str(r[0]).upper().strip()
                if sym in COMMODITY_SYMS:
                    br = rows[i+1] if i+1 < len(rows) else (None,)*9
                    sr = rows[i+3] if i+3 < len(rows) else (None,)*9
                    def _v(row, idx, default=0):
                        try: return float(row[idx]) if row[idx] is not None else default
                        except Exception: return default
                    result[sym] = {
                        "buy_above":  _v(br,3), "T1": _v(br,4), "T2": _v(br,5), "T3": _v(br,6),
                        "sell_below": _v(sr,3), "ST1":_v(sr,4), "ST2":_v(sr,5), "ST3":_v(sr,6),
                        "x_mult": 0.0, "prev_close": _v(br,0)
                    }
                    i += 7; continue
            i += 1
        log.info("_load_commodity_levels_xlsx: loaded %d symbols from %s", len(result), cands[0])
        return result
    except Exception as e:
        log.debug("_load_commodity_levels_xlsx: %s", e); return {}

def _load_hist_generic(trade_dir,pattern,pnl_key_net="net",pnl_key_gross="gross",capital_per_trade=100000):
    import re as _re
    result={}
    if not os.path.isdir(trade_dir): return result
    for fname in os.listdir(trade_dir):
        m=_re.match(pattern,fname)
        if not m: continue
        ds=m.group(1); path=os.path.join(trade_dir,fname)
        trades=_read_jsonl(path)
        if not trades: continue
        nt=len(trades)
        net=sum(float(t.get(pnl_key_net,t.get("net",0)) or 0) for t in trades)
        gross=sum(float(t.get(pnl_key_gross,t.get("gross",0)) or 0) for t in trades)
        wins=sum(1 for t in trades if float(t.get(pnl_key_net,t.get("net",0)) or 0)>0)
        cap=nt*capital_per_trade
        result[ds]={"total_trades":nt,"gross_pnl":round(gross,2),"net_pnl":round(net,2),
                    "pct_return":round(net/cap*100,4) if cap>0 else 0.0,
                    "win_count":wins,"loss_count":nt-wins,
                    "win_rate":round(wins/nt*100,1) if nt else 0.0,
                    "capital_used":cap,"trades":trades}
    return result

def _load_opt_csv(ds):
    try:
        import pandas as pd
        d="x_optimizer_results"
        if not os.path.isdir(d): return None,9999.0
        live=os.path.join(d,f"xopt_live_{ds}.csv")
        if os.path.exists(live):
            age=time.time()-os.path.getmtime(live)
            if age<300: return pd.read_csv(live),round(age,1)
        fs=sorted([f for f in os.listdir(d) if f.startswith("xopt_ranked_") and f.endswith(".csv")],reverse=True)
        if fs:
            p=os.path.join(d,fs[0]); age=time.time()-os.path.getmtime(p)
            return pd.read_csv(p),round(age,1)
    except Exception: pass
    return None,9999.0

def get_lan_ip():
    try:
        s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM); s.connect(("8.8.8.8",80))
        ip=s.getsockname()[0]; s.close(); return ip
    except Exception: return "127.0.0.1"

def _age_label(age_s):
    a=float(age_s) if age_s is not None else 9999.0
    if a<2:    return "Live ●"
    if a<10:   return f"{a:.0f}s ago"
    if a<60:   return f"{a:.0f}s ago ⚠"
    if a<300:  return f"{int(a//60)}m {int(a%60)}s ago ⚠"
    if a<3600: return f"{int(a//60)}m ago — check feed"
    if a<86400:return f"{int(a//3600)}h ago — prev session"
    return "No data — offline"

_EXIT_PFX  = ("T1","T2","T3","T4","T5","ST1","ST2","ST3","ST4","ST5",
              "BUY_SL","SELL_SL","BUY_RETREAT","SELL_RETREAT","EOD","MANUAL")
_ENTRY_PFX = ("BUY_ENTRY","SELL_ENTRY","BUY_REENTRY","SELL_REENTRY")

def _compute_pnl(events):
    BROK=20.0; sym={}; tg=tn=0.0
    for ev in events:
        et=(ev.get("event_type") or "").upper()
        if any(et.startswith(p) for p in _ENTRY_PFX): continue
        if not any(et.startswith(p) for p in _EXIT_PFX): continue
        ep=float(ev.get("entry_price") or 0); px=float(ev.get("price") or 0)
        qty=float(ev.get("qty") or 0); sd=(ev.get("side") or "").upper()
        if ep<=0 or qty<=0: continue
        g=(px-ep)*qty if sd=="BUY" else (ep-px)*qty; n=g-BROK
        tg+=g; tn+=n; s=ev.get("symbol","?")
        if s not in sym: sym[s]={"symbol":s,"gross":0.0,"net":0.0,"trades":0,"wins":0}
        sym[s]["gross"]+=g; sym[s]["net"]+=n; sym[s]["trades"]+=1
        if n>0: sym[s]["wins"]+=1
    return round(tg,2),round(tn,2),list(sym.values())

def _closed_trades(events):
    BROK=20.0; opn={}; closed=[]
    for ev in sorted(events,key=lambda e:e.get("timestamp","")):
        sym=ev.get("symbol","?"); et=(ev.get("event_type") or "").upper()
        if any(et.startswith(p) for p in _ENTRY_PFX): opn[sym]=ev
        elif any(et.startswith(p) for p in _EXIT_PFX):
            ent=opn.pop(sym,None)
            ep=float(ev.get("entry_price") or (ent or {}).get("price") or 0)
            xp=float(ev.get("price") or 0); qty=float(ev.get("qty") or (ent or {}).get("qty") or 0)
            sd=(ev.get("side") or (ent or {}).get("side") or "").upper()
            if ep<=0 or xp<=0 or qty<=0: continue
            g=(xp-ep)*qty if sd=="BUY" else (ep-xp)*qty; n=g-BROK
            lbl=("Stop Loss" if "SL" in et else "Retreat" if "RETREAT" in et
                 else "EOD" if "EOD" in et else "Manual" if "MANUAL" in et
                 else et.split("_")[0] if "_" in et else et[:3])
            ts=ev.get("timestamp","")
            closed.append({"symbol":sym,"side":sd,"entry_price":ep,"exit_price":xp,
                           "qty":int(qty),"gross":round(g,2),"net":round(n,2),
                           "exit_type":lbl,"time":ts[11:19] if len(ts)>=19 else ts})
    return list(reversed(closed))

def _open_positions(events,prices):
    pos={}
    for ev in sorted(events,key=lambda e:e.get("timestamp","")):
        sym=ev.get("symbol","?"); et=(ev.get("event_type") or "").upper()
        if any(et.startswith(p) for p in _ENTRY_PFX): pos[sym]=ev
        elif any(et.startswith(p) for p in _EXIT_PFX): pos.pop(sym,None)
    rows=[]
    for sym,ev in pos.items():
        ep=float(ev.get("price") or 0); qty=float(ev.get("qty") or 0)
        sd=(ev.get("side") or "").upper(); lp=prices.get(sym.upper(),0)
        upnl=((lp-ep)*qty if sd=="BUY" else (ep-lp)*qty) if lp and ep else 0
        rows.append({"Symbol":sym,"Side":sd,"Entry":f"₹{ep:,.2f}",
                     "Live":f"₹{lp:,.2f}" if lp else "—","Qty":str(int(qty)),
                     "uPnL":f"₹{upnl:+,.2f}"})
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  AI AGENT  — Google Gemini API (primary) → Anthropic API (fallback) → Offline KB
# ══════════════════════════════════════════════════════════════════════════════
ALGOSTACK_KB = """
=== AlgoStack v10.7 — Complete Knowledge Base ===

WHAT IS ALGOSTACK?
AlgoStack is a fully automated algorithmic trading system by Ridhaant Ajoy Thackur.
It trades simultaneously across 3 markets:
  - NSE Equity (38 stocks + Nifty/BankNifty)  — session 09:30–15:11 IST
  - MCX Commodity (Gold/Silver/Crude/NatGas/Copper) — session 09:00–23:30 IST
  - Binance Crypto (BTC/ETH/BNB/SOL/ADA) — 24/7, re-anchors every 6h
Target: 0.30% daily ROI on deployed capital (₹1 lakh per trade).

THE X MULTIPLIER (CORE CONCEPT):
X is the single number that controls how far from yesterday's closing price
the system waits before entering a trade.

  buy_above  = prev_close + (prev_close × X)   → long entry trigger
  sell_below = prev_close − (prev_close × X)   → short entry trigger
  T1..T5     = progressive profit targets above buy_above
  ST1..ST5   = progressive profit targets below sell_below
  Stop Loss  = back at prev_close (= X × prev_close rupees risk)

Current equity X = 0.008575 (0.8575% of prev_close)
Why this number? It represents ~₹85 distance on a ₹10,000 stock —
enough buffer to filter out random noise but close enough for intraday moves.

HOW THE 3 SCANNERS WORK:
The scanners test thousands of X values simultaneously on live data
to find which X would have been most profitable TODAY.

  Scanner 1 (Narrow):  1,000 X values  — range 0.0080–0.0090  — fine-tune near live X
  Scanner 2 (Dual):   13,000 X values  — range 0.001–0.007 + 0.009–0.016
  Scanner 3 (WideDual):31,000 X values — range 0.001–0.032 — extreme range
  TOTAL: 45,000 equity variations per trading day

For commodity: 97,500 variations/day across 3 scanners
For crypto: 97,500 variations/6h window across 3 scanners
Grand total: 244,000+ variations tested every day.

WHEN TO CHANGE LIVE X:
  < 5% difference from live X = near-optimal, no change needed
  5–20% = consider testing in paper mode for 3–5 days first
  > 20% = significant divergence — investigate if market regime changed

EQUITY TRADING LOGIC:
  - Pre-market 09:00–09:30: system uses prev_close to set levels
  - 09:30 re-anchor: recalculates ALL levels using the actual 09:30 open price
  - 09:30–09:35 blackout: no new entries for 5 minutes (avoids gap-fill whipsaws)
  - Entry: price crosses buy_above (long) or sell_below (short)
  - Retreat exit: if price reverses 25% back toward entry from best point → auto-exit
    (This guarantees a small profit even if the trade reverses)
  - Re-entry: if price touches the level again after a retreat exit → new entry
  - EOD square-off: all positions closed at 15:11 IST

COMMODITY TRADING (MCX):
  - 5 symbols: GOLD, SILVER, CRUDE, NATURALGAS, COPPER
  - Each has its own calibrated X multiplier (one size does NOT fit all):
    GOLD X ≈ 0.003430, SILVER X ≈ 0.005145, NATURALGAS X ≈ 0.000857,
    CRUDE X ≈ 0.000602, COPPER X ≈ 0.004000
  - Session 09:00–23:30 IST (much longer than equity!)
  - Check the Commodity page in the evening — it's still trading at 9pm

CRYPTO TRADING (Binance):
  - 5 coins: BTC, ETH, BNB, SOL, ADA
  - 24/7 trading — never stops
  - 6-hour re-anchor: every 6h recalculates anchor prices + buy/sell levels
  - Uses same X logic but in USD (0.008575 = 0.8575% from anchor price)
  - P&L displayed in ₹ (USDT P&L × 84 exchange rate)
  - CryptoScanners were crashing (exit code=1) due to a bug in sweep_core.py
    that has been fixed in v10.7 — PriceSubscriber now accepts the 'topic' parameter

OPTIMIZER PAGE:
  Combines rankings from all 3 scanners → single best X leaderboard
  Red dashed line = current live X for comparison
  "vs Live X %" column shows divergence
  Use this to decide if you should change your live X

BEST-X TRADER:
  Runs paper trades simultaneously using TODAY's best X from the optimizer
  NOT real money — used to validate if the best X actually works live
  Compare Best-X Trader P&L vs Equity page P&L to see if a change is beneficial

HISTORY PAGE:
  Daily P&L for every past trading day
  Progress bar = how close to 0.30% target
  Green bar = hit target, Red = below target
  Export to CSV for Excel analysis

PERFORMANCE PAGE:
  Bar chart of daily returns vs 0.30% target (last 21 trading days)
  Monthly projection = avg% × 21 days × avg_trades × ₹1,00,000
  Win rate, trade count, drawdown analysis

INTEL PAGE:
  India news (ET Markets, Moneycontrol, Business Standard)
  Global news (Reuters, CNBC)
  Market signals: Nifty RSI, 20DMA/50DMA, trend indicator
  FII/DII daily flows from NSE
  LIVE crypto prices (BTC/ETH/BNB/SOL/ADA) — NEW in v10.0
  LIVE MCX commodity prices (GOLD/SILVER/CRUDE/NATGAS/COPPER) — NEW in v10.0

SYSTEM PAGE:
  All 15 processes: 1 Algofinal + 3 equity scanners + 1 XOptimizer + 1 BestX =7 equity
                   + 1 CommodityEngine + 3 CommScanners = 4 commodity
                   + 1 CryptoEngine + 3 CryptoScanners = 4 crypto
  Live/LAN access links
  Internet speed + ping
  Price feed age (equity/commodity/crypto)

WATCHDOG (autohealer.py):
  Monitors all 15 processes. Auto-restarts any that exit unexpectedly.
  On weekends: CommScanners exit code=0 (MCX closed) — THIS IS NORMAL, not a bug.
  CryptoScanners: were crashing (exit code=1) due to PriceSubscriber bug — now FIXED.
  Restart loop detection: alerts Telegram if process restarts >5 times in an hour.

TELEGRAM ALERTS:
  3 separate bots:
  - Equity bot (7587...): equity trade alerts + daily level Excel + startup URL
  - Commodity bot (8340...): MCX trade alerts + startup URL
  - Crypto bot (8710...): Binance trade alerts + startup URL
  ALL 3 bots now receive the UNIFIED dashboard URL (port 8055) at startup.
  (Old bug: Algofinal was sending separate 8050 URL — FIXED in v10.0)

BROKERAGE: ₹20 per round-trip (₹10 entry + ₹10 exit)
RETURN %: net_pnl / (num_trades × ₹1,00,000) × 100
EXAMPLE: 3 trades, net ₹350 → 350/300000 × 100 = 0.117%

ZMQ PRICE FEED (internal):
  Algofinal publishes live prices on tcp://127.0.0.1:28081
  Topics: "prices" (equity), "commodity" (MCX), "crypto" (Binance)
  Price latency: 1–5s from exchange to all scanners
  CryptoScanner crash fix: sweep_core.PriceSubscriber now accepts topic=b"crypto"

COMMON QUESTIONS:
  Q: CommScanners show exit code=0 every weekend — is this a bug?
  A: No! MCX is closed on weekends. Code=0 means clean exit. They restart Monday 09:00.

  Q: CryptoScanners show exit code=1 — crash?
  A: Was a bug in v9.x — PriceSubscriber didn't accept the 'topic' parameter.
     FIXED in v10.7 by adding topic:bytes=None to sweep_core.py

  Q: Equity page was showing GOLD and SILVER prices
  A: Fixed in v10.7 — levels table now filtered to EQUITY_STOCKS set only.

  Q: Optimizer page was empty / not working
  A: Fixed in v10.7 — stacked decorator bug removed.

  Q: Old dashboard URL (8050) in Telegram
  A: Fixed in v10.7 — all 3 bots receive unified dash URL (8055).

TIPS FOR BEGINNERS:
  1. During market hours (09:30–15:11) watch the equity page for live trades
  2. Check scanner pages to see X testing in real-time
  3. After market close: check History page for daily P&L
  4. After 5 days of scanner data: check Optimizer to see if X should be updated
  5. Commodity runs until 23:30 — check it in evenings
  6. Crypto is 24/7 — re-anchors at 00:00, 06:00, 12:00, 18:00 UTC
  7. Performance page: if 0.30% target hit 15+ of 21 days, system is healthy
  8. If Best-X Trader consistently outperforms live X for 5+ days, consider updating X
"""

_AI_HIST: List[dict] = []
_AI_LOCK = threading.Lock()
_GEMINI_LAST_ERROR: List[str] = [""]  # mutable container for last error

def _ai_respond(question: str) -> str:
    global _AI_HIST
    with _AI_LOCK:
        _AI_HIST.append({"role":"user","content":question})
        history = list(_AI_HIST[-12:])

    SYSTEM_PROMPT = (
        "You are the AlgoStack AI assistant — an expert on this algorithmic trading system "
        "built by Ridhaant Ajoy Thackur. Help traders understand the X multiplier, scanners, "
        "equity/commodity/crypto trading logic, and dashboard usage. Be concise and practical. "
        "Use Indian trading terminology (₹, NSE, MCX, Nifty, etc.).\n\n"
        + ALGOSTACK_KB
    )

    # ── 1. Google Gemini API (primary) ──────────────────────────────────────
    if GEMINI_API_KEY:
        try:
            # Build contents array: system turn + conversation history
            contents = []
            # Gemini doesn't have a system role — prepend KB as first user/model pair
            contents.append({"role":"user",  "parts":[{"text": SYSTEM_PROMPT}]})
            contents.append({"role":"model", "parts":[{"text":"Understood. I'm the AlgoStack AI assistant. How can I help?"}]})
            for msg in history[:-1]:  # all turns except the new question
                role = "user" if msg["role"] == "user" else "model"
                contents.append({"role": role, "parts": [{"text": msg["content"]}]})
            # Add current question
            contents.append({"role":"user","parts":[{"text":question}]})

            payload = json.dumps({
                "contents": contents,
                "generationConfig": {
                    "temperature": 0.7,
                    "maxOutputTokens": 900,
                    "topP": 0.9,
                },
                "safetySettings": [
                    {"category":"HARM_CATEGORY_HARASSMENT","threshold":"BLOCK_NONE"},
                    {"category":"HARM_CATEGORY_HATE_SPEECH","threshold":"BLOCK_NONE"},
                    {"category":"HARM_CATEGORY_SEXUALLY_EXPLICIT","threshold":"BLOCK_NONE"},
                    {"category":"HARM_CATEGORY_DANGEROUS_CONTENT","threshold":"BLOCK_NONE"},
                ],
            }).encode()
            url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
                   f"gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}")
            req = urllib.request.Request(url, data=payload,
                                         headers={"Content-Type":"application/json"},
                                         method="POST")
            with urllib.request.urlopen(req, timeout=20) as resp:
                data = json.loads(resp.read())
                answer = data["candidates"][0]["content"]["parts"][0]["text"]
                with _AI_LOCK: _AI_HIST.append({"role":"assistant","content":answer})
                return answer
        except Exception as e:
            log.warning("Gemini API error: %s", e)
            # Store error for display in AI page
            _GEMINI_LAST_ERROR[0] = str(e)[:120]

    # ── 2. Anthropic API (fallback if ANTHROPIC_API_KEY set in .env) ────────
    if ANTHROPIC_API_KEY:
        try:
            payload = json.dumps({
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 900,
                "system": SYSTEM_PROMPT,
                "messages": history,
            }).encode()
            req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages", data=payload,
                headers={"Content-Type":"application/json","x-api-key":ANTHROPIC_API_KEY,
                         "anthropic-version":"2023-06-01"}, method="POST")
            with urllib.request.urlopen(req, timeout=18) as resp:
                data = json.loads(resp.read())
                answer = data["content"][0]["text"]
                with _AI_LOCK: _AI_HIST.append({"role":"assistant","content":answer})
                return answer
        except Exception as e:
            log.debug("Anthropic API fallback: %s", e)

    # Rich offline fallback — keyword matching against knowledge base
    q = question.lower()
    KB = [
        (["what is x","x multiplier","how does x work","explain x","what is the x"],
         "X is the entry distance parameter. buy_above = prev_close + (prev_close × X). "
         "Current equity X = 0.008575 (0.8575%). Example: RELIANCE at ₹2,800 → buy above ₹2,824, "
         "sell below ₹2,776. Scanners test 45,000 X values daily to find the most profitable one."),
        (["scanner","what do scanners do","s1 s2 s3","how scanner works"],
         "3 scanners test different X ranges on live data:\n"
         "• S1 Narrow: 1,000 variations (0.0080–0.0090) — fine-tune near live X\n"
         "• S2 Dual: 13,000 variations (0.001–0.007 + 0.009–0.016)\n"
         "• S3 Wide: 31,000 variations (0.001–0.032)\n"
         "Total: 45,000 equity + 97,500 commodity + 97,500 crypto = 244,000/day."),
        (["commodity","mcx","gold","silver","crude","naturalgas","copper","why comm"],
         "Commodity trades GOLD/SILVER/CRUDE/NATURALGAS/COPPER on MCX.\n"
         "Session: 09:00–23:30 IST (much longer than equity!).\n"
         "Each commodity has its own X: GOLD≈0.003430, SILVER≈0.005145, NatGas≈0.000857.\n"
         "CommScanners exit code=0 on weekends = NORMAL (MCX closed). They restart Monday 09:00."),
        (["crypto","bitcoin","btc","eth","binance","cryptoscanner","exit code 1"],
         "Crypto trades BTC/ETH/BNB/SOL/ADA on Binance 24/7.\n"
         "Re-anchors every 6h (00:00, 06:00, 12:00, 18:00 UTC).\n"
         "CryptoScanner crash (exit code=1) was a bug — FIXED in v10.0.\n"
         "Root cause: sweep_core.PriceSubscriber didn't accept 'topic' parameter.\n"
         "P&L shown in ₹ (USDT × 84 exchange rate)."),
        (["history","performance","daily pnl","track record","how many days"],
         "History page: daily P&L for every past trading day.\n"
         "Progress bar shows % toward 0.30% target. Green=hit target, Red=below.\n"
         "Performance page: bar chart comparing 21 days with monthly projection.\n"
         "Monthly projection = avg% × 21 × avg_trades × ₹1,00,000."),
        (["optimizer","best x","which x is best","should i change x","what x to use"],
         "Optimizer combines all 3 scanners → single best X leaderboard.\n"
         "< 5% from live X = no change needed.\n"
         "5–20% = test in paper mode for 3–5 days.\n"
         "> 20% = significant divergence, check if market regime changed.\n"
         "Best-X Trader runs paper trades to validate before you change live X."),
        (["retreat","exit","stop loss","target","t1","t2"],
         "Retreat exit: if price reverses 25% back from entry → auto-exit (small profit guaranteed).\n"
         "T1–T5: progressive profit targets after entry.\n"
         "Stop loss = prev_close (= X × prev_close risk in ₹).\n"
         "Re-entry: if price re-touches level after retreat → new entry triggered."),
        (["telegram","alert","notification","url","8050","8055"],
         "3 bots send alerts: equity bot, commodity bot, crypto bot.\n"
         "ALL 3 now receive the unified dashboard URL (port 8055) at startup.\n"
         "Old bug: Algofinal was sending a separate 8050 URL — FIXED in v10.0.\n"
         "Trade alerts sent immediately on entry/exit. No spam."),
        (["weekend","saturday","sunday","market closed","closed"],
         "NSE equity: closed Saturday & Sunday.\n"
         "MCX commodity: closed Saturday & Sunday. CommScanners exit code=0 = NORMAL.\n"
         "Binance crypto: 24/7, never closes. CryptoScanners run all weekend.\n"
         "Dashboard (port 8055): always accessible, shows last session data."),
        (["brokerage","commission","20 rupees","charges","cost"],
         "Brokerage: ₹20 per round-trip (₹10 entry + ₹10 exit).\n"
         "Return% = net_pnl / (num_trades × ₹1,00,000) × 100.\n"
         "Example: 3 trades, gross ₹370 → net ₹350 (₹20 brok) → 0.117% return."),
        (["watchdog","autohealer","restart","process","15 processes"],
         "autohealer.py monitors all 15 processes:\n"
         "7 equity (Algofinal + 3 scanners + XOptimizer + BestX + NewsDash)\n"
         "4 commodity (CommodityEngine + 3 CommScanners)\n"
         "4 crypto (CryptoEngine + 3 CryptoScanners)\n"
         "Auto-restarts any crash. Alerts Telegram if restarts >5× in 1 hour."),
        (["intel","news","fii","dii","nifty rsi","signals"],
         "Intel page shows:\n"
         "• India news (ET Markets, Moneycontrol, Business Standard)\n"
         "• Global news (Reuters, CNBC)\n"
         "• Nifty RSI, 20DMA/50DMA trend signals\n"
         "• FII/DII daily flows from NSE\n"
         "• LIVE crypto prices (BTC/ETH/BNB/SOL/ADA)\n"
         "• LIVE MCX commodity prices — both NEW in v10.7"),
        (["system page","all processes","status","port","lan"],
         "System page shows all 15 processes with live/stale/offline status,\n"
         "public ngrok tunnel URL, LAN IP URL, internet speed, and price feed ages.\n"
         "Green dot = live, Yellow = stale, Red = offline.\n"
         "244K variations/day: 45K equity + 97.5K commodity + 97.5K crypto."),
        (["getting started","beginner","how to use","explain","what should i watch"],
         "Getting started with AlgoStack:\n"
         "1. 09:30–15:11 IST: watch Equity page for live trades\n"
         "2. Watch scanner pages to see X testing in real-time\n"
         "3. After close: check History for daily P&L\n"
         "4. Evenings: check Commodity page (MCX runs until 23:30)\n"
         "5. Crypto page: runs 24/7, check anytime\n"
         "6. After 5 trading days: check Optimizer for X recommendations\n"
         "7. Add ANTHROPIC_API_KEY to .env for full AI chat capabilities"),
    ]
    for keywords, answer in KB:
        if any(k in q for k in keywords):
            with _AI_LOCK: _AI_HIST.append({"role":"assistant","content":answer})
            return answer

    fallback = ("I'm the AlgoStack AI assistant. I can explain:\n"
                "• The X multiplier and how entry levels work\n"
                "• Scanner 1/2/3 purpose and X ranges\n"
                "• Equity / Commodity / Crypto trading logic\n"
                "• History, Performance, Optimizer pages\n"
                "• Telegram alerts and system processes\n"
                "• Why CommScanners exit on weekends (it's normal!)\n\n"
                "💡 Gemini API active — if you see this, check logs for API errors.")
    with _AI_LOCK: _AI_HIST.append({"role":"assistant","content":fallback})
    return fallback


# ══════════════════════════════════════════════════════════════════════════════
#  DASH APP
# ══════════════════════════════════════════════════════════════════════════════
def build_app():
    from dash import Dash, dcc, html, Input, Output, State
    from dash import dash_table
    import plotly.graph_objects as go

    # v10.9: Dash() constructor — update_title removed in Dash 2.18+; use try/except
    _dash_kwargs = dict(
        suppress_callback_exceptions=True,
        title="AlgoStack v10.8",
        meta_tags=[{"name":"viewport","content":"width=device-width,initial-scale=1"}]
    )
    # update_title removed in Dash 2.18 — only add if supported
    try:
        import dash as _dash_ver
        _dv = tuple(int(x) for x in _dash_ver.__version__.split(".")[:2])
        if _dv < (2, 18):
            _dash_kwargs["update_title"] = None
    except Exception:
        pass  # skip update_title if version check fails
    try:
        app = Dash(__name__, **_dash_kwargs)
    except TypeError as _te:
        # Older/newer Dash may not accept some kwargs — retry with minimal args
        log.warning("Dash() failed with full kwargs (%s) — retrying minimal", _te)
        app = Dash(__name__, suppress_callback_exceptions=True,
                   title="AlgoStack v10.9",
                   meta_tags=[{"name":"viewport","content":"width=device-width,initial-scale=1"}])

    CSS = f"""
*{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%;overflow-x:hidden;-webkit-tap-highlight-color:transparent}}
body{{background:{BG};color:{TEXT};font-family:{FONT};font-size:15px;line-height:1.5}}
a{{color:{ACCENT};text-decoration:none}}
input,button{{outline:none}}
::-webkit-scrollbar{{width:5px;height:5px}}
::-webkit-scrollbar-track{{background:{SB}}}
::-webkit-scrollbar-thumb{{background:#30363d;border-radius:3px}}
::-webkit-scrollbar-thumb:hover{{background:#484f58}}
@keyframes marquee{{from{{transform:translateX(0)}} to{{transform:translateX(-50%)}}}}
@keyframes pulse{{0%,100%{{opacity:1}} 50%{{opacity:.3}}}}
@keyframes pulseGreen{{0%,100%{{box-shadow:0 0 0 0 {GREEN}44}} 70%{{box-shadow:0 0 0 6px {GREEN}00}}}}
@keyframes fadeIn{{from{{opacity:0;transform:translateY(6px)}} to{{opacity:1;transform:translateY(0)}}}}
@keyframes slideIn{{from{{opacity:0;transform:translateX(-10px)}} to{{opacity:1;transform:translateX(0)}}}}
@keyframes shimmer{{0%{{background-position:-200% 0}} 100%{{background-position:200% 0}}}}
@keyframes spin{{to{{transform:rotate(360deg)}}}}
@keyframes flashRed{{0%,100%{{background:transparent}} 50%{{background:{RED}18}}}}
.page{{display:none}}.page.active{{display:block;animation:fadeIn .18s ease}}
.card{{background:{CARD};border:1px solid {BORDER};border-radius:12px;padding:16px;margin-bottom:14px;transition:border-color .2s,box-shadow .2s}}
.card:hover{{border-color:#30363d;box-shadow:0 2px 12px rgba(0,0,0,.35)}}
.sg{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px}}
.sc{{background:{CARD};border:1px solid {BORDER};border-radius:12px;padding:14px 16px;transition:transform .15s,border-color .15s,box-shadow .15s}}
.sc:hover{{border-color:#30363d;transform:translateY(-2px);box-shadow:0 4px 16px rgba(0,0,0,.3)}}
.sl{{font-size:10px;color:{DIM};text-transform:uppercase;letter-spacing:.08em;font-weight:600}}
.sv{{font-size:22px;font-weight:700;margin-top:4px;line-height:1.2}}
.ss{{font-size:11px;color:{DIM};margin-top:3px}}
.sec{{font-size:11px;color:{DIM};text-transform:uppercase;letter-spacing:.1em;margin-bottom:10px;padding-bottom:8px;border-bottom:1px solid {BORDER};font-weight:700;display:flex;align-items:center;gap:6px}}
.top-bar{{position:sticky;top:0;z-index:100;background:rgba(13,17,23,.96);border-bottom:1px solid {BORDER};
           padding:9px 16px;display:flex;align-items:center;justify-content:space-between;
           backdrop-filter:blur(16px);-webkit-backdrop-filter:blur(16px)}}
.top-title{{font-size:16px;font-weight:700;color:{ACCENT};flex-shrink:0;min-width:90px;letter-spacing:-.01em}}
.top-time{{font-size:11px;color:{DIM};text-align:right;flex-shrink:0;min-width:90px;line-height:1.4}}
.ticker-wrap{{flex:1;overflow:hidden;margin:0 12px;height:20px;position:relative}}
.subnav{{position:fixed;bottom:58px;left:0;right:0;z-index:199;
         background:rgba(13,17,23,0.98);border-top:1px solid {BORDER};
         display:none;flex-direction:row;overflow-x:auto;height:44px;
         -webkit-overflow-scrolling:touch;scrollbar-width:none;backdrop-filter:blur(8px)}}
.subnav::-webkit-scrollbar{{display:none}}
.subnav-link{{flex:0 0 auto;display:flex;flex-direction:column;align-items:center;
             justify-content:center;padding:2px 12px;min-width:52px;text-decoration:none;
             color:{DIM};font-size:9px;font-weight:600;white-space:nowrap;gap:1px;
             border-bottom:2px solid transparent;transition:all .12s}}
.subnav-link:hover{{color:{TEXT}}}
.subnav-link.on{{color:{ACCENT};border-bottom-color:{ACCENT}}}
.subnav-ic{{font-size:14px;line-height:1}}
.tab-bar{{position:fixed;bottom:0;left:0;right:0;z-index:200;background:rgba(13,17,23,.98);
           border-top:1px solid {BORDER};display:flex;height:58px;overflow-x:auto;
           -webkit-overflow-scrolling:touch;backdrop-filter:blur(12px)}}
.tab-btn{{flex:0 0 auto;display:flex;flex-direction:column;align-items:center;
           justify-content:center;cursor:pointer;border:none;background:transparent;
           color:{DIM};font-size:9px;font-family:inherit;padding:4px 11px;min-height:44px;
           min-width:58px;text-decoration:none;white-space:nowrap;transition:color .15s,background .15s;
           border-top:2px solid transparent}}
.tab-btn:hover{{color:{ACCENT};background:rgba(88,166,255,.06)}}
.tab-btn.on{{color:{ACCENT};background:rgba(88,166,255,.1);border-top-color:{ACCENT}}}
.tab-icon{{font-size:17px;line-height:1;margin-bottom:2px}}
.main{{padding:14px 14px 112px;max-width:1240px;margin:0 auto}}
.pg-title{{font-size:19px;font-weight:700;margin-bottom:4px;letter-spacing:-.02em}}
.pg-sub{{color:{DIM};font-size:13px;margin-bottom:14px}}
.sidebar{{display:none;position:fixed;left:0;top:0;bottom:0;width:232px;
           flex-direction:column;background:rgba(13,17,23,.97);border-right:1px solid {BORDER};
           overflow-y:auto;z-index:150;backdrop-filter:blur(12px)}}
.sb-brand{{padding:18px 16px 14px;border-bottom:1px solid {BORDER}}}
.sb-title{{font-size:15px;font-weight:700;color:{ACCENT};letter-spacing:-.01em}}
.sb-sub{{font-size:10px;color:{DIM};margin-top:3px}}
.sb-sec{{padding:8px 16px 4px;font-size:10px;font-weight:700;letter-spacing:.1em;
          text-transform:uppercase;color:{DIM};border-top:1px solid {BORDER};margin-top:6px}}
.sb-lnk{{display:flex;align-items:center;gap:10px;padding:8px 16px;color:{DIM};
          font-size:12px;border-left:3px solid transparent;transition:all .12s;text-decoration:none}}
.sb-lnk:hover,.sb-lnk.on{{background:rgba(255,255,255,.04)}}
.sb-lnk.eq:hover,.sb-lnk.eq.on{{border-left-color:{EQ_COL};color:{EQ_COL}}}
.sb-lnk.cm:hover,.sb-lnk.cm.on{{border-left-color:{COMM_COL};color:{COMM_COL}}}
.sb-lnk.cr:hover,.sb-lnk.cr.on{{border-left-color:{CRYPTO_COL};color:{CRYPTO_COL}}}
.sb-lnk.ut:hover,.sb-lnk.ut.on{{border-left-color:{GREEN};color:{GREEN}}}
.sb-ic{{width:18px;text-align:center;font-size:13px}}
.pos-row{{padding:10px 0;border-bottom:1px solid {BORDER}}}
.pos-sym{{font-weight:700;font-size:14px}}
.pos-meta{{color:{DIM};font-size:12px;margin-top:2px}}
.pup{{color:{GREEN};font-size:13px;font-weight:600}}
.pdn{{color:{RED};font-size:13px;font-weight:600}}
.err{{background:#2d1519;border:1px solid {RED};border-radius:8px;padding:10px 14px;
       margin-bottom:12px;font-size:12px;color:{RED}}}
.warn-banner{{background:#2d2205;border:1px solid {AMBER};border-radius:8px;padding:10px 14px;
              margin-bottom:12px;font-size:12px;color:{AMBER};display:flex;align-items:center;gap:8px}}
.info-banner{{background:#0d1f33;border:1px solid {ACCENT};border-radius:8px;padding:10px 14px;
              margin-bottom:12px;font-size:12px;color:{ACCENT};display:flex;align-items:center;gap:8px}}
.tw{{overflow-x:auto;-webkit-overflow-scrolling:touch;border-radius:8px}}
table{{border-collapse:collapse;width:100%}}
th{{background:{SB};color:{DIM};font-size:11px;font-weight:700;text-align:left;
    padding:8px 10px;border-bottom:2px solid {BORDER};white-space:nowrap}}
td{{padding:7px 10px;border-bottom:1px solid {BORDER};font-size:13px;vertical-align:middle}}
tr:hover td{{background:rgba(255,255,255,.025)}}
.ai-chat{{display:flex;flex-direction:column;min-height:320px;max-height:520px;
           overflow-y:auto;padding:12px;background:{BG};border:1px solid {BORDER};
           border-radius:8px;margin-bottom:12px;gap:6px;scroll-behavior:smooth}}
.ai-u{{background:#1a2c40;border-radius:12px 12px 4px 12px;padding:10px 14px;
        font-size:13px;max-width:82%;align-self:flex-end;line-height:1.55;white-space:pre-wrap;
        animation:slideIn .12s ease}}
.ai-b{{background:{CARD};border:1px solid {BORDER};border-radius:12px 12px 12px 4px;
        padding:10px 14px;font-size:13px;max-width:82%;align-self:flex-start;
        line-height:1.65;white-space:pre-wrap}}
.live-dot{{display:inline-block;width:7px;height:7px;border-radius:50%;
            background:{GREEN};animation:pulse 1.8s infinite;margin-right:5px;vertical-align:middle}}
.live-ring{{display:inline-block;width:9px;height:9px;border-radius:50%;
             background:{GREEN};animation:pulseGreen 2s infinite;margin-right:5px;vertical-align:middle}}
.stale-dot{{display:inline-block;width:7px;height:7px;border-radius:50%;
             background:{YELLOW};animation:pulse 2.5s infinite;margin-right:5px;vertical-align:middle}}
.dead-dot{{display:inline-block;width:7px;height:7px;border-radius:50%;
            background:{RED};margin-right:5px;vertical-align:middle}}
.price-chip{{display:inline-flex;align-items:center;gap:5px;background:{CARD};
              border:1px solid {BORDER};border-radius:20px;padding:3px 10px;
              font-size:12px;font-weight:600;margin:2px 2px 2px 0;white-space:nowrap;
              transition:border-color .2s}}
.price-chip.up{{border-color:{GREEN}66;color:{GREEN}}}
.price-chip.dn{{border-color:{RED}66;color:{RED}}}
.price-chip.live{{border-color:{GREEN}44;animation:pulseGreen 3s infinite}}
.badge{{display:inline-block;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:700}}
.badge-green{{background:{GREEN}22;color:{GREEN}}}
.badge-red{{background:{RED}22;color:{RED}}}
.badge-amber{{background:{AMBER}22;color:{AMBER}}}
.badge-blue{{background:{ACCENT}22;color:{ACCENT}}}
.badge-purple{{background:{PURPLE}22;color:{PURPLE}}}
.loading{{background:linear-gradient(90deg,{CARD} 25%,{BORDER} 50%,{CARD} 75%);
           background-size:200% 100%;animation:shimmer 1.5s infinite;border-radius:8px;
           height:40px;margin-bottom:8px}}
.pnl-pos{{color:{GREEN};font-weight:700}}
.pnl-neg{{color:{RED};font-weight:700}}
.market-banner-closed{{background:linear-gradient(135deg,#2d0a0a,#1a0505);
  border:1px solid {RED}66;border-radius:8px;padding:10px 16px;
  margin-bottom:14px;text-align:center;font-size:13px;color:{RED};font-weight:600}}
.market-banner-open{{background:linear-gradient(135deg,#0a2d0a,#051a05);
  border:1px solid {GREEN}66;border-radius:8px;padding:10px 16px;
  margin-bottom:14px;text-align:center;font-size:13px;color:{GREEN};font-weight:600}}
.spin{{animation:spin 1.5s linear infinite}}
@media(min-width:768px){{
  .tab-bar{{display:none}}
  .subnav{{display:none!important}}
  .sidebar{{display:flex!important}}
  .main{{padding:20px 28px 20px 252px}}
  .sg{{grid-template-columns:repeat(4,1fr)}}
  .sv{{font-size:24px}}
  .ai-chat{{max-height:600px}}
}}
@media(min-width:1100px){{.sg{{grid-template-columns:repeat(5,1fr)}}}}
@media(max-width:480px){{
  .top-title{{font-size:13px}}
  .sv{{font-size:18px}}
  .card{{padding:12px}}
  th,td{{padding:5px 6px;font-size:11px}}
  .tw{{overflow-x:auto;-webkit-overflow-scrolling:touch}}
  .sg{{grid-template-columns:1fr 1fr}}
}}
@media(hover:none){{
  .sc:hover,.card:hover{{transform:none;box-shadow:none;border-color:{BORDER}}}
}}
"""

    # ── Navigation (FIX 4: 3 grouped sections) ──────────────────────────────
    NAV_EQ = [
        ("/",            "📊","Equity",     "eq"),
        ("/history",     "📈","Eq History", "eq"),
        ("/performance", "🎯","Eq Perf",    "eq"),
        ("/s1",          "🔬","Eq-S1",      "eq"),
        ("/s2",          "🔭","Eq-S2",      "eq"),
        ("/s3",          "🌐","Eq-S3",      "eq"),
        ("/opt",         "⚡","Eq-Opt",     "eq"),
        ("/bestx",       "🤖","Eq-BestX",   "eq"),
    ]
    NAV_CM = [
        ("/comm",        "🥇","Commodity",  "cm"),
        ("/hist-comm",   "📈","Cm History", "cm"),
        ("/perf-comm",   "🎯","Cm Perf",    "cm"),
        ("/cs1",         "🔬","Cm-S1",      "cm"),
        ("/cs2",         "🔭","Cm-S2",      "cm"),
        ("/cs3",         "🌐","Cm-S3",      "cm"),
        ("/copt",        "⚡","Cm-Opt",     "cm"),
        ("/cbestx",      "🤖","Cm-BestX",   "cm"),
    ]
    NAV_CR = [
        ("/crypto",      "₿", "Crypto",     "cr"),
        ("/hist-crypto", "📈","Cr History", "cr"),
        ("/perf-crypto", "🎯","Cr Perf",    "cr"),
        ("/cr1",         "🔬","Cr-S1",      "cr"),
        ("/cr2",         "🔭","Cr-S2",      "cr"),
        ("/cr3",         "🌐","Cr-S3",      "cr"),
        ("/cropt",       "⚡","Cr-Opt",     "cr"),
        ("/crbestx",     "🤖","Cr-BestX",   "cr"),
    ]
    NAV_UT = [
        ("/charts", "📈","Charts",  "ut"),
        ("/intel", "📰","Intel",  "ut"),
        ("/sys",   "⚙", "System", "ut"),
        ("/ai",    "🤖","AI Help","ut"),
    ]
    NAV_ALL = NAV_EQ + NAV_CM + NAV_CR + NAV_UT

    # ── ID maps ──────────────────────────────────────────────────────────────
    PAGE_IDS = [
        "pg-eq","pg-hist","pg-perf",
        "pg-s1","pg-s2","pg-s3","pg-opt","pg-bx",
        "pg-comm","pg-hcm","pg-pcm",
        "pg-cs1","pg-cs2","pg-cs3","pg-copt","pg-cbx",
        "pg-crypto","pg-hcr","pg-pcr",
        "pg-cr1","pg-cr2","pg-cr3","pg-cropt","pg-crbx",
        "pg-intel","pg-sys","pg-ai","pg-charts",
    ]
    PATH_TO_ID = {
        "/":"pg-eq", "/history":"pg-hist", "/performance":"pg-perf",
        "/s1":"pg-s1","/s2":"pg-s2","/s3":"pg-s3","/opt":"pg-opt","/bestx":"pg-bx",
        "/comm":"pg-comm","/hist-comm":"pg-hcm","/perf-comm":"pg-pcm",
        "/cs1":"pg-cs1","/cs2":"pg-cs2","/cs3":"pg-cs3","/copt":"pg-copt","/cbestx":"pg-cbx",
        "/crypto":"pg-crypto","/hist-crypto":"pg-hcr","/perf-crypto":"pg-pcr",
        "/cr1":"pg-cr1","/cr2":"pg-cr2","/cr3":"pg-cr3","/cropt":"pg-cropt","/crbestx":"pg-crbx",
        "/intel":"pg-intel","/sys":"pg-sys","/ai":"pg-ai","/charts":"pg-charts",
    }

    # ── Helpers ──────────────────────────────────────────────────────────────
    def sc(label, val, color=TEXT, sub=""):
        return html.Div([
            html.Div(label,className="sl"),
            html.Div(val if not isinstance(val,(int,float)) else f"{val}",
                     className="sv",style={"color":color,"fontWeight":"700"}),
            *([html.Div(sub,className="ss")] if sub else []),
        ],className="sc")


    EQ_LINKS  = [("/","📊","Main"),("/s1","🔬","Scan 1"),("/s2","🔭","Scan 2"),("/s3","🌐","Scan 3"),
                 ("/opt","⚡","Optimizer"),("/bestx","🤖","BestX"),("/history","📈","History"),("/performance","🏆","Performance")]
    CM_LINKS  = [("/comm","🥇","Main"),("/cs1","🔬","CS1"),("/cs2","🔭","CS2"),("/cs3","🌐","CS3"),
                 ("/copt","⚡","Optimizer"),("/cbestx","🤖","BestX"),("/hist-comm","📈","History"),("/perf-comm","🏆","Performance")]
    CR_LINKS  = [("/crypto","₿","Main"),("/cr1","🔬","CR1"),("/cr2","🔭","CR2"),("/cr3","🌐","CR3"),
                 ("/cropt","⚡","Optimizer"),("/crbestx","🤖","BestX"),("/hist-crypto","📈","History"),("/perf-crypto","🏆","Performance")]


    def pos_row(sym, side, entry, live, qty, upnl):
        v=float(upnl.replace("₹","").replace(",","").replace("+",""))
        return html.Div([
            html.Div([html.Span(sym,className="pos-sym"),
                      html.Span(f"  {side}",style={"color":GREEN if side=="BUY" else RED,
                                                     "fontSize":"12px","fontWeight":"600"})]),
            html.Div(f"Entry:{entry}  Qty:{qty}",className="pos-meta"),
            html.Div([html.Span(f"Live:{live}  ",style={"color":DIM,"fontSize":"12px"}),
                      html.Span(f"uPnL:{upnl}",className="pup" if v>=0 else "pdn")],
                     style={"marginTop":"2px"}),
        ],className="pos-row")

    def table_wrap(headers, rows_data):
        trs=[html.Tr([html.Th(h,style={"padding":"7px 10px","color":DIM,"fontSize":"11px",
                     "borderBottom":f"1px solid {BORDER}","textAlign":"left" if i==0 else "right"})
                     for i,h in enumerate(headers)])]
        for row in rows_data:
            cells=[]
            for i,(val,col) in enumerate(row):
                cells.append(html.Td(val,style={"padding":"6px 10px","color":col,"textAlign":"left" if i==0 else "right"}))
            trs.append(html.Tr(cells,style={"borderBottom":f"1px solid {BORDER}"}))
        return html.Table(trs,style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"})

    def hist_cards(history, target_pct=0.30):
        if not history:
            return html.Div("No trade logs found yet.",style={"color":DIM,"padding":"12px"},className="card")
        cards=[]
        for ds in sorted(history.keys(),reverse=True):
            d=history[ds]; pct=d["pct_return"]; pnl=d["net_pnl"]; wr=d["win_rate"]
            try:
                dt=__import__("datetime").datetime.strptime(ds,"%Y%m%d"); lbl=dt.strftime("%a %d %b %Y")
            except: lbl=ds
            pc=GREEN if pct>=target_pct else(YELLOW if pct>=0.10 else RED)
            bw=min(max(pct/target_pct*100,0),100) if pct>0 else 0
            cards.append(html.Div([
                html.Div([html.Span(lbl,style={"fontWeight":"700"}),
                          html.Div([html.Div(style={"height":"4px","width":f"{bw:.1f}%","background":pc,"borderRadius":"3px"})],
                                   style={"background":BORDER,"borderRadius":"3px","height":"4px","flex":"1","margin":"0 10px"}),
                          html.Span(f"{pct:+.3f}%",style={"color":pc,"fontWeight":"700","fontFamily":"monospace","fontSize":"13px"})],
                         style={"display":"flex","alignItems":"center","marginBottom":"10px"}),
                html.Div([
                    html.Div([html.Div(f"₹{pnl:+,.0f}",style={"color":GREEN if pnl>=0 else RED,"fontWeight":"700","fontSize":"18px"}),html.Div("Net P&L",style={"color":DIM,"fontSize":"11px"})]),
                    html.Div([html.Div(str(d["total_trades"]),style={"color":ACCENT,"fontWeight":"700","fontSize":"18px"}),html.Div("Trades",style={"color":DIM,"fontSize":"11px"})]),
                    html.Div([html.Div(f"{wr:.1f}%",style={"color":GREEN if wr>=60 else AMBER,"fontWeight":"700","fontSize":"18px"}),html.Div("Win Rate",style={"color":DIM,"fontSize":"11px"})]),
                ],style={"display":"flex","gap":"24px"}),
            ],className="card",style={"borderLeft":f"3px solid {pc}","marginBottom":"10px"}))
        return html.Div(cards)

    def perf_charts(history, sc_col=ACCENT, target_pct=0.30):
        if not history:
            return [html.Div("No performance data.",className="card",style={"color":DIM,"padding":"20px"})]*4
        days=sorted(history.keys())[-21:]; pcts=[history[d]["pct_return"] for d in days]
        try: labels=[__import__("datetime").datetime.strptime(d,"%Y%m%d").strftime("%d %b") for d in days]
        except: labels=days
        fig=go.Figure(go.Bar(x=labels,y=pcts,
                             marker_color=[GREEN if p>=target_pct else(AMBER if p>=0 else RED) for p in pcts],
                             text=[f"{p:+.3f}%" for p in pcts],textposition="outside"))
        fig.add_hline(y=target_pct,line_dash="dot",line_color="#ffd700",annotation_text=f"{target_pct:.2f}% Target",annotation_font_color="#ffd700")
        fig.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
                          font_color=TEXT,height=300,margin=dict(l=40,r=20,t=30,b=40),
                          xaxis=dict(gridcolor=BORDER),yaxis=dict(gridcolor=BORDER,tickformat=".3f",ticksuffix="%"))
        chart=html.Div([dcc.Graph(figure=fig,config={"displayModeBar":False})],className="card")
        avg=sum(pcts)/len(pcts) if pcts else 0
        avg_t=sum(history[d]["total_trades"] for d in days)/len(days) if days else 5
        proj=avg/100*21*avg_t*100_000; tgt_proj=target_pct/100*21*avg_t*100_000
        hits=sum(1 for p in pcts if p>=target_pct)
        proj_c=html.Div([
            html.Div("🔮 Monthly Projection",className="sec"),
            html.Div([
                html.Div([html.Div("Current X",style={"color":DIM,"fontSize":"11px","marginBottom":"6px"}),
                          html.Div(f"₹{proj:,.0f}/mo",style={"color":GREEN if avg>=target_pct else AMBER,"fontSize":"22px","fontWeight":"800"}),
                          html.Div(f"{avg:+.3f}% avg × 21d × {avg_t:.0f}t",style={"color":DIM,"fontSize":"11px","fontFamily":"monospace"}),
                          html.Div(f"Target days hit: {hits}/{len(pcts)}",style={"color":AMBER,"fontSize":"11px","marginTop":"4px"})],
                         className="card",style={"flex":"1","textAlign":"center","borderTop":f"3px solid {sc_col}"}),
                html.Div([html.Div(f"{target_pct:.2f}% Target",style={"color":DIM,"fontSize":"11px","marginBottom":"6px"}),
                          html.Div(f"₹{tgt_proj:,.0f}/mo",style={"color":"#ffd700","fontSize":"22px","fontWeight":"800"}),
                          html.Div(f"{target_pct:.3f}% × 21d × {avg_t:.0f}t",style={"color":DIM,"fontSize":"11px","fontFamily":"monospace"})],
                         className="card",style={"flex":"1","textAlign":"center","borderTop":"3px solid #ffd700"}),
            ],style={"display":"flex","gap":"16px","flexWrap":"wrap"}),
        ],className="card")
        tbl_rows=[]
        for d in reversed(days):
            h=history[d]; p=h["pct_return"]
            try: lbl=__import__("datetime").datetime.strptime(d,"%Y%m%d").strftime("%d %b")
            except: lbl=d
            pc=GREEN if p>=target_pct else(AMBER if p>=0 else RED)
            tbl_rows.append(html.Tr([
                html.Td(lbl,style={"padding":"6px 10px"}),
                html.Td(str(h["total_trades"]),style={"padding":"6px 10px","textAlign":"right"}),
                html.Td(f"₹{h['net_pnl']:+,.0f}",style={"padding":"6px 10px","textAlign":"right","color":GREEN if h["net_pnl"]>=0 else RED,"fontWeight":"700"}),
                html.Td(f"{p:+.3f}%",style={"padding":"6px 10px","textAlign":"right","color":pc,"fontWeight":"700","fontFamily":"monospace"}),
                html.Td(f"{h['win_rate']:.1f}%",style={"padding":"6px 10px","textAlign":"right"}),
            ]))
        tbl=html.Div([html.Div("Daily Returns",className="sec"),
                      html.Table([html.Thead(html.Tr([html.Th(h,style={"padding":"8px 10px","color":DIM,"fontSize":"11px","borderBottom":f"1px solid {BORDER}","textAlign":"left" if i==0 else "right"}) for i,h in enumerate(("Date","Trades","Net P&L","Return%","Win%"))])),
                                  html.Tbody(tbl_rows)],style={"width":"100%","borderCollapse":"collapse"})],
                     className="card",style={"overflowX":"auto"})
        return tbl,chart,proj_c,html.Div()

    # ── Page builders ────────────────────────────────────────────────────────
    def pg_eq():
        return html.Div([
            html.Div("Equity Trading",className="pg-title",style={"color":EQ_COL}),
            html.Div("NSE live levels · open positions · day P&L · 1-5s prices",className="pg-sub"),
            html.Div(id="eq-mkt"),
            html.Div(id="eq-s",className="sg"),
            html.Div(id="eq-pos"),
            html.Div(id="eq-closed"),
            html.Div(id="eq-lvl"),
            html.Div(id="eq-chart"),
        ],id="pg-eq",className="page")

    def pg_hist(pid,title,sub,xid,cid,btnid,dlid,sc=EQ_COL):
        _hn = (EQ_LINKS if sc==EQ_COL else CM_LINKS if sc==COMM_COL else CR_LINKS)
        _ha = ("/" if sc==EQ_COL else "/comm" if sc==COMM_COL else "/crypto")
        return html.Div([
            html.Div(title,className="pg-title",style={"color":sc}),
            html.Div(sub,className="pg-sub"),
            html.Div(id=xid,className="card",style={"marginBottom":"14px"}),
            html.Div(id=cid),
            dcc.Download(id=dlid),
            html.Button("⬇ Export CSV",id=btnid,style={"marginTop":"12px","background":ACCENT,"color":"#000","border":"none","borderRadius":"6px","padding":"8px 18px","cursor":"pointer","fontSize":"12px"}),
        ],id=f"pg-{pid}",className="page")

    def pg_perf(pid,title,sub,t1,t2,t3,t4,btnid,dlid,sc=EQ_COL):
        return html.Div([
            html.Div(title,className="pg-title",style={"color":sc}),
            html.Div(sub,className="pg-sub"),
            html.Div(id=t1),html.Div(id=t2),html.Div(id=t3),html.Div(id=t4),
            dcc.Download(id=dlid),
            html.Button("⬇ Export CSV",id=btnid,style={"marginTop":"12px","background":ACCENT,"color":"#000","border":"none","borderRadius":"6px","padding":"8px 18px","cursor":"pointer","fontSize":"12px"}),
        ],id=f"pg-{pid}",className="page")

    def pg_sc(sid,col=EQ_COL):
        info={"1":"Narrow (0.008–0.009, 1K)","2":"Dual-Band (0.001–0.016, 13K)","3":"WideDual (0.001–0.032, 31K)"}
        return html.Div([
            html.Div(f"Equity Scanner {sid} — {info[str(sid)]}",className="pg-title",style={"color":col}),
            html.Div(f"Testing X variations on all 38 equity stocks live",className="pg-sub"),
            html.Div(id=f"s{sid}-s",className="sg"),
            html.Div(id=f"s{sid}-c"),
            html.Div(id=f"s{sid}-t"),
        ],id=f"pg-s{sid}",className="page")

    def pg_opt():
        return html.Div([
            html.Div("Equity X-Optimizer",className="pg-title",style={"color":EQ_COL}),
            html.Div("Combined leaderboard from all 3 equity scanners (45K variations/day)",className="pg-sub"),
            html.Div(id="op-s",className="sg"),
            html.Div(id="op-c"),
            html.Div(id="op-t"),
        ],id="pg-opt",className="page")

    def pg_bx():
        return html.Div([
            html.Div("Equity Best-X Paper Trader",className="pg-title",style={"color":EQ_COL}),
            html.Div("Paper trades using today's optimizer best X",className="pg-sub"),
            html.Div(id="bx-s",className="sg"),
            html.Div(id="bx-open"),
            html.Div(id="bx-closed"),
        ],id="pg-bx",className="page")

    def pg_comm():
        return html.Div([
            html.Div("Commodity Trading — MCX Live",className="pg-title",style={"color":COMM_COL}),
            html.Div("Gold · Silver · Crude · NatGas · Copper | Session 09:00–23:30 IST | 1-5s prices",className="pg-sub"),
            html.Div(id="cm-s",className="sg"),
            html.Div(id="cm-lvl"),
            html.Div(id="cm-pos"),
            html.Div(id="cm-trades"),
        ],id="pg-comm",className="page")

    def pg_cs(sid):
        return html.Div([
            html.Div(f"MCX Scanner {sid}",className="pg-title",style={"color":COMM_COL}),
            html.Div("X-factor sweep across 5 MCX commodities",className="pg-sub"),
            html.Div(id=f"cs{sid}-s",className="sg"),
            html.Div(id=f"cs{sid}-t"),
        ],id=f"pg-cs{sid}",className="page")

    def pg_copt():
        return html.Div([
            html.Div("MCX Optimizer",className="pg-title",style={"color":COMM_COL}),
            html.Div("Best X per commodity across 97,500 variations",className="pg-sub"),
            html.Div(id="co-c"),
        ],id="pg-copt",className="page")

    def pg_cbx():
        return html.Div([
            html.Div("Commodity Best-X Trader",className="pg-title",style={"color":COMM_COL}),
            html.Div("Paper trades on MCX using optimizer best X",className="pg-sub"),
            html.Div(id="cbx-c"),
        ],id="pg-cbx",className="page")

    def pg_crypto():
        return html.Div([
            html.Div("Crypto Trading — Binance Live",className="pg-title",style={"color":CRYPTO_COL}),
            html.Div("BTC · ETH · BNB · SOL · ADA | 24/7 · Re-anchors every 6h",className="pg-sub"),
            html.Div(id="cr-s",className="sg"),
            html.Div(id="cr-anchor",style={"padding":"8px 14px","background":"#161b22","borderRadius":"6px",
                     "color":YELLOW,"fontSize":"13px","marginBottom":"12px"}),
            html.Div(id="cr-lvl"),
            html.Div(id="cr-pos"),
            html.Div(id="cr-trades"),
        ],id="pg-crypto",className="page")

    def pg_cr(sid):
        return html.Div([
            html.Div(f"Crypto Scanner {sid}",className="pg-title",style={"color":CRYPTO_COL}),
            html.Div("X-factor sweep across BTC/ETH/BNB/SOL/ADA | 24/7 rolling 6h windows",className="pg-sub"),
            html.Div(id=f"cr{sid}-s",className="sg"),
            html.Div(id=f"cr{sid}-t"),
        ],id=f"pg-cr{sid}",className="page")

    def pg_cropt():
        return html.Div([
            html.Div("Crypto Optimizer",className="pg-title",style={"color":CRYPTO_COL}),
            html.Div("Best X per coin across 97,500 crypto variations",className="pg-sub"),
            html.Div(id="cro-c"),
        ],id="pg-cropt",className="page")

    def pg_crbx():
        return html.Div([
            html.Div("Crypto Best-X Trader",className="pg-title",style={"color":CRYPTO_COL}),
            html.Div("Paper trades on Binance using optimizer best X",className="pg-sub"),
            html.Div(id="crbx-c"),
        ],id="pg-crbx",className="page")

    def pg_intel():
        return html.Div([
            html.Div("Market Intelligence",className="pg-title"),
            html.Div("Live news · Indices · FII/DII · Crypto prices · Commodity prices",className="pg-sub"),
            html.Div(id="in-idx",className="sg"),
            html.Div([
                html.Div(id="in-cpx",style={"flex":"1","minWidth":"240px"}),
                html.Div(id="in-mpx",style={"flex":"1","minWidth":"240px"}),
            ],style={"display":"flex","gap":"12px","flexWrap":"wrap","marginBottom":"14px"}),
            html.Div([
                html.Div(id="in-india",style={"flex":"1","minWidth":"280px"}),
                html.Div(id="in-intl", style={"flex":"1","minWidth":"280px"}),
                html.Div(id="in-sig",  style={"flex":"1","minWidth":"280px"}),
            ],style={"display":"flex","gap":"12px","flexWrap":"wrap"}),
        ],id="pg-intel",className="page")

    def pg_sys():
        return html.Div([
            html.Div("System Status",className="pg-title"),
            html.Div("All 15 processes · URLs · Prices · Coverage",className="pg-sub"),
            html.Div(id="sys-net"),
            html.Div(id="sys-c"),
        ],id="pg-sys",className="page")

    def pg_ai():
        qs=["How does the X multiplier work?","Explain the 3 scanners",
            "Why do CommScanners exit on weekends?","What is retreat exit?",
            "How does crypto trading work?","How to read the Performance page?"]
        return html.Div([
            html.Div("🤖 AI Assistant",className="pg-title"),
            html.Div("Ask anything about AlgoStack — trading logic, X multiplier, scanners, dashboard",className="pg-sub"),
            html.Div([
                html.Div(id="ai-box",className="ai-chat",children=[
                    html.Div("👋 Hello! I'm the AlgoStack AI assistant. I can explain the X multiplier, "
                             "scanners, equity/commodity/crypto trading logic, and help you read the dashboard.\n\n"
                             "What would you like to know?",className="ai-b"),
                ]),
                html.Div([
                    dcc.Input(id="ai-in",type="text",debounce=False,
                              placeholder="Ask about AlgoStack...",
                              style={"flex":"1","background":CARD,"border":f"1px solid {BORDER}",
                                     "borderRadius":"8px","padding":"10px 14px","color":TEXT,
                                     "fontFamily":FONT,"fontSize":"13px","outline":"none"}),
                    html.Button("Ask ›",id="ai-btn",n_clicks=0,
                                style={"background":ACCENT,"color":"#000","border":"none","borderRadius":"8px",
                                       "padding":"10px 20px","cursor":"pointer","fontWeight":"700","marginLeft":"8px"}),
                ],style={"display":"flex","alignItems":"center"}),
                html.Div([
                    html.Div("Quick questions:",style={"color":DIM,"fontSize":"11px","marginTop":"14px","marginBottom":"8px"}),
                    html.Div([html.Button(q,id=f"aiq-{i}",n_clicks=0,
                              style={"background":CARD,"color":ACCENT,"border":f"1px solid {BORDER}",
                                     "borderRadius":"16px","padding":"4px 12px","cursor":"pointer",
                                     "fontSize":"11px","marginRight":"6px","marginBottom":"6px"})
                              for i,q in enumerate(qs)]),
                ]),
                html.Div(id="ai-status",style={"color":DIM,"fontSize":"11px","marginTop":"6px"}),
                html.Div(id="ai-api-status",style={"marginTop":"4px"}),
            ],className="card"),
        ],id="pg-ai",className="page")

    # ── Charts Page ──────────────────────────────────────────────────────────
    def page_charts():
        """Charts page: in-app lightweight charts from live engine prices."""

        def _live_chart_card(title, graph_id, color, height=260):
            return html.Div([
                html.Div(title, style={"color":color,"fontWeight":"700","fontSize":"12px",
                                       "padding":"8px 12px","background":SB,
                                       "borderBottom":f"1px solid {BORDER}",
                                       "borderRadius":"8px 8px 0 0"}),
                dcc.Graph(
                    id=graph_id,
                    config={"displayModeBar": False},
                    style={"height": f"{height}px"},
                ),
            ], style={"flex":"1","minWidth":"280px","border":f"1px solid {BORDER}",
                      "borderRadius":"10px","overflow":"hidden","background":CARD,
                      "marginBottom":"0"})

        equity_section = html.Div([
            html.Div("📊  NSE Equity Charts", className="sec"),
            html.Div([
                _live_chart_card("NIFTY", "ch-eq-nifty", EQ_COL, 260),
                _live_chart_card("BANKNIFTY", "ch-eq-banknifty", EQ_COL, 260),
                _live_chart_card("INFY", "ch-eq-infy", EQ_COL, 200),
                _live_chart_card("RELIANCE", "ch-eq-reliance", EQ_COL, 200),
            ], style={"display":"flex","gap":"12px","flexWrap":"wrap","marginBottom":"12px"}),
        ], className="card")

        commodity_section = html.Div([
            html.Div("🥇  MCX Commodity Charts", className="sec"),
            html.Div([
                _live_chart_card("GOLD", "ch-comm-gold", COMM_COL, 230),
                _live_chart_card("SILVER", "ch-comm-silver", COMM_COL, 230),
                _live_chart_card("CRUDE OIL", "ch-comm-crude", COMM_COL, 230),
                _live_chart_card("NATURAL GAS", "ch-comm-naturalgas", COMM_COL, 230),
                _live_chart_card("COPPER", "ch-comm-copper", COMM_COL, 200),
            ], style={"display":"flex","gap":"14px","flexWrap":"wrap"}),
        ], className="card")

        crypto_section = html.Div([
            html.Div("₿  Binance Crypto Charts", className="sec"),
            html.Div([
                _live_chart_card("BTC / USDT", "ch-crypto-btc", CRYPTO_COL, 230),
                _live_chart_card("ETH / USDT", "ch-crypto-eth", CRYPTO_COL, 230),
                _live_chart_card("BNB / USDT", "ch-crypto-bnb", CRYPTO_COL, 230),
                _live_chart_card("SOL / USDT", "ch-crypto-sol", CRYPTO_COL, 230),
            ], style={"display":"flex","gap":"14px","flexWrap":"wrap"}),
        ], className="card")

        # Quick levels reference (pulled directly from DataStore — no callback needed)
        lvls   = _DS.get("levels",[])
        eq_px  = _DS.get("live_prices",{})
        def _fmt(v):
            if v is None: return "—"
            try: return f"₹{float(v):,.2f}"
            except: return str(v)

        eq_rows = [html.Tr([
            html.Th(h,style={"color":DIM,"fontSize":"10px","padding":"4px 8px",
                             "textAlign":"left","borderBottom":f"1px solid {BORDER}"})
            for h in ["Stock","Sell ↓","Current","Buy ↑"]])]
        for l in lvls[:18]:
            sym=l.get("SYMBOL",""); px_v=eq_px.get(sym,0)
            eq_rows.append(html.Tr([
                html.Td(sym, style={"color":ACCENT,"fontWeight":"700","padding":"4px 8px","fontSize":"12px"}),
                html.Td(_fmt(l.get("SELL BELOW")), style={"color":RED,"padding":"4px 6px","textAlign":"right","fontSize":"12px"}),
                html.Td(f"₹{px_v:,.2f}" if px_v else "—",
                        style={"color":YELLOW,"fontWeight":"700","padding":"4px 8px","textAlign":"center",
                               "background":"rgba(227,179,65,0.06)","fontSize":"12px"}),
                html.Td(_fmt(l.get("BUY ABOVE")), style={"color":GREEN,"padding":"4px 6px","fontSize":"12px"}),
            ], style={"borderBottom":f"1px solid {BORDER}"}))

        levels_section = html.Div([
            html.Div([
                html.Span("📋 Quick Levels Reference", className="sec"),
                html.Div([
                    html.A("Full equity →", href="/",
                           style={"color":EQ_COL,"textDecoration":"none","fontSize":"12px","marginRight":"14px"}),
                    html.A("MCX levels →", href="/comm",
                           style={"color":COMM_COL,"textDecoration":"none","fontSize":"12px","marginRight":"14px"}),
                    html.A("Crypto levels →", href="/crypto",
                           style={"color":CRYPTO_COL,"textDecoration":"none","fontSize":"12px"}),
                ], style={"display":"flex","marginBottom":"10px","flexWrap":"wrap","gap":"4px"}),
            ]),
            html.Div(html.Table(eq_rows,
                style={"width":"100%","borderCollapse":"collapse"}),
                className="tw"),
        ], className="card")

        return html.Div([
            html.Div([
                html.Div("📈  Charts", style={"fontSize":"20px","fontWeight":"700","color":ACCENT,"marginBottom":"4px"}),
                html.Div("Live charts from AlgoStack engine feeds · NSE Equity · MCX Commodities · Binance Crypto",
                         style={"color":DIM,"fontSize":"13px","marginBottom":"16px"}),
                html.Div([
                    html.Span("🔴 Sell Zone", style={"color":RED,"fontWeight":"700","marginRight":"16px"}),
                    html.Span("🟡 Current Price", style={"color":YELLOW,"fontWeight":"700","marginRight":"16px"}),
                    html.Span("🟢 Buy Zone", style={"color":GREEN,"fontWeight":"700"}),
                    html.Div(style={"flex":"1"}),
                    html.Span("No TradingView dependency — runs from your live engine data",
                              style={"color":DIM,"fontSize":"11px"}),
                ], style={"background":CARD,"border":f"1px solid {BORDER}","borderRadius":"8px",
                          "padding":"10px 14px","display":"flex","alignItems":"center",
                          "flexWrap":"wrap","gap":"6px","marginBottom":"14px","fontSize":"13px"}),
                equity_section,
                commodity_section,
                crypto_section,
                levels_section,
            ], style={"padding":"0 4px"}),
        ], id="pg-charts", className="page")

    # ── Sidebar ──────────────────────────────────────────────────────────────
    def sidebar():
        lnks=[]
        for sec,nav in [("── EQUITY",NAV_EQ),("── COMMODITY",NAV_CM),("── CRYPTO",NAV_CR),("── TOOLS",NAV_UT)]:
            lnks.append(html.Div(sec,className="sb-sec"))
            for hr,ic,lb,cls in nav:
                lnks.append(dcc.Link([html.Span(ic,className="sb-ic"),html.Span(lb)],
                                     href=hr,className=f"sb-lnk {cls}"))
        return html.Div([
            html.Div([
                html.Div([
                    html.Span("AlgoStack",className="sb-title"),
                    html.Span(" v10.7",style={"color":ACCENT,"fontSize":"13px","fontWeight":"400","opacity":"0.7"}),
                ],style={"display":"flex","alignItems":"baseline","gap":"0"}),
                html.Div("Ridhaant Ajoy Thackur",className="sb-sub"),
                html.Div([
                    html.Span(id="sb-mkt-dot",style={"display":"inline-block","width":"6px","height":"6px",
                                                      "borderRadius":"50%","background":GREEN,
                                                      "marginRight":"5px","verticalAlign":"middle"}),
                    html.Span(id="sb-mkt-lbl",style={"fontSize":"9px","color":DIM,"letterSpacing":"0.06em"})
                ],style={"marginTop":"5px","display":"flex","alignItems":"center"}),
            ],className="sb-brand"),
            *lnks,
            html.Div(id="sb-time",style={"padding":"14px 16px","fontSize":"11px","color":DIM,"borderTop":f"1px solid {BORDER}"}),
        ],className="sidebar")

    def tabbar():
        """Bottom tab bar + subnav strips — NO JavaScript (avoids ngrok redirect loops).
        Active state handled by Dash callback via CSS class updates."""
        mob = [("/","📊","Equity"),("/comm","🥇","Commod."),("/crypto","₿","Crypto"),
               ("/charts","📈","Charts"),("/intel","📰","Intel"),("/sys","⚙","System"),("/ai","🤖","AI")]
        tab_ids = ["eq","cm","cr","ch","in","sy","ai"]

        # Sub-section links per section — shown/hidden via Dash callback (no JS)
        sub_eq=[("/","📊","Main"),("/s1","🔬","Scan1"),("/s2","🔭","Scan2"),("/s3","🌐","Scan3"),
                ("/opt","⚡","Opt"),("/bestx","🤖","BestX"),("/history","📈","Hist"),("/performance","🏆","Perf")]
        sub_cm=[("/comm","🥇","Main"),("/cs1","🔬","CS1"),("/cs2","🔭","CS2"),("/cs3","🌐","CS3"),
                ("/copt","⚡","Opt"),("/cbestx","🤖","BestX"),("/hist-comm","📈","Hist"),("/perf-comm","🏆","Perf")]
        sub_cr=[("/crypto","₿","Main"),("/cr1","🔬","CR1"),("/cr2","🔭","CR2"),("/cr3","🌐","CR3"),
                ("/cropt","⚡","Opt"),("/crbestx","🤖","BestX"),("/hist-crypto","📈","Hist"),("/perf-crypto","🏆","Perf")]

        def _strip(nav_id, links):
            return html.Div([
                dcc.Link([html.Div(ic,className="subnav-ic"),html.Div(lb)],
                         href=hr, className="subnav-link")
                for hr,ic,lb in links
            ], id=nav_id, className="subnav", style={"display":"flex" if nav_id=="subnav-eq" else "none"})

        return html.Div([
            _strip("subnav-eq", sub_eq),
            _strip("subnav-cm", sub_cm),
            _strip("subnav-cr", sub_cr),
            html.Nav([
                dcc.Link(
                    [html.Div(ic,className="tab-icon"),html.Div(lb,className="tab-lbl")],
                    href=hr,
                    id="tab-"+tab_ids[i],
                    className="tab-btn",
                )
                for i,(hr,ic,lb) in enumerate(mob)
            ], className="tab-bar"),
            # NO JavaScript - active state via Dash callback below
        ])

    # ── Layout ───────────────────────────────────────────────────────────────
    app.index_string = (
        "<!DOCTYPE html><html lang='en'><head>{%metas%}<title>{%title%}</title>"
        "{%favicon%}{%css%}<style>"+CSS+"</style>"
        "<script>(function(){try{"
        "if(!window._as_no_reload){window._as_no_reload=true;"
        "var o=window.location.reload.bind(window.location);"
        "window.location.reload=function(hard){"
        "if(!hard){console.info('AlgoStack: page reload suppressed (WS reconnect)');return;}"
        "return o(hard);};}"
        "}catch(e){}})();</script>"
        "</head>"
        "<body>{%app_entry%}<footer>{%config%}{%scripts%}{%renderer%}</footer></body></html>"
    )

    app.layout = html.Div([
        html.Div([
            html.Div("AlgoStack",className="top-title"),
            html.Div(id="ticker",className="ticker-wrap"),
            html.Div(id="tb-time",className="top-time"),
        ],className="top-bar"),
        sidebar(), tabbar(),
        dcc.Location(id="url",refresh=False),
        dcc.Store(id="is-mobile", storage_type="memory"),
        html.Div([
            html.Div(id="err-banner"),
            pg_eq(),
            pg_hist("hist","Equity History","Daily P&L track record","hx-x","hx-c","hx-btn","hx-dl",EQ_COL),
            pg_perf("perf","Equity Performance","21-day returns + projection","pp1","pp2","pp3","pp4","pp-btn","pp-dl",EQ_COL),
            pg_sc(1),pg_sc(2),pg_sc(3),
            pg_opt(),pg_bx(),
            pg_comm(),
            pg_hist("hcm","Commodity History","MCX daily P&L","hcm-x","hcm-c","hcm-btn","hcm-dl",COMM_COL),
            pg_perf("pcm","Commodity Performance","MCX returns + projection","pcm1","pcm2","pcm3","pcm4","pcm-btn","pcm-dl",COMM_COL),
            pg_cs(1),pg_cs(2),pg_cs(3),
            pg_copt(),pg_cbx(),
            pg_crypto(),
            pg_hist("hcr","Crypto History","24/7 Binance P&L","hcr-x","hcr-c","hcr-btn","hcr-dl",CRYPTO_COL),
            pg_perf("pcr","Crypto Performance","Crypto returns + projection","pcr1","pcr2","pcr3","pcr4","pcr-btn","pcr-dl",CRYPTO_COL),
            pg_cr(1),pg_cr(2),pg_cr(3),
            pg_cropt(),pg_crbx(),
            pg_intel(),pg_sys(),pg_ai(),page_charts(),
        ],className="main"),
        dcc.Interval(id="t2", interval=1500,  n_intervals=0),  # v10.9: 1.5s prices (was 1s — too aggressive)
        dcc.Interval(id="t10",interval=4000,  n_intervals=0),  # v10.9: 4s scanners (was 2.5s)
        dcc.Interval(id="t30",interval=10000, n_intervals=0),  # v10.9: 10s market status (was 6s)
        dcc.Interval(id="t60",interval=30000, n_intervals=0),  # v10.9: 30s history/net (was 20s)
        html.Div(["AlgoStack v10.8  |  Ridhaant Ajoy Thackur  |  NSE + MCX + Binance  |  Target: 0.30%/day"],
                 style={"textAlign":"center","padding":"14px","borderTop":f"1px solid {BORDER}",
                        "background":SB,"marginTop":"40px","color":DIM,"fontSize":"11px"}),
    ],style={"background":BG,"minHeight":"100vh"})

    # ══════════════════════════════════════════════════════════════════════════
    #  ROUTER
    # ══════════════════════════════════════════════════════════════════════════
    @app.callback([Output(pid,"className") for pid in PAGE_IDS], Input("url","pathname"))
    def _route(path):
        active=PATH_TO_ID.get(path or "/","pg-eq")
        return ["page active" if pid==active else "page" for pid in PAGE_IDS]

    # ══════════════════════════════════════════════════════════════════════════
    #  MOBILE-LITE: throttle intervals on mobile Safari/phones
    # ══════════════════════════════════════════════════════════════════════════
    try:
        app.clientside_callback(
            """
            function(pathname){
              try{
                var ua = (navigator && navigator.userAgent) ? navigator.userAgent : "";
                var w  = (window && window.innerWidth) ? window.innerWidth : 9999;
                var isIOS = /iPhone|iPad|iPod/i.test(ua);
                var isMobile = isIOS || /Android/i.test(ua) || w <= 720;
                return isMobile;
              }catch(e){ return false; }
            }
            """,
            Output("is-mobile", "data"),
            Input("url", "pathname"),
        )
    except Exception:
        pass

    @app.callback(
        Output("t2", "interval"), Output("t10", "interval"),
        Output("t30", "interval"), Output("t60", "interval"),
        Input("is-mobile", "data"),
    )
    def _mobile_intervals(is_mobile):
        # Default (desktop)
        if not is_mobile:
            return 1200, 2500, 10000, 30000
        # Mobile-lite (reduce DOM churn + reconnect pressure)
        return 3000, 5000, 20000, 60000

    # ══════════════════════════════════════════════════════════════════════════
    #  SUBNAV + TAB ACTIVE STATE  (replaces all JS - zero redirect risk)
    # ══════════════════════════════════════════════════════════════════════════
    _EQ_PATHS  = {'','s1','s2','s3','opt','bestx','history','performance'}
    _CM_PATHS  = {'comm','cs1','cs2','cs3','copt','cbestx','hist-comm','perf-comm'}
    _CR_PATHS  = {'crypto','cr1','cr2','cr3','cropt','crbestx','hist-crypto','perf-crypto'}
    _TAB_MAP   = {'eq':'/','cm':'/comm','cr':'/crypto','ch':'/charts','in':'/intel','sy':'/sys','ai':'/ai'}

    @app.callback(
        Output("subnav-eq","style"), Output("subnav-cm","style"), Output("subnav-cr","style"),
        Output("tab-eq","className"), Output("tab-cm","className"), Output("tab-cr","className"),
        Output("tab-ch","className"), Output("tab-in","className"), Output("tab-sy","className"),
        Output("tab-ai","className"),
        Input("url","pathname"),
    )
    def _nav_active(path):
        p = (path or "/").lstrip("/")
        # Which section is active?
        if p in _EQ_PATHS or any(p.startswith(x+"/") for x in _EQ_PATHS if x):
            sec = "eq"
        elif p in _CM_PATHS or any(p.startswith(x+"/") for x in _CM_PATHS):
            sec = "cm"
        elif p in _CR_PATHS or any(p.startswith(x+"/") for x in _CR_PATHS):
            sec = "cr"
        else:
            sec = None  # intel/sys/ai/charts — no subnav

        _show = {"display":"flex"}
        _hide = {"display":"none"}
        nav_eq = _show if sec=="eq" else _hide
        nav_cm = _show if sec=="cm" else _hide
        nav_cr = _show if sec=="cr" else _hide

        def _tab_cls(key):
            tp = _TAB_MAP[key].lstrip("/")
            active = (tp=="" and p=="") or (tp and (p==tp or p.startswith(tp+"/")))
            return "tab-btn on" if active else "tab-btn"

        return (nav_eq, nav_cm, nav_cr,
                _tab_cls("eq"), _tab_cls("cm"), _tab_cls("cr"),
                _tab_cls("ch"), _tab_cls("in"), _tab_cls("sy"), _tab_cls("ai"))

    def _mk_live_chart_fig(sym: str, color: str, title: str):
        import plotly.graph_objects as go
        pts = list(_CHART_HIST.get(sym, []))
        if not pts:
            fig = go.Figure()
            fig.add_annotation(
                x=0.5, y=0.5, xref="paper", yref="paper",
                text=f"{title}: waiting for live ticks...",
                showarrow=False, font=dict(color=DIM, size=12),
            )
            fig.update_layout(
                paper_bgcolor=CARD, plot_bgcolor=CARD,
                margin=dict(l=28, r=14, t=10, b=26),
                xaxis=dict(visible=False), yaxis=dict(visible=False),
            )
            return fig
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        uniq = len({round(float(v), 8) for v in ys})
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=xs, y=ys, mode="lines+markers",
            line=dict(color=color, width=2),
            marker=dict(size=3, color=color),
            name=title
        ))
        if uniq <= 1:
            fig.add_annotation(
                x=0.5, y=0.92, xref="paper", yref="paper",
                text="Live feed OK - awaiting price movement",
                showarrow=False, font=dict(color=DIM, size=10),
            )
        fig.update_layout(
            paper_bgcolor=CARD, plot_bgcolor=CARD,
            margin=dict(l=28, r=14, t=10, b=26),
            showlegend=False,
            xaxis=dict(showgrid=False, color=DIM, tickfont=dict(size=9)),
            yaxis=dict(showgrid=True, gridcolor="rgba(110,118,129,0.18)", color=DIM),
        )
        return fig

    @app.callback(
        Output("ch-eq-nifty", "figure"),
        Output("ch-eq-banknifty", "figure"),
        Output("ch-eq-infy", "figure"),
        Output("ch-eq-reliance", "figure"),
        Output("ch-comm-gold", "figure"),
        Output("ch-comm-silver", "figure"),
        Output("ch-comm-crude", "figure"),
        Output("ch-comm-naturalgas", "figure"),
        Output("ch-comm-copper", "figure"),
        Output("ch-crypto-btc", "figure"),
        Output("ch-crypto-eth", "figure"),
        Output("ch-crypto-bnb", "figure"),
        Output("ch-crypto-sol", "figure"),
        Input("t2", "n_intervals"),
        prevent_initial_call=True,
    )
    def _live_chart_refresh(_n):
        try:
            eq = _DS.get("live_prices", {}) or {}
            cm = _DS.get("commodity_prices", {}) or {}
            cr = _DS.get("crypto_prices", {}) or {}

            _chart_push("NIFTY", float(eq.get("NIFTY", 0) or 0))
            _chart_push("BANKNIFTY", float(eq.get("BANKNIFTY", 0) or 0))
            _chart_push("INFY", float(eq.get("INFY", 0) or 0))
            _chart_push("RELIANCE", float(eq.get("RELIANCE", 0) or 0))
            _chart_push("GOLD", float(cm.get("GOLD", 0) or 0))
            _chart_push("SILVER", float(cm.get("SILVER", 0) or 0))
            _chart_push("CRUDE", float(cm.get("CRUDE", 0) or 0))
            _chart_push("NATURALGAS", float(cm.get("NATURALGAS", 0) or 0))
            _chart_push("COPPER", float(cm.get("COPPER", 0) or 0))
            _chart_push("BTC", float(cr.get("BTC", 0) or 0))
            _chart_push("ETH", float(cr.get("ETH", 0) or 0))
            _chart_push("BNB", float(cr.get("BNB", 0) or 0))
            _chart_push("SOL", float(cr.get("SOL", 0) or 0))

            return (
                _mk_live_chart_fig("NIFTY", EQ_COL, "NIFTY"),
                _mk_live_chart_fig("BANKNIFTY", EQ_COL, "BANKNIFTY"),
                _mk_live_chart_fig("INFY", EQ_COL, "INFY"),
                _mk_live_chart_fig("RELIANCE", EQ_COL, "RELIANCE"),
                _mk_live_chart_fig("GOLD", COMM_COL, "GOLD"),
                _mk_live_chart_fig("SILVER", COMM_COL, "SILVER"),
                _mk_live_chart_fig("CRUDE", COMM_COL, "CRUDE"),
                _mk_live_chart_fig("NATURALGAS", COMM_COL, "NATURALGAS"),
                _mk_live_chart_fig("COPPER", COMM_COL, "COPPER"),
                _mk_live_chart_fig("BTC", CRYPTO_COL, "BTC/USDT"),
                _mk_live_chart_fig("ETH", CRYPTO_COL, "ETH/USDT"),
                _mk_live_chart_fig("BNB", CRYPTO_COL, "BNB/USDT"),
                _mk_live_chart_fig("SOL", CRYPTO_COL, "SOL/USDT"),
            )
        except Exception as exc:
            log.warning("chart refresh callback fallback: %s", exc)
            _empty = {
                "data": [],
                "layout": {
                    "paper_bgcolor": CARD, "plot_bgcolor": CARD,
                    "margin": {"l": 28, "r": 14, "t": 10, "b": 26},
                    "xaxis": {"visible": False}, "yaxis": {"visible": False},
                    "annotations": [{
                        "text": "Chart fallback - updating",
                        "xref": "paper", "yref": "paper", "x": 0.5, "y": 0.5,
                        "showarrow": False, "font": {"color": DIM, "size": 11},
                    }],
                },
            }
            return (_empty, _empty, _empty, _empty, _empty, _empty, _empty, _empty, _empty, _empty, _empty, _empty, _empty)

    # ══════════════════════════════════════════════════════════════════════════
    #  CLOCK + TICKER
    # ══════════════════════════════════════════════════════════════════════════
    @app.callback(Output("tb-time","children"),Output("sb-time","children"),Output("ticker","children"),
                  Output("sb-mkt-dot","style"),Output("sb-mkt-lbl","children"),
                  Input("t2","n_intervals"),prevent_initial_call=True)
    def _clock(_):
        try:
            n=datetime.now(IST); t=n.hour*60+n.minute
            if   9*60+30<=t<=15*60+11: sess,col="NSE OPEN",GREEN
            elif 9*60+15<=t< 9*60+30:  sess,col="PRE-MKT", YELLOW
            elif 9*60<=t<=23*60+30:    sess,col="MCX OPEN",AMBER
            else:                      sess,col="CLOSED",  DIM
            ts=n.strftime("%H:%M:%S IST")
            px=_DS.get("live_prices",{}); cpx=_DS.get("crypto_prices",{}); mpx=_DS.get("commodity_prices",{})
            items=[]
            for sym in ["NIFTY","BANKNIFTY"]:
                v=px.get(sym)
                if v: items.append((EQ_COL,f"{sym}: ₹{float(v):,.2f}"))
            for sym,usd_fmt in [("BTC","${:,.0f}"),("ETH","${:,.0f}"),("BNB","${:,.1f}"),("SOL","${:,.2f}"),("ADA","${:,.4f}")]:
                v=cpx.get(sym)
                if v:
                    usd_v=float(v)
                    inr_v=usd_v*_usdt_to_inr()
                    # Show USD and INR
                    items.append((CRYPTO_COL,f"{sym}: {usd_fmt.format(usd_v)} / ₹{inr_v:,.0f}"))
            for sym in ["GOLD","SILVER","CRUDE","NATURALGAS","COPPER"]:
                v=mpx.get(sym)
                if v:
                    usd_v=float(v)
                    inr_v=usd_v*_usdt_to_inr()
                    items.append((COMM_COL,f"{sym}: ${usd_v:,.2f} / ₹{inr_v:,.0f}"))
            # Live status
            eq_age=float(_DS.get("price_age",9999) or 9999); cr_age=float(_DS.get("crypto_age",9999) or 9999)
            cm_age=float(_DS.get("commodity_age",9999) or 9999)
            status_items=[]
            if eq_age<5:  status_items.append((GREEN,"NSE ●"))
            elif eq_age<30: status_items.append((AMBER,f"NSE {eq_age:.0f}s"))
            if cr_age<5:  status_items.append((GREEN,"Binance ●"))
            elif cr_age<30: status_items.append((AMBER,f"Binance {cr_age:.0f}s"))
            if cm_age<15: status_items.append((GREEN,"MCX ●"))
            elif cm_age<60: status_items.append((AMBER,f"MCX {cm_age:.0f}s"))
            items=status_items+items
            if not items: items=[(DIM,"AlgoStack v10.7 — 240K calculations/day")]
            spans=[]
            for c,txt in items*3:
                spans.append(html.Span(f"◆ {txt}   ",style={"color":c,"marginRight":"32px","fontSize":"12px","fontWeight":"600"}))
            ticker=html.Div(html.Div(spans,style={"whiteSpace":"nowrap","animation":"marquee 45s linear infinite","display":"inline-block","paddingLeft":"100%"}),style={"overflow":"hidden","height":"20px"})
            t_disp=[html.Div(ts),html.Div(sess,style={"color":col,"fontWeight":"700"})]
            # Sidebar market dot colour
            sb_dot_col = GREEN if col==GREEN else (AMBER if col==AMBER else (YELLOW if col==YELLOW else DIM))
            dot_style = {"display":"inline-block","width":"6px","height":"6px","borderRadius":"50%",
                         "background":sb_dot_col,"marginRight":"5px","verticalAlign":"middle"}
            return t_disp, t_disp, ticker, dot_style, sess
        except Exception:
            _dot = {"display":"inline-block","width":"6px","height":"6px","borderRadius":"50%","background":DIM,"marginRight":"5px"}
            return "—","—",html.Div(), _dot, "—"

    # ══════════════════════════════════════════════════════════════════════════
    #  EQUITY PAGE
    # ══════════════════════════════════════════════════════════════════════════
    @app.callback(Output("eq-mkt","children"),Input("t30","n_intervals"),prevent_initial_call=True)
    def _eq_mkt(_):
        try:
            n=datetime.now(IST)
            if _MC_AVAILABLE and not MarketCalendar.is_trading_day(n):
                nd=MarketCalendar.next_trading_day(n)
                return html.Div(
                    f"🔴  MARKET CLOSED — {n.strftime('%A %d %b %Y')} is not a trading day.   "
                    f"Next: {nd.strftime('%a %d %b %Y')}",
                    className="market-banner-closed")
            t=n.hour*60+n.minute
            if 9*60+30<=t<=15*60+11:
                return html.Div("🟢  NSE OPEN — Equity trading active 09:30–15:11 IST",
                                className="market-banner-open")
        except Exception: pass
        return html.Div()

    @app.callback(Output("eq-s","children"),Output("eq-pos","children"),
                  Output("eq-closed","children"),Output("eq-lvl","children"),Output("eq-chart","children"),
                  Input("t2","n_intervals"))
    def _eq_upd(_):
        try:
            px=_DS.get("live_prices",{}); age=float(_DS.get("price_age",9999.0) or 9999.0)
            evs=_DS.get("events",[])
            # FIX 1: filter levels to equity stocks only
            lvls=[l for l in _DS.get("levels",[]) if str(l.get("SYMBOL","")).upper().strip() not in COMMODITY_SYMS|CRYPTO_SYMS]
            gross,net,sym_pnl=_compute_pnl(evs)
            nt=sum(s.get("trades",0) for s in sym_pnl)
            # Price freshness color
            age_col = GREEN if age<5 else (AMBER if age<30 else RED)
            age_txt = _age_label(age)
            # v10.9: Prominent staleness banner on main dashboard
            age_banner = html.Div([
                html.Span("● " if age < 5 else "⚠ " if age < 30 else "🔴 ", style={"fontSize":"14px"}),
                html.Span(f"NSE data: {age_txt}", style={"fontWeight":"700","color":age_col}),
                html.Span(f"  |  {len(px)} stocks live" if age < 30 else "  |  STALE — check Algofinal",
                          style={"color":DIM,"fontSize":"12px","marginLeft":"6px"}),
            ], style={"background":f"rgba({','.join(['88,166,255' if age<5 else '255,166,0' if age<30 else '255,80,80'])},0.1)",
                      "border":f"1px solid {age_col}","borderRadius":"8px","padding":"8px 14px",
                      "marginBottom":"12px","fontSize":"13px","color":age_col,"display":"flex","alignItems":"center"})
            stats=[sc("Day Net P&L",f"₹{net:+,.0f}",GREEN if net>=0 else RED),
                   sc("Day Gross",  f"₹{gross:+,.0f}",GREEN if gross>=0 else RED),
                   sc("Trades",     str(nt),ACCENT),
                   html.Div([
                       html.Div("EQUITY PRICES",className="sl"),
                       html.Div([
                           html.Span(f"{len(px)} stocks",style={"color":age_col,"fontWeight":"700","fontSize":"15px"}),
                           html.Div(age_txt,style={"color":age_col,"fontSize":"11px","marginTop":"2px","fontStyle":"italic"}),
                       ],className="sv"),
                   ],className="sc")]
            # Open positions
            ops=_open_positions(evs,px)
            pos_div=(html.Div([html.Div("Open Positions",className="sec"),
                               html.Div([pos_row(o["Symbol"],o["Side"],o["Entry"],o["Live"],o["Qty"],o["uPnL"]) for o in ops])],className="card")
                     if ops else
                     html.Div([html.Div("Open Positions",className="sec"),
                                html.Div("No open positions",style={"color":DIM,"fontSize":"13px"})],className="card"))
            # Closed trades
            cl=_closed_trades(evs)
            EXC={" Stop Loss":RED,"Retreat":AMBER,"EOD":DIM,"Manual":YELLOW}
            if cl:
                rows=[html.Div([
                    html.Div([html.Span(c["symbol"],style={"fontWeight":"700","fontSize":"14px"}),
                              html.Span(f" {c['side']}",style={"color":GREEN if c["side"]=="BUY" else RED,"fontSize":"12px","fontWeight":"600"}),
                              html.Span(f"  → {c['exit_type']}",style={"color":next((v for k,v in EXC.items() if k in c["exit_type"]),GREEN),"fontSize":"12px","fontWeight":"700","marginLeft":"6px"}),
                              html.Span(f"  {c['time']}",style={"color":DIM,"fontSize":"11px","marginLeft":"8px"})]),
                    html.Div(f"Entry ₹{c['entry_price']:,.2f} → Exit ₹{c['exit_price']:,.2f}  Qty:{c['qty']}",style={"color":DIM,"fontSize":"12px","marginTop":"2px"}),
                    html.Div([html.Span("Net: ",style={"color":DIM}),
                              html.Span(f"₹{c['net']:+,.2f}",style={"color":GREEN if c["net"]>=0 else RED,"fontWeight":"700"})],style={"marginTop":"2px"}),
                ],style={"padding":"10px 0","borderBottom":f"1px solid {BORDER}"}) for c in cl]
                cl_div=html.Div([html.Div(f"Completed Trades ({len(cl)})",className="sec"),html.Div(rows)],className="card")
            else:
                cl_div=html.Div([html.Div("Completed Trades",className="sec"),html.Div("No completed trades today",style={"color":DIM})],className="card")
            # Levels table
            if lvls:
                def _fmt(v):
                    if v is None: return "—"
                    try: return f"₹{float(v):,.2f}"
                    except: return str(v)
                # v10.6: show file age to indicate if levels are from prev session
                lvl_age_s = ""
                try:
                    cands=[os.path.join(LEVELS_DIR,f) for f in os.listdir(LEVELS_DIR)
                           if f.endswith(".xlsx") and "initial_levels" in f]
                    if cands:
                        cands.sort(key=os.path.getmtime,reverse=True)
                        lvl_age = round(time.time()-os.path.getmtime(cands[0]))
                        if lvl_age < 86400: lvl_age_s = f"  (updated {_age_label(lvl_age)})"
                        else: lvl_age_s = "  ⚠ prev session data"
                except Exception: pass
                disp=[{
                    # ── SELL SIDE (left, red) ──────────────────────────────────
                    "Sell SL":  _fmt(l.get("SL")),          # stop loss for shorts
                    "ST1":      _fmt(l.get("ST1")),
                    "Sell Tgt": _fmt(l.get("SELL BELOW")),  # sell trigger level
                    # ── CURRENT PRICE (centre, yellow) ─────────────────────────
                    "Stock":    l.get("SYMBOL",""),
                    "Current":  f"₹{px.get(str(l.get('SYMBOL','')).upper(),0):,.2f}"
                                if px.get(str(l.get("SYMBOL","")).upper()) else "—",
                    # ── BUY SIDE (right, green) ────────────────────────────────
                    "Buy Tgt":  _fmt(l.get("BUY ABOVE")),   # buy trigger level
                    "T1":       _fmt(l.get("T1")),
                    "Buy SL":   _fmt(l.get("SL")),          # stop loss for longs
                } for l in lvls[:40]]
                lvl_div=html.Div([html.Div([html.Span("Trading Levels (Equity Only)"),
                                            html.Span(lvl_age_s,style={"color":AMBER,"fontSize":"11px","marginLeft":"8px"}),
                                            html.Span(" · Sell Below  →  Current  →  Buy Above",
                                                      style={"color":DIM,"fontSize":"10px","marginLeft":"12px"})],
                                            className="sec"),
                                  html.Div([dash_table.DataTable(
                                      data=disp,columns=[{"name":c,"id":c} for c in disp[0]],
                                      sort_action="native",page_size=40,filter_action="native",
                                      style_header={"backgroundColor":SB,"color":EQ_COL,"fontWeight":"bold","border":f"1px solid {BORDER}"},
                                      style_cell={"backgroundColor":BG,"color":TEXT,"border":f"1px solid {BORDER}","fontFamily":FONT,"fontSize":"12px","padding":"8px"},
                                      style_cell_conditional=[
                                          {"if":{"column_id":"Stock"},"textAlign":"left","minWidth":"90px"},
                                          {"if":{"column_id":"Current"},"textAlign":"center","minWidth":"95px"},
                                      ],
                                      style_data_conditional=[
                                          # Sell side — red gradient
                                          {"if":{"column_id":"Sell Tgt"},"color":RED,"fontWeight":"700"},
                                          {"if":{"column_id":"ST1"},    "color":RED,"fontSize":"11px"},
                                          {"if":{"column_id":"Sell SL"},"color":AMBER,"fontSize":"11px","fontWeight":"700"},
                                          # Current price — yellow highlight
                                          {"if":{"column_id":"Stock"},  "color":ACCENT,"fontWeight":"bold"},
                                          {"if":{"column_id":"Current"},"color":YELLOW,"fontWeight":"700",
                                           "backgroundColor":"rgba(227,179,65,0.06)","borderLeft":f"2px solid {YELLOW}","borderRight":f"2px solid {YELLOW}"},
                                          # Buy side — green gradient
                                          {"if":{"column_id":"Buy Tgt"},"color":GREEN,"fontWeight":"700"},
                                          {"if":{"column_id":"T1"},     "color":GREEN,"fontSize":"11px"},
                                          {"if":{"column_id":"Buy SL"}, "color":AMBER,"fontSize":"11px","fontWeight":"700"},
                                      ])],className="tw")],className="card")
            else:
                lvl_div=html.Div([
                    html.Div("Trading Levels", className="sec"),
                    html.Div([
                        html.Div("📊 Loading equity levels…", style={"color":AMBER,"fontWeight":"600","marginBottom":"6px"}),
                        html.Div("Levels are written by Algofinal when the NSE market opens (Mon–Fri ~09:15 IST).", style={"color":DIM,"fontSize":"13px"}),
                        html.Div("Previous session levels will be shown if available.", style={"color":DIM,"fontSize":"12px","marginTop":"4px"}),
                    ], style={"padding":"16px"})
                ], className="card")
            # P&L chart
            chart=html.Div()
            if sym_pnl:
                ss=sorted(sym_pnl,key=lambda x:x["net"],reverse=True)
                fig=go.Figure(go.Bar(x=[s["symbol"] for s in ss],y=[s["net"] for s in ss],
                                     marker_color=[GREEN if s["net"]>=0 else RED for s in ss],
                                     text=[f"₹{s['net']:+,.0f}" for s in ss],textposition="outside"))
                fig.update_layout(paper_bgcolor=BG,plot_bgcolor=CARD,font_color=TEXT,
                                  xaxis=dict(gridcolor=BORDER),yaxis=dict(title="Net P&L (₹)",gridcolor=BORDER),
                                  margin=dict(l=50,r=10,t=20,b=50),height=240)
                chart=html.Div([html.Div("P&L by Symbol",className="sec"),dcc.Graph(figure=fig,style={"height":"250px"},config={"displayModeBar":False})],className="card")
            return [age_banner]+stats,pos_div,cl_div,lvl_div,chart
        except Exception as e:
            log.warning("eq_upd: %s",e); err=html.Div(f"Error: {e}",className="err"); return [],err,err,err,err

    # ══════════════════════════════════════════════════════════════════════════
    #  EQUITY SCANNERS
    # ══════════════════════════════════════════════════════════════════════════
    def _make_sc_cb(sid):
        @app.callback(Output(f"s{sid}-s","children"),Output(f"s{sid}-c","children"),Output(f"s{sid}-t","children"),
                      Input("t10","n_intervals"),prevent_initial_call=True)
        def _cb(_,_s=sid):
            try:
                with _DS._lock:
                    state=_DS._d["scanner"].get(_s); age=_DS._d["scanner_age"].get(_s,9999.0)
                if not state:
                    # v10.8: show informative status — on weekends market is closed
                    import datetime as _dtt
                    wd = _dtt.datetime.now().weekday()  # 5=Sat, 6=Sun
                    if wd >= 5:
                        msg = f"📅 Weekend — Equity Scanner {_s} resumes Monday 09:15 IST"
                    else:
                        msg = f"⏳ Equity Scanner {_s} starting up — please wait (takes ~2 min)"
                    return [sc("Status", "Offline" if wd >= 5 else "Starting", AMBER)]*4, \
                           html.Div(msg, style={"color": AMBER, "padding": "18px", "fontSize": "14px",
                                                "background": CARD, "borderRadius": "8px",
                                                "border": f"1px solid {BORDER}"}), html.Div()
                sweeps={}
                if "bands" in state:
                    for bdata in state["bands"].values():
                        for sym,sd in bdata.get("sweeps",{}).items():
                            if not sd.get("has_trades", True): continue  # v10.5: skip no-trade sweeps
                            xv=sd.get("x_values",[]); pnl=sd.get("total_pnl",[])
                            if not xv: continue
                            tc=sd.get("trade_count",[])
                            if not any(t>0 for t in tc if isinstance(tc,list)): continue
                            ex=sweeps.get(sym)
                            cur_best = sd.get("best_pnl", max((p for p in pnl if isinstance(p,(int,float))),default=-1e9))
                            prev_best = ex.get("best_pnl", -1e9) if ex else -1e9
                            if not ex or cur_best > prev_best:
                                sweeps[sym]=sd
                elif "sweeps" in state: sweeps=state["sweeps"]
                rows=[]
                for sym,sd in sweeps.items():
                    if not sd.get("has_trades", True): continue  # v10.7 guard
                    # Use pre-computed best_x from dump_state if available
                    pre_bx  = sd.get("best_x", 0)
                    pre_pnl = sd.get("best_pnl", 0)
                    pre_tc  = sd.get("best_trade_count", 0)
                    if pre_bx and pre_tc:
                        rows.append({"sym":sym,"x":float(pre_bx),"pnl":float(pre_pnl),
                                     "tc":int(pre_tc),
                                     "vs":f"{(float(pre_bx)-CURRENT_X)/CURRENT_X*100:+.1f}%"})
                        continue
                    # Fallback: compute from arrays with composite score
                    xv=sd.get("x_values",[]); pnl=sd.get("total_pnl",[]); tc=sd.get("trade_count",[])
                    if not xv or not any(t>0 for t in tc if isinstance(tc,list)): continue
                    try:
                        import numpy as _np3
                        pnl_a=_np3.array(pnl[:len(xv)],dtype=float)
                        tc_a =_np3.array(tc[:len(xv)], dtype=float) if isinstance(tc,list) else _np3.zeros(len(xv))
                        wc_a =_np3.array(sd.get("win_count",[0]*len(xv))[:len(xv)],dtype=float) if isinstance(sd.get("win_count",[]),list) else _np3.zeros(len(xv))
                        wr_a = _np3.where(tc_a>0, wc_a/tc_a, 0.0)
                        ar_a = _np3.where(tc_a>0, pnl_a/tc_a, 0.0)
                        pr=pnl_a.max()-pnl_a.min(); pn=(pnl_a-pnl_a.min())/pr if pr>1e-9 else _np3.zeros(len(xv))
                        ar_r=ar_a.max()-ar_a.min(); an=(ar_a-ar_a.min())/ar_r if ar_r>1e-9 else _np3.zeros(len(xv))
                        bi=int(_np3.argmax(0.50*pn+0.30*wr_a+0.20*an))
                    except Exception: bi=max(range(len(xv)),key=lambda i:pnl[i] if i<len(pnl) else 0)
                    rows.append({"sym":sym,"x":float(xv[bi]),"pnl":float(pnl[bi] if bi<len(pnl) else 0),
                                 "tc":int(tc[bi] if isinstance(tc,list) and bi<len(tc) else 0),
                                 "vs":f"{(float(xv[bi])-CURRENT_X)/CURRENT_X*100:+.1f}%"})
                rows.sort(key=lambda r:r["pnl"],reverse=True)
                if not rows:
                    return [
                    sc("STATUS","Pre-Market",AMBER),
                    sc("OPENS AT","09:30 IST",DIM),
                    sc("NEXT SESSION",datetime.now(IST).strftime("%d %b"),DIM),
                    sc("SCANNER",f"S{_s} Ready",GREEN),
                ], html.Div([
                    html.Div(f"⏳  Scanner {_s} — NSE opens 09:30 IST (Mon–Fri)",
                             style={"color":AMBER,"padding":"14px","fontWeight":"600","fontSize":"14px"}),
                    html.Div("Scanner is running and will start producing results once the market opens. "
                             "All X-variations will be tested live during the trading session.",
                             style={"color":DIM,"padding":"0 14px 14px","fontSize":"13px"}),
                ],className="card"), html.Div()
                bx=rows[0]["x"]; bp=rows[0]["pnl"]
                vsd=(bx-CURRENT_X)/CURRENT_X*100; ac=GREEN if age<60 else(AMBER if age<180 else RED)
                xc=GREEN if abs(vsd)<5 else(AMBER if abs(vsd)<20 else RED)
                stats=[sc("Best X",f"{bx:.6f}",EQ_COL),sc("Best P&L",f"₹{bp:+,.0f}",GREEN if bp>=0 else RED),
                       sc("vs Live X",f"{vsd:+.1f}%",xc),sc("Data Age",f"{age:.0f}s",ac)]
                top=rows[:30]
                fig=go.Figure(go.Bar(x=[f"{r['x']:.5f}" for r in top],y=[r["pnl"] for r in top],
                                     marker_color=[GREEN if r["pnl"]>=0 else RED for r in top],
                                     text=[f"₹{r['pnl']:+,.0f}" for r in top],textposition="outside"))
                fig.update_layout(paper_bgcolor=BG,plot_bgcolor=CARD,font_color=TEXT,height=280,
                                  xaxis=dict(title="X Multiplier",gridcolor=BORDER,tickangle=-45),
                                  yaxis=dict(title="P&L (₹)",gridcolor=BORDER),margin=dict(l=60,r=10,t=20,b=60))
                chart=html.Div([html.Div("X-Value P&L Distribution",className="sec"),
                                dcc.Graph(figure=fig,style={"height":"290px"},config={"displayModeBar":False})],className="card")

                # ── Per-symbol sweep progress cards ──────────────────────────
                sym_cards=[]
                for r in rows[:20]:
                    xv_arr=r.get("xv",[]); tc_arr=r.get("tc_arr",[])
                    n_total=len(xv_arr)
                    n_tested=sum(1 for v in tc_arr if isinstance(v,(int,float)) and v>0) if isinstance(tc_arr,list) else 0
                    pct_done=n_tested/n_total if n_total>0 else 0
                    wr=r.get("wr",0)
                    pnl_c=GREEN if r["pnl"]>=0 else RED
                    wr_c=GREEN if wr>=55 else(AMBER if wr>=40 else RED)
                    vs_v=float(r["vs"].replace("%","").replace("+","")) if r["vs"]!="—" else 0
                    vs_c=GREEN if abs(vs_v)<5 else(AMBER if abs(vs_v)<20 else RED)
                    sym_cards.append(html.Div([
                        html.Div([
                            html.Span(r["sym"],style={"fontWeight":"700","color":ACCENT,"fontSize":"14px","minWidth":"90px"}),
                            html.Span(f"Best X: {r['x']:.6f}",style={"color":EQ_COL,"fontSize":"12px","fontFamily":"monospace","marginLeft":"8px"}),
                            html.Span(f"  {r['vs']}",style={"color":vs_c,"fontSize":"11px","marginLeft":"6px"}),
                        ],style={"display":"flex","alignItems":"center","flexWrap":"wrap"}),
                        html.Div([
                            html.Span("P&L: ",style={"color":DIM,"fontSize":"11px"}),
                            html.Span(f"₹{r['pnl']:+,.0f}",style={"color":pnl_c,"fontWeight":"700","fontSize":"13px","marginRight":"16px"}),
                            html.Span(f"Trades: {r['tc']}",style={"color":TEXT,"fontSize":"11px","marginRight":"16px"}),
                            html.Span(f"Win: {wr:.0f}%",style={"color":wr_c,"fontSize":"11px","marginRight":"16px"}),
                            html.Span(f"Tested: {n_tested}/{n_total} X values",style={"color":DIM,"fontSize":"10px"}),
                        ],style={"marginTop":"5px","display":"flex","alignItems":"center","flexWrap":"wrap"}),
                        html.Div([
                            html.Span("Sweep",style={"color":DIM,"fontSize":"9px","marginRight":"6px","minWidth":"38px"}),
                            html.Div([html.Div(style={"width":f"{pct_done*100:.0f}%","height":"5px",
                                                      "background":GREEN if pct_done>0.8 else ACCENT,
                                                      "borderRadius":"3px","transition":"width 0.4s ease"})],
                                     style={"flex":"1","height":"5px","background":BORDER,"borderRadius":"3px"}),
                            html.Span(f"{pct_done*100:.0f}%",style={"color":DIM,"fontSize":"9px","marginLeft":"6px"}),
                        ],style={"display":"flex","alignItems":"center","marginTop":"6px","gap":"4px"}),
                    ],style={"background":CARD,"border":f"1px solid {BORDER}","borderRadius":"8px",
                             "padding":"10px 14px","marginBottom":"8px","borderLeft":f"3px solid {pnl_c}"}))

                tbl=html.Div([
                    html.Div("Per-Symbol Sweep Results",className="sec"),
                    html.Div(sym_cards),
                    html.Div([
                        html.Div("Full Ranked Table",className="sec",style={"marginTop":"16px"}),
                        html.Div([dash_table.DataTable(
                            data=[{"#":f"#{i+1}","Symbol":r["sym"],"Best X":f"{r['x']:.6f}",
                                   "P&L (₹)":f"₹{r['pnl']:+,.0f}","Trades":str(r["tc"]),
                                   "Win%":f"{r.get('wr',0):.0f}%","vs Live X":r["vs"]}
                                  for i,r in enumerate(rows)],
                            columns=[{"name":c,"id":c} for c in ["#","Symbol","Best X","P&L (₹)","Trades","Win%","vs Live X"]],
                            sort_action="native",page_size=20,
                            style_header={"backgroundColor":SB,"color":EQ_COL,"fontWeight":"bold","border":f"1px solid {BORDER}"},
                            style_cell={"backgroundColor":BG,"color":TEXT,"border":f"1px solid {BORDER}",
                                        "fontFamily":FONT,"fontSize":"12px","padding":"6px"},
                            style_data_conditional=[
                                {"if":{"column_id":"Best X"},"color":EQ_COL,"fontFamily":"monospace"},
                                {"if":{"column_id":"P&L (₹)","filter_query":"{P&L (₹)} contains '+'"},"color":GREEN,"fontWeight":"700"},
                                {"if":{"column_id":"P&L (₹)","filter_query":"{P&L (₹)} contains '-'"},"color":RED,"fontWeight":"700"},
                            ])],className="tw")],className="card"),
                ])
                return ([cm_age_banner]+stats if 'cm_age_banner' in dir() and age>=15 else stats),chart,tbl
            except Exception as e:
                err=html.Div(f"Scanner {_s}: {e}",className="err"); return [],err,err
    for _sid in (1,2,3): _make_sc_cb(_sid)

    # ══════════════════════════════════════════════════════════════════════════
    #  EQUITY OPTIMIZER  (FIX 2: stacked decorator bug removed)
    # ══════════════════════════════════════════════════════════════════════════
    @app.callback(Output("op-s","children"),Output("op-c","children"),Output("op-t","children"),
                  Input("t10","n_intervals"),prevent_initial_call=True)
    def _opt_upd(_):
        try:
            import pandas as pd
            df=_DS.get("optimizer_df"); age=_DS.get("opt_age",9999.0)
            if df is None or (hasattr(df,"empty") and df.empty):
                # Build from scanner states when no CSV
                rows=[]
                with _DS._lock:
                    for sid in (1,2,3):
                        state=_DS._d["scanner"].get(sid)
                        if not state: continue
                        sweeps_src={}
                        if "bands" in state:
                            for bdata in state["bands"].values():
                                for sym,sd in bdata.get("sweeps",{}).items():
                                    xv=sd.get("x_values",[]); pnl=sd.get("total_pnl",[])
                                    if not xv: continue
                                    bi=max(range(len(xv)),key=lambda i:pnl[i] if i<len(pnl) else -1e9)
                                    np2=pnl[bi] if bi<len(pnl) else 0
                                    ex=sweeps_src.get(sym)
                                    if not ex or np2>ex.get("pnl",-1e9): sweeps_src[sym]={"x":float(xv[bi]),"pnl":float(np2)}
                        elif "sweeps" in state:
                            for sym,sd in state["sweeps"].items():
                                xv=sd.get("x_values",[]); pnl=sd.get("total_pnl",[]); tc=sd.get("trade_count",[])
                                if not xv: continue
                                bi=max(range(len(xv)),key=lambda i:pnl[i] if i<len(pnl) else -1e9)
                                rows.append({"x_value":float(xv[bi]),"total_pnl":pnl[bi] if bi<len(pnl) else 0,"total_trades":tc[bi] if bi<len(tc) else 0})
                        for d in sweeps_src.values():
                            rows.append({"x_value":d["x"],"total_pnl":d["pnl"],"total_trades":0})
                if not rows:
                    return [sc("Status","Waiting for scanners",AMBER)]*4,html.Div("Optimizer shows data after scanners run during market hours.",className="card",style={"color":DIM,"padding":"20px"}),html.Div()
                df=pd.DataFrame(rows).groupby("x_value",as_index=False).agg(total_pnl=("total_pnl","sum"),total_trades=("total_trades","sum")).sort_values("total_pnl",ascending=False).reset_index(drop=True)
            df=df.copy()
            for c in ["x_value","total_pnl","total_trades"]:
                if c in df.columns: df[c]=pd.to_numeric(df[c],errors="coerce")
            df=df.dropna(subset=["x_value","total_pnl"]).reset_index(drop=True)
            if df.empty:
                return [sc("Status","No trades yet",AMBER)]*4,html.Div(),html.Div()
            b=df.iloc[0]; bx=float(b.get("x_value",0)); bp=float(b.get("total_pnl",0))
            vsd=(bx-CURRENT_X)/CURRENT_X*100; xc=GREEN if abs(vsd)<5 else(AMBER if abs(vsd)<20 else RED)
            raw_age=float(age) if age is not None else 9999.0; ac=GREEN if raw_age<60 else(AMBER if raw_age<300 else RED)
            stats=[sc("Best X",f"{bx:.6f}",EQ_COL),sc("Best P&L",f"₹{bp:+,.0f}",GREEN if bp>=0 else RED),
                   sc("vs Live X",f"{vsd:+.1f}%",xc),sc("Data Age",_age_label(raw_age),ac)]
            top=df.head(50); xv=top["x_value"].astype(float).tolist(); yv=top["total_pnl"].astype(float).tolist()
            fig=go.Figure()
            fig.add_trace(go.Bar(x=xv,y=yv,marker_color=[GREEN if v>=0 else RED for v in yv],
                                 text=[f"₹{v:,.0f}" for v in yv],textposition="outside",
                                 hovertemplate="X=%{x:.6f}<br>P&L=₹%{y:,.0f}<extra></extra>"))
            fig.add_vline(x=CURRENT_X,line_dash="dash",line_color=RED,line_width=2,
                          annotation_text=f"Live X={CURRENT_X:.5f}",annotation_font_color=RED,annotation_position="top right")
            fig.update_layout(paper_bgcolor=BG,plot_bgcolor=CARD,font_color=TEXT,height=280,
                              xaxis=dict(title="X Value",gridcolor=BORDER,tickformat=".5f",tickangle=-45),
                              yaxis=dict(title="Combined P&L (₹)",gridcolor=BORDER),margin=dict(l=60,r=10,t=20,b=60),bargap=0.1)
            chart=html.Div([html.Div("Top 50 X Values — All 3 Scanners Combined",className="sec"),dcc.Graph(figure=fig,style={"height":"290px"},config={"displayModeBar":False})],className="card")
            disp=df[["x_value","total_pnl","total_trades"]].head(50).copy()
            disp["total_pnl"]=disp["total_pnl"].apply("₹{:,.0f}".format)
            tbl=html.Div([html.Div("Full Leaderboard",className="sec"),
                          html.Div([dash_table.DataTable(data=disp.to_dict("records"),columns=[{"name":c.replace("_"," ").title(),"id":c} for c in disp.columns],
                                    sort_action="native",page_size=15,style_header={"backgroundColor":SB,"color":EQ_COL,"fontWeight":"bold","border":f"1px solid {BORDER}"},
                                    style_cell={"backgroundColor":BG,"color":TEXT,"border":f"1px solid {BORDER}","fontFamily":FONT,"fontSize":"12px","padding":"6px"})],className="tw")],className="card")
            return stats,chart,tbl
        except Exception as e:
            log.warning("opt_upd: %s",e); err=html.Div(f"Optimizer error: {e}",className="err"); return [],err,err

    # ══════════════════════════════════════════════════════════════════════════
    #  BEST-X TRADER
    # ══════════════════════════════════════════════════════════════════════════
    @app.callback(Output("bx-s","children"),Output("bx-open","children"),Output("bx-closed","children"),Input("t2","n_intervals"),prevent_initial_call=True)
    def _bx_upd(_):
        try:
            state=_rj(BESTX_FILE) or {}
            if not state:
                m=html.Div([html.Div("Best-X Trader offline.",style={"color":AMBER}),html.Div("Start: python best_x_trader.py",style={"color":DIM,"fontSize":"12px"})],className="card")
                return [sc("Status","Offline",RED)]*4,m,m
            bx=float(state.get("best_x",CURRENT_X)); dn=float(state.get("day_net",0))
            nt=int(state.get("n_trades",0)); pa=float(state.get("price_age",9999))
            stats=[sc("Best X",f"{bx:.6f}",EQ_COL),sc("Day Net",f"₹{dn:+,.0f}",GREEN if dn>=0 else RED),
                   sc("Trades",str(nt),ACCENT),sc("Prices",_age_label(pa),GREEN if pa<5 else AMBER)]
            ops=state.get("open",[])
            od=(html.Div([html.Div("Open Positions",className="sec"),
                          html.Div([pos_row(o["symbol"],o["side"],f"₹{o['entry']:,.2f}",f"₹{o.get('live',0):,.2f}",str(o["qty"]),f"₹{float(o.get('upnl',0)):+,.2f}") for o in ops])],className="card")
                if ops else html.Div([html.Div("Open Positions",className="sec"),html.Div("No open positions",style={"color":DIM})],className="card"))
            trd=state.get("trades",[])
            if trd:
                rows=[html.Div([html.Div([html.Span(t["symbol"],style={"fontWeight":"700"}),
                                          html.Span(f" {t['side']}",style={"color":GREEN if t["side"]=="BUY" else RED,"fontSize":"12px"}),
                                          html.Span(f"  → {t.get('exit_type','?')}",style={"color":GREEN,"fontSize":"12px","marginLeft":"6px"}),
                                          html.Span(f"  {t.get('time','')}",style={"color":DIM,"fontSize":"11px","marginLeft":"8px"})]),
                                html.Div([html.Span("Net: ",style={"color":DIM}),html.Span(f"₹{float(t['net']):+,.2f}",style={"color":GREEN if float(t['net'])>=0 else RED,"fontWeight":"700"})],style={"marginTop":"2px"}),
                               ],style={"padding":"8px 0","borderBottom":f"1px solid {BORDER}"}) for t in reversed(trd[-20:])]
                cd=html.Div([html.Div(f"Trades ({len(trd)})",className="sec"),html.Div(rows)],className="card")
            else:
                cd=html.Div([html.Div("Trades",className="sec"),html.Div("No trades yet",style={"color":DIM})],className="card")
            return stats,od,cd
        except Exception as e: err=html.Div(f"Error: {e}",className="err"); return [],err,err

    # ══════════════════════════════════════════════════════════════════════════
    #  EQUITY HISTORY & PERFORMANCE
    # ══════════════════════════════════════════════════════════════════════════
    @app.callback(Output("hx-x","children"),Output("hx-c","children"),Input("t60","n_intervals"),prevent_initial_call=True)
    def _hist_eq(_):
        try:
            h=_DS.get("history",{})
            x=html.Div([html.Div("Current Live X",className="sl"),html.Div(f"{CURRENT_X:.6f}",style={"fontSize":"28px","fontWeight":"800","color":EQ_COL}),html.Div("via CURRENT_X_MULTIPLIER in .env",style={"color":DIM,"fontSize":"11px","marginTop":"4px"})],className="card")
            return x,hist_cards(h)
        except Exception as e: return html.Div(),html.Div(f"Error: {e}",className="err")

    @app.callback(Output("hx-dl","data"),Input("hx-btn","n_clicks"),prevent_initial_call=True)
    def _hist_eq_dl(_):
        import io,csv as _c; buf=io.StringIO(); w=_c.writer(buf)
        w.writerow(["Date","Trades","NetPnL","Return%","WinRate%"])
        for ds,d in sorted(_DS.get("history",{}).items(),reverse=True): w.writerow([ds,d["total_trades"],d["net_pnl"],d["pct_return"],d["win_rate"]])
        return dict(content=buf.getvalue(),filename="algostack_equity_history.csv")

    @app.callback(Output("pp1","children"),Output("pp2","children"),Output("pp3","children"),Output("pp4","children"),Input("t60","n_intervals"),prevent_initial_call=True)
    def _perf_eq(_):
        try: r=perf_charts(_DS.get("history",{}),EQ_COL,0.30); return r if len(r)==4 else (r[0],)*4
        except Exception as e: err=html.Div(f"Error: {e}",className="err"); return err,err,err,err

    @app.callback(Output("pp-dl","data"),Input("pp-btn","n_clicks"),prevent_initial_call=True)
    def _perf_eq_dl(_):
        import io,csv as _c; buf=io.StringIO(); w=_c.writer(buf)
        w.writerow(["Date","Trades","NetPnL","Return%","WinRate%"])
        for ds,d in sorted(_DS.get("history",{}).items(),reverse=True): w.writerow([ds,d["total_trades"],d["net_pnl"],d["pct_return"],d["win_rate"]])
        return dict(content=buf.getvalue(),filename="algostack_equity_performance.csv")

    # ══════════════════════════════════════════════════════════════════════════
    #  COMMODITY PAGE
    # ══════════════════════════════════════════════════════════════════════════
    # ── CONSOLIDATED: all 4 commodity outputs in one callback (reduces HTTP round-trips) ──
    @app.callback(
        Output("cm-s","children"), Output("cm-lvl","children"),
        Output("cm-pos","children"), Output("cm-trades","children"),
        Input("t2","n_intervals")
    )
    def _cm_all(_):
        prices=_DS.get("commodity_prices",{}); evs=_DS.get("commodity_events",[])
        age=float(_DS.get("commodity_age",9999) or 9999)
        age=age if isinstance(age,float) else 9999.0
        # Stats
        try:
            closed=[e for e in evs if e.get("exit_px") is not None]
            net=sum(float(e.get("net_pnl",0) or 0) for e in closed)
            n=datetime.now(IST); t=n.hour*60+n.minute
            open_=9*60<=t<23*60+30; s_col=GREEN if open_ else RED
            is_live=age<15; is_cached=(age>=9000)
            price_label = ("🟢 Live — just now" if age<2 else
                           ("🟢 " + _age_label(age)) if age<15 else
                           ("🟡 " + _age_label(age)) if age<60 else
                           ("⚫ Cached (prev session)" if is_cached else
                            ("🔴 " + _age_label(age) + " — STALE")))
            price_col = GREEN if age<15 else (AMBER if age<60 else (DIM if is_cached else RED))
            # v10.9: prominent data-age banner
            cm_age_banner = html.Div([
                html.Span("● " if age<15 else "⚠ " if age<60 else "🔴 "),
                html.Span(f"MCX data: {price_label}",style={"fontWeight":"700","color":price_col}),
            ],style={"background":f"rgba(255,166,0,{0.05 if age<60 else 0.15})","border":f"1px solid {price_col}",
                     "borderRadius":"8px","padding":"8px 14px","marginBottom":"12px","fontSize":"13px",
                     "color":price_col,"display":"flex","alignItems":"center"}) if age >= 15 else html.Div()
            stats=[sc("MCX Net P&L",f"₹{net:+,.0f}",GREEN if net>=0 else RED),
                   sc("MCX Trades",str(len(closed)),ACCENT),
                   html.Div([html.Div("MCX STATUS",className="sl"),
                             html.Div(["MCX OPEN" if open_ else "MCX CLOSED"],className="sv",style={"color":s_col})],className="sc"),
                   html.Div([html.Div([html.Span("MCX PRICES",className="sl"),
                                       html.Div(price_label,style={"fontSize":"11px","color":price_col,"marginTop":"2px","fontStyle":"italic"})],
                                      style={"display":"flex","flexDirection":"column","gap":"2px"}),
                             html.Div([html.Span(
                                 [html.Span("●",style={"color":price_col,"fontSize":"9px","marginRight":"4px"}),
                                  f"{s}: ${float(prices[s]):,.2f}"],
                                 className=f"price-chip {'live' if is_live else ''}")
                                 for s in ["GOLD","SILVER","CRUDE","NATURALGAS","COPPER"] if s in prices
                             ] or [html.Span("Waiting for CommodityEngine…",style={"color":AMBER,"fontSize":"12px"})],
                             style={"display":"flex","flexWrap":"wrap","marginTop":"4px"})],className="sc")]
        except Exception: stats=[html.Div("—")]
        # Levels
        try:
            levels=_DS.get("commodity_levels",{})
            if not levels:
                lvl_div=html.Div([
                    html.Div("MCX Trading Levels", className="sec"),
                    html.Div([
                        html.Div("📊 Loading commodity levels…", style={"color":AMBER,"fontWeight":"600","marginBottom":"6px"}),
                        html.Div("Levels are calculated by CommodityEngine when MCX opens (Mon–Fri ~09:00 IST).", style={"color":DIM,"fontSize":"13px"}),
                        html.Div("Previous session levels will be shown if available.", style={"color":DIM,"fontSize":"12px","marginTop":"4px"}),
                    ], style={"padding":"16px"})
                ], className="card")
            else:
                # v10.8: match equity column order — Sell side (left,red) → Current (centre,yellow) → Buy side (right,green)
                hdr_style = {"padding":"8px 6px","fontSize":"11px","fontWeight":"700","borderBottom":f"1px solid {BORDER}","whiteSpace":"nowrap"}
                rows=[html.Tr([
                    html.Th("Sell SL",  style={**hdr_style,"color":RED,"textAlign":"right"}),
                    html.Th("ST2",      style={**hdr_style,"color":RED,  "textAlign":"right","opacity":"0.7"}),
                    html.Th("ST1",      style={**hdr_style,"color":RED,  "textAlign":"right"}),
                    html.Th("Sell ↓",   style={**hdr_style,"color":RED,  "textAlign":"right","fontWeight":"800"}),
                    html.Th("Commodity",style={**hdr_style,"color":COMM_COL,"textAlign":"left","minWidth":"110px"}),
                    html.Th("Current",  style={**hdr_style,"color":YELLOW,"textAlign":"center","minWidth":"110px"}),
                    html.Th("Buy ↑",    style={**hdr_style,"color":GREEN,"textAlign":"left","fontWeight":"800"}),
                    html.Th("T1",       style={**hdr_style,"color":GREEN,"textAlign":"left"}),
                    html.Th("T2",       style={**hdr_style,"color":GREEN,"textAlign":"left","opacity":"0.7"}),
                    html.Th("Buy SL",   style={**hdr_style,"color":RED,"textAlign":"left"}),
                ])]
                for sym in ["GOLD","SILVER","CRUDE","NATURALGAS","COPPER"]:
                    lv=levels.get(sym,{}) if isinstance(levels,dict) else {}
                    if not lv: continue
                    px_val = float(prices.get(sym,0) or 0)
                    in_pos = any(e.get("symbol")==sym and e.get("exit_px") is None for e in evs)
                    row_bg = f"rgba({','.join(str(int(COMM_COL.lstrip('#')[i:i+2],16)) for i in (0,2,4))},0.06)" if in_pos else "transparent"
                    def _p(v, dec=2): return f"${float(v):,.{dec}f}" if v else "—"
                    sell_sl  = lv.get("prev_close", lv.get("sell_sl", 0))
                    buy_sl   = lv.get("prev_close", lv.get("buy_sl",  0))
                    # near-level highlight
                    def _near_c(level, base_col):
                        return YELLOW if px_val>0 and level>0 and abs(px_val-level)/level<0.003 else base_col
                    rows.append(html.Tr([
                        html.Td(_p(sell_sl),           style={"color":RED, "fontWeight":"700","padding":"8px 7px","textAlign":"right","fontSize":"12px"}),
                        html.Td(_p(lv.get("ST2",0)),   style={"color":RED,   "padding":"8px 7px","textAlign":"right","fontSize":"11px","opacity":"0.7"}),
                        html.Td(_p(lv.get("ST1",0)),   style={"color":RED,   "padding":"8px 7px","textAlign":"right","fontSize":"12px"}),
                        html.Td(_p(lv.get("sell_below",0)), style={"color":_near_c(lv.get("sell_below",0),RED),"fontWeight":"800","padding":"8px 8px","textAlign":"right","fontSize":"13px","borderRight":f"2px solid {BORDER}"}),
                        # Centre: commodity name
                        html.Td([html.Span("● " if in_pos else "",style={"color":GREEN,"fontSize":"9px"}),sym],
                                style={"fontWeight":"700","color":COMM_COL,"padding":"8px 9px","textAlign":"left","borderRight":f"2px solid {YELLOW}","borderLeft":f"2px solid {YELLOW}"}),
                        # Centre: current price
                        html.Td(f"${px_val:,.2f}" if px_val else "—",
                                style={"color":YELLOW,"fontWeight":"700","padding":"8px 8px","textAlign":"center",
                                       "background":"rgba(227,179,65,0.08)","fontSize":"13px","borderRight":f"2px solid {BORDER}"}),
                        # Buy side
                        html.Td(_p(lv.get("buy_above",0)), style={"color":_near_c(lv.get("buy_above",0),GREEN),"fontWeight":"800","padding":"8px 8px","textAlign":"left","fontSize":"13px"}),
                        html.Td(_p(lv.get("T1",0)),    style={"color":GREEN,"padding":"8px 7px","textAlign":"left","fontSize":"12px"}),
                        html.Td(_p(lv.get("T2",0)),    style={"color":GREEN,"padding":"8px 7px","textAlign":"left","fontSize":"11px","opacity":"0.7"}),
                        html.Td(_p(buy_sl),            style={"color":RED,"fontWeight":"700","padding":"8px 7px","textAlign":"left","fontSize":"12px"}),
                    ], style={"borderBottom":f"1px solid {BORDER}","background":row_bg}))
                lvl_div=html.Div([
                    html.Div([
                        html.Span("MCX Trading Levels"),
                        html.Span(" · Sell Below  →  Current  →  Buy Above",
                                  style={"color":DIM,"fontSize":"10px","marginLeft":"12px"}),
                    ], className="sec"),
                    html.Div(html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"}),className="tw")
                ], className="card")
        except Exception as e: lvl_div=html.Div(f"Error: {e}",className="err")
        # Positions
        try:
            ops=[e for e in evs if e.get("exit_px") is None and e.get("entry_px")]
            if not ops:
                pos_div=html.Div([html.Div("Open MCX Positions",className="sec"),
                                  html.Div("No open positions",style={"color":DIM,"fontSize":"13px"})],className="card")
            else:
                rows=[html.Tr([html.Th(h,style={"padding":"7px 8px","color":DIM,"fontSize":"11px"}) for h in ["Symbol","Side","Entry","Current","Unrealized","Time"]])]
                for p in ops:
                    sym=p.get("symbol","—"); side=p.get("side","—")
                    ep=float(p.get("entry_px",0) or 0); px=float(prices.get(sym,0) or ep)
                    qty=int(p.get("qty",1) or 1); unr=(px-ep)*qty if side=="BUY" else (ep-px)*qty
                    rows.append(html.Tr([
                        html.Td(sym,style={"fontWeight":"700","color":COMM_COL}),
                        html.Td(side,style={"color":GREEN if side=="BUY" else RED,"fontWeight":"600"}),
                        html.Td(f"${ep:,.2f}"),html.Td(f"${px:,.2f}",style={"color":ACCENT}),
                        html.Td(f"₹{unr*_usdt_to_inr():+,.0f}",style={"color":GREEN if unr>=0 else RED,"fontWeight":"700"}),
                        html.Td(p.get("ts","—"),style={"fontSize":"11px","color":DIM})],
                        style={"borderBottom":f"1px solid {BORDER}"}))
                pos_div=html.Div([html.Div(f"Open MCX Positions ({len(ops)})",className="sec"),
                                  html.Div(html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"}),className="tw")],className="card")
        except Exception as e: pos_div=html.Div(f"Error: {e}",className="err")
        # Trades
        try:
            cl=[e for e in evs if e.get("exit_px") is not None][-25:]
            if not cl:
                trades_div=html.Div([html.Div("Completed MCX Trades",className="sec"),
                                     html.Div("No completed trades today",style={"color":DIM})],className="card")
            else:
                EXC={"SL_HIT":"🛑","RETREAT":"↩️","EOD_2330":"🌙"}
                rows=[html.Tr([html.Th(h,style={"padding":"7px 6px","color":DIM,"fontSize":"11px"}) for h in ["Time","Symbol","Side","Entry","Exit","Net P&L","Reason"]])]
                for e in reversed(cl):
                    net=float(e.get("net_pnl",0) or 0); reason=e.get("reason","—")
                    em=EXC.get(reason,"✅") if "HIT" in reason else EXC.get(reason,"📤")
                    rows.append(html.Tr([
                        html.Td(e.get("ts","—"),style={"fontSize":"11px","color":DIM}),
                        html.Td(e.get("symbol","—"),style={"fontWeight":"700","color":COMM_COL}),
                        html.Td(e.get("side","—"),style={"color":GREEN if e.get("side")=="BUY" else RED}),
                        html.Td(f"${float(e.get('entry_px',0)):,.2f}"),
                        html.Td(f"${float(e.get('exit_px',0)):,.2f}"),
                        html.Td(f"₹{net:+,.0f}",style={"color":GREEN if net>=0 else RED,"fontWeight":"700"}),
                        html.Td(f"{em} {reason}",style={"fontSize":"11px"})],
                        style={"borderBottom":f"1px solid {BORDER}"}))
                trades_div=html.Div([html.Div(f"Completed MCX Trades ({len(cl)})",className="sec"),
                                     html.Div(html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"}),className="tw")],className="card")
        except Exception: trades_div=html.Div("Error",className="err")
        return stats,lvl_div,pos_div,trades_div

    # ── Commodity Scanners ───────────────────────────────────────────────────
    def _cs_stats(sid):
        try:
            state=_DS.get("comm_scanner",{}).get(sid)
            age_raw=_DS.get("comm_scanner_age",{}).get(sid,9999)
            # age can be a list or float — normalise
            age=float(age_raw[0] if isinstance(age_raw,list) else age_raw or 9999)
            if not state:
                import datetime as _dtt
                wd = _dtt.datetime.now().weekday()
                if wd >= 5:
                    msg = f"📅 Weekend — MCX Commodity Scanner {sid} resumes Monday ~09:00 IST"
                else:
                    msg = f"⏳ MCX Scanner {sid} starts when CommodityEngine detects market open (~09:00 IST)"
                return [html.Div(msg, style={"color": AMBER, "padding": "14px", "fontSize": "13px"})]
            merged=state.get("merged_best",state.get("sweeps",{}))
            if not merged: return [html.Div("No sweep data yet — accumulating…",style={"color":DIM})]
            def _s(v, d=0):
                if isinstance(v, list): return float(v[0]) if v else d
                try: return float(v) if v is not None else d
                except: return d
            # Find best symbol safely
            best_sym = max(merged, key=lambda k: _s(merged[k].get("pnl",merged[k].get("best_pnl",0))), default=None)
            if best_sym and best_sym in merged:
                bd=merged[best_sym]
                bp=_s(bd.get("pnl",bd.get("best_pnl",0)))
                bx=_s(bd.get("best_x",0))
            else:
                best_sym="—"; bp=0.0; bx=0.0
            ac=GREEN if age<60 else(AMBER if age<300 else RED)
            return [sc("BEST X",f"{bx:.6f}" if bx else "—",COMM_COL),
                    sc(f"BEST P&L ({best_sym})",f"₹{(bp*_usdt_to_inr()):+,.0f}",GREEN if bp>=0 else RED),
                    sc("DATA AGE",f"{age:.0f}s",ac),
                    sc("SYMS",str(len(merged)),ACCENT)]
        except Exception as e: return [html.Div(f"Comm Scanner {sid}: {e}",style={"color":RED})]

    def _cs_tbl(sid):
        try:
            state=_DS.get("comm_scanner",{}).get(sid)
            if not state:
                return html.Div([
                    html.Div("📊 MCX Commodity Scanner", style={"fontWeight":"700","marginBottom":"8px","color":COMM_COL}),
                    html.Div("Scans GOLD, SILVER, CRUDE, NATURALGAS, COPPER for optimal entry levels.", style={"color":DIM,"fontSize":"13px","marginBottom":"6px"}),
                    html.Div("Data appears automatically once MCX opens (Mon–Fri 09:00–23:30 IST).", style={"color":DIM,"fontSize":"13px"}),
                ], style={"background":CARD,"borderRadius":"8px","padding":"20px","border":f"1px solid {BORDER}"})
            merged=state.get("merged_best",state.get("sweeps",{}))
            if not merged: return html.Div("No sweep data yet",style={"color":DIM,"padding":"12px"})

            def _scalar(v, default=0):
                if isinstance(v, list): return float(v[0]) if v else default
                try: return float(v) if v is not None else default
                except: return default

            sym_cards=[]
            table_rows=[]
            for sym,d in sorted(merged.items(), key=lambda x: _scalar(x[1].get("pnl",x[1].get("best_pnl",0))), reverse=True):
                if not isinstance(d,dict): continue
                pnl   = _scalar(d.get("pnl",d.get("best_pnl",0)))
                bx    = _scalar(d.get("best_x",0))
                wr    = _scalar(d.get("win_rate",d.get("best_win_rate",0)))
                tc    = int(_scalar(d.get("trade_count",d.get("best_trade_count",0))))
                xv    = d.get("x_values",[])
                tc_arr= d.get("trade_count",[]) if isinstance(d.get("trade_count",[]),list) else []
                n_total = len(xv) if isinstance(xv,list) else 0
                n_tested= sum(1 for v in tc_arr if isinstance(v,(int,float)) and v>0) if tc_arr else (1 if tc>0 else 0)
                tcv = d.get("tick_count", 0)
                has_ticks = (isinstance(tcv, list) and any(isinstance(v, (int, float)) and v > 0 for v in tcv)) or (
                    isinstance(tcv, (int, float)) and tcv > 0
                )
                if n_total > 0 and n_tested == 0 and has_ticks:
                    n_tested = n_total
                pct   = n_tested/n_total if n_total>0 else (1.0 if tc>0 else 0.0)
                pnl_c = GREEN if pnl>=0 else RED
                wr_c  = GREEN if wr>=55 else (AMBER if wr>=40 else RED)

                sym_cards.append(html.Div([
                    html.Div([
                        html.Span(sym, style={"fontWeight":"700","color":COMM_COL,"fontSize":"14px","minWidth":"110px"}),
                        html.Span(f"Best X: {bx:.6f}" if bx else "Best X: sweeping…",
                                  style={"color":GREEN if bx else DIM,"fontSize":"12px","fontFamily":"monospace","marginLeft":"8px"}),
                    ],style={"display":"flex","alignItems":"center","flexWrap":"wrap"}),
                    html.Div([
                        html.Span("P&L: ",style={"color":DIM,"fontSize":"11px"}),
                        html.Span(f"₹{(pnl*_usdt_to_inr()):+,.0f}",style={"color":pnl_c,"fontWeight":"700","fontSize":"13px","marginRight":"16px"}),
                        html.Span(f"Trades: {tc}",style={"color":TEXT,"fontSize":"11px","marginRight":"16px"}),
                        html.Span(f"Win: {wr:.0f}%",style={"color":wr_c,"fontSize":"11px","marginRight":"16px"}),
                        html.Span(f"Tested: {n_tested}/{n_total} X values" if n_total>0 else "Accumulating data…",
                                  style={"color":DIM,"fontSize":"10px"}),
                    ],style={"marginTop":"5px","display":"flex","alignItems":"center","flexWrap":"wrap"}),
                    html.Div([
                        html.Span("Sweep",style={"color":DIM,"fontSize":"9px","marginRight":"6px","minWidth":"38px"}),
                        html.Div([html.Div(style={"width":f"{pct*100:.0f}%","height":"5px",
                                                  "background":COMM_COL if pct<1.0 else GREEN,
                                                  "borderRadius":"3px","transition":"width 0.4s ease"})],
                                 style={"flex":"1","height":"5px","background":BORDER,"borderRadius":"3px"}),
                        html.Span(f"{pct*100:.0f}%",style={"color":DIM,"fontSize":"9px","marginLeft":"6px"}),
                    ],style={"display":"flex","alignItems":"center","marginTop":"6px","gap":"4px"}),
                ],style={"background":CARD,"border":f"1px solid {BORDER}","borderRadius":"8px",
                         "padding":"10px 14px","marginBottom":"8px","borderLeft":f"3px solid {pnl_c}"}))
                table_rows.append({"Symbol":sym,"Best X":f"{bx:.6f}" if bx else "—",
                                   "P&L (₹)":f"₹{(pnl*_usdt_to_inr()):+,.0f}","Win%":f"{wr:.1f}%","Trades":str(tc)})

            return html.Div([
                html.Div("Per-Commodity Sweep Results",className="sec"),
                html.Div(sym_cards),
                html.Div([
                    html.Div("Full Table",className="sec",style={"marginTop":"14px"}),
                    html.Div(html.Table(
                        [html.Tr([html.Th(h,style={"padding":"7px 10px","color":DIM,"fontSize":"11px","fontWeight":"700","borderBottom":f"1px solid {BORDER}","textAlign":"left"}) for h in ["Symbol","Best X","P&L (₹)","Win%","Trades"]])] +
                        [html.Tr([html.Td(r[h],style={"padding":"6px 10px","color":COMM_COL if h=="Symbol" else (GREEN if "+" in str(r.get("P&L (₹)","")) else (RED if "-" in str(r.get("P&L (₹)","")) else TEXT)),"fontWeight":"700" if h in ("Symbol","P&L (₹)") else "400","borderBottom":f"1px solid {BORDER}"}) for h in ["Symbol","Best X","P&L (₹)","Win%","Trades"]]) for r in table_rows],
                        style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"}),className="tw")
                ],className="card"),
            ])
        except Exception as e: return html.Div(f"Error: {e}",style={"color":RED})

    for _sid in (1,2,3):
        (lambda sid: (
            app.callback(Output(f"cs{sid}-s","children"),Input("t10","n_intervals"))(lambda _,s=sid: _cs_stats(s)),
            app.callback(Output(f"cs{sid}-t","children"),Input("t10","n_intervals"))(lambda _,s=sid: _cs_tbl(s)),
        ))(_sid)

    @app.callback(Output("co-c","children"),Input("t10","n_intervals"),prevent_initial_call=True)
    def _copt(_):
        try:
            syms=["GOLD","SILVER","CRUDE","NATURALGAS","COPPER"]
            rows=[html.Tr([html.Th(h,style={"padding":"6px"}) for h in ["Symbol","Best X","Source","P&L (₹)","Win%","Trades"]])]
            for sym in syms:
                best=None
                for sid,lbl in [(1,"CS1"),(2,"CS2"),(3,"CS3")]:
                    state=_DS.get("comm_scanner",{}).get(sid)
                    if not state: continue
                    d=state.get("merged_best",state.get("sweeps",{})).get(sym,{})
                    if isinstance(d,dict) and d.get("best_x",0):
                        pnl=float(d.get("pnl",d.get("best_pnl",-1e9)) or -1e9)
                        if best is None or pnl>best["pnl"]: best={"x":d["best_x"],"pnl":pnl,"src":lbl,"wr":d.get("win_rate",d.get("best_win_rate",0)),"tc":d.get("trade_count",d.get("best_trade_count",0))}
                if best:
                    rows.append(html.Tr([html.Td(sym,style={"fontWeight":"700","color":COMM_COL}),html.Td(f"{best['x']:.6f}",style={"color":GREEN}),
                        html.Td(best["src"],style={"color":YELLOW}),html.Td(f"₹{(best['pnl']*_usdt_to_inr()):+,.0f}",style={"color":GREEN if best["pnl"]>=0 else RED}),
                        html.Td(f"{float(best['wr'] or 0):.1f}%"),html.Td(str(best["tc"] or 0))],style={"borderBottom":f"1px solid {BORDER}"}))
                else:
                    rows.append(html.Tr([html.Td(sym,style={"color":DIM})]+[html.Td("—")]*5))
            return html.Div([html.Div("MCX Optimizer — Best X per commodity",className="sec"),html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"})],className="card")
        except Exception as e: return html.Div(f"Error: {e}",style={"color":RED})

    @app.callback(Output("cbx-c","children"),Input("t2","n_intervals"),prevent_initial_call=True)
    def _cbx(_):
        try:
            evs=_DS.get("commodity_events",[]); cl=[e for e in evs if e.get("exit_px") is not None]
            net=sum(float(e.get("net_pnl",0) or 0) for e in cl)
            rows=[html.Tr([html.Th(h,style={"padding":"6px"}) for h in ["Symbol","Side","Entry","Exit","Net P&L","Reason","Time"]])]
            for e in reversed(cl[-15:]):
                n=float(e.get("net_pnl",0) or 0)
                rows.append(html.Tr([html.Td(e.get("symbol","—"),style={"fontWeight":"700"}),html.Td(e.get("side","—"),style={"color":GREEN if e.get("side")=="BUY" else RED}),
                    html.Td(f"${float(e.get('entry_px',0)):,.2f}"),html.Td(f"${float(e.get('exit_px',0)):,.2f}"),
                    html.Td(f"₹{n:+,.0f}",style={"color":GREEN if n>=0 else RED,"fontWeight":"600"}),html.Td(e.get("reason","—")),html.Td(e.get("ts","—"))],style={"borderBottom":f"1px solid {BORDER}"}))
            return html.Div([html.Div([html.Span(f"MCX Trades: {len(cl)}  |  Net: "),html.Span(f"₹{net:+,.0f}",style={"color":GREEN if net>=0 else RED,"fontWeight":"700"})],style={"marginBottom":"12px","fontSize":"14px"}),html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"})])
        except Exception as e: return html.Div(f"Error: {e}",style={"color":RED})

    # Commodity History & Performance  (FIX 5)
    @app.callback(Output("hcm-x","children"),Output("hcm-c","children"),Input("t60","n_intervals"),prevent_initial_call=True)
    def _hist_cm(_):
        try:
            h=_DS.get("comm_history",{})
            x=html.Div([html.Div("MCX Commodity History",className="sl"),html.Div(f"{len(h)} trading days",style={"fontSize":"20px","fontWeight":"700","color":COMM_COL})],className="card")
            return x,hist_cards(h)
        except Exception as e: return html.Div(),html.Div(f"Error: {e}",className="err")

    @app.callback(Output("hcm-dl","data"),Input("hcm-btn","n_clicks"),prevent_initial_call=True)
    def _hist_cm_dl(_):
        import io,csv as _c; buf=io.StringIO(); w=_c.writer(buf); w.writerow(["Date","Trades","NetPnL","Return%","WinRate%"])
        for ds,d in sorted(_DS.get("comm_history",{}).items(),reverse=True): w.writerow([ds,d["total_trades"],d["net_pnl"],d["pct_return"],d["win_rate"]])
        return dict(content=buf.getvalue(),filename="algostack_commodity_history.csv")

    @app.callback(Output("pcm1","children"),Output("pcm2","children"),Output("pcm3","children"),Output("pcm4","children"),Input("t60","n_intervals"),prevent_initial_call=True)
    def _perf_cm(_):
        try: r=perf_charts(_DS.get("comm_history",{}),COMM_COL,0.30); return r if len(r)==4 else (r[0],)*4
        except Exception as e: err=html.Div(f"Error: {e}",className="err"); return err,err,err,err

    @app.callback(Output("pcm-dl","data"),Input("pcm-btn","n_clicks"),prevent_initial_call=True)
    def _perf_cm_dl(_):
        import io,csv as _c; buf=io.StringIO(); w=_c.writer(buf); w.writerow(["Date","Trades","NetPnL","Return%","WinRate%"])
        for ds,d in sorted(_DS.get("comm_history",{}).items(),reverse=True): w.writerow([ds,d["total_trades"],d["net_pnl"],d["pct_return"],d["win_rate"]])
        return dict(content=buf.getvalue(),filename="algostack_commodity_performance.csv")

    # ══════════════════════════════════════════════════════════════════════════
    #  CRYPTO PAGE
    # ══════════════════════════════════════════════════════════════════════════
    # ── CONSOLIDATED: all crypto page outputs in one callback ─────────────────
    @app.callback(
        Output("cr-s","children"), Output("cr-anchor","children"),
        Output("cr-lvl","children"), Output("cr-pos","children"),
        Output("cr-trades","children"),
        Input("t2","n_intervals")
    )
    def _cr_all(_):
        prices=_DS.get("crypto_prices",{}); evs=_DS.get("crypto_events",[])
        age=float(_DS.get("crypto_age",9999) or 9999); lv=_DS.get("crypto_levels",{})
        # Stats
        try:
            cl=[e for e in evs if e.get("exit_px") is not None]
            net=sum(float(e.get("net_pnl_inr",e.get("net_pnl",0)) or 0) for e in cl)
            age_src = ("🟢 Binance WS — just now" if age<3 else
                       ("🟢 " + _age_label(age)) if age<10 else
                       ("🟡 " + _age_label(age)) if age<30 else
                       ("🔴 " + _age_label(age) + " — STALE"))
            age_col_cr = GREEN if age<10 else (AMBER if age<30 else RED)
            chips=[html.Span(
                [html.Span("●",style={"color":age_col_cr,"fontSize":"9px","marginRight":"4px"}),
                 f"{s}: ${float(prices[s]):,.2f}  ₹{float(prices[s])*_usdt_to_inr():,.0f}"],
                className="price-chip") for s in ["BTC","ETH","BNB","SOL","ADA"] if s in prices]
            # v10.9: prominent crypto data-age banner
            cr_age_banner = html.Div([
                html.Span("● " if age<10 else "⚠ " if age<30 else "🔴 "),
                html.Span(f"Binance WS: {age_src}",style={"fontWeight":"700","color":age_col_cr}),
            ],style={"background":f"rgba(88,166,255,{0.05 if age<10 else 0.15})","border":f"1px solid {age_col_cr}",
                     "borderRadius":"8px","padding":"8px 14px","marginBottom":"12px","fontSize":"13px",
                     "color":age_col_cr,"display":"flex","alignItems":"center"}) if age >= 10 else html.Div()
            stats=[sc("Crypto Net P&L",f"₹{net:+,.0f}",GREEN if net>=0 else RED),
                   sc("Crypto Trades",str(len(cl)),ACCENT),
                   html.Div([
                       html.Div("CRYPTO PRICES",className="sl"),
                       html.Div([
                           html.Div(age_src,style={"color":age_col_cr,"fontSize":"11px","fontStyle":"italic","marginTop":"2px"}),
                       ],className="sv"),
                   ],className="sc"),
                   html.Div([html.Div("LIVE PRICES",className="sl"),
                             html.Div(chips or [html.Span("Waiting for Binance…",style={"color":AMBER,"fontSize":"12px"})],
                                      style={"display":"flex","flexWrap":"wrap","marginTop":"4px"})],className="sc")]
        except Exception: stats=[html.Div("—")]
        # Anchor info
        try:
            at=_DS.get("crypto_anchor_time","")
            btc=lv.get("BTC",{}).get("anchor",0) if isinstance(lv,dict) else 0
            anchor_str=f"Anchor @ {at[:16] if at else '—'} UTC  |  BTC=${btc:,.2f}  |  X={CRYPTO_X:.6f}  |  Re-anchors every 6h"
        except: anchor_str="Crypto anchor unavailable"
        # Levels table
        try:
            if not lv:
                lvl_div=html.Div([
                    html.Div("⏳ Waiting for CryptoEngine to write levels file…",
                             style={"color":AMBER,"padding":"12px","fontSize":"13px"}),
                    html.Div("crypto_initial_levels_latest.json not found yet",
                             style={"color":DIM,"padding":"0 12px 12px","fontSize":"11px"})
                ])
            else:
                # v10.8: match equity/commodity column order — Sell side → Current → Buy side
                hdr_c = {"padding":"7px 6px","fontSize":"11px","fontWeight":"700","borderBottom":f"1px solid {BORDER}","whiteSpace":"nowrap"}
                rows=[html.Tr([
                    html.Th("Sell SL($)",  style={**hdr_c,"color":RED,         "textAlign":"right"}),
                    html.Th("ST2",         style={**hdr_c,"color":RED,         "textAlign":"right","opacity":"0.7"}),
                    html.Th("ST1",         style={**hdr_c,"color":RED,         "textAlign":"right"}),
                    html.Th("Sell ↓($)",   style={**hdr_c,"color":RED,         "textAlign":"right","fontWeight":"800"}),
                    html.Th("Coin",        style={**hdr_c,"color":CRYPTO_COL,  "textAlign":"left","minWidth":"70px"}),
                    html.Th("Current",     style={**hdr_c,"color":YELLOW,      "textAlign":"center","minWidth":"105px"}),
                    html.Th("Buy ↑($)",    style={**hdr_c,"color":GREEN,       "textAlign":"left","fontWeight":"800"}),
                    html.Th("T1",          style={**hdr_c,"color":GREEN,       "textAlign":"left"}),
                    html.Th("T2",          style={**hdr_c,"color":GREEN,       "textAlign":"left","opacity":"0.7"}),
                    html.Th("Buy SL($)",   style={**hdr_c,"color":RED,         "textAlign":"left"}),
                    html.Th("T1(₹)",       style={**hdr_c,"color":GREEN,       "textAlign":"left","fontSize":"10px"}),
                ])]
                for sym in ["BTC","ETH","BNB","SOL","ADA"]:
                    d=lv.get(sym,{}) if isinstance(lv,dict) else {}
                    if not d: continue
                    px_v=float(prices.get(sym,0) or 0)
                    in_pos=any(e.get("symbol")==sym and e.get("exit_px") is None for e in evs)
                    row_bg="rgba(163,113,247,0.08)" if in_pos else "transparent"
                    sl_usd  = d.get("stop_loss", d.get("buy_sl", d.get("anchor",0)))
                    t1_usd  = d.get("T1",0)
                    t1_inr  = d.get("T1_inr", round(t1_usd * USDT_TO_INR, 0) if t1_usd else 0)
                    ba_usd  = d.get("buy_above",0)
                    sb_usd  = d.get("sell_below",0)
                    def _nearp(level):
                        return YELLOW if px_v>0 and level>0 and abs(px_v-level)/level<0.005 else None
                    def _fc(v, dec=4): return f"${float(v):,.{dec}f}" if v else "—"
                    rows.append(html.Tr([
                        html.Td(_fc(sl_usd), style={"color":RED,"fontWeight":"700","padding":"7px 6px","textAlign":"right","fontSize":"11px"}),
                        html.Td(_fc(d.get("ST2",0)), style={"color":RED,"padding":"7px 6px","textAlign":"right","fontSize":"10px","opacity":"0.7"}),
                        html.Td(_fc(d.get("ST1",0)), style={"color":RED,"padding":"7px 6px","textAlign":"right","fontSize":"11px"}),
                        html.Td(_fc(sb_usd), style={"color":_nearp(sb_usd) or RED,"fontWeight":"800","padding":"7px 7px","textAlign":"right","fontSize":"12px","borderRight":f"2px solid {BORDER}"}),
                        # Centre: coin name
                        html.Td([html.Span("●",style={"color":GREEN if in_pos else DIM,"fontSize":"9px","marginRight":"4px"}),sym],
                                style={"fontWeight":"700","color":CRYPTO_COL,"padding":"7px 8px","textAlign":"left","borderRight":f"2px solid {YELLOW}","borderLeft":f"2px solid {YELLOW}"}),
                        # Centre: current price
                        html.Td(f"${px_v:,.4f}" if px_v else "—",
                                style={"color":YELLOW,"fontWeight":"700","padding":"7px 7px","textAlign":"center",
                                       "background":"rgba(227,179,65,0.08)","fontSize":"12px","borderRight":f"2px solid {BORDER}"}),
                        # Buy side
                        html.Td(_fc(ba_usd), style={"color":_nearp(ba_usd) or GREEN,"fontWeight":"800","padding":"7px 7px","textAlign":"left","fontSize":"12px"}),
                        html.Td(_fc(t1_usd), style={"color":GREEN,"padding":"7px 6px","textAlign":"left","fontSize":"11px"}),
                        html.Td(_fc(d.get("T2",0)), style={"color":GREEN,"padding":"7px 6px","textAlign":"left","fontSize":"10px","opacity":"0.7"}),
                        html.Td(_fc(sl_usd), style={"color":RED,"fontWeight":"700","padding":"7px 6px","textAlign":"left","fontSize":"11px"}),
                        html.Td(f"₹{t1_inr:,.0f}" if t1_inr else "—", style={"color":GREEN,"padding":"7px 6px","textAlign":"left","fontSize":"11px","fontWeight":"600"}),
                    ], style={"borderBottom":f"1px solid {BORDER}","background":row_bg}))
                lvl_div=html.Div([
                    html.Div([
                        html.Span("Crypto Anchor Levels"),
                        html.Span(" · Sell Below  →  Current  →  Buy Above",
                                  style={"color":DIM,"fontSize":"10px","marginLeft":"12px"}),
                    ], className="sec"),
                    html.Div(html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"12px"}),className="tw")
                ], className="card")
        except Exception as e: lvl_div=html.Div(f"Error: {e}",className="err")
        # Positions
        try:
            ops=[e for e in evs if e.get("exit_px") is None and e.get("entry_px")]
            if not ops:
                pos_div=html.Div([html.Div("Open Crypto Positions",className="sec"),
                                  html.Div("No open positions",style={"color":DIM,"fontSize":"13px"})],className="card")
            else:
                rows=[html.Tr([html.Th(h,style={"padding":"7px 6px","color":DIM,"fontSize":"11px"}) for h in ["Coin","Side","Entry($)","Current($)","Unrealized(₹)","Qty","Time"]])]
                for p in ops:
                    sym=p.get("symbol","—"); side=p.get("side","—")
                    ep=float(p.get("entry_px",0) or 0); qty=float(p.get("qty",0) or 0)
                    px=float(prices.get(sym,0) or ep)
                    unr=((px-ep)*qty if side=="BUY" else (ep-px)*qty)*_usdt_to_inr()
                    rows.append(html.Tr([
                        html.Td(sym,style={"fontWeight":"700","color":CRYPTO_COL}),
                        html.Td(side,style={"color":GREEN if side=="BUY" else RED,"fontWeight":"600"}),
                        html.Td(f"${ep:,.4f}"),html.Td(f"${px:,.4f}",style={"color":ACCENT}),
                        html.Td(f"₹{unr:+,.0f}",style={"color":GREEN if unr>=0 else RED,"fontWeight":"700"}),
                        html.Td(f"{qty:.6f}",style={"fontSize":"11px","color":DIM}),
                        html.Td(p.get("ts","—"),style={"fontSize":"11px","color":DIM})],
                        style={"borderBottom":f"1px solid {BORDER}"}))
                pos_div=html.Div([html.Div(f"Open Crypto Positions ({len(ops)})",className="sec"),
                                  html.Div(html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"}),className="tw")],className="card")
        except Exception as e: pos_div=html.Div(f"Error: {e}",className="err")
        # Trades
        try:
            cl2=[e for e in evs if e.get("exit_px") is not None][-25:]
            if not cl2:
                trades_div=html.Div([html.Div("Crypto Trades",className="sec"),
                                     html.Div("No closed trades yet",style={"color":DIM})],className="card")
            else:
                EXC={"SL_HIT":"🛑","RETREAT":"↩️","RE_ANCHOR":"🔄"}
                rows=[html.Tr([html.Th(h,style={"padding":"7px 6px","color":DIM,"fontSize":"11px"}) for h in ["Time","Coin","Side","Entry($)","Exit($)","Net(₹)","Reason"]])]
                for e in reversed(cl2):
                    net=float(e.get("net_pnl_inr",e.get("net_pnl",0)) or 0)
                    reason=e.get("reason","—")
                    em=EXC.get(reason,"✅") if any(x in reason for x in ["HIT","ENTRY"]) else EXC.get(reason,"📤")
                    rows.append(html.Tr([
                        html.Td(e.get("ts","—"),style={"fontSize":"11px","color":DIM}),
                        html.Td(e.get("symbol","—"),style={"fontWeight":"700","color":CRYPTO_COL}),
                        html.Td(e.get("side","—"),style={"color":GREEN if e.get("side")=="BUY" else RED}),
                        html.Td(f"${float(e.get('entry_px',0)):,.4f}"),
                        html.Td(f"${float(e.get('exit_px',0)):,.4f}"),
                        html.Td(f"₹{net:+,.0f}",style={"color":GREEN if net>=0 else RED,"fontWeight":"700"}),
                        html.Td(f"{em} {reason}",style={"fontSize":"11px"})],
                        style={"borderBottom":f"1px solid {BORDER}"}))
                trades_div=html.Div([html.Div(f"Crypto Trades ({len(cl2)})",className="sec"),
                                     html.Div(html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"12px"}),className="tw")],className="card")
        except Exception: trades_div=html.Div("Error",className="err")
        return stats,anchor_str,lvl_div,pos_div,trades_div

    # ── Crypto Scanners ──────────────────────────────────────────────────────
    def _cr_stat(sid):
        try:
            state=_DS.get("crypto_scanner",{}).get(sid); age=_DS.get("crypto_scanner_age",{}).get(sid,9999)
            if not state:
                return [html.Div(
                    f"🔄 Crypto Scanner {sid} — 24/7 scanning. First results appear after a 6-hour window completes (~wait up to 6h).",
                    style={"color":DIM,"padding":"12px","fontSize":"13px"}
                )]
            merged=state.get("merged_best",state.get("sweeps",{}))
            if not merged: return [html.Div("Waiting for sweep data...",style={"color":DIM})]
            def _s(v, d=0):
                if isinstance(v, list): return float(v[0]) if v else d
                try: return float(v) if v is not None else d
                except: return d
            def _best_pnl(d2): return _s(d2.get("best_pnl",d2.get("pnl",0)))
            best_sym=max(merged,key=lambda k:_best_pnl(merged[k])*_usdt_to_inr(),default="—")
            bd=merged.get(best_sym,{}); bpi=_best_pnl(bd)*_usdt_to_inr()
            bx=_s(bd.get("best_x",0)); win=state.get("date_tag","—")
            ac=GREEN if float(age)<60 else(AMBER if float(age)<300 else RED)
            return [sc("BEST X",f"{bx:.6f}" if bx else "—",CRYPTO_COL),
                    sc(f"BEST P&L ({best_sym})",f"₹{bpi:+,.0f}",GREEN if bpi>=0 else RED),
                    sc("6H WINDOW",win or "—",ACCENT),
                    sc("DATA AGE",f"{float(age):.0f}s",ac)]
        except Exception as e: return [html.Div(f"Cr-Scanner {sid}: {e}",style={"color":RED})]
    def _cr_tbl(sid):
        try:
            state=_DS.get("crypto_scanner",{}).get(sid)
            if not state:
                return html.Div([
                    html.Div("₿ Crypto Scanner (24/7)", style={"fontWeight":"700","marginBottom":"8px","color":CRYPTO_COL}),
                    html.Div("Scans BTC, ETH, BNB, SOL, ADA across 6-hour rolling windows.", style={"color":DIM,"fontSize":"13px","marginBottom":"6px"}),
                    html.Div("Results appear after the first 6-hour sweep completes. No action needed.", style={"color":DIM,"fontSize":"13px"}),
                ], style={"background":CARD,"borderRadius":"8px","padding":"20px","border":f"1px solid {BORDER}"})

            merged=state.get("merged_best",state.get("sweeps",{}))
            if not merged: return html.Div("Sweep in progress — no completed windows yet",style={"color":DIM,"padding":"12px"})

            def _sc(v, default=0):
                if isinstance(v, list): return float(v[0]) if v else default
                try: return float(v) if v is not None else default
                except: return default

            def _best_from_sweep(d):
                if not d.get("has_trades", True): return None
                xv=d.get("x_values",[]); pnl=d.get("total_pnl",[]); tc=d.get("trade_count",[]); wc=d.get("win_count",[])
                if not (xv and pnl and isinstance(pnl,list)): return None
                n=len(xv)
                has_any=any(tc[i]>0 for i in range(min(n,len(tc)))) if isinstance(tc,list) else False
                if not has_any: return None
                import numpy as _np2
                pnl_a=_np2.array(pnl[:n],dtype=float)
                tc_a=_np2.array(tc[:n],dtype=float) if isinstance(tc,list) else _np2.zeros(n)
                wc_a=_np2.array(wc[:n],dtype=float) if isinstance(wc,list) else _np2.zeros(n)
                wr_a=_np2.where(tc_a>0,wc_a/tc_a,0.0)
                ar_a=_np2.where(tc_a>0,pnl_a/tc_a,0.0)
                pr=pnl_a.max()-pnl_a.min()
                pn=(pnl_a-pnl_a.min())/pr if pr>1e-9 else _np2.zeros(n)
                ar_r=ar_a.max()-ar_a.min()
                an=(ar_a-ar_a.min())/ar_r if ar_r>1e-9 else _np2.zeros(n)
                bi=int(_np2.argmax(0.50*pn+0.30*wr_a+0.20*an))
                tc_bi=int(tc_a[bi]); wc_bi=int(wc_a[bi])
                return {"best_x":float(xv[bi]),"best_pnl":float(pnl_a[bi]),
                        "trade_count":tc_bi,"win_rate":round(wc_bi/tc_bi*100,1) if tc_bi>0 else 0,
                        "n_tested":sum(1 for v in tc if isinstance(v,(int,float)) and v>0) if isinstance(tc,list) else 0,
                        "n_total":n}

            sym_cards=[]; table_rows=[]
            for sym,d in sorted(merged.items(), key=lambda x: _sc(x[1].get("best_pnl",x[1].get("pnl",0)))*_usdt_to_inr(), reverse=True):
                if not isinstance(d,dict): continue
                best=_best_from_sweep(d)
                if best:
                    bx=best["best_x"]; pu=best["best_pnl"]; tc=best["trade_count"]
                    wr=best["win_rate"]; n_tested=best["n_tested"]; n_total=best["n_total"]
                else:
                    bx=_sc(d.get("best_x",0)); pu=_sc(d.get("best_pnl",d.get("pnl",0)))
                    tc=int(_sc(d.get("trade_count",d.get("best_trade_count",0))))
                    wr=_sc(d.get("win_rate",d.get("best_win_rate",0)))
                    xv=d.get("x_values",[]); tc_arr=d.get("trade_count",[])
                    n_total=len(xv) if isinstance(xv,list) else 0
                    n_tested=sum(1 for v in tc_arr if isinstance(v,(int,float)) and v>0) if isinstance(tc_arr,list) else (1 if tc>0 else 0)
                    tcv = d.get("tick_count", 0)
                    has_ticks = (isinstance(tcv, list) and any(isinstance(v, (int, float)) and v > 0 for v in tcv)) or (
                        isinstance(tcv, (int, float)) and tcv > 0
                    )
                    if n_total > 0 and n_tested == 0 and has_ticks:
                        n_tested = n_total

                pi=pu*_usdt_to_inr(); pct=n_tested/n_total if n_total>0 else (1.0 if tc>0 else 0.0)
                pnl_c=GREEN if pi>=0 else RED; wr_c=GREEN if wr>=55 else(AMBER if wr>=40 else RED)
                vs=f"{(bx-CRYPTO_X)/CRYPTO_X*100:+.2f}%" if bx else "—"
                vs_c=GREEN if bx and abs((bx-CRYPTO_X)/CRYPTO_X*100)<5 else(AMBER if bx and abs((bx-CRYPTO_X)/CRYPTO_X*100)<20 else RED)

                sym_cards.append(html.Div([
                    html.Div([
                        html.Span(sym,style={"fontWeight":"700","color":CRYPTO_COL,"fontSize":"14px","minWidth":"60px"}),
                        html.Span(f"Best X: {bx:.6f}" if bx else "sweeping…",
                                  style={"color":GREEN if bx else DIM,"fontSize":"12px","fontFamily":"monospace","marginLeft":"8px"}),
                        html.Span(f"  {vs}",style={"color":vs_c,"fontSize":"11px","marginLeft":"6px"}),
                    ],style={"display":"flex","alignItems":"center","flexWrap":"wrap"}),
                    html.Div([
                        html.Span("P&L: ",style={"color":DIM,"fontSize":"11px"}),
                        html.Span(f"₹{pi:+,.0f}",style={"color":pnl_c,"fontWeight":"700","fontSize":"13px","marginRight":"12px"}),
                        html.Span(f"(${pu:+.4f})",style={"color":DIM,"fontSize":"10px","marginRight":"12px"}),
                        html.Span(f"Trades: {tc}",style={"color":TEXT,"fontSize":"11px","marginRight":"12px"}),
                        html.Span(f"Win: {wr:.0f}%",style={"color":wr_c,"fontSize":"11px","marginRight":"12px"}),
                        html.Span(f"Tested: {n_tested}/{n_total}",style={"color":DIM,"fontSize":"10px"}),
                    ],style={"marginTop":"5px","display":"flex","alignItems":"center","flexWrap":"wrap"}),
                    html.Div([
                        html.Span("Sweep",style={"color":DIM,"fontSize":"9px","marginRight":"6px","minWidth":"38px"}),
                        html.Div([html.Div(style={"width":f"{pct*100:.0f}%","height":"5px",
                                                  "background":CRYPTO_COL if pct<1.0 else GREEN,
                                                  "borderRadius":"3px","transition":"width 0.4s ease"})],
                                 style={"flex":"1","height":"5px","background":BORDER,"borderRadius":"3px"}),
                        html.Span(f"{pct*100:.0f}%",style={"color":DIM,"fontSize":"9px","marginLeft":"6px"}),
                    ],style={"display":"flex","alignItems":"center","marginTop":"6px","gap":"4px"}),
                ],style={"background":CARD,"border":f"1px solid {BORDER}","borderRadius":"8px",
                         "padding":"10px 14px","marginBottom":"8px","borderLeft":f"3px solid {pnl_c}"}))
                table_rows.append({"Coin":sym,"Best X":f"{bx:.6f}" if bx else "—",
                                   "P&L (₹)":f"₹{pi:+,.0f}","P&L ($)":f"${pu:+.4f}" if abs(pu)>0.0001 else "—",
                                   "Win%":f"{wr:.1f}%","Trades":str(tc),"vs Live X":vs})

            return html.Div([
                html.Div("Per-Coin Sweep Results",className="sec"),
                html.Div(sym_cards),
                html.Div([
                    html.Div("Full Table",className="sec",style={"marginTop":"14px"}),
                    html.Div([dash_table.DataTable(
                        data=table_rows,
                        columns=[{"name":c,"id":c} for c in ["Coin","Best X","P&L (₹)","P&L ($)","Win%","Trades","vs Live X"]],
                        sort_action="native",page_size=10,
                        style_header={"backgroundColor":SB,"color":CRYPTO_COL,"fontWeight":"bold","border":f"1px solid {BORDER}"},
                        style_cell={"backgroundColor":BG,"color":TEXT,"border":f"1px solid {BORDER}","fontFamily":FONT,"fontSize":"12px","padding":"6px"},
                        style_data_conditional=[
                            {"if":{"column_id":"Best X"},"color":GREEN,"fontFamily":"monospace"},
                            {"if":{"column_id":"P&L (₹)","filter_query":"{P&L (₹)} contains '+'"},"color":GREEN,"fontWeight":"700"},
                            {"if":{"column_id":"P&L (₹)","filter_query":"{P&L (₹)} contains '-'"},"color":RED,"fontWeight":"700"},
                            {"if":{"column_id":"Coin"},"color":CRYPTO_COL,"fontWeight":"700"},
                        ])],className="tw")
                ],className="card"),
            ])
        except Exception as e: return html.Div(f"Scanner {sid}: {e}",style={"color":RED,"padding":"12px","background":"#2d1519","borderRadius":"8px"})

    for _sid in (1,2,3):
        (lambda sid: (
            app.callback(Output(f"cr{sid}-s","children"),Input("t10","n_intervals"))(lambda _,s=sid: _cr_stat(s)),
            app.callback(Output(f"cr{sid}-t","children"),Input("t10","n_intervals"))(lambda _,s=sid: _cr_tbl(s)),
        ))(_sid)

    @app.callback(Output("cro-c","children"),Input("t10","n_intervals"),prevent_initial_call=True)
    def _cropt(_):
        try:
            syms=["BTC","ETH","BNB","SOL","ADA"]
            rows=[html.Tr([html.Th(h,style={"padding":"6px"}) for h in ["Coin","Best X","Source","P&L (₹)","Win%","Trades"]])]
            for sym in syms:
                best=None
                for sid,lbl in [(1,"CR1"),(2,"CR2"),(3,"CR3")]:
                    state=_DS.get("crypto_scanner",{}).get(sid)
                    if not state: continue
                    d=state.get("merged_best",state.get("sweeps",{})).get(sym,{})
                    if isinstance(d,dict) and d.get("best_x",0):
                        pi=float((d.get("pnl",d.get("best_pnl",-1e9)) or -1e9))*_usdt_to_inr()
                        if best is None or pi>best["pi"]: best={"x":d["best_x"],"pi":pi,"src":lbl,"wr":d.get("win_rate",d.get("best_win_rate",0)),"tc":d.get("trade_count",d.get("best_trade_count",0))}
                if best:
                    rows.append(html.Tr([html.Td(sym,style={"fontWeight":"700","color":CRYPTO_COL}),html.Td(f"{best['x']:.6f}",style={"color":GREEN}),
                        html.Td(best["src"],style={"color":YELLOW}),html.Td(f"₹{best['pi']:+,.0f}",style={"color":GREEN if best["pi"]>=0 else RED}),
                        html.Td(f"{float(best['wr'] or 0):.1f}%"),html.Td(str(best["tc"] or 0))],style={"borderBottom":f"1px solid {BORDER}"}))
                else:
                    rows.append(html.Tr([html.Td(sym,style={"color":DIM})]+[html.Td("—")]*5))
            return html.Div([html.Div("Crypto Optimizer — Best X per coin",className="sec"),html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"})],className="card")
        except Exception as e: return html.Div(f"Error: {e}",style={"color":RED})

    @app.callback(Output("crbx-c","children"),Input("t2","n_intervals"),prevent_initial_call=True)
    def _crbx(_):
        try:
            evs=_DS.get("crypto_events",[]); cl=[e for e in evs if e.get("exit_px") is not None]
            net=sum(float(e.get("net_pnl_inr",e.get("net_pnl",0)) or 0) for e in cl)
            rows=[html.Tr([html.Th(h,style={"padding":"6px"}) for h in ["Coin","Side","Entry($)","Exit($)","Net(₹)","Reason","Time"]])]
            for e in reversed(cl[-15:]):
                n=float(e.get("net_pnl_inr",e.get("net_pnl",0)) or 0)
                rows.append(html.Tr([html.Td(e.get("symbol","—"),style={"fontWeight":"700"}),html.Td(e.get("side","—"),style={"color":GREEN if e.get("side")=="BUY" else RED}),
                    html.Td(f"${float(e.get('entry_px',0)):,.4f}"),html.Td(f"${float(e.get('exit_px',0)):,.4f}"),
                    html.Td(f"₹{n:+,.0f}",style={"color":GREEN if n>=0 else RED,"fontWeight":"600"}),html.Td(e.get("reason","—")),html.Td(e.get("ts","—"))],style={"borderBottom":f"1px solid {BORDER}"}))
            return html.Div([html.Div([html.Span(f"Crypto (24/7): {len(cl)} trades  |  Net: "),html.Span(f"₹{net:+,.0f}",style={"color":GREEN if net>=0 else RED,"fontWeight":"700"})],style={"marginBottom":"12px","fontSize":"14px"}),html.Table(rows,style={"width":"100%","borderCollapse":"collapse","fontSize":"13px"})])
        except Exception as e: return html.Div(f"Error: {e}",style={"color":RED})

    # Crypto History & Performance  (FIX 6)
    @app.callback(Output("hcr-x","children"),Output("hcr-c","children"),Input("t60","n_intervals"),prevent_initial_call=True)
    def _hist_cr(_):
        try:
            h=_DS.get("crypto_history",{})
            x=html.Div([html.Div("Crypto History (24/7)",className="sl"),html.Div(f"{len(h)} days recorded",style={"fontSize":"20px","fontWeight":"700","color":CRYPTO_COL})],className="card")
            return x,hist_cards(h)
        except Exception as e: return html.Div(),html.Div(f"Error: {e}",className="err")

    @app.callback(Output("hcr-dl","data"),Input("hcr-btn","n_clicks"),prevent_initial_call=True)
    def _hist_cr_dl(_):
        import io,csv as _c; buf=io.StringIO(); w=_c.writer(buf); w.writerow(["Date","Trades","NetPnL_INR","Return%","WinRate%"])
        for ds,d in sorted(_DS.get("crypto_history",{}).items(),reverse=True): w.writerow([ds,d["total_trades"],d["net_pnl"],d["pct_return"],d["win_rate"]])
        return dict(content=buf.getvalue(),filename="algostack_crypto_history.csv")

    @app.callback(Output("pcr1","children"),Output("pcr2","children"),Output("pcr3","children"),Output("pcr4","children"),Input("t60","n_intervals"),prevent_initial_call=True)
    def _perf_cr(_):
        try: r=perf_charts(_DS.get("crypto_history",{}),CRYPTO_COL,0.30); return r if len(r)==4 else (r[0],)*4
        except Exception as e: err=html.Div(f"Error: {e}",className="err"); return err,err,err,err

    @app.callback(Output("pcr-dl","data"),Input("pcr-btn","n_clicks"),prevent_initial_call=True)
    def _perf_cr_dl(_):
        import io,csv as _c; buf=io.StringIO(); w=_c.writer(buf); w.writerow(["Date","Trades","NetPnL_INR","Return%","WinRate%"])
        for ds,d in sorted(_DS.get("crypto_history",{}).items(),reverse=True): w.writerow([ds,d["total_trades"],d["net_pnl"],d["pct_return"],d["win_rate"]])
        return dict(content=buf.getvalue(),filename="algostack_crypto_performance.csv")

    # ══════════════════════════════════════════════════════════════════════════
    #  INTEL PAGE  (FIX 7: crypto + commodity prices added)
    # ══════════════════════════════════════════════════════════════════════════
    @app.callback(Output("in-idx","children"),Input("t30","n_intervals"),prevent_initial_call=True)
    def _intel_idx(_):
        try:
            import yfinance as yf, warnings
            cards=[]
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                for name,ticker in [("Nifty50","^NSEI"),("Sensex","^BSESN"),("USD/INR","INR=X"),("Crude","CL=F")]:
                    try:
                        t=yf.Ticker(ticker); px=t.fast_info.get("lastPrice") or 0; prev=t.fast_info.get("previousClose") or px or 1
                        chg=(float(px)-float(prev))/float(prev)*100 if prev else 0; c=GREEN if chg>=0 else RED
                        cards.append(html.Div([html.Div(name,className="sl"),html.Div(f"{float(px):,.2f}",style={"fontSize":"20px","fontWeight":"700","color":TEXT}),html.Div(f"{'+' if chg>=0 else ''}{chg:.2f}%",style={"color":c,"fontSize":"13px","fontWeight":"600"})],className="sc"))
                    except Exception: pass
            return cards
        except Exception: return [html.Div("Market data loading...",style={"color":DIM})]

    # ── CONSOLIDATED: Intel live prices (crypto + commodity) in one callback ──
    @app.callback(Output("in-cpx","children"), Output("in-mpx","children"),
                  Input("t2","n_intervals"),prevent_initial_call=True)
    def _intel_prices(_):
        cprices=_DS.get("crypto_prices",{}); cage=float(_DS.get("crypto_age",9999) or 9999)
        mprices=_DS.get("commodity_prices",{}); mage=float(_DS.get("commodity_age",9999) or 9999)
        # Crypto panel
        try:
            c_live=float(cage)<10
            ICONS={"BTC":"₿","ETH":"Ξ","BNB":"◈","SOL":"◎","ADA":"₳"}
            crow=[html.Div([html.Span("●",style={"color":GREEN if c_live else AMBER,"fontSize":"10px","marginRight":"6px"}),
                html.Span("CRYPTO PRICES (Binance)",style={"color":CRYPTO_COL,"fontSize":"11px","fontWeight":"700","letterSpacing":"0.08em"}),
                html.Span(f"  {_age_label(cage)}",style={"color":DIM,"fontSize":"10px","marginLeft":"auto"})],
                style={"marginBottom":"10px","display":"flex","alignItems":"center"})]
            for sym in ["BTC","ETH","BNB","SOL","ADA"]:
                icon=ICONS.get(sym,"$"); px_v=cprices.get(sym,0)
                if px_v:
                    crow.append(html.Div([
                        html.Span(f"{icon} {sym}",style={"color":CRYPTO_COL,"fontWeight":"700","width":"60px","display":"inline-block"}),
                        html.Span(f"${float(px_v):,.2f}" if sym!="ADA" else f"${float(px_v):.4f}",style={"color":TEXT,"fontWeight":"600","fontSize":"14px"}),
                        html.Span(f"  ≈₹{float(px_v)*_usdt_to_inr():,.0f}",style={"color":DIM,"fontSize":"11px","marginLeft":"6px"})],
                        style={"padding":"6px 0","borderBottom":f"1px solid {BORDER}","display":"flex","alignItems":"center"}))
                else:
                    crow.append(html.Div([html.Span(f"{icon} {sym}",style={"color":CRYPTO_COL,"fontWeight":"700","width":"60px","display":"inline-block"}),
                        html.Span("Waiting...",style={"color":AMBER,"fontSize":"12px"})],
                        style={"padding":"6px 0","borderBottom":f"1px solid {BORDER}"}))
            if not cprices: crow.append(html.Div("⚠ crypto_engine.py not running",style={"color":AMBER,"fontSize":"12px","padding":"8px 0"}))
            cpx_div=html.Div(crow,style={"background":CARD,"border":f"1px solid {'#2ea04380' if c_live else BORDER}","borderRadius":"8px","padding":"16px"})
        except Exception as e: cpx_div=html.Div(f"Crypto: {e}",style={"color":DIM,"background":CARD,"borderRadius":"8px","padding":"16px"})
        # Commodity panel
        try:
            m_live=float(mage)<15
            UNITS={"GOLD":"₹/10g","SILVER":"₹/kg","CRUDE":"₹/bbl","NATURALGAS":"₹/mmBtu","COPPER":"₹/kg"}
            mrow=[html.Div([html.Span("●",style={"color":GREEN if m_live else AMBER,"fontSize":"10px","marginRight":"6px"}),
                html.Span("MCX PRICES",style={"color":COMM_COL,"fontSize":"11px","fontWeight":"700","letterSpacing":"0.08em"}),
                html.Span(f"  {_age_label(mage)}",style={"color":DIM,"fontSize":"10px","marginLeft":"auto"})],
                style={"marginBottom":"10px","display":"flex","alignItems":"center"})]
            for sym in ["GOLD","SILVER","CRUDE","NATURALGAS","COPPER"]:
                px_v=mprices.get(sym,0); unit=UNITS.get(sym,"₹")
                if px_v:
                    mrow.append(html.Div([
                        html.Span(sym,style={"color":COMM_COL,"fontWeight":"700","width":"100px","display":"inline-block","fontSize":"13px"}),
                        html.Span(f"₹{float(px_v):,.2f}",style={"color":TEXT,"fontWeight":"600","fontSize":"14px"}),
                        html.Span(f"  {unit}",style={"color":DIM,"fontSize":"10px","marginLeft":"4px"})],
                        style={"padding":"6px 0","borderBottom":f"1px solid {BORDER}","display":"flex","alignItems":"center"}))
                else:
                    mrow.append(html.Div([html.Span(sym,style={"color":COMM_COL,"fontWeight":"700","width":"100px","display":"inline-block"}),
                        html.Span("Waiting...",style={"color":AMBER,"fontSize":"12px"})],
                        style={"padding":"6px 0","borderBottom":f"1px solid {BORDER}"}))
            if not mprices: mrow.append(html.Div("⚠ commodity_engine.py not running",style={"color":AMBER,"fontSize":"12px","padding":"8px 0"}))
            mpx_div=html.Div(mrow,style={"background":CARD,"border":f"1px solid {'#d2992280' if m_live else BORDER}","borderRadius":"8px","padding":"16px"})
        except Exception as e: mpx_div=html.Div(f"MCX: {e}",style={"color":DIM,"background":CARD,"borderRadius":"8px","padding":"16px"})
        return cpx_div, mpx_div

    @app.callback(Output("in-india","children"),Input("t30","n_intervals"),prevent_initial_call=True)
    def _intel_india(_):
        try:
            import xml.etree.ElementTree as ET, requests as _req
            hdrs={"User-Agent":"Mozilla/5.0 AlgoStack/10.0"}
            feeds=[("ET Markets","https://economictimes.indiatimes.com/markets/rssfeeds/1977021501.cms"),
                   ("Moneycontrol","https://www.moneycontrol.com/rss/latestnews.xml"),
                   ("Biz Standard","https://www.business-standard.com/rss/markets-106.rss")]
            items=[]
            for src,url in feeds:
                try:
                    root=ET.fromstring(_req.get(url,headers=hdrs,timeout=6).content)
                    for it in root.findall(".//item")[:4]:
                        t=(it.findtext("title") or "").strip(); l=(it.findtext("link") or "#").strip()
                        if t: items.append((src,t,l,(it.findtext("pubDate") or "")[:16]))
                except Exception: pass
            rows=[html.Div("INDIA MARKET NEWS",style={"color":ACCENT,"fontSize":"11px","fontWeight":"700","letterSpacing":"0.08em","marginBottom":"10px"})]
            for src,t,l,pub in items[:12]:
                rows.append(html.Div([html.Span(src,style={"color":AMBER,"fontSize":"10px","fontWeight":"600","background":AMBER+"22","borderRadius":"3px","padding":"1px 5px","marginRight":"6px"}),html.Span(pub,style={"color":DIM,"fontSize":"10px"}),html.Br(),html.A(t,href=l,target="_blank",style={"color":TEXT,"fontSize":"12px","textDecoration":"none","fontWeight":"600","lineHeight":"1.4"})],style={"padding":"7px 0","borderBottom":f"1px solid {BORDER}"}))
            if not items: rows.append(html.Div("Feeds unavailable",style={"color":DIM}))
            return html.Div(rows,style={"background":CARD,"border":f"1px solid {BORDER}","borderRadius":"8px","padding":"16px"})
        except Exception as e: return html.Div(f"India news: {e}",style={"color":DIM,"background":CARD,"borderRadius":"8px","padding":"16px"})

    @app.callback(Output("in-intl","children"),Input("t30","n_intervals"),prevent_initial_call=True)
    def _intel_intl(_):
        try:
            import xml.etree.ElementTree as ET, requests as _req
            hdrs={"User-Agent":"Mozilla/5.0 AlgoStack/10.0"}
            feeds=[("Reuters","https://feeds.reuters.com/reuters/businessNews"),("CNBC","https://www.cnbc.com/id/100003114/device/rss/rss.html")]
            items=[]
            for src,url in feeds:
                try:
                    root=ET.fromstring(_req.get(url,headers=hdrs,timeout=6).content)
                    for it in root.findall(".//item")[:5]:
                        t=(it.findtext("title") or "").strip(); l=(it.findtext("link") or "#").strip()
                        if t: items.append((src,t,l,(it.findtext("pubDate") or "")[:16]))
                except Exception: pass
            rows=[html.Div("GLOBAL MARKET NEWS",style={"color":ACCENT,"fontSize":"11px","fontWeight":"700","letterSpacing":"0.08em","marginBottom":"10px"})]
            for src,t,l,pub in items[:12]:
                rows.append(html.Div([html.Span(src,style={"color":PURPLE,"fontSize":"10px","fontWeight":"600","background":PURPLE+"22","borderRadius":"3px","padding":"1px 5px","marginRight":"6px"}),html.Span(pub,style={"color":DIM,"fontSize":"10px"}),html.Br(),html.A(t,href=l,target="_blank",style={"color":TEXT,"fontSize":"12px","textDecoration":"none","fontWeight":"600","lineHeight":"1.4"})],style={"padding":"7px 0","borderBottom":f"1px solid {BORDER}"}))
            if not items: rows.append(html.Div("Feeds unavailable",style={"color":DIM}))
            return html.Div(rows,style={"background":CARD,"border":f"1px solid {BORDER}","borderRadius":"8px","padding":"16px"})
        except Exception as e: return html.Div(f"Global news: {e}",style={"color":DIM,"background":CARD,"borderRadius":"8px","padding":"16px"})

    @app.callback(Output("in-sig","children"),Input("t30","n_intervals"),prevent_initial_call=True)
    def _intel_sig(_):
        try:
            import yfinance as yf, warnings, numpy as np
            rows=[html.Div("MARKET SIGNALS",style={"color":ACCENT,"fontSize":"11px","fontWeight":"700","letterSpacing":"0.08em","marginBottom":"10px"})]
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                df=yf.download("^NSEI",period="3mo",interval="1d",progress=False)
                if df is not None and len(df)>=20:
                    closes=df["Close"].dropna().values.astype(float).flatten()
                    last=float(closes[-1]); dma20=float(closes[-20:].mean())
                    r5=float((last-closes[-5])/closes[-5]*100) if len(closes)>=5 else 0.0
                    deltas=np.diff(closes[-15:])
                    gains_arr=np.where(deltas>0,deltas,0.0); losses_arr=np.where(deltas<0,-deltas,0.0)
                    gains=float(gains_arr.mean()) if len(gains_arr)>0 else 0.0
                    losses=float(losses_arr.mean()) if len(losses_arr)>0 else 0.0001
                    rsi=float(100-(100/(1+gains/max(losses,0.0001))))
                    trend="BULLISH" if last>dma20 and r5>0 else("BEARISH" if last<dma20 and r5<0 else "NEUTRAL")
                    tc=GREEN if trend=="BULLISH" else(RED if trend=="BEARISH" else AMBER)
                    rows+=[html.Div([html.Span("Nifty 50 ",style={"fontWeight":"700","fontSize":"14px"}),html.Span(trend,style={"color":tc,"fontWeight":"700"})],style={"marginBottom":"8px"}),
                           html.Div(f"Last:{float(last):,.0f}  20DMA:{float(dma20):,.0f}",style={"color":DIM,"fontSize":"12px"}),
                           html.Div(f"5D:{float(r5):+.2f}%",style={"color":GREEN if r5>0 else RED,"fontSize":"12px","marginTop":"4px"}),
                           html.Div(f"RSI(14):{float(rsi):.1f} "+("⚠Overbought" if rsi>70 else "✓Oversold" if rsi<30 else "Neutral"),style={"color":RED if rsi>70 else(GREEN if rsi<30 else DIM),"fontSize":"12px","marginTop":"4px"})]
                try:
                    import requests as _req
                    sess=_req.Session(); sess.get("https://www.nseindia.com",headers={"User-Agent":"Mozilla/5.0"},timeout=3)
                    r=sess.get("https://www.nseindia.com/api/fiidiiTradeReact",headers={"User-Agent":"Mozilla/5.0"},timeout=5)
                    d=r.json()[0]; fii=float(d.get("fiiNetVal",0)); dii=float(d.get("diiNetVal",0))
                    rows+=[html.Div("FII / DII Flows",style={"color":DIM,"fontSize":"11px","marginTop":"12px","fontWeight":"600"}),
                           html.Div(f"FII: ₹{fii:+,.0f} Cr",style={"color":GREEN if fii>0 else RED,"fontSize":"13px","marginTop":"4px"}),
                           html.Div(f"DII: ₹{dii:+,.0f} Cr",style={"color":GREEN if dii>0 else RED,"fontSize":"13px","marginTop":"2px"})]
                except Exception: rows.append(html.Div("FII/DII: NSE API unavailable",style={"color":DIM,"fontSize":"11px","marginTop":"8px"}))
            return html.Div(rows,style={"background":CARD,"border":f"1px solid {BORDER}","borderRadius":"8px","padding":"16px"})
        except Exception as e: return html.Div(f"Signals: {e}",style={"color":DIM,"background":CARD,"borderRadius":"8px","padding":"16px"})

    # ══════════════════════════════════════════════════════════════════════════
    #  SYSTEM PAGE  (FIX 8: all 15 processes)
    # ══════════════════════════════════════════════════════════════════════════
    # ── CONSOLIDATED: system page (net + process status in one callback) ─────
    @app.callback(Output("sys-net","children"), Output("sys-c","children"),
                  Input("t2","n_intervals"),prevent_initial_call=True)
    def _sys_all(_):
        # Network status
        try:
            import wifi_keepalive as _wk
            wk=getattr(_wk,"_GLOBAL_KEEPALIVE",None)
            if wk:
                s=wk.status; ok=s.get("internet_up",False)
                dl=float(s.get("download_mbps",0) or 0); ul=float(s.get("upload_mbps",0) or 0)
                ping=float(s.get("ping_ms",0) or 0); col=GREEN if dl>=5 else(AMBER if dl>0 else RED)
                net_div=html.Div([
                    html.Span("🌐 ",style={"fontSize":"14px"}),
                    html.Span("UP" if ok else "DOWN",style={"color":GREEN if ok else RED,"fontWeight":"700","marginRight":"12px"}),
                    html.Span(f"↓{dl:.2f}Mbps  ↑{ul:.2f}Mbps  ping={ping:.0f}ms",style={"color":col,"fontSize":"12px"})],
                    style={"padding":"6px 14px","marginBottom":"8px","background":"#0d1117","borderRadius":"6px","border":f"1px solid {BORDER}"})
            else:
                net_div=html.Div("🌐 Internet: checking...",style={"color":DIM,"fontSize":"12px","padding":"4px 0"})
        except: net_div=html.Div("🌐 Internet status unavailable",style={"color":DIM,"fontSize":"12px","padding":"4px"})

        # Processes and status
        try:
            pub=_DS.get("pub_url"); lip=get_lan_ip()
            px=_DS.get("live_prices",{}); age=float(_DS.get("price_age",9999) or 9999)
            mpx=_DS.get("commodity_prices",{}); cpx=_DS.get("crypto_prices",{}); cage=float(_DS.get("crypto_age",9999) or 9999)
            ds=datetime.now(IST).strftime("%Y%m%d")

            def _port(p):
                try: urllib.request.urlopen(f"http://127.0.0.1:{p}",timeout=0.5); return True
                except: return False

            def _file_age(path):
                if os.path.exists(path):
                    a=round(time.time()-os.path.getmtime(path),0)
                    return a<120, f"{'live' if a<120 else 'stale'} ({a:.0f}s)"
                return False,"no file"

            def _sc_file(sid, dirs):
                for d in dirs:
                    p=os.path.join(d,ds,"live_state.json")
                    if os.path.exists(p):
                        a=round(time.time()-os.path.getmtime(p),0)
                        return a<90, f"{'live' if a<90 else 'stale'} ({a:.0f}s)"
                # Check without date subfolder too (crypto scanners use date-range dirs)
                for d in dirs:
                    if os.path.isdir(d):
                        subs=sorted([f for f in os.listdir(d) if os.path.isdir(os.path.join(d,f))],reverse=True)
                        for sub in subs[:3]:
                            p=os.path.join(d,sub,"live_state.json")
                            if os.path.exists(p):
                                a=round(time.time()-os.path.getmtime(p),0)
                                return a<300, f"{'live' if a<300 else 'stale'} ({a:.0f}s)"
                return False,"no data"

            def row(lbl,ok,detail="",warn=False,href=None):
                c=GREEN if ok else(AMBER if warn else RED)
                dot=html.Span("●" if ok else("◐" if warn else "○"),style={"color":c,"fontSize":"15px","marginRight":"6px"})
                lbl_el=html.A(lbl,href=href,style={"fontWeight":"600","marginRight":"8px","color":ACCENT}) if href else html.Span(lbl,style={"fontWeight":"600","marginRight":"8px"})
                return html.Div([dot,lbl_el,html.Span(detail,style={"color":DIM,"fontSize":"12px"})],
                                style={"padding":"7px 0","borderBottom":f"1px solid {BORDER}","display":"flex","alignItems":"center"})

            s1ok,s1d=_sc_file(1,[SCANNER_DIRS.get(1,""),SCANNER_DIRS_LEGACY.get(1,"")])
            s2ok,s2d=_sc_file(2,[SCANNER_DIRS.get(2,""),SCANNER_DIRS_LEGACY.get(2,"")])
            s3ok,s3d=_sc_file(3,[SCANNER_DIRS.get(3,""),SCANNER_DIRS_LEGACY.get(3,"")])
            co_ok,co_d=_file_age(os.path.join(LEVELS_DIR,f"commodity_initial_levels_{ds}.json"))
            cs1ok,cs1d=_sc_file(1,[COMM_SCANNER_DIRS.get(1,"")])
            cs2ok,cs2d=_sc_file(2,[COMM_SCANNER_DIRS.get(2,"")])
            cs3ok,cs3d=_sc_file(3,[COMM_SCANNER_DIRS.get(3,"")])
            cr_ok,cr_d=_file_age(os.path.join(LEVELS_DIR,"crypto_initial_levels_latest.json"))
            cr1ok,cr1d=_sc_file(1,[CRYPTO_SCANNER_DIRS.get(1,"")])
            cr2ok,cr2d=_sc_file(2,[CRYPTO_SCANNER_DIRS.get(2,"")])
            cr3ok,cr3d=_sc_file(3,[CRYPTO_SCANNER_DIRS.get(3,"")])
            eq_live=age<30 and len(px)>0
            cr_live=cage<30 and len(cpx)>0

            eq_bl=html.Div([
                html.Div("📈 EQUITY (NSE 09:30–15:11 IST)",style={"color":EQ_COL,"fontSize":"11px","fontWeight":"700","letterSpacing":"0.1em","padding":"10px 0 6px"}),
                row("Algofinal :8050",_port(8050),"Equity engine + ZMQ PUB + daily Excel",href="http://localhost:8050"),
                row("Scanner 1 (Narrow 1K)",s1ok,f"X:0.0080–0.0090 — {s1d}"),
                row("Scanner 2 (Dual 13K)",s2ok,f"X:0.001–0.007+0.009–0.016 — {s2d}"),
                row("Scanner 3 (Wide 31K)",s3ok,f"X:0.001–0.032 — {s3d}"),
                row("XOptimizer :8063",_port(8063),"Cross-scanner X leaderboard",href="http://localhost:8063"),
                row("BestXTrader",os.path.exists(BESTX_FILE),"Paper trades with best X"),
            ],className="card")

            cm_bl=html.Div([
                html.Div("🥇 COMMODITY (MCX 09:00–23:30 IST)",style={"color":COMM_COL,"fontSize":"11px","fontWeight":"700","letterSpacing":"0.1em","padding":"10px 0 6px"}),
                row("CommodityEngine",co_ok,f"MCX prices via TradingView WS — {co_d}"),
                row("CommScanner 1",cs1ok,f"MCX narrow sweep — {cs1d}",warn=not cs1ok),
                row("CommScanner 2",cs2ok,f"MCX dual sweep — {cs2d}",warn=not cs2ok),
                row("CommScanner 3",cs3ok,f"MCX wide sweep — {cs3d}",warn=not cs3ok),
                html.Div("ℹ CommScanners exit code=0 on weekends = NORMAL (MCX closed Mon–Fri only)",
                         style={"color":AMBER,"fontSize":"11px","padding":"6px 0","fontStyle":"italic"}),
            ],className="card")

            cr_bl=html.Div([
                html.Div("₿ CRYPTO (Binance — 24/7, re-anchor every 6h)",style={"color":CRYPTO_COL,"fontSize":"11px","fontWeight":"700","letterSpacing":"0.1em","padding":"10px 0 6px"}),
                row("CryptoEngine",cr_ok,f"Binance WS → ZMQ 'crypto' topic — {cr_d}"),
                row("CryptoScanner 1",cr1ok,f"Crypto narrow sweep — {cr1d}"),
                row("CryptoScanner 2",cr2ok,f"Crypto dual sweep — {cr2d}"),
                row("CryptoScanner 3",cr3ok,f"Crypto wide sweep — {cr3d}"),
            ],className="card")

            # v10.6: AlertMonitor status
            am_ok = os.path.exists(os.path.join("logs","alert_monitor.log"))
            am_age_s = ""
            if am_ok:
                am_age = round(time.time()-os.path.getmtime(os.path.join("logs","alert_monitor.log")),0)
                am_age_s = f"active ({am_age:.0f}s ago)" if am_age<120 else f"stale ({am_age:.0f}s)"

            db_bl=html.Div([
                html.Div("⚙ DASHBOARD & MONITORING",style={"color":ACCENT,"fontSize":"11px","fontWeight":"700","letterSpacing":"0.1em","padding":"10px 0 6px"}),
                row(f"UnifiedDash v10.7 :{DASH_PORT}",_port(DASH_PORT),"Equity + Commodity + Crypto + AI",href=f"http://localhost:{DASH_PORT}"),
                row("AlertMonitor v10.7",am_ok,f"Staleness · Tunnel · SL · P&L milestones — {am_age_s}" if am_age_s else "Starting…"),
            ],className="card")

            url_bl=html.Div([
                html.Div("ACCESS LINKS",className="sec"),
                html.Div([
                    html.Div([html.Span("🌐 Public: ",style={"color":DIM,"fontSize":"12px","fontWeight":"600"}),
                              html.A(pub or "(tunnel starting…)",href=pub or "#",
                                     style={"color":ACCENT,"wordBreak":"break-all","fontWeight":"700",
                                            "fontSize":"13px"})],style={"marginBottom":"10px"}),
                    html.Div([html.Span("🏠 LAN:    ",style={"color":DIM,"fontSize":"12px","fontWeight":"600"}),
                              html.A(f"http://{lip}:{DASH_PORT}",href=f"http://{lip}:{DASH_PORT}",
                                     style={"color":ACCENT,"fontWeight":"600"})],style={"marginBottom":"6px"}),
                    html.Div([html.Span("💻 Local:  ",style={"color":DIM,"fontSize":"12px","fontWeight":"600"}),
                              html.A(f"http://localhost:{DASH_PORT}",href=f"http://localhost:{DASH_PORT}",
                                     style={"color":DIM})]),
                ]),
            ],className="card")

            # v10.6: enhanced price feed status with age bars
            def _age_bar(age_s, max_s=30):
                a=float(age_s) if age_s is not None else 9999.0
                if a>=9000: return html.Span("cached",style={"color":AMBER,"fontSize":"10px","marginLeft":"6px"})
                pct=min(a/max_s*100,100)
                col=GREEN if a<5 else (YELLOW if a<15 else (AMBER if a<30 else RED))
                return html.Span([
                    html.Span(f"{a:.1f}s",style={"color":col,"fontSize":"10px","marginLeft":"6px"}),
                    html.Div(style={"display":"inline-block","width":f"{max(2,100-pct):.0f}px",
                                     "height":"3px","background":col,"borderRadius":"2px",
                                     "marginLeft":"4px","verticalAlign":"middle","opacity":"0.7"}),
                ])

            px_bl=html.Div([
                html.Div("⚡ PRICE FEED STATUS",className="sec"),
                html.Div([
                    html.Div([
                        html.Span("📈 Equity",style={"color":EQ_COL,"fontWeight":"700","fontSize":"12px"}),
                        _age_bar(age,60),
                        html.Span(f"  {len(px)} stocks",style={"color":DIM,"fontSize":"11px","marginLeft":"8px"}),
                    ],style={"padding":"6px 0","borderBottom":f"1px solid {BORDER}","display":"flex","alignItems":"center"}),
                    html.Div([
                        html.Span("🥇 MCX",style={"color":COMM_COL,"fontWeight":"700","fontSize":"12px"}),
                        _age_bar(_DS.get("commodity_age",9999),30),
                        html.Span(f"  {len(mpx)} symbols",style={"color":DIM,"fontSize":"11px","marginLeft":"8px"}),
                    ],style={"padding":"6px 0","borderBottom":f"1px solid {BORDER}","display":"flex","alignItems":"center"}),
                    html.Div([
                        html.Span("₿ Crypto",style={"color":CRYPTO_COL,"fontWeight":"700","fontSize":"12px"}),
                        _age_bar(cage,15),
                        html.Span(f"  {len(cpx)} coins",style={"color":DIM,"fontSize":"11px","marginLeft":"8px"}),
                    ],style={"padding":"6px 0","display":"flex","alignItems":"center"}),
                ],style={"marginBottom":"10px"}),
                html.Div("ZMQ PUB: tcp://127.0.0.1:28081 | Fast: 0.5s | XLSX: 30s | Scanners: 1.5s",
                         style={"color":DIM,"fontSize":"11px","fontFamily":"monospace","marginBottom":"6px"}),
                html.Div("Variations/day: 45K equity + 97.5K MCX + 97.5K crypto = 240K total",
                         style={"color":DIM,"fontSize":"11px","fontFamily":"monospace"}),
            ],className="card")

            sys_div=html.Div([eq_bl,cm_bl,cr_bl,db_bl,url_bl,px_bl])
        except Exception as e: sys_div=html.Div(f"System error: {e}",className="err")
        return net_div, sys_div

    # ══════════════════════════════════════════════════════════════════════════
    #  AI AGENT PAGE  (FIX 9)
    # ══════════════════════════════════════════════════════════════════════════
    QUICK_QS=["How does the X multiplier work?","Explain the 3 scanners",
              "Why do CommScanners exit on weekends?","What is retreat exit?",
              "How does crypto trading work?","How to read the Performance page?"]

    @app.callback(
        Output("ai-box","children"), Output("ai-status","children"), Output("ai-in","value"),
        Output("ai-api-status","children"),
        Input("ai-btn","n_clicks"),
        *[Input(f"aiq-{i}","n_clicks") for i in range(6)],
        State("ai-in","value"), State("ai-box","children"),
        prevent_initial_call=True,
    )
    def _ai_chat(send_clicks,*args):
        try:
            from dash import ctx as _ctx
            from dash.exceptions import PreventUpdate
            qclicks=args[:6]; q_input=args[6]; msgs=list(args[7] or [])
            triggered=getattr(_ctx,"triggered_id","")
            question=""
            if triggered=="ai-btn" and q_input and q_input.strip(): question=q_input.strip()
            elif triggered and triggered.startswith("aiq-"):
                idx=int(triggered.split("-")[-1])
                if idx<len(QUICK_QS): question=QUICK_QS[idx]
            if not question: raise PreventUpdate
            msgs.append(html.Div(question,className="ai-u"))
            msgs.append(html.Div("⏳ Thinking…",className="ai-b",style={"color":DIM,"fontStyle":"italic"}))
            result=[None]
            def _get(): result[0]=_ai_respond(question)
            t=threading.Thread(target=_get); t.start(); t.join(timeout=25)
            answer=result[0] or "⚠ Response timeout. If Gemini is active, the campus network may be blocking Google APIs. Check logs."
            msgs[-1]=html.Div(answer,className="ai-b")
            # Build API status
            err=_GEMINI_LAST_ERROR[0] if _GEMINI_LAST_ERROR else ""
            if GEMINI_API_KEY and not err:
                api_stat=html.Div([html.Span("API: ",style={"color":DIM,"fontSize":"11px"}),
                                   html.Span("Gemini 2.0 Flash ✓",style={"color":GREEN,"fontSize":"11px"})],style={"marginTop":"4px"})
            elif GEMINI_API_KEY and err:
                api_stat=html.Div([html.Span("API: ",style={"color":DIM,"fontSize":"11px"}),
                                   html.Span(f"Gemini error → Offline KB: {err[:80]}",style={"color":AMBER,"fontSize":"11px"})],style={"marginTop":"4px"})
            else:
                api_stat=html.Div([html.Span("API: ",style={"color":DIM,"fontSize":"11px"}),
                                   html.Span("Offline KB (no Gemini key)",style={"color":AMBER,"fontSize":"11px"})],style={"marginTop":"4px"})
            status=f"✓ {'Gemini 2.0 Flash' if GEMINI_API_KEY and not err else 'Offline KB'}"
            return msgs,status,"",api_stat
        except Exception as e:
            if "PreventUpdate" in str(type(e)): raise
            err_div=html.Div(f"Error: {e}",className="ai-b",style={"color":RED})
            api_stat=html.Div(f"Error: {str(e)[:80]}",style={"color":RED,"fontSize":"11px"})
            return list(args[7] or [])+[err_div],f"Error: {e}","",api_stat

    try:
        _ncb = len(app.callback_map)
    except Exception:
        _ncb = len([k for k in dir(app) if 'callback' in k.lower()])
    log.info("Dash app built (%d callbacks)", _ncb)

    # v10.9: Anti-refresh — suppress Dash's automatic page reload on WebSocket
    # timeout. On ngrok free tier, WS drops every ~30s causing mobile page reloads.
    # This override keeps the page alive and lets Dash reconnect silently.
    try:
        app.clientside_callback(
            """
            function(n) {
                if(n > 0 && !window._as_no_reload) {
                    window._as_no_reload = true;
                    var orig = window.location.reload.bind(window.location);
                    window.location.reload = function(hard) {
                        if(!hard) {
                            console.info("AlgoStack: page reload suppressed - reconnecting WebSocket");
                            return;
                        }
                        return orig(hard);
                    };
                }
                return window.dash_clientside.no_update;
            }
            """,
            Output("t60", "disabled"),
            Input("t60", "n_intervals"),
            prevent_initial_call=True,
        )
        log.info("Anti-refresh override registered")
    except Exception as _are:
        log.debug("Anti-refresh cb: %s", _are)

    return app


# ══════════════════════════════════════════════════════════════════════════════
#  TUNNEL
# ══════════════════════════════════════════════════════════════════════════════
def _try_ssh_tunnel(port, host, ssh_port=22, timeout=40):
    """SSH-based tunnel — works on university networks blocking ngrok/cloudflare.
    Pinggy uses port 443 which bypasses almost all firewalls.
    """
    import re as _re
    global _SSH_PROC, _SSH_URL
    try:
        if _SSH_PROC is not None and _SSH_PROC.poll() is None and _SSH_URL:
            return _SSH_URL
        cmd = ["ssh",
               "-p", str(ssh_port),
               "-o", "StrictHostKeyChecking=no",
               "-o", "UserKnownHostsFile=/dev/null",
               "-o", f"ConnectTimeout={'10' if ssh_port==443 else '15'}",
               "-o", "ServerAliveInterval=30",
               "-R", f"0:localhost:{port}",   # 0 = let server pick port
               host]
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT, text=True, bufsize=1)
        found = None; ev = threading.Event()
        pat = _re.compile(
            r"https?://[\w\-\.]+\."
            r"(?:localhost\.run|serveo\.net|lhr\.life|a\.pinggy\.io|pinggy\.link|pinggy\.io)"
            r"[^\s]*"
        )
        def _d():
            nonlocal found
            try:
                for line in proc.stdout:
                    if not found:
                        m = pat.search(line)
                        if m:
                            found = m.group(0).rstrip("/")
                            if not found.startswith("https"):
                                found = found.replace("http://","https://")
                            ev.set()
            except Exception: pass
            if not ev.is_set(): ev.set()
        threading.Thread(target=_d, daemon=True).start()
        ev.wait(timeout)
        if found:
            _SSH_PROC = proc
            _SSH_URL = found
            log.info("SSH tunnel (%s:%s) → %s", host, ssh_port, found)
            return found
        proc.terminate()
    except FileNotFoundError:
        log.debug("ssh not in PATH")
    except Exception as e:
        log.debug("SSH tunnel %s: %s", host, e)
    return None


def _try_localtunnel(port, timeout=30):
    """Localtunnel via npx — zero-account, zero-install beyond Node.js."""
    global _LT_PROC
    try:
        if _LT_PROC is not None and _LT_PROC.poll() is None:
            return None
        # Prefer globally installed `lt` (more stable than spawning via npx each restart).
        try:
            proc = subprocess.Popen(
                ["lt.cmd", "--port", str(port)],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1)
        except FileNotFoundError:
            proc = subprocess.Popen(
                ["npx.cmd", "localtunnel", "--port", str(port)],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1)
        found = None; ev = threading.Event()
        import re as _re2
        pat = _re2.compile(r"https?://[\w\-\.]+\.loca\.lt[^\s]*")
        def _r():
            nonlocal found
            try:
                for line in proc.stdout:
                    if not found:
                        m = pat.search(line)
                        if m: found = m.group(0).rstrip("/"); ev.set()
            except Exception as e:
                log.debug("localtunnel reader: %s", e)
            if not ev.is_set(): ev.set()
        threading.Thread(target=_r, daemon=True).start()
        ev.wait(timeout)
        if found:
            _LT_PROC = proc
            log.info("✓ Tunnel (localtunnel/npx) → %s", found)
            return found
        try: proc.terminate()
        except Exception: pass
    except FileNotFoundError:
        log.debug("npx not in PATH — skipping localtunnel")
    except Exception as e:
        log.debug("localtunnel: %s", e)
    return None


def open_tunnel(port, timeout=50):
    """
    v10.11 — tunnel failsafe. Tries each method in order, returns first URL found.
    1) pyngrok        (Python library — uses NGROK_TOKEN authtoken)
    2) ngrok CLI      (subprocess — bypasses pyngrok quirks)
    3) cloudflared    (Cloudflare free tunnel — proc kept alive in _CF_PROC)
    4) Pinggy SSH     (port 443 — bypasses university/hostel firewalls that block 22/80)
    5) localtunnel    (npx — Node.js required, zero account needed)

    NOTE:
    localhost.run and serveo are intentionally disabled because they can
    redirect to account/admin pages, which breaks public dashboard sharing.
    """
    global _CF_PROC, _NGROK_PROC, _LT_PROC, _SSH_PROC, _SSH_URL
    with _TUNNEL_LOCK:
        stable_mode = os.getenv("TUNNEL_STABLE_MODE", "1").strip() not in ("0", "false", "False")

        # Preferred stable chain
        if stable_mode:
            # 1) Pinggy over SSH:443 (Cloudflare-blacklist path, stable free option)
            u = _try_ssh_tunnel(port, "a.pinggy.io", ssh_port=443, timeout=30)
            if u:
                log.info("✓ Tunnel (Pinggy SSH:443) → %s", u)
                return u

            # 2) localtunnel
            u = _try_localtunnel(port, timeout=45)
            if u:
                return u

            # Retry pinggy once
            time.sleep(6)
            u = _try_ssh_tunnel(port, "a.pinggy.io", ssh_port=443, timeout=30)
            if u:
                return u

        # Pyngrok
        if not DISABLE_PYNGROK:
            try:
                from pyngrok import ngrok as _ngrok, conf as _ngrok_conf
                try:
                    _ngrok_conf.get_default().auth_token = NGROK_TOKEN
                except Exception:
                    _ngrok.set_auth_token(NGROK_TOKEN)
                try:
                    existing = _ngrok.get_tunnels()
                    for t in existing:
                        taddr = str(getattr(t, "config", {}).get("addr", "") or getattr(t, "public_url", "") or "")
                        if str(port) in taddr:
                            try:
                                _ngrok.disconnect(t.public_url)
                            except Exception:
                                pass
                except Exception:
                    pass
                try:
                    _t = _ngrok.connect(addr=port, proto="http", bind_tls=True)
                except TypeError:
                    try:
                        _t = _ngrok.connect(port, proto="http", options={"bind_tls": True})
                    except TypeError:
                        _t = _ngrok.connect(port)
                _u = getattr(_t, "public_url", None) or str(_t)
                if _u and _u.startswith("http"):
                    _u = _u.replace("http://", "https://")
                    log.info("✓ Tunnel (pyngrok) → %s", _u)
                    return _u
            except Exception:
                pass

        # ngrok CLI
        try:
            if _NGROK_PROC and _NGROK_PROC.poll() is None:
                try:
                    _NGROK_PROC.terminate()
                except Exception:
                    pass
                time.sleep(1)
            _NGROK_PROC = subprocess.Popen(
                ["ngrok", "http", str(port), "--authtoken", NGROK_TOKEN, "--log", "stdout", "--log-format", "json"],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1
            )
            found_ngrok = None
            ev_ngrok = threading.Event()

            def _read_ngrok():
                nonlocal found_ngrok
                try:
                    for line in _NGROK_PROC.stdout:
                        if not found_ngrok:
                            try:
                                d = json.loads(line)
                                u = d.get("url", "") or d.get("public_url", "")
                            except Exception:
                                import re as _re
                                m = _re.search(r"https://[\w\-\.]+\.ngrok[\w\-\.]*\.(?:io|app|dev|free\.app)[^\s]*", line)
                                u = m.group(0) if m else ""
                            if u and u.startswith("http"):
                                found_ngrok = u.replace("http://", "https://")
                                ev_ngrok.set()
                except Exception:
                    pass
                if not ev_ngrok.is_set():
                    ev_ngrok.set()

            threading.Thread(target=_read_ngrok, daemon=True, name="NgrokCLI-Reader").start()
            ev_ngrok.wait(25)
            if found_ngrok:
                log.info("✓ Tunnel (ngrok CLI) → %s", found_ngrok)
                return found_ngrok
        except Exception:
            pass

        # cloudflared fallback (disabled by default for this setup)
        if not DISABLE_CLOUDFLARE:
            try:
                if _CF_PROC is not None and _CF_PROC.poll() is None:
                    try:
                        _CF_PROC.terminate()
                    except Exception:
                        pass
                    time.sleep(1)
                _CF_PROC = subprocess.Popen(
                    [CLOUDFLARED, "tunnel", "--url", f"http://localhost:{port}", "--no-autoupdate", "--metrics", "localhost:0"],
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1
                )
                found_cf = None
                ev_cf = threading.Event()
                import re as _re_cf

                def _read_cf():
                    nonlocal found_cf
                    try:
                        for line in _CF_PROC.stdout:
                            if not found_cf:
                                m = _re_cf.search(r"https://[\w\-\.]+\.trycloudflare\.com", line)
                                if m:
                                    found_cf = m.group(0)
                                    ev_cf.set()
                    except Exception:
                        pass
                    if not ev_cf.is_set():
                        ev_cf.set()

                threading.Thread(target=_read_cf, daemon=True, name="CF-Reader").start()
                ev_cf.wait(timeout)
                if found_cf:
                    log.info("✓ Tunnel (cloudflared) → %s", found_cf)
                    return found_cf
            except Exception:
                pass

        # final fallbacks
        u = _try_ssh_tunnel(port, "a.pinggy.io", ssh_port=443, timeout=30)
        if u:
            log.info("✓ Tunnel (Pinggy SSH:443) → %s", u)
            return u
        u = _try_localtunnel(port, timeout=60)
        if u:
            return u

        log.warning("⚠ All tunnel methods failed — LAN only: http://%s:%d", get_lan_ip(), port)
        return None


def _tunnel_guardian(port, lip):
    """
    v10.8 — permanent background thread that keeps the tunnel alive.
    Checks every 60s (was 90s): URL liveness + cloudflared/ngrok proc health.
    Restarts full tunnel chain if dead. Updates dashboard_url.json silently.
    """
    last_url = [_DS.get("pub_url","")]

    def _is_alive(url):
        if not url or not url.startswith("http"): return False
        # localtunnel often blocks generic health probes; trust running process instead.
        if ".loca.lt" in url:
            return (_LT_PROC is not None and _LT_PROC.poll() is None)
        # trycloudflare can intermittently fail local probes (1033) while tunnel process
        # is still active; avoid aggressive URL churn and trust cloudflared process health.
        if ".trycloudflare.com" in url:
            return (_CF_PROC is not None and _CF_PROC.poll() is None)
        if any(h in url for h in (".a.pinggy.io", ".pinggy.link", ".pinggy.io")):
            return (_SSH_PROC is not None and _SSH_PROC.poll() is None)
        try:
            urllib.request.urlopen(url, timeout=6)
            return True
        except Exception:
            return False

    def _save(url):
        try:
            gate_pw = _public_ipv4() if ".loca.lt" in str(url or "") else ""
            os.makedirs(LEVELS_DIR, exist_ok=True)
            with open(os.path.join(LEVELS_DIR,"dashboard_url.json"),"w") as f:
                json.dump({
                    "public_url": url,
                    "port": port,
                    "lan": f"http://{lip}:{port}",
                    "app_password": PUBLIC_LINK_PASSWORD,
                    "tunnel_gate_password": gate_pw,
                }, f)
        except Exception: pass

    dead_streak = 0

    while True:
        time.sleep(60)
        try:
            cur = last_url[0]
            # Check ngrok local API for URL rotations (free-tier occasionally rotates),
            # but only when this dashboard is actually on an ngrok URL.
            if "ngrok" in (cur or "") or (_NGROK_PROC is not None):
                try:
                    # ngrok web API may be on 4040 or fallback 4041.
                    ngrok_api = None
                    for _api in ("http://127.0.0.1:4040/api/tunnels", "http://127.0.0.1:4041/api/tunnels"):
                        try:
                            ngrok_api = urllib.request.urlopen(_api, timeout=2)
                            break
                        except Exception:
                            ngrok_api = None
                    if ngrok_api is not None:
                        for tun in json.loads(ngrok_api.read()).get("tunnels", []):
                            u = tun.get("public_url", "").replace("http://", "https://")
                            # Only accept if this tunnel points to OUR port (8055)
                            cfg_addr = str(tun.get("config", {}).get("addr", ""))
                            if str(port) not in cfg_addr:
                                continue
                            if u.startswith("https://") and u != cur:
                                log.info("Tunnel URL rotated → %s", u)
                                last_url[0] = u; _DS._set(pub_url=u); _save(u); cur = u
                except Exception:
                    pass

            # Check if tunnel process died
            cf_dead    = (_CF_PROC    is not None and _CF_PROC.poll()    is not None)
            ngrok_dead = (_NGROK_PROC is not None and _NGROK_PROC.poll() is not None)
            lt_dead    = (_LT_PROC    is not None and _LT_PROC.poll()    is not None)
            ssh_dead   = (_SSH_PROC   is not None and _SSH_PROC.poll()   is not None)
            proc_dead  = cf_dead or ngrok_dead or lt_dead or ssh_dead

            # ngrok free URLs can transiently fail health checks from localhost.
            # Avoid aggressive restarts unless process died or repeated failures.
            if _is_alive(cur) and not proc_dead:
                dead_streak = 0
                continue
            # Avoid aggressive churn for localtunnel links unless process actually died.
            if ".loca.lt" in (cur or "") and not lt_dead:
                dead_streak = 0
                continue
            # Same for cloudflared quick tunnels: keep URL unless process has died.
            if ".trycloudflare.com" in (cur or "") and not cf_dead:
                dead_streak = 0
                continue
            # Same for Pinggy: trust process health over URL probe noise.
            if any(h in (cur or "") for h in (".a.pinggy.io", ".pinggy.link", ".pinggy.io")) and not ssh_dead:
                dead_streak = 0
                continue
            if not proc_dead:
                dead_streak += 1
                if dead_streak < 5:  # 5 minutes of consecutive failures before recycle
                    log.info("Tunnel probe failed (%d/5); deferring restart", dead_streak)
                    continue

            reason = "process died" if proc_dead else f"URL dead ({cur[:50]})"
            log.warning("Tunnel %s — restarting full chain", reason)

            new_url = open_tunnel(port)
            if new_url:
                dead_streak = 0
                last_url[0] = new_url
                _DS._set(pub_url=new_url)
                _save(new_url)
                log.info("✓ Tunnel revived → %s", new_url)
                # Re-notify Telegram with new URL
                try:
                    _gate = _public_ipv4() if ".loca.lt" in str(new_url or "") else ""
                    _msg = (
                        f"🔄 AlgoStack tunnel restarted\n{new_url}\n"
                        f"Dashboard password: {PUBLIC_LINK_PASSWORD}"
                    )
                    if _gate:
                        _msg += f"\nLocaltunnel gate password: {_gate}"
                    _tg(_msg)
                except Exception: pass
        except Exception as e:
            log.debug("tunnel_guardian: %s", e)


def _tg(text):
    def _go():
        for cid in TG_CHATS:
            try:
                data=urllib.parse.urlencode({"chat_id":cid,"text":text}).encode()
                urllib.request.urlopen(f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage",
                                       data=data,timeout=12)
            except: pass
    threading.Thread(target=_go,daemon=True).start()


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("[STARTUP] main() called — Python and imports OK", flush=True)
    import logging as _lg
    _lg.getLogger("werkzeug").setLevel(_lg.WARNING)
    _lg.getLogger("dash").setLevel(_lg.WARNING)
    lip = get_lan_ip()
    print(f"[STARTUP] LAN IP: {lip}, Port: {DASH_PORT}", flush=True)
    log.info("Building AlgoStack v10.9 dashboard on port %d…", DASH_PORT)

    # v10.8 FIX: Clear any stale public URL from previous session immediately.
    # The old URL may point to Algofinal's port-8050 dashboard or an expired tunnel.
    # We overwrite with LAN URL first; tunnel thread will update once alive.
    try:
        os.makedirs(LEVELS_DIR, exist_ok=True)
        with open(os.path.join(LEVELS_DIR,"dashboard_url.json"),"w") as _f:
            json.dump({"public_url": f"http://{lip}:{DASH_PORT}",
                       "port": DASH_PORT,
                       "lan": f"http://{lip}:{DASH_PORT}",
                       "_note": "tunnel starting…"}, _f)
        _DS._set(pub_url=f"http://{lip}:{DASH_PORT}")
    except Exception: pass
    try:
        app = build_app()
    except Exception as e:
        import sys as _sys
        full_tb = traceback.format_exc()
        err_msg = f"\n{'='*60}\nFATAL build_app crashed:\n{full_tb}\n{'='*60}"
        try: log.error("FATAL build_app: %s\n%s", e, full_tb)
        except Exception: pass
        print(err_msg, flush=True)
        print(err_msg, file=_sys.stderr, flush=True)
        raise SystemExit(1)

    # v10.9 performance: enable response compression + caching
    try:
        _srv = app.server
        _srv.config["COMPRESS_MIMETYPES"] = [
            "text/html","text/css","text/javascript",
            "application/javascript","application/json",
        ]
        _srv.config["COMPRESS_LEVEL"] = 6
        _srv.config["COMPRESS_MIN_SIZE"] = 400
        from flask_compress import Compress
        Compress(_srv)
        log.info("Flask-Compress: gzip level-6 active")
    except ImportError:
        log.info("flask-compress not installed")
    except Exception as _fce:
        log.warning("flask-compress init: %s", _fce)

    # v10.9: Use waitress WSGI server if available — much faster than Flask dev server for tunnels
    try:
        from waitress import serve as _waitress_serve
        log.info("Using waitress WSGI server (8 threads — faster for remote access)")
        threading.Thread(
            target=lambda: _waitress_serve(
                app.server,
                host="0.0.0.0", port=DASH_PORT,
                threads=16,
                connection_limit=300,
                channel_timeout=120,
            ),
            daemon=True, name="DashWaitress").start()
    except ImportError:
        log.info("waitress not installed — using Flask/Dash dev server")
        def _run_flask():
            # app.run() in Dash 2.x; app.run_server() in older Dash
            try:
                app.run(host="0.0.0.0", port=DASH_PORT, debug=False, use_reloader=False)
            except TypeError:
                try:
                    app.run_server(host="0.0.0.0", port=DASH_PORT, debug=False, use_reloader=False)
                except Exception as _fe:
                    log.error("Flask start failed: %s", _fe)
        threading.Thread(target=_run_flask, daemon=True, name="DashFlask").start()
    except Exception as _wsgi_err:
        log.warning("waitress failed (%s) — falling back to Dash dev server", _wsgi_err)
        def _run_flask2():
            try:
                app.run(host="0.0.0.0", port=DASH_PORT, debug=False, use_reloader=False)
            except TypeError:
                app.run_server(host="0.0.0.0", port=DASH_PORT, debug=False, use_reloader=False)
        threading.Thread(target=_run_flask2, daemon=True, name="DashFlask").start()

    # Wait for port to be ready
    deadline = time.monotonic() + 35; ready = False
    while time.monotonic() < deadline:
        try:
            urllib.request.urlopen(f"http://127.0.0.1:{DASH_PORT}", timeout=1)
            ready = True; break
        except Exception: time.sleep(0.4)
    if not ready:
        log.error("Port %d never ready in 35s", DASH_PORT)

    def _bg():
        global _STARTUP_URL_SENT
        if DISABLE_PUBLIC_TUNNEL:
            pub = None
            url = f"http://{lip}:{DASH_PORT}"
            log.info("Public tunnel disabled by DISABLE_PUBLIC_TUNNEL=1")
        else:
            pub = open_tunnel(DASH_PORT)
            url = pub or f"http://{lip}:{DASH_PORT}"
        try:
            gate_pw = _public_ipv4() if ".loca.lt" in str(url or "") else ""
            os.makedirs(LEVELS_DIR, exist_ok=True)
            with open(os.path.join(LEVELS_DIR,"dashboard_url.json"),"w") as f:
                json.dump({
                    "public_url": url,
                    "port": DASH_PORT,
                    "lan": f"http://{lip}:{DASH_PORT}",
                    "app_password": PUBLIC_LINK_PASSWORD,
                    "tunnel_gate_password": gate_pw,
                }, f)
        except Exception: pass
        _DS._set(pub_url=url)

        # Send startup URL once to all 3 bots
        with _STARTUP_URL_LOCK:
            if not _STARTUP_URL_SENT:
                _gate = _public_ipv4() if ".loca.lt" in str(url or "") else ""
                _msg = (
                    f"🟢 AlgoStack v10.7 LIVE\n{url}\n"
                    f"Dashboard password: {PUBLIC_LINK_PASSWORD}\n"
                    f"Equity + Commodity + Crypto | History + Performance | AI Agent"
                )
                if _gate:
                    _msg += f"\nLocaltunnel gate password: {_gate}"
                try:
                    from tg_async import send_startup_url
                    send_startup_url(url)
                    _tg(_msg)
                except Exception:
                    _tg(_msg)
                _STARTUP_URL_SENT = True

        # Start permanent tunnel guardian only when tunnel mode is enabled.
        if not DISABLE_PUBLIC_TUNNEL:
            threading.Thread(target=_tunnel_guardian, args=(DASH_PORT, lip),
                             daemon=True, name="TunnelGuardian").start()

    threading.Thread(target=_bg, daemon=True, name="Tunnel").start()

    log.info("=" * 62)
    log.info("  AlgoStack v10.7 | Author: Ridhaant Ajoy Thackur")
    if DISABLE_PUBLIC_TUNNEL:
        log.info("  http://%s:%d  |  Public host mode (no tunnel)", lip, DASH_PORT)
    else:
        log.info("  http://%s:%d  |  Tunnel → watch Telegram (~30s)", lip, DASH_PORT)
    log.info("  Price refresh: 0.5s fast / 5s slow / 1.5s scanners")
    log.info("  Alert watchdog: equity/MCX/crypto staleness + tunnel monitor")
    log.info("  AI Agent: %s",
             "Gemini 2.0 Flash ✓" if GEMINI_API_KEY else
             ("Anthropic fallback" if ANTHROPIC_API_KEY else "Offline KB only"))
    log.info("=" * 62)

    # v10.6: start AlertMonitor
    try:
        from alert_monitor import start_in_background as _am_start
        _am_start()
        log.info("AlertMonitor v10.7 active")
    except Exception as _ame:
        log.warning("AlertMonitor not loaded: %s", _ame)

    try:
        while True: time.sleep(60)
    except KeyboardInterrupt:
        log.info("Stopped.")


if __name__ == "__main__":
    main()