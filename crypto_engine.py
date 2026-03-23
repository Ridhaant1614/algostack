# ═══════════════════════════════════════════════════════════════════════
# © 2026 Ridhaant Ajoy Thackur. All rights reserved.
# AlgoStack™ is proprietary software. Unauthorised copying or distribution is prohibited.
# AlgoStack v9.0 | Author: Ridhaant Ajoy Thackur
# crypto_engine.py — Crypto trading engine (Binance WS, 24/7, 6h re-anchor)
# ═══════════════════════════════════════════════════════════════════════
"""
crypto_engine.py  v9.0
======================
Crypto trading engine using Binance WebSocket.

SYMBOLS:   BTC, ETH, BNB, SOL, ADA  (all vs USDT)
HOURS:     24/7 — never closes, runs across weekends
X VALUE:   0.008575 — SAME as equity (identical constant)
RE-ANCHOR: Every 6 hours from startup
PRICE:     Binance WS (1s) → CoinCap WS → CoinGecko REST
           Published via ZMQ "crypto" topic → shared with all scanners in 1-5s

Level formula (IDENTICAL to equity — DO NOT CHANGE):
  anchor_price = current Binance price at startup / re-anchor
  x_val        = anchor_price * 0.008575
  buy_above    = anchor_price + x_val
  sell_below   = anchor_price - x_val
  step         = x_val
  T1..T5       = buy_above + step * 1..5
  ST1..ST5     = sell_below - step * 1..5
  buy_sl       = anchor_price (= buy_above - x_val)
  sell_sl      = anchor_price (= sell_below + x_val)

Budget per position:
  budget_usdt = CRYPTO_BUDGET_INR / USDT_TO_INR  (100000 / 84 ≈ $1190)
  qty         = budget_usdt / current_price_usdt

P&L:
  gross_pnl_usdt = (exit_px - entry_px) * qty  (BUY)
  gross_pnl_inr  = gross_pnl_usdt * USDT_TO_INR
  brokerage_inr  = abs(gross_pnl_inr) * 0.001  (0.1%)
  net_pnl_inr    = gross_pnl_inr - brokerage_inr
"""

from __future__ import annotations

import json
import logging
import os
import threading
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional

import pytz
import requests as _req_module

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [CryptoEngine] %(levelname)s — %(message)s",
)
log = logging.getLogger("crypto_engine")
try:
    import urllib3; urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception: pass
import warnings as _w; _w.filterwarnings("ignore", message="Unverified HTTPS")

IST = pytz.timezone("Asia/Kolkata")
UTC = pytz.utc

from config import cfg

# ── Constants ─────────────────────────────────────────────────────────────────
CRYPTO_X         = cfg.CRYPTO_X_MULTIPLIER   # 0.008575
BUDGET_INR       = cfg.CRYPTO_BUDGET_INR      # 100000
# v10.9: Live USD/INR rate (updated at startup; falls back to config value)
def _fetch_usdt_inr() -> float:
    try:
        import urllib.request as _ur, json as _jj
        r = _ur.Request("https://cdn.jsdelivr.net/npm/@fawazahmed0/currency-api@latest/v1/currencies/usd.json",
                        headers={"User-Agent":"AlgoStack/10.9"})
        with _ur.urlopen(r, timeout=5) as resp:
            rate = float(_jj.loads(resp.read())["usd"]["inr"])
            if 70 < rate < 120: return round(rate, 2)
    except Exception:
        pass
    return cfg.USDT_TO_INR
USDT_TO_INR = _fetch_usdt_inr()   # fetched live at startup, falls back to 84.0
BROKERAGE_PCT    = cfg.CRYPTO_BROKERAGE_PCT   # 0.001
REANCHOR_HOURS   = 6

SYMBOLS: List[str] = ["BTC", "ETH", "BNB", "SOL", "ADA"]

BINANCE_SYMBOLS: Dict[str, str] = {
    "BTCUSDT": "BTC",
    "ETHUSDT": "ETH",
    "BNBUSDT": "BNB",
    "SOLUSDT": "SOL",
    "ADAUSDT": "ADA",
}

COINCAP_ASSETS: Dict[str, str] = {
    "bitcoin":      "BTC",
    "ethereum":     "ETH",
    "binance-coin": "BNB",
    "solana":       "SOL",
    "cardano":      "ADA",
}

COINGECKO_IDS: Dict[str, str] = {
    "BTC": "bitcoin",
    "ETH": "ethereum",
    "BNB": "binancecoin",
    "SOL": "solana",
    "ADA": "cardano",
}

# ── Shared state ──────────────────────────────────────────────────────────────
_CRYPTO_PRICES:    Dict[str, float] = {}   # {BTC: 85234.50, ...} in USDT
_CRYPTO_ANCHOR:    Dict[str, float] = {}   # anchor prices per 6h window
_CRYPTO_LEVELS:    Dict[str, dict]  = {}   # calculated levels per symbol
_CRYPTO_POSITIONS: Dict[str, Optional[dict]] = {s: None for s in SYMBOLS}
_CRYPTO_EXITED:    Dict[str, bool]  = {s: False for s in SYMBOLS}
_CRYPTO_TRADES:    List[dict]       = []
_PRICE_LOCK        = threading.Lock()
_PRICE_AGE:        Dict[str, float] = {}   # monotonic time of last update

# Expose for scanners via import
CRYPTO_ANCHOR: Dict[str, float] = {}    # scanners import this as prev_close equivalent
_STARTUP_TIME: Optional[datetime] = None
_NEXT_ANCHOR:  Optional[datetime] = None


# ════════════════════════════════════════════════════════════════════════════
# TIME HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _now_utc() -> datetime:
    return datetime.now(UTC)


def _now_ist() -> datetime:
    return datetime.now(IST)


# ════════════════════════════════════════════════════════════════════════════
# LEVEL CALCULATOR  (identical formula to equity)
# ════════════════════════════════════════════════════════════════════════════

def _calc_levels(sym: str, anchor: float) -> dict:
    """Calculate all trading levels. Formula IDENTICAL to equity.
    v10.7 FIX: Added full INR fields for T1-T5, ST1-ST5, stop_loss, stop_loss_inr.
    """
    xval = anchor * CRYPTO_X
    step = xval
    ba   = anchor + xval
    sb   = anchor - xval
    t1   = round(ba + step * 1, 4)
    t2   = round(ba + step * 2, 4)
    t3   = round(ba + step * 3, 4)
    t4   = round(ba + step * 4, 4)
    t5   = round(ba + step * 5, 4)
    st1  = round(sb - step * 1, 4)
    st2  = round(sb - step * 2, 4)
    st3  = round(sb - step * 3, 4)
    st4  = round(sb - step * 4, 4)
    st5  = round(sb - step * 5, 4)
    sl   = round(anchor, 4)   # stop-loss = anchor price (buy_above - xval = anchor)
    return {
        "symbol":           sym,
        "anchor":           round(anchor, 4),
        "x_val":            round(xval, 4),
        "x_mult":           CRYPTO_X,
        "buy_above":        round(ba, 4),
        "sell_below":       round(sb, 4),
        "buy_sl":           sl,
        "sell_sl":          sl,
        "stop_loss":        sl,
        "stop_loss_inr":    round(sl * USDT_TO_INR, 2),
        "step":             round(step, 4),
        "T1":  t1,  "T2":  t2,  "T3":  t3,  "T4":  t4,  "T5":  t5,
        "ST1": st1, "ST2": st2, "ST3": st3, "ST4": st4, "ST5": st5,
        # ── INR equivalents (dashboard display) ──────────────────────────────
        "buy_above_inr":    round(ba  * USDT_TO_INR, 2),
        "sell_below_inr":   round(sb  * USDT_TO_INR, 2),
        "T1_inr":  round(t1  * USDT_TO_INR, 2),
        "T2_inr":  round(t2  * USDT_TO_INR, 2),
        "T3_inr":  round(t3  * USDT_TO_INR, 2),
        "T4_inr":  round(t4  * USDT_TO_INR, 2),
        "T5_inr":  round(t5  * USDT_TO_INR, 2),
        "ST1_inr": round(st1 * USDT_TO_INR, 2),
        "ST2_inr": round(st2 * USDT_TO_INR, 2),
        "ST3_inr": round(st3 * USDT_TO_INR, 2),
        "ST4_inr": round(st4 * USDT_TO_INR, 2),
        "ST5_inr": round(st5 * USDT_TO_INR, 2),
        # ── Retreat thresholds ────────────────────────────────────────────────
        "retreat_65": round(ba + 0.65 * step, 4),
        "retreat_45": round(ba + 0.45 * step, 4),
        "retreat_25": round(ba + 0.25 * step, 4),
    }


def _crypto_qty(price_usdt: float) -> float:
    """Position size: budget_usdt / current_price."""
    budget = BUDGET_INR / USDT_TO_INR
    return round(budget / price_usdt, 8) if price_usdt > 0 else 0.0


# ════════════════════════════════════════════════════════════════════════════
# PRICE FEED — Binance WebSocket (primary, 1s updates)
# ════════════════════════════════════════════════════════════════════════════

def _start_binance_ws() -> None:
    t = threading.Thread(target=_binance_ws_loop, daemon=True, name="Binance-WS")
    t.start()


def _binance_ws_loop() -> None:
    try:
        import websocket
    except ImportError:
        log.warning("websocket-client not installed — falling back to REST polling")
        _start_coingecko_poll()
        return

    url = "wss://stream.binance.com:9443/ws/!miniTicker@arr"

    def on_message(ws, message):
        try:
            tickers = json.loads(message)
            if not isinstance(tickers, list):
                return
            now_t = time.monotonic()
            updated = {}
            for t in tickers:
                sym_map = BINANCE_SYMBOLS.get(t.get("s", ""))
                if sym_map:
                    price = float(t.get("c", 0))
                    if price > 0:
                        with _PRICE_LOCK:
                            _CRYPTO_PRICES[sym_map] = price
                            _PRICE_AGE[sym_map] = now_t
                        updated[sym_map] = price
            if updated:
                _publish_prices(updated)
        except Exception:
            pass

    def on_error(ws, err):
        log.debug("Binance WS error: %s", err)

    def on_close(ws, *_):
        log.debug("Binance WS closed — reconnecting in 5s")

    _reconnect_delay = 2
    _MAX_RECONNECT_DELAY = 60
    _last_msg_t = [time.monotonic()]

    while True:
        try:
            def on_open_reset(ws):
                nonlocal _reconnect_delay
                _reconnect_delay = 2  # reset backoff on successful connect
                log.info("Binance WS connected — streaming %d symbols", len(BINANCE_SYMBOLS))

            _orig_on_msg = on_message
            def on_msg_track(ws, message):
                _last_msg_t[0] = time.monotonic()
                global _LAST_WS_PRICE_TIME
                _LAST_WS_PRICE_TIME = time.monotonic()  # update for price_loop failsafe
                _orig_on_msg(ws, message)

            ws = websocket.WebSocketApp(
                url,
                on_message=on_msg_track,
                on_error=on_error,
                on_close=on_close,
                on_open=on_open_reset,
            )
            ws.run_forever(ping_interval=15, ping_timeout=8,
                           skip_utf8_validation=True)   # 10% faster parsing
        except Exception as exc:
            log.debug("Binance WS loop error: %s", exc)
        # Exponential backoff
        log.info("Binance WS reconnecting in %ds...", _reconnect_delay)
        time.sleep(_reconnect_delay)
        _reconnect_delay = min(_reconnect_delay * 2, _MAX_RECONNECT_DELAY)
        # If very stale (>30s no data), launch CoinCap WS in parallel
        if time.monotonic() - _last_msg_t[0] > 30:
            _try_coincap_fallback()


def _try_coincap_fallback() -> None:
    """Try CoinCap WebSocket if Binance is down."""
    try:
        import websocket
        symbols_str = ",".join(COINCAP_ASSETS.keys())
        url = f"wss://ws.coincap.io/prices?assets={symbols_str}"

        def on_msg(ws, msg):
            data = json.loads(msg)
            now_t = time.monotonic()
            for asset, price_str in data.items():
                sym = COINCAP_ASSETS.get(asset)
                if sym:
                    try:
                        price = float(price_str)
                        if price > 0:
                            with _PRICE_LOCK:
                                _CRYPTO_PRICES[sym] = price
                                _PRICE_AGE[sym] = now_t
                    except ValueError:
                        pass

        ws = websocket.WebSocketApp(url, on_message=on_msg)
        ws.run_forever(ping_interval=20, ping_timeout=10)
    except Exception:
        pass


def _start_coingecko_poll() -> None:
    """CoinGecko REST polling fallback (every 5s, respects 30 calls/min limit)."""
    t = threading.Thread(target=_coingecko_loop, daemon=True, name="CoinGecko-Poll")
    t.start()


def _coingecko_loop() -> None:
    ids_str = ",".join(COINGECKO_IDS.values())
    url = (f"https://api.coingecko.com/api/v3/simple/price"
           f"?ids={ids_str}&vs_currencies=usd")
    sym_by_id = {v: k for k, v in COINGECKO_IDS.items()}
    while True:
        try:
            r = _req_module.get(url, timeout=5)
            if r.status_code == 200:
                data = r.json()
                now_t = time.monotonic()
                updated = {}
                for cg_id, prices in data.items():
                    sym = sym_by_id.get(cg_id)
                    if sym:
                        price = float(prices.get("usd", 0))
                        if price > 0:
                            with _PRICE_LOCK:
                                _CRYPTO_PRICES[sym] = price
                                _PRICE_AGE[sym] = now_t
                            updated[sym] = price
                if updated:
                    _publish_prices(updated)
        except Exception as exc:
            log.debug("CoinGecko poll error: %s", exc)
        time.sleep(5)


# ════════════════════════════════════════════════════════════════════════════
# ZMQ PUBLISHER  (shares prices with all crypto scanners in 1-5s)
# ════════════════════════════════════════════════════════════════════════════

_pub = None


def _init_zmq() -> None:
    global _pub
    try:
        from ipc_bus import PricePublisher
        _pub = PricePublisher()
        log.info("ZMQ PUB initialised — publishing on 'crypto' topic")
    except Exception as exc:
        log.warning("ZMQ init failed: %s", exc)


def _publish_prices(prices: Dict[str, float]) -> None:
    """Publish crypto prices via ZMQ 'crypto' topic AND update live_prices.json."""
    if _pub is None:
        # Even without ZMQ, write JSON so scanners can find prices
        try:
            import json as _j, os as _o
            _o.makedirs("levels", exist_ok=True)
            lp = _o.path.join("levels", "live_prices.json")
            try:
                existing = _j.load(open(lp, encoding="utf-8")) if _o.path.exists(lp) else {}
            except Exception:
                existing = {}
            ts_now = _now_ist().strftime("%Y-%m-%d %H:%M:%S")
            existing["crypto_prices"] = prices
            existing["crypto_ts"] = ts_now
            existing["ts"] = ts_now
            # DO NOT touch "prices" or "equity_prices" keys — those belong to price_service
            tmp = lp + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                _j.dump(existing, f, separators=(",", ":"))
            _o.replace(tmp, lp)
        except Exception:
            pass
        return
    try:
        # Publish on ZMQ "crypto" topic (scanners subscribe to this)
        _pub.publish(prices, _now_ist(), topic="crypto")
        # Also write crypto_prices section to live_prices.json for scanner fallback
        try:
            import json as _j, os as _o
            _o.makedirs("levels", exist_ok=True)
            lp = _o.path.join("levels", "live_prices.json")
            try:
                existing = _j.load(open(lp, encoding="utf-8")) if _o.path.exists(lp) else {}
            except Exception:
                existing = {}
            ts_now = _now_ist().strftime("%Y-%m-%d %H:%M:%S")
            existing["crypto_prices"] = prices
            existing["crypto_ts"] = ts_now
            existing["ts"] = ts_now
            # DO NOT touch "prices" or "equity_prices" keys — those belong to price_service
            tmp = lp + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                _j.dump(existing, f, separators=(",", ":"))
            _o.replace(tmp, lp)
        except Exception:
            pass
    except Exception:
        pass


# ════════════════════════════════════════════════════════════════════════════
# TRADE LOGGING
# ════════════════════════════════════════════════════════════════════════════

def _log_trade(event: dict) -> None:
    _CRYPTO_TRADES.append(event)
    date_str = _now_utc().strftime("%Y%m%d")
    os.makedirs("trade_logs", exist_ok=True)
    path = os.path.join("trade_logs", f"crypto_trade_events_{date_str}.jsonl")
    try:
        with open(path, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(event) + "\n")
    except Exception:
        pass


# ════════════════════════════════════════════════════════════════════════════
# EXCEL WRITERS
# ════════════════════════════════════════════════════════════════════════════

def _write_levels_xlsx(path: str, ts: datetime) -> None:
    """Write crypto initial levels Excel (prices in USD + INR equivalent)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Crypto Levels"
        hdr = ["Symbol", "Anchor ($)", "X Mult", "Buy Above ($)", "T1 ($)", "T2 ($)", "T3 ($)",
               "Sell Below ($)", "ST1 ($)", "ST2 ($)", "ST3 ($)",
               "Buy Above (Rs)", "Sell Below (Rs)", "USDT/INR"]
        for col, h in enumerate(hdr, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1A3A6B")
        for sym in SYMBOLS:
            lv = _CRYPTO_LEVELS.get(sym, {})
            ws.append([
                sym,
                lv.get("anchor", 0),
                lv.get("x_mult", CRYPTO_X),
                lv.get("buy_above", 0),
                lv.get("T1", 0), lv.get("T2", 0), lv.get("T3", 0),
                lv.get("sell_below", 0),
                lv.get("ST1", 0), lv.get("ST2", 0), lv.get("ST3", 0),
                lv.get("buy_above_inr", 0),
                lv.get("sell_below_inr", 0),
                USDT_TO_INR,
            ])
        ws.append([])
        ws.append(["Author: Ridhaant Ajoy Thackur", "", f"AlgoStack v9.0 | {ts.strftime('%Y-%m-%d %H:%M')} UTC"])
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        tmp = path + ".tmp"
        wb.save(tmp)
        os.replace(tmp, path)
        log.info("Crypto levels XLSX → %s", path)
    except Exception as exc:
        log.warning("Crypto XLSX write failed: %s", exc)


def _write_trade_analysis_xlsx(path: str, ts: datetime, trades: List[dict]) -> None:
    """Write 6h trade analysis Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Crypto Trades"
        hdr = ["Symbol", "Side", "Entry ($)", "Exit ($)", "Qty",
               "Gross (USDT)", "Gross (Rs)", "Brokerage (Rs)", "Net (Rs)",
               "Exit Type", "Time", "X Used"]
        for col, h in enumerate(hdr, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1A3A6B")
        for t in trades:
            ws.append([
                t.get("symbol"), t.get("side"),
                t.get("entry_px"), t.get("exit_px"), t.get("qty"),
                t.get("gross_pnl_usdt"), t.get("gross_pnl_inr"),
                t.get("brokerage_inr"), t.get("net_pnl_inr"),
                t.get("reason"), t.get("ts"), t.get("x_val"),
            ])
        ws.append([])
        ws.append(["Author: Ridhaant Ajoy Thackur", "", f"AlgoStack v9.0"])
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        tmp = path + ".tmp"
        wb.save(tmp)
        os.replace(tmp, path)
    except Exception as exc:
        log.warning("Trade analysis XLSX write failed: %s", exc)


# ════════════════════════════════════════════════════════════════════════════
# TRADING LOGIC  (identical to equity logic)
# ════════════════════════════════════════════════════════════════════════════

# Per-symbol last-processed price for dedup
_LAST_PROCESSED: Dict[str, float] = {}
_MIN_PRICE_CHANGE_PCT = 0.00001  # 0.001% — skip sub-cent noise
_EXITED_RESET_LOGGED: bool = False   # guard for logging

# v10.6: Per-symbol re-entry watch state (threshold + retouch, mirrors Algofinal)
_CRYPTO_REENTRY: Dict[str, Optional[dict]] = {s: None for s in SYMBOLS}

# v10.6: Flat brokerage ₹20/round-trip (was % of trade — incorrect for crypto)
BROKERAGE_FLAT_INR = 20.0  # ₹10 entry + ₹10 exit

def _maybe_reset_exited() -> None:
    """Reset _CRYPTO_EXITED for any symbol where no position is open.
    Called each price loop tick so entries are never permanently locked
    between re-anchor windows (fixes 'no crypto trades taken' bug).
    """
    global _EXITED_RESET_LOGGED
    reset_syms = []
    for sym in SYMBOLS:
        if _CRYPTO_EXITED[sym] and _CRYPTO_POSITIONS[sym] is None:
            # Only reset if we've been in exited state and no re-anchor imminent
            if _NEXT_ANCHOR is not None:
                now = _now_utc()
                mins_to_anchor = (_NEXT_ANCHOR - now).total_seconds() / 60
                # Reset exit flag if more than 30 min until next re-anchor
                # This allows re-entry in the same session if levels are still valid
                if mins_to_anchor > 30:
                    _CRYPTO_EXITED[sym] = False
                    reset_syms.append(sym)
    if reset_syms and not _EXITED_RESET_LOGGED:
        log.info("Re-entry unlocked for: %s (30m+ until re-anchor)", reset_syms)
        _EXITED_RESET_LOGGED = True


# ════════════════════════════════════════════════════════════════════════════
# UNIFIED ALERT BUILDER  (exact mirror of Algofinal.build_simple_alert)
# ════════════════════════════════════════════════════════════════════════════

def _cr_fmt_pct(v: float) -> str:
    return f"{'+' if v >= 0 else ''}{v:.2f}%"


def _build_crypto_alert(
    title: str, symbol: str, status_line: str,
    *, anchor: float, current_price: float,
    x_val: float, quantity, side: str, lv: dict, ts_str: str,
) -> str:
    """
    Exact mirror of Algofinal.build_simple_alert() for crypto.
    Uses anchor as 'Previous Close' equivalent.
    Alert types fired: Entry, Exit (Target/SL/Retreat/Re-anchor), Re-entry Armed, Re-entry Entry.
    NO proximity/warning alerts — only when levels are actually breached.
    """
    change_pct = (current_price - anchor) / anchor * 100.0 if anchor else 0.0
    # INR equivalent for display
    cp_inr = current_price * USDT_TO_INR
    lines = [f"🚨 {symbol} [₿ Crypto] — {title} at {ts_str}", ""]
    lines += [
        f"Anchor Price: ${anchor:,.4f}  (≈₹{anchor*USDT_TO_INR:,.0f})",
        f"Current Price: ${current_price:,.4f}  (≈₹{cp_inr:,.0f})",
        f"Change: {_cr_fmt_pct(change_pct)}",
        f"Deviation (X): {x_val:.6f}",
    ]
    if quantity is not None:
        lines.append(f"Quantity: {quantity:.6f}")
    lines += [f"Status: {status_line}", "", "📊 Technical Analysis:"]
    if side == "BUY":
        lines.append("📈 Buy Levels:")
        lines.append(f"Buy Above: ${lv['buy_above']:,.4f}")
        for i in range(1, 6):
            tgt = lv.get(f"T{i}", 0)
            pct = (tgt - current_price) / current_price * 100
            lines.append(f"Target {i}: ${tgt:,.4f}  ({_cr_fmt_pct(pct)})")
        sl  = lv.get("buy_sl", anchor)
        pct = (sl - current_price) / current_price * 100
        lines.append(f"Stop Loss: ${sl:,.4f}  ({_cr_fmt_pct(pct)})")
    else:
        lines.append("📉 Sell Levels:")
        lines.append(f"Sell Below: ${lv['sell_below']:,.4f}")
        for i in range(1, 6):
            tgt = lv.get(f"ST{i}", 0)
            pct = (tgt - current_price) / current_price * 100
            lines.append(f"Target {i}: ${tgt:,.4f}  ({_cr_fmt_pct(pct)})")
        sl  = lv.get("sell_sl", anchor)
        pct = (sl - current_price) / current_price * 100
        lines.append(f"Stop Loss: ${sl:,.4f}  ({_cr_fmt_pct(pct)})")
    return "\n".join(lines)


def _send_alert(msg: str) -> None:
    try:
        from tg_async import send_alert
        send_alert(msg, asset_class="crypto")
    except Exception:
        pass


# ════════════════════════════════════════════════════════════════════════════
# 6-HOUR RE-ANCHOR
# ════════════════════════════════════════════════════════════════════════════

def _process_price(sym: str, price: float, ts: datetime) -> None:
    """Process one price tick. Mirrors Algofinal logic exactly:
    v10.6: Unified alert format, flat ₹20 brokerage, full re-entry watch.
    Alert types: Entry, Target Hit, SL Hit, Retreat Exit, Re-entry Armed, Re-entry Entry.
    NO proximity/warning alerts — only when levels are actually breached.
    """
    if price <= 0:
        return
    last = _LAST_PROCESSED.get(sym, 0.0)
    if last > 0 and abs(price - last) / last < _MIN_PRICE_CHANGE_PCT:
        return
    _LAST_PROCESSED[sym] = price

    lv = _CRYPTO_LEVELS.get(sym)
    if not lv:
        return

    pos    = _CRYPTO_POSITIONS[sym]
    exited = _CRYPTO_EXITED[sym]
    ts_str = ts.strftime("%Y-%m-%d %H:%M:%S IST+0530")
    qty    = _crypto_qty(price)
    anchor = lv.get("anchor", price)
    step   = lv.get("step", anchor * CRYPTO_X)

    def _alert_kw():
        return dict(anchor=anchor, current_price=price,
                    x_val=CRYPTO_X, lv=lv, ts_str=ts_str)

    # ── Re-entry watch ────────────────────────────────────────────────────────
    rw = _CRYPTO_REENTRY[sym]
    if pos is None and rw is not None:
        td = rw["threshold_direction"]
        if td == "ABOVE" and price >= rw["threshold_price"]:
            _CRYPTO_REENTRY[sym] = None
            _arm_reentry_position(sym, rw, "threshold", price, qty, ts_str, lv, anchor, step)
            return
        elif td == "BELOW" and price <= rw["threshold_price"]:
            _CRYPTO_REENTRY[sym] = None
            _arm_reentry_position(sym, rw, "threshold", price, qty, ts_str, lv, anchor, step)
            return
        if td == "ABOVE" and price <= rw["target_price"]:
            _CRYPTO_REENTRY[sym] = None
            _arm_reentry_position(sym, rw, "retouch", price, qty, ts_str, lv, anchor, step)
            return
        elif td == "BELOW" and price >= rw["target_price"]:
            _CRYPTO_REENTRY[sym] = None
            _arm_reentry_position(sym, rw, "retouch", price, qty, ts_str, lv, anchor, step)
            return

    # ── Fresh entry ────────────────────────────────────────────────────────────
    if pos is None and not exited:
        if qty <= 0:
            return
        if price >= lv["buy_above"]:
            _CRYPTO_POSITIONS[sym] = {"side": "BUY", "entry_px": price, "qty": qty,
                "targets": [lv[f"T{i}"] for i in range(1, 6)],
                "buy_sl": lv["buy_sl"], "retreat_peak_reached": False, "ts": ts_str}
            cap_inr = qty * price * USDT_TO_INR
            _send_alert(_build_crypto_alert("Entry", sym,
                f"BUY TRIGGERED | Capital ₹{cap_inr:,.0f}", side="BUY",
                quantity=qty, **_alert_kw()))
            _log_trade({"ts": ts_str, "symbol": sym, "side": "BUY",
                "entry_px": price, "exit_px": None, "qty": qty,
                "gross_pnl_usdt": None, "gross_pnl_inr": None,
                "brokerage_inr": None, "net_pnl_inr": None,
                "reason": "ENTRY", "x_val": CRYPTO_X,
                "asset_class": "crypto", "currency": "USDT"})
            log.info("BUY %s @ $%.4f qty=%.6f", sym, price, qty)
        elif price <= lv["sell_below"]:
            _CRYPTO_POSITIONS[sym] = {"side": "SELL", "entry_px": price, "qty": qty,
                "targets": [lv[f"ST{i}"] for i in range(1, 6)],
                "sell_sl": lv["sell_sl"], "retreat_peak_reached": False, "ts": ts_str}
            cap_inr = qty * price * USDT_TO_INR
            _send_alert(_build_crypto_alert("Entry", sym,
                f"SELL TRIGGERED | Capital ₹{cap_inr:,.0f}", side="SELL",
                quantity=qty, **_alert_kw()))
            _log_trade({"ts": ts_str, "symbol": sym, "side": "SELL",
                "entry_px": price, "exit_px": None, "qty": qty,
                "gross_pnl_usdt": None, "gross_pnl_inr": None,
                "brokerage_inr": None, "net_pnl_inr": None,
                "reason": "ENTRY", "x_val": CRYPTO_X,
                "asset_class": "crypto", "currency": "USDT"})
            log.info("SELL %s @ $%.4f qty=%.6f", sym, price, qty)
        return

    if pos is None:
        return

    side     = pos["side"]
    entry_px = pos["entry_px"]
    qty      = pos["qty"]

    def _close(exit_px: float, reason: str,
                arm_re: bool = False, tgt_lbl: str = "", tgt_px: float = 0.0):
        gross_u = (exit_px-entry_px)*qty if side=="BUY" else (entry_px-exit_px)*qty
        gross_i = gross_u * USDT_TO_INR
        net_i   = gross_i - BROKERAGE_FLAT_INR
        pct     = net_i / BUDGET_INR * 100
        _CRYPTO_POSITIONS[sym] = None; _CRYPTO_EXITED[sym] = True; _CRYPTO_REENTRY[sym] = None
        emj = {"T1_HIT":"✅","T2_HIT":"✅","T3_HIT":"✅","T4_HIT":"✅","T5_HIT":"✅",
               "ST1_HIT":"✅","ST2_HIT":"✅","ST3_HIT":"✅","ST4_HIT":"✅","ST5_HIT":"✅",
               "SL_HIT":"🛑","RETREAT":"↩️","RE_ANCHOR":"🔄"}.get(reason,"📤")
        status = (f"{emj} {reason} | Side: {side}\n"
                  f"Gross: ${gross_u:+.4f}  (₹{gross_i:+,.0f}) | "
                  f"Brokerage: ₹{BROKERAGE_FLAT_INR:.2f} | "
                  f"Net: ₹{net_i:+,.0f}  ({pct:+.3f}%)")
        _send_alert(_build_crypto_alert("Exit", sym, status, side=side, quantity=qty,
            anchor=anchor, current_price=exit_px, x_val=CRYPTO_X, lv=lv, ts_str=ts_str))
        _log_trade({"ts": ts_str, "symbol": sym, "side": side,
            "entry_px": entry_px, "exit_px": exit_px, "qty": qty,
            "gross_pnl_usdt": round(gross_u, 6), "gross_pnl_inr": round(gross_i, 2),
            "brokerage_inr": BROKERAGE_FLAT_INR, "net_pnl_inr": round(net_i, 2),
            "reason": reason, "x_val": CRYPTO_X, "asset_class": "crypto", "currency": "USDT"})
        log.info("%s %s @ $%.4f net_inr=%.2f", reason, sym, exit_px, net_i)
        if arm_re and tgt_lbl and tgt_px > 0:
            _schedule_crypto_reentry(sym, side, tgt_lbl, tgt_px, step, lv, CRYPTO_X)

    if side == "BUY":
        for i, tgt in enumerate(pos["targets"], 1):
            if price >= tgt:
                _close(tgt, f"T{i}_HIT", arm_re=True, tgt_lbl=f"T{i}", tgt_px=tgt); return
        sl = pos.get("buy_sl", lv["buy_sl"])
        if price <= sl:
            g = (price-entry_px)*qty; n = g - BROKERAGE_FLAT_INR
            _close(price, "SL_HIT", arm_re=False); return
        if price >= lv["buy_above"] + 0.65*step: pos["retreat_peak_reached"] = True
        if pos["retreat_peak_reached"] and price <= lv["buy_above"] + 0.25*step:
            _close(price, "RETREAT", arm_re=False); return
    else:
        for i, tgt in enumerate(pos["targets"], 1):
            if price <= tgt:
                _close(tgt, f"ST{i}_HIT", arm_re=True, tgt_lbl=f"ST{i}", tgt_px=tgt); return
        sl = pos.get("sell_sl", lv["sell_sl"])
        if price >= sl:
            _close(price, "SL_HIT", arm_re=False); return
        if price <= lv["sell_below"] - 0.65*step: pos["retreat_peak_reached"] = True
        if pos["retreat_peak_reached"] and price >= lv["sell_below"] - 0.25*step:
            _close(price, "RETREAT", arm_re=False); return


def _schedule_crypto_reentry(sym: str, side: str, tgt_lbl: str, tgt_px: float,
                              step: float, lv: dict, x_val: float) -> None:
    """Arm re-entry watch and send 'Re-entry monitoring armed' alert."""
    buffer = 0.75 * x_val
    if side == "BUY":
        rw = {"threshold_direction": "ABOVE",
              "threshold_price": tgt_px + buffer,
              "threshold_entry_side": "BUY",
              "threshold_entry_price": tgt_px + step,
              "threshold_entry_sl": tgt_px,
              "threshold_entry_target": tgt_px + step * 2,
              "retouch_entry_side": "SELL",
              "retouch_entry_price": tgt_px,
              "retouch_entry_sl": tgt_px + step,
              "retouch_entry_target": tgt_px - step,
              "target_price": tgt_px}
    else:
        rw = {"threshold_direction": "BELOW",
              "threshold_price": tgt_px - buffer,
              "threshold_entry_side": "SELL",
              "threshold_entry_price": tgt_px - step,
              "threshold_entry_sl": tgt_px,
              "threshold_entry_target": tgt_px - step * 2,
              "retouch_entry_side": "BUY",
              "retouch_entry_price": tgt_px,
              "retouch_entry_sl": tgt_px - step,
              "retouch_entry_target": tgt_px + step,
              "target_price": tgt_px}
    _CRYPTO_REENTRY[sym] = rw
    td = "≥" if side=="BUY" else "≤"; rd = "≤" if side=="BUY" else "≥"
    armed = (f"🔁 {sym} [₿ Crypto] — Re-entry monitoring armed after {tgt_lbl}\n\n"
             f"LSP=${tgt_px:,.4f} (≈₹{tgt_px*USDT_TO_INR:,.0f}) | "
             f"New BA=${rw['threshold_entry_price']:,.4f} | "
             f"New SB=${rw['retouch_entry_price']:,.4f}\n"
             f"▶ Threshold: price {td} ${rw['threshold_price']:,.4f} "
             f"→ {rw['threshold_entry_side']} @ ${rw['threshold_entry_price']:,.4f}  "
             f"SL ${rw['threshold_entry_sl']:,.4f}  Tgt ${rw['threshold_entry_target']:,.4f}\n"
             f"▶ Retouch: price {rd} ${tgt_px:,.4f} "
             f"→ {rw['retouch_entry_side']} @ ${rw['retouch_entry_price']:,.4f}  "
             f"SL ${rw['retouch_entry_sl']:,.4f}  Tgt ${rw['retouch_entry_target']:,.4f}")
    _send_alert(armed)


def _arm_reentry_position(sym: str, rw: dict, mode: str,
                           price: float, qty: float, ts_str: str,
                           lv: dict, anchor: float, step: float) -> None:
    """Open re-entry position."""
    if mode == "threshold":
        sr = rw["threshold_entry_side"]; er = rw["threshold_entry_price"]
        slr = rw["threshold_entry_sl"]; tr = rw["threshold_entry_target"]
    else:
        sr = rw["retouch_entry_side"]; er = rw["retouch_entry_price"]
        slr = rw["retouch_entry_sl"]; tr = rw["retouch_entry_target"]
    targets = ([tr + step*i for i in range(0,5)] if sr=="BUY"
               else [tr - step*i for i in range(0,5)])
    _CRYPTO_POSITIONS[sym] = {"side": sr, "entry_px": er, "qty": qty,
        "targets": targets, f"{'buy' if sr=='BUY' else 'sell'}_sl": slr,
        "retreat_peak_reached": False, "ts": ts_str, "is_reentry": True}
    mode_l = mode.upper()
    _send_alert(_build_crypto_alert("Re-entry Entry", sym,
        f"{mode_l} RE-ENTRY {sr} @ ${er:,.4f} | SL ${slr:,.4f} | Tgt ${tr:,.4f}",
        side=sr, quantity=qty, anchor=anchor, current_price=er,
        x_val=CRYPTO_X, lv=lv, ts_str=ts_str))
    _log_trade({"ts": ts_str, "symbol": sym, "side": sr,
        "entry_px": er, "exit_px": None, "qty": qty,
        "gross_pnl_usdt": None, "gross_pnl_inr": None,
        "brokerage_inr": None, "net_pnl_inr": None,
        "reason": f"{sr}_REENTRY_{mode_l}", "x_val": CRYPTO_X,
        "asset_class": "crypto", "currency": "USDT"})
    log.info("REENTRY_%s %s %s @ $%.4f", mode_l, sr, sym, er)



def _do_reanchor(ts: datetime, prev_trades: list) -> None:
    """Perform 6-hour re-anchor: square-off open positions, recalculate levels."""
    global _CRYPTO_EXITED
    from tg_async import send_alert, send_document_alert, get_dashboard_url

    ts_str   = ts.strftime("%H:%M:%S")
    ts_tag   = ts.strftime("%Y%m%d_%H%M")
    date_str = ts.strftime("%Y%m%d")
    log.info("═" * 50)
    log.info("CRYPTO RE-ANCHOR @ %s UTC", ts_str)

    # Square off any open positions at current price
    for sym in SYMBOLS:
        pos = _CRYPTO_POSITIONS[sym]
        if pos is None:
            continue
        with _PRICE_LOCK:
            price = _CRYPTO_PRICES.get(sym, pos["entry_px"])
        side     = pos["side"]
        entry_px = pos["entry_px"]
        qty      = pos["qty"]
        gross_u  = (price - entry_px) * qty if side == "BUY" else (entry_px - price) * qty
        gross_i  = gross_u * USDT_TO_INR
        brok_i   = BROKERAGE_FLAT_INR  # v10.6: flat ₹20 (was % of gross)
        net_i    = gross_i - brok_i
        _CRYPTO_POSITIONS[sym] = None
        _log_trade({"ts": ts_str, "symbol": sym, "side": side,
                    "entry_px": entry_px, "exit_px": price, "qty": qty,
                    "gross_pnl_usdt": round(gross_u, 6),
                    "gross_pnl_inr": round(gross_i, 2),
                    "brokerage_inr": round(brok_i, 2),
                    "net_pnl_inr": round(net_i, 2),
                    "reason": "RE_ANCHOR",
                    "x_val": CRYPTO_X, "asset_class": "crypto", "currency": "USDT"})
        log.info("RE-ANCHOR square-off: %s %s @ $%.4f net_inr=%.2f",
                 side, sym, price, net_i)

    # Write 6h trade analysis
    if prev_trades:
        xl_path = os.path.join("trade_analysis",
                               f"crypto_trade_analysis_{ts_tag}.xlsx")
        _write_trade_analysis_xlsx(xl_path, ts, prev_trades)
        prev_start = prev_trades[0].get("ts", "?") if prev_trades else "?"
        send_document_alert(
            xl_path,
            f"Crypto 6h report {prev_start}–{ts_str} ({date_str})",
            asset_class="crypto"
        )

    # Recalculate levels with new anchor prices
    for sym in SYMBOLS:
        with _PRICE_LOCK:
            anchor = _CRYPTO_PRICES.get(sym, 0.0)
        if anchor > 0:
            _CRYPTO_ANCHOR[sym] = anchor
            CRYPTO_ANCHOR[sym]  = anchor
            _CRYPTO_LEVELS[sym] = _calc_levels(sym, anchor)
            _CRYPTO_EXITED[sym] = False  # allow fresh entries after re-anchor
            log.info("RE-ANCHOR %s @ $%.4f", sym, anchor)

    # Write new levels JSON + XLSX
    _write_anchor_json(ts)
    xl_path = os.path.join("levels", f"crypto_initial_levels_{ts_tag}.xlsx")
    _write_levels_xlsx(xl_path, ts)
    # Also update "latest" symlink/copy
    latest = os.path.join("levels", "crypto_initial_levels_latest.json")
    _write_anchor_json(ts, path=latest)

    btc = _CRYPTO_PRICES.get("BTC", 0)
    eth = _CRYPTO_PRICES.get("ETH", 0)
    send_document_alert(
        xl_path,
        f"Crypto re-anchor {date_str} {ts_str} — new levels active",
        asset_class="crypto"
    )
    log.info("Re-anchor complete. Next anchor in %dh.", REANCHOR_HOURS)


def _write_anchor_json(ts: datetime, path: Optional[str] = None) -> None:
    ts_tag = ts.strftime("%Y%m%d_%H%M")
    if path is None:
        path = os.path.join("levels", f"crypto_initial_levels_{ts_tag}.json")
    data = {
        "anchor_time": ts.isoformat(),
        "author": "Ridhaant Ajoy Thackur",
        "x_multiplier": CRYPTO_X,
        "usdt_to_inr": USDT_TO_INR,
        "levels": _CRYPTO_LEVELS,
    }
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    tmp = path + ".tmp"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        os.replace(tmp, path)
    except Exception as exc:
        log.debug("Anchor JSON write error: %s", exc)


# ════════════════════════════════════════════════════════════════════════════
# MAIN PRICE LOOP
# ════════════════════════════════════════════════════════════════════════════

_LAST_WS_PRICE_TIME: float = 0.0  # monotonic time of last Binance WS price update

def _price_loop() -> None:
    global _NEXT_ANCHOR, _LAST_WS_PRICE_TIME
    log.info("Crypto price loop started (24/7, re-anchor every %dh)", REANCHOR_HOURS)
    _rest_fallback_active = False

    while True:
        try:
            now = _now_utc()
            
            # Failsafe: if Binance WS hasn't updated in 30s, force REST fetch
            staleness = time.monotonic() - _LAST_WS_PRICE_TIME
            if staleness > 30 and not _rest_fallback_active:
                log.warning("Binance WS stale for %.0fs — activating REST fallback", staleness)
                fallback = _fetch_initial_binance_prices()
                if fallback:
                    with _PRICE_LOCK:
                        for sym, price in fallback.items():
                            _CRYPTO_PRICES[sym] = price
                            _PRICE_AGE[sym] = time.monotonic()
                    _LAST_WS_PRICE_TIME = time.monotonic()  # reset
                    log.info("REST fallback fetched %d prices", len(fallback))
                _rest_fallback_active = staleness > 60  # REST mode if WS dead >60s

            # Check for 6h re-anchor
            if _NEXT_ANCHOR is not None and now >= _NEXT_ANCHOR:
                prev = list(_CRYPTO_TRADES)   # snapshot before clearing window
                _do_reanchor(now, prev)
                _NEXT_ANCHOR = now + timedelta(hours=REANCHOR_HOURS)
                global _EXITED_RESET_LOGGED
                _EXITED_RESET_LOGGED = False  # allow logging again after re-anchor

            # Allow re-entry 30+ min before next re-anchor (prevents permanent lock)
            _maybe_reset_exited()

            # Process prices for each symbol
            with _PRICE_LOCK:
                prices_snap = dict(_CRYPTO_PRICES)

            for sym in SYMBOLS:
                px = prices_snap.get(sym)
                if px and px > 0 and _CRYPTO_LEVELS.get(sym):
                    _process_price(sym, px, now)

            # Publish to ZMQ + JSON every 2s reliably regardless of price changes
            # This ensures live_prices.json always has fresh crypto_prices for dashboard
            _publish_prices(prices_snap)
            
            time.sleep(2)
        except Exception as exc:
            log.error("Crypto price loop error: %s", exc)
            time.sleep(5)


# ════════════════════════════════════════════════════════════════════════════
# STARTUP
# ════════════════════════════════════════════════════════════════════════════

def _fetch_initial_binance_prices() -> Dict[str, float]:
    """
    Fetch current crypto prices via multiple sources with retry.
    Returns {} only if ALL sources fail after retries.
    Sources tried in order:
      1. Binance REST /ticker/price (each symbol individually — avoids array param issue)
      2. Binance REST /ticker/24hr (richer data, different endpoint)
      3. CoinGecko REST (free tier, 30/min limit)
      4. CoinCap REST (backup)
    """
    prices: Dict[str, float] = {}
    
    # Source 1: Binance individual price endpoint (most reliable)
    for b_sym, our_sym in BINANCE_SYMBOLS.items():
        try:
            r = _req_module.get(
                "https://api.binance.com/api/v3/ticker/price",
                params={"symbol": b_sym},
                timeout=8,
            )
            if r.status_code == 200:
                prices[our_sym] = float(r.json()["price"])
        except Exception:
            pass
    
    if len(prices) == len(SYMBOLS):
        log.info("Startup prices from Binance REST (individual): %s",
                 {k: f"${v:,.2f}" for k, v in prices.items()})
        return prices
    
    # Source 2: Binance batch endpoint (array format)
    try:
        syms_json = json.dumps(list(BINANCE_SYMBOLS.keys()))
        r = _req_module.get(
            "https://api.binance.com/api/v3/ticker/price",
            params={"symbols": syms_json},
            timeout=10,
        )
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list):
                for item in data:
                    sym = BINANCE_SYMBOLS.get(item.get("symbol",""))
                    if sym and float(item.get("price", 0)) > 0:
                        prices[sym] = float(item["price"])
    except Exception as exc:
        log.debug("Binance batch REST failed: %s", exc)
    
    if len(prices) == len(SYMBOLS):
        log.info("Startup prices from Binance REST (batch): %s",
                 {k: f"${v:,.2f}" for k, v in prices.items()})
        return prices
    
    # Source 3: CoinGecko
    try:
        ids = ",".join(COINGECKO_IDS.values())
        r = _req_module.get(
            f"https://api.coingecko.com/api/v3/simple/price?ids={ids}&vs_currencies=usd",
            timeout=12,
        )
        if r.status_code == 200:
            sym_by_id = {v: k for k, v in COINGECKO_IDS.items()}
            for cg_id, px_data in r.json().items():
                sym = sym_by_id.get(cg_id)
                if sym and float(px_data.get("usd", 0)) > 0:
                    prices[sym] = float(px_data["usd"])
    except Exception as exc:
        log.debug("CoinGecko REST failed: %s", exc)
    
    if prices:
        log.info("Startup prices from CoinGecko: %s",
                 {k: f"${v:,.2f}" for k, v in prices.items()})
        return prices
    
    # Source 4: CoinCap REST
    try:
        cap_map = {"bitcoin":"BTC","ethereum":"ETH","binance-coin":"BNB",
                   "solana":"SOL","cardano":"ADA"}
        r = _req_module.get(
            "https://api.coincap.io/v2/assets",
            params={"ids": ",".join(cap_map.keys())},
            timeout=10,
        )
        if r.status_code == 200:
            for item in r.json().get("data", []):
                sym = cap_map.get(item.get("id",""))
                px  = float(item.get("priceUsd", 0) or 0)
                if sym and px > 0:
                    prices[sym] = px
    except Exception as exc:
        log.debug("CoinCap REST failed: %s", exc)
    
    if prices:
        log.info("Startup prices from CoinCap: %s",
                 {k: f"${v:,.2f}" for k, v in prices.items()})
    else:
        log.error("ALL price sources failed — crypto engine cannot start")
    return prices


def startup() -> None:
    """Initialize crypto engine — call once at process start."""
    global _STARTUP_TIME, _NEXT_ANCHOR

    now = _now_utc()
    _STARTUP_TIME = now
    _NEXT_ANCHOR  = now + timedelta(hours=REANCHOR_HOURS)

    log.info("═" * 60)
    log.info("AlgoStack v9.0 — Crypto Engine (Binance WS)")
    log.info("Author: Ridhaant Ajoy Thackur")
    log.info("Symbols: %s", ", ".join(SYMBOLS))
    log.info("X Multiplier: %.6f (same as equity)", CRYPTO_X)
    log.info("Re-anchor: every %dh | Next: %s UTC",
             REANCHOR_HOURS, _NEXT_ANCHOR.strftime("%H:%M"))
    log.info("Price sharing: Binance WS → ZMQ 'crypto' topic (1-5s latency)")
    log.info("═" * 60)

    # Fetch startup anchor prices — retry until at least 3 symbols have prices
    _retry_count = 0
    initial: Dict[str, float] = {}
    while not initial or len(initial) < 3:
        initial = _fetch_initial_binance_prices()
        if len(initial) >= 3:
            break
        _retry_count += 1
        if _retry_count >= 6:
            log.error("Price fetch failed after %d attempts — using last partial result", _retry_count)
            break
        log.warning("Only %d/%d prices — retry %d/6 in 5s", len(initial), len(SYMBOLS), _retry_count)
        time.sleep(5)
    if not initial:
        log.error("Could not fetch any prices — crypto engine standby (will retry)")
        time.sleep(30)
        # Tail-call to retry startup
        return startup()

    # Set initial prices and calculate levels
    with _PRICE_LOCK:
        for sym, price in initial.items():
            _CRYPTO_PRICES[sym] = price
            _PRICE_AGE[sym]     = time.monotonic()
    for sym in SYMBOLS:
        anchor = initial.get(sym, 0.0)
        if anchor > 0:
            _CRYPTO_ANCHOR[sym] = anchor
            CRYPTO_ANCHOR[sym]  = anchor
            _CRYPTO_LEVELS[sym] = _calc_levels(sym, anchor)
            log.info("  %-6s anchor=$%.4f  buy_above=$%.4f  sell_below=$%.4f",
                     sym, anchor,
                     _CRYPTO_LEVELS[sym]["buy_above"],
                     _CRYPTO_LEVELS[sym]["sell_below"])

    # Write crypto_initial_levels_latest.json IMMEDIATELY (before XLSX)
    # so crypto_scanner*.py can find anchor prices quickly on startup
    _write_anchor_json(now, path=os.path.join("levels", "crypto_initial_levels_latest.json"))
    log.info("Crypto anchor prices written to crypto_initial_levels_latest.json")
    
    # Also write live_prices.json crypto_prices right away for scanner fallback
    _publish_prices({sym: _CRYPTO_ANCHOR[sym] for sym in SYMBOLS if sym in _CRYPTO_ANCHOR})

    # Write initial JSON + XLSX
    ts_tag   = now.strftime("%Y%m%d_%H%M")
    date_str = now.strftime("%Y%m%d")
    os.makedirs("levels", exist_ok=True)
    _write_anchor_json(now)
    _write_anchor_json(now, path=os.path.join("levels", "crypto_initial_levels_latest.json"))

    xl_path = os.path.join("levels", f"crypto_initial_levels_{ts_tag}.xlsx")
    _write_levels_xlsx(xl_path, now)

    btc = initial.get("BTC", 0)
    eth = initial.get("ETH", 0)
    try:
        from tg_async import send_document_alert
        send_document_alert(
            xl_path,
            f"Crypto anchor: BTC=${btc:,.0f} ETH=${eth:,.0f} X={CRYPTO_X:.6f}",
            asset_class="crypto"
        )
    except Exception:
        pass

    # Start Binance WebSocket price feed
    _start_binance_ws()

    # Start ZMQ publisher
    _init_zmq()

    log.info("Crypto engine ready. Starting 24/7 price loop...")


def main() -> None:
    startup()
    _price_loop()


if __name__ == "__main__":
    main()
