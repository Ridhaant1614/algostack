# ═══════════════════════════════════════════════════════════════════════
# © 2026 Ridhaant Ajoy Thackur. All rights reserved.
# AlgoStack™ is proprietary software. Unauthorised copying or distribution is prohibited.
# AlgoStack v9.0 | Author: Ridhaant Ajoy Thackur
# commodity_engine.py — MCX commodity trading engine (mirrors Algofinal)
# ═══════════════════════════════════════════════════════════════════════
"""
commodity_engine.py
===================
MCX commodity trading engine. Mirrors Algofinal.py for 5 MCX symbols.

SYMBOLS: GOLD, SILVER, CRUDE, NATURALGAS, COPPER
HOURS:   09:00 – 23:30 IST (Mon–Fri)
EOD:     23:30 IST square-off
X_MULTS: from config.cfg.COMM_X dict

Price sources (priority):
  1. TradingView WebSocket (wss://data.tradingview.com) — real-time INR
  2. Investing.com JSON API — ~1-2s latency, reliable MCX INR
  3. Goodreturns.in scrape
  4. MoneyControl scrape
  5. In-memory cache (stale fallback)

Level formula (IDENTICAL to Algofinal — DO NOT CHANGE):
  x_val      = prev_close * X_MULTIPLIER
  buy_above  = prev_close + x_val
  sell_below = prev_close - x_val
  step       = x_val
  T1..T5     = buy_above + step * 1..5
  ST1..ST5   = sell_below - step * 1..5
  buy_sl     = prev_close  (= buy_above - x_val)
  sell_sl    = prev_close
  qty        = 1 lot (MCX lot sizes vary)
  brokerage  = Rs 20 per round-trip
"""

from __future__ import annotations

import json
import logging
import os
import threading
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional

import pytz
import requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [CommodityEngine] %(levelname)s — %(message)s",
)
log = logging.getLogger("commodity_engine")
try:
    import urllib3; urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception: pass
import warnings as _w; _w.filterwarnings("ignore", message="Unverified HTTPS")

IST = pytz.timezone("Asia/Kolkata")

# ── Config ────────────────────────────────────────────────────────────────────
from config import cfg

# Commodity internal currency:
# - User requirement: keep commodities in USDT for internal levels/calcs.
# - TradingView/yfinance/REST sources may return INR or USD-like values.
# We normalize everything to USDT with a simple heuristic, then convert to INR
# only when computing/storing P&L for alerts/UI.
USDT_TO_INR: float = float(getattr(cfg, "USDT_TO_INR", 84.0) or 84.0)

def _normalise_price_to_usdt(px: float) -> float:
    """Best-effort: if value looks like INR, convert to USDT by /USDT_TO_INR."""
    try:
        p = float(px)
    except Exception:
        return 0.0
    if p <= 0:
        return 0.0
    # Heuristic: INR commodity prices are usually in the tens of thousands.
    # USDT prices are usually <= a few thousand.
    if p > 5000 and USDT_TO_INR > 0:
        return p / USDT_TO_INR
    return p

def _usdt_to_inr(px_usdt: float) -> float:
    try:
        return float(px_usdt) * USDT_TO_INR
    except Exception:
        return 0.0

SYMBOLS: List[str] = ["GOLD", "SILVER", "CRUDE", "NATURALGAS", "COPPER"]
LOT_SIZES: Dict[str, int] = {
    "GOLD":       1,
    "SILVER":     1,
    "CRUDE":      1,
    "NATURALGAS": 1,
    "COPPER":     1,
}
BROKERAGE = 20.0  # Rs per round-trip

# TradingView MCX symbol map
_TV_MCX_SYMBOLS: Dict[str, str] = {
    "MCX:GOLD1!":       "GOLD",
    "MCX:SILVER1!":     "SILVER",
    "MCX:CRUDEOIL1!":   "CRUDE",
    "MCX:NATURALGAS1!": "NATURALGAS",
    "MCX:COPPER1!":     "COPPER",
}

# yfinance proxy symbols for prev_close
_YF_SYMBOLS: Dict[str, str] = {
    "GOLD":       "GC=F",
    "SILVER":     "SI=F",
    "CRUDE":      "CL=F",
    "NATURALGAS": "NG=F",
    "COPPER":     "HG=F",
}

# USD futures -> MCX-like INR proxy factors (then normalized back to USDT).
_USD_TO_MCX_MULT: Dict[str, float] = {
    "GOLD": 27.34,
    "SILVER": 2734.2,
    "CRUDE": 85.0,
    "NATURALGAS": 85.0,
    "COPPER": 187.4,
}

# Shared state
_COMM_PRICES:     Dict[str, float] = {}
_COMM_PREV_CLOSE: Dict[str, float] = {}
_COMM_LEVELS:     Dict[str, dict]  = {}
_COMM_POSITIONS:  Dict[str, Optional[dict]] = {s: None for s in SYMBOLS}
_COMM_EXITED:     Dict[str, bool]  = {s: False for s in SYMBOLS}
_COMM_TRADES:     List[dict]       = []
_COMM_ANCHOR:     Dict[str, float] = {}   # anchor prices for scanners
_PRICE_LOCK       = threading.Lock()

# v10.6: Re-entry watch state (mirrors Algofinal threshold+retouch logic)
_COMM_REENTRY: Dict[str, Optional[dict]] = {s: None for s in SYMBOLS}

# Expose anchor for commodity scanners
COMMODITY_ANCHOR: Dict[str, float] = {}


# ════════════════════════════════════════════════════════════════════════════
# TIME HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _now_ist() -> datetime:
    return datetime.now(IST)


def _in_session(dt: datetime) -> bool:
    t = dt.hour * 60 + dt.minute
    return 9 * 60 + 0 <= t < 23 * 60 + 30


def _in_930_blackout(dt: datetime) -> bool:
    t = dt.hour * 60 + dt.minute
    return 9 * 60 + 30 <= t < 9 * 60 + 35


def _is_eod(dt: datetime) -> bool:
    return dt.hour == 23 and dt.minute >= 30


# ════════════════════════════════════════════════════════════════════════════
# PREV CLOSE LOADER
# ════════════════════════════════════════════════════════════════════════════

def _fetch_prev_close_yfinance(symbol: str) -> Optional[float]:
    try:
        import yfinance as yf
        yf_sym = _YF_SYMBOLS.get(symbol, f"{symbol}=F")
        t = yf.Ticker(yf_sym)
        df = t.history(period="5d", interval="1d")
        if df is not None and not df.empty:
            # yfinance returns commodity prices in USD-like units for futures.
            # Normalize to USDT for internal ladder calculations.
            price = float(df["Close"].iloc[-1]) * _USD_TO_MCX_MULT.get(symbol, 84.0)
            return _normalise_price_to_usdt(price)
    except Exception as exc:
        log.debug("yfinance prev_close failed for %s: %s", symbol, exc)
    return None


def _load_prev_closes() -> Dict[str, float]:
    """Load previous closes for MCX commodities."""
    result: Dict[str, float] = {}

    # Try Algofinal's commodity_initial_levels JSON first
    date_str = _now_ist().strftime("%Y%m%d")
    for fname in (
        f"commodity_prev_closes_{date_str}.json",
        f"prev_closes_persistent_{date_str}.json",
    ):
        path = os.path.join("levels", fname)
        if os.path.exists(path):
            try:
                with open(path, encoding="utf-8") as f:
                    data = json.load(f)
                for sym in SYMBOLS:
                    val = data.get(sym) or data.get(sym.upper())
                    if isinstance(val, dict):
                        val = val.get("prev_close")
                    if val:
                        result[sym] = _normalise_price_to_usdt(val)
                if len(result) == len(SYMBOLS):
                    log.info("Loaded commodity prev_closes from %s", fname)
                    return result
            except Exception:
                pass

    # yfinance fallback for each symbol
    for sym in SYMBOLS:
        if sym in result:
            continue
        pc = _fetch_prev_close_yfinance(sym)
        if pc and pc > 0:
            result[sym] = _normalise_price_to_usdt(pc)
            log.info("  %s prev_close=%.2f (yfinance)", sym, pc)
        else:
            log.warning("  %s prev_close unavailable — using placeholder", sym)
            result[sym] = 0.0

    return result


# ════════════════════════════════════════════════════════════════════════════
# LEVEL CALCULATOR
# ════════════════════════════════════════════════════════════════════════════

def _calc_levels(sym: str, prev_close: float) -> dict:
    """Calculate all trading levels for a commodity symbol."""
    xm   = cfg.COMM_X.get(sym, 0.004)
    xval = prev_close * xm
    step = xval  # no 0.6 factor for commodities (no special symbols)
    ba   = prev_close + xval
    sb   = prev_close - xval
    return {
        "symbol":      sym,
        "prev_close":  round(prev_close, 2),
        "x_val":       round(xval, 4),
        "x_mult":      round(xm, 6),
        "buy_above":   round(ba, 2),
        "sell_below":  round(sb, 2),
        "buy_sl":      round(prev_close, 2),
        "sell_sl":     round(prev_close, 2),
        "step":        round(step, 4),
        "T1": round(ba + step * 1, 2),
        "T2": round(ba + step * 2, 2),
        "T3": round(ba + step * 3, 2),
        "T4": round(ba + step * 4, 2),
        "T5": round(ba + step * 5, 2),
        "ST1": round(sb - step * 1, 2),
        "ST2": round(sb - step * 2, 2),
        "ST3": round(sb - step * 3, 2),
        "ST4": round(sb - step * 4, 2),
        "ST5": round(sb - step * 5, 2),
        "retreat_65": round(ba + 0.65 * step, 2),
        "retreat_45": round(ba + 0.45 * step, 2),
        "retreat_25": round(ba + 0.25 * step, 2),
    }


# ════════════════════════════════════════════════════════════════════════════
# PRICE FEED — TradingView WebSocket
# ════════════════════════════════════════════════════════════════════════════

def _start_tradingview_ws() -> None:
    """Start TradingView WebSocket streamer in daemon thread."""
    thread = threading.Thread(target=_tv_ws_loop, daemon=True, name="TV-MCX-WS")
    thread.start()


def _tv_ws_loop() -> None:
    import json as _json
    try:
        import websocket
    except ImportError:
        log.warning("websocket-client not installed — using REST fallback only")
        return

    tv_symbols = list(_TV_MCX_SYMBOLS.keys())
    session_id = f"qs_{''.join(['abcdefghijklmnopqrstuvwxyz'[hash(s)%26] for s in tv_symbols[:10]])}"

    def _send(ws, obj):
        msg = f"~m~{len(_json.dumps(obj))}~m~{_json.dumps(obj)}"
        ws.send(msg)

    def on_message(ws, message):
        try:
            # TradingView frames: ~m~<len>~m~<json>
            parts = message.split("~m~")
            for i in range(0, len(parts) - 1, 2):
                try:
                    data = _json.loads(parts[i + 1])
                    if data.get("m") == "qsd":
                        for p in data.get("p", []):
                            if isinstance(p, dict) and "n" in p and "v" in p:
                                tv_sym = p["n"]
                                our_sym = _TV_MCX_SYMBOLS.get(tv_sym)
                                if our_sym and isinstance(p["v"], dict):
                                    price = p["v"].get("lp") or p["v"].get("ch")
                                    if price and float(price) > 0:
                                        with _PRICE_LOCK:
                                            _COMM_PRICES[our_sym] = _normalise_price_to_usdt(price)
                except Exception:
                    pass
        except Exception:
            pass

    def on_open(ws):
        log.info("TradingView WS connected for MCX")
        _send(ws, {"m": "set_auth_token", "p": ["unauthorized_user_token"]})
        _send(ws, {"m": "chart_create_session", "p": [session_id, ""]})
        for tv_sym in tv_symbols:
            _send(ws, {"m": "resolve_symbol", "p": [session_id, tv_sym,
                       f'={{"symbol":"{tv_sym}","adjustment":"splits"}}']})
            _send(ws, {"m": "create_study", "p": [session_id, tv_sym,
                       "sds_sym_1", tv_sym, {"text": ""}]})

    def on_error(ws, error):
        log.debug("TradingView WS error: %s", error)

    _tv_backoff = 5
    while True:
        try:
            ws = websocket.WebSocketApp(
                "wss://data.tradingview.com/socket.io/websocket",
                on_message=on_message,
                on_open=on_open,
                on_error=on_error,
            )
            ws.run_forever(ping_interval=20, ping_timeout=8,
                           skip_utf8_validation=True)
            _tv_backoff = 5  # reset on clean exit
        except Exception as exc:
            log.debug("TradingView WS loop error: %s", exc)
        log.debug("TradingView WS reconnecting in %ds", _tv_backoff)
        time.sleep(_tv_backoff)
        _tv_backoff = min(_tv_backoff * 2, 60)


def _fetch_commodity_price_rest(sym: str) -> Optional[float]:
    """REST fallback using USD futures only, normalized to USDT."""
    _USD_FUT = {"GOLD":"GC=F","SILVER":"SI=F","CRUDE":"CL=F","NATURALGAS":"NG=F","COPPER":"HG=F"}
    # USD → MCX INR conversion factors (approximate, validated against MCX rates)
    # GOLD: $/troy_oz × 85(INR/USD) / 31.1(g/troy_oz) × 10(g per MCX unit)
    # SILVER: $/troy_oz × 85 / 31.1 × 1000 (₹/kg)
    # CRUDE: $/bbl × 85 = ₹/bbl
    # NATURALGAS: $/mmBtu × 85 = ₹/mmBtu
    # COPPER: $/lb × 85 × 2.205 = ₹/kg
    try:
        import yfinance as yf
        t = yf.Ticker(_USD_FUT.get(sym, sym+"=F"))
        fi = t.fast_info
        usd = getattr(fi,"last_price",None) or getattr(fi,"regularMarketPrice",None)
        if usd and float(usd) > 0:
            # _MULT here is USD→INR based; normalize the INR-ish result back to USDT.
            return _normalise_price_to_usdt(float(usd) * _USD_TO_MCX_MULT.get(sym, 84.0))
    except Exception:
        pass
    with _PRICE_LOCK:
        return _COMM_PRICES.get(sym)


def _publish_comm_prices(prices: Dict[str, float]) -> None:
    """Merge commodity prices into live_prices.json (preserve equity/crypto keys)."""
    try:
        import json as _j, os as _o
        _o.makedirs("levels", exist_ok=True)
        lp = _o.path.join("levels", "live_prices.json")
        try:
            existing = _j.load(open(lp, encoding="utf-8")) if _o.path.exists(lp) else {}
        except Exception:
            existing = {}
        ts_now = _now_ist().strftime("%Y-%m-%d %H:%M:%S")
        existing["commodity_prices"] = prices
        existing["commodity_ts"] = ts_now
        existing["ts"] = ts_now   # update master ts so dashboard sees fresh data
        # DO NOT touch "prices" or "equity_prices" keys
        tmp = lp + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            _j.dump(existing, f, separators=(",", ":"))
        _o.replace(tmp, lp)
    except Exception:
        pass


def _start_commodity_rest_poll() -> None:
    """Fallback REST poller: fetch MCX prices every 3s if TradingView WS has no data."""
    threading.Thread(target=_commodity_rest_loop, daemon=True, name="Comm-REST").start()


def _commodity_rest_loop() -> None:
    # Use USD futures only in REST fallback to avoid noisy/invalid *.MCX yfinance symbols.
    _USD_FUT = {"GOLD":"GC=F","SILVER":"SI=F","CRUDE":"CL=F","NATURALGAS":"NG=F","COPPER":"HG=F"}
    # USD → MCX INR conversion factors (approximate, validated against MCX rates)
    # GOLD: $/troy_oz × 85(INR/USD) / 31.1(g/troy_oz) × 10(g per MCX unit)
    # SILVER: $/troy_oz × 85 / 31.1 × 1000 (₹/kg)
    # CRUDE: $/bbl × 85 = ₹/bbl
    # NATURALGAS: $/mmBtu × 85 = ₹/mmBtu
    # COPPER: $/lb × 85 × 2.205 = ₹/kg
    log.info("Commodity REST poll started (3s fallback loop)")
    try:
        import yfinance as yf
    except ImportError:
        log.warning("yfinance not available — commodity REST poll disabled")
        return
    while True:
        try:
            updated: Dict[str, float] = {}
            for sym in SYMBOLS:
                px: Optional[float] = None
                try:
                    t = yf.Ticker(_USD_FUT.get(sym, sym+"=F"))
                    fi = t.fast_info
                    v = getattr(fi,"last_price",None) or getattr(fi,"regularMarketPrice",None)
                    if v and float(v) > 0:
                        px = float(v) * _USD_TO_MCX_MULT.get(sym, 84.0)
                except Exception:
                    pass
                if px and px > 0:
                    px_u = _normalise_price_to_usdt(px)
                    with _PRICE_LOCK:
                        _COMM_PRICES[sym] = px_u
                    updated[sym] = px_u
            if updated:
                _publish_comm_prices(updated)
        except Exception as exc:
            log.debug("Commodity REST loop error: %s", exc)
        # v10.7 PERF: adaptive sleep — 1s during MCX session, 10s off-hours
        _now_poll = _now_ist()
        _poll_sleep = 1 if _in_session(_now_poll) else 10
        time.sleep(_poll_sleep)


# ════════════════════════════════════════════════════════════════════════════
# TRADE LOGGING
# ════════════════════════════════════════════════════════════════════════════

def _log_trade(event: dict) -> None:
    """Append trade event to JSONL file and in-memory list."""
    _COMM_TRADES.append(event)
    date_str = _now_ist().strftime("%Y%m%d")
    os.makedirs("trade_logs", exist_ok=True)
    path = os.path.join("trade_logs", f"commodity_trade_events_{date_str}.jsonl")
    try:
        tmp = path + ".tmp"
        with open(path, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(event) + "\n")
    except Exception as exc:
        log.debug("Trade log write error: %s", exc)


# ════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ════════════════════════════════════════════════════════════════════════════

def _write_initial_levels_xlsx(date_str: str, path: str) -> None:
    """Write initial levels Excel file for commodity."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "MCX Levels"
        hdr = ["Symbol", "Prev Close", "X Mult", "X Val",
               "Buy Above", "T1", "T2", "T3", "T4", "T5",
               "Sell Below", "ST1", "ST2", "ST3", "ST4", "ST5",
               "Buy SL", "Sell SL"]
        for col, h in enumerate(hdr, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="1F4E79")
            c.font = Font(bold=True, color="FFFFFF")
        for row_idx, sym in enumerate(SYMBOLS, 2):
            lv = _COMM_LEVELS.get(sym, {})
            ws.append([
                sym,
                lv.get("prev_close", 0),
                lv.get("x_mult", 0),
                lv.get("x_val", 0),
                lv.get("buy_above", 0),
                lv.get("T1", 0), lv.get("T2", 0), lv.get("T3", 0),
                lv.get("T4", 0), lv.get("T5", 0),
                lv.get("sell_below", 0),
                lv.get("ST1", 0), lv.get("ST2", 0), lv.get("ST3", 0),
                lv.get("ST4", 0), lv.get("ST5", 0),
                lv.get("buy_sl", 0),
                lv.get("sell_sl", 0),
            ])
        ws.append([])
        ws.append(["Author: Ridhaant Ajoy Thackur", "", "AlgoStack v9.0"])
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        tmp = path + ".tmp"
        wb.save(tmp)
        os.replace(tmp, path)
        log.info("Commodity initial levels XLSX → %s", path)
    except Exception as exc:
        log.warning("XLSX write failed: %s", exc)


def _write_eod_xlsx(date_str: str, path: str) -> None:
    """Write EOD trade analysis Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "MCX Trades"
        hdr = ["Symbol", "Side", "Entry", "Exit", "Qty",
               "Gross (Rs)", "Brokerage (Rs)", "Net (Rs)", "Exit Type", "Time", "X Used"]
        for col, h in enumerate(hdr, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
        for r, t in enumerate(_COMM_TRADES, 2):
            ws.append([
                t.get("symbol"), t.get("side"), t.get("entry_px"),
                t.get("exit_px"), t.get("qty"),
                t.get("gross_pnl"), t.get("brokerage", BROKERAGE),
                t.get("net_pnl"), t.get("reason"), t.get("ts"),
                t.get("x_val"),
            ])
        ws.append([])
        ws.append(["Author: Ridhaant Ajoy Thackur", "", "AlgoStack v9.0"])
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        tmp = path + ".tmp"
        wb.save(tmp)
        os.replace(tmp, path)
    except Exception as exc:
        log.warning("EOD XLSX write failed: %s", exc)


# ════════════════════════════════════════════════════════════════════════════
# ZMQ PUBLISHER
# ════════════════════════════════════════════════════════════════════════════

_pub: Optional["ipc_bus.PricePublisher"] = None


def _init_zmq() -> None:
    global _pub
    try:
        from ipc_bus import PricePublisher
        _pub = PricePublisher()
    except Exception as exc:
        log.warning("ZMQ publisher init failed: %s", exc)


# ════════════════════════════════════════════════════════════════════════════
# TRADING LOGIC
# ════════════════════════════════════════════════════════════════════════════

def _fmt_pct(v: float) -> str:
    return f"{'+' if v >= 0 else ''}{v:.2f}%"

BROKERAGE_FLAT = 20.0  # ₹10 entry + ₹10 exit = ₹20 per round-trip (mirrors Algofinal)

def _build_alert(
    title: str, symbol: str, status_line: str,
    *, prev_close: float, current_price: float,
    x_val: float, quantity, side: str,
    buy_above: float, sell_below: float,
    targets_buy: list, targets_sell: list,
    buy_sl: float, sell_sl: float,
    ts_str: str,
) -> str:
    """Exact mirror of Algofinal.build_simple_alert() — [🏅 MCX Commodity] tag."""
    change_pct = (current_price - prev_close) / prev_close * 100.0 if prev_close else 0.0
    lines = [f"🚨 {symbol} [🏅 MCX Commodity] — {title} at {ts_str}", ""]
    lines += [f"Previous Close: ${prev_close:,.2f}",
              f"Current Price: ${current_price:,.2f}",
              f"Change: {_fmt_pct(change_pct)}",
              f"Deviation (X): {x_val:.4f}"]
    if quantity is not None:
        lines.append(f"Quantity: {quantity}")
    lines += [f"Status: {status_line}", "", "📊 Technical Analysis:"]
    if side == "BUY":
        lines.append("📈 Buy Levels:")
        lines.append(f"Buy Above: ${buy_above:,.2f}")
        for i, tgt in enumerate(targets_buy, 1):
            pct = (tgt - current_price) / current_price * 100
            lines.append(f"Target {i}: ${tgt:,.2f} ({_fmt_pct(pct)})")
        pct_sl = (buy_sl - current_price) / current_price * 100
        lines.append(f"Stop Loss: ${buy_sl:,.2f} ({_fmt_pct(pct_sl)})")
    else:
        lines.append("📉 Sell Levels:")
        lines.append(f"Sell Below: ${sell_below:,.2f}")
        for i, tgt in enumerate(targets_sell, 1):
            pct = (tgt - current_price) / current_price * 100
            lines.append(f"Target {i}: ${tgt:,.2f} ({_fmt_pct(pct)})")
        pct_sl = (sell_sl - current_price) / current_price * 100
        lines.append(f"Stop Loss: ${sell_sl:,.2f} ({_fmt_pct(pct_sl)})")
    return "\n".join(lines)


def _send_alert(msg: str) -> None:
    try:
        from tg_async import send_alert
        send_alert(msg, asset_class="commodity")
    except Exception:
        pass


def _lv_alert_kwargs(lv: dict, price: float, ts_str: str = "") -> dict:
    """Extract _build_alert kwargs from a level dict."""
    return dict(
        prev_close=lv.get("prev_close", 0),
        current_price=price,
        x_val=lv.get("x_val", lv.get("x_mult", 0)),
        buy_above=lv.get("buy_above", 0),
        sell_below=lv.get("sell_below", 0),
        targets_buy=[lv.get(f"T{i}", 0) for i in range(1, 6)],
        targets_sell=[lv.get(f"ST{i}", 0) for i in range(1, 6)],
        buy_sl=lv.get("buy_sl", lv.get("prev_close", 0)),
        sell_sl=lv.get("sell_sl", lv.get("prev_close", 0)),
        ts_str=ts_str,
    )


def _process_price(sym: str, price: float, ts: datetime) -> None:
    """Process one price tick for a commodity symbol.
    v10.6: Unified alert format + re-entry watch (threshold + retouch mirrors Algofinal).
    Alert types: Entry, Target Hit, Stop Loss Hit, Retreat Exit, Re-entry Armed, Re-entry Entry.
    NO proximity/warning alerts — only triggered when levels are actually breached.
    """
    if price <= 0:
        return
    lv = _COMM_LEVELS.get(sym)
    if not lv:
        return

    pos    = _COMM_POSITIONS[sym]
    exited = _COMM_EXITED[sym]
    ts_str = ts.strftime("%Y-%m-%d %H:%M:%S IST+0530")
    qty    = LOT_SIZES.get(sym, 1)
    step   = lv.get("step", lv.get("x_val", 0))
    kw     = _lv_alert_kwargs(lv, price, ts_str)

    # ── Re-entry watch (no open position, exited today) ────────────────────
    rw = _COMM_REENTRY[sym]
    if pos is None and rw is not None:
        thresh_dir = rw["threshold_direction"]
        if thresh_dir == "ABOVE" and price >= rw["threshold_price"]:
            # Threshold continuation re-entry
            _COMM_REENTRY[sym] = None
            side_r   = rw["threshold_entry_side"]
            entry_r  = rw["threshold_entry_price"]
            sl_r     = rw["threshold_entry_sl"]
            tgt_r    = rw["threshold_entry_target"]
            _COMM_POSITIONS[sym] = {
                "side": side_r, "entry_px": entry_r, "qty": qty,
                "targets": ([tgt_r] + [tgt_r + step * i for i in range(1, 5)])
                            if side_r == "BUY" else
                            ([tgt_r] + [tgt_r - step * i for i in range(1, 5)]),
                "buy_sl": sl_r, "sell_sl": sl_r,
                "retreat_peak_reached": False, "ts": ts_str,
                "is_reentry": True,
            }
            status = f"THRESHOLD RE-ENTRY {side_r} @ {entry_r:,.2f} | SL {sl_r:,.2f} | Target {tgt_r:,.2f}"
            _send_alert(_build_alert("Re-entry Entry", sym, status, side=side_r, quantity=qty, **kw))
            _log_trade({"ts": ts_str, "symbol": sym, "side": side_r, "entry_px": entry_r,
                        "exit_px": None, "qty": qty, "gross_pnl": None, "net_pnl": None,
                        "reason": f"{side_r}_REENTRY_THRESHOLD", "x_val": lv.get("x_mult", 0),
                        "asset_class": "commodity"})
            return
        elif thresh_dir == "BELOW" and price <= rw["threshold_price"]:
            _COMM_REENTRY[sym] = None
            side_r  = rw["threshold_entry_side"]
            entry_r = rw["threshold_entry_price"]
            sl_r    = rw["threshold_entry_sl"]
            tgt_r   = rw["threshold_entry_target"]
            _COMM_POSITIONS[sym] = {
                "side": side_r, "entry_px": entry_r, "qty": qty,
                "targets": ([tgt_r - step * i for i in range(0, 5)]),
                "buy_sl": sl_r, "sell_sl": sl_r,
                "retreat_peak_reached": False, "ts": ts_str,
                "is_reentry": True,
            }
            status = f"THRESHOLD RE-ENTRY {side_r} @ {entry_r:,.2f} | SL {sl_r:,.2f} | Target {tgt_r:,.2f}"
            _send_alert(_build_alert("Re-entry Entry", sym, status, side=side_r, quantity=qty, **kw))
            _log_trade({"ts": ts_str, "symbol": sym, "side": side_r, "entry_px": entry_r,
                        "exit_px": None, "qty": qty, "gross_pnl": None, "net_pnl": None,
                        "reason": f"{side_r}_REENTRY_THRESHOLD", "x_val": lv.get("x_mult", 0),
                        "asset_class": "commodity"})
            return
        # Retouch check (price returns to original target level)
        if rw["threshold_direction"] == "ABOVE" and price <= rw["target_price"]:
            _COMM_REENTRY[sym] = None
            side_r  = rw["retouch_entry_side"]
            entry_r = rw["retouch_entry_price"]
            sl_r    = rw["retouch_entry_sl"]
            tgt_r   = rw["retouch_entry_target"]
            _COMM_POSITIONS[sym] = {
                "side": side_r, "entry_px": entry_r, "qty": qty,
                "targets": ([tgt_r - step * i for i in range(0, 5)]),
                "buy_sl": sl_r, "sell_sl": sl_r,
                "retreat_peak_reached": False, "ts": ts_str,
                "is_reentry": True,
            }
            status = f"RETOUCH RE-ENTRY {side_r} @ {entry_r:,.2f} | SL {sl_r:,.2f} | Target {tgt_r:,.2f}"
            _send_alert(_build_alert("Re-entry Entry", sym, status, side=side_r, quantity=qty, **kw))
            _log_trade({"ts": ts_str, "symbol": sym, "side": side_r, "entry_px": entry_r,
                        "exit_px": None, "qty": qty, "gross_pnl": None, "net_pnl": None,
                        "reason": f"{side_r}_REENTRY_RETOUCH", "x_val": lv.get("x_mult", 0),
                        "asset_class": "commodity"})
            return
        elif rw["threshold_direction"] == "BELOW" and price >= rw["target_price"]:
            _COMM_REENTRY[sym] = None
            side_r  = rw["retouch_entry_side"]
            entry_r = rw["retouch_entry_price"]
            sl_r    = rw["retouch_entry_sl"]
            tgt_r   = rw["retouch_entry_target"]
            _COMM_POSITIONS[sym] = {
                "side": side_r, "entry_px": entry_r, "qty": qty,
                "targets": ([tgt_r + step * i for i in range(0, 5)]),
                "buy_sl": sl_r, "sell_sl": sl_r,
                "retreat_peak_reached": False, "ts": ts_str,
                "is_reentry": True,
            }
            status = f"RETOUCH RE-ENTRY {side_r} @ {entry_r:,.2f} | SL {sl_r:,.2f} | Target {tgt_r:,.2f}"
            _send_alert(_build_alert("Re-entry Entry", sym, status, side=side_r, quantity=qty, **kw))
            _log_trade({"ts": ts_str, "symbol": sym, "side": side_r, "entry_px": entry_r,
                        "exit_px": None, "qty": qty, "gross_pnl": None, "net_pnl": None,
                        "reason": f"{side_r}_REENTRY_RETOUCH", "x_val": lv.get("x_mult", 0),
                        "asset_class": "commodity"})
            return

    # ── Fresh entry (no position, not exited today) ────────────────────────
    if pos is None and not exited:
        if _in_930_blackout(ts):
            return
        if price >= lv["buy_above"]:
            _COMM_POSITIONS[sym] = {
                "side": "BUY", "entry_px": price, "qty": qty,
                "buy_sl": lv["buy_sl"], "sell_sl": lv["sell_sl"],
                "targets": [lv[f"T{i}"] for i in range(1, 6)],
                "retreat_peak_reached": False, "ts": ts_str,
            }
            chg_pct = (price - lv["prev_close"]) / lv["prev_close"] * 100 if lv.get("prev_close") else 0
            status = f"BUY TRIGGERED | Capital ₹{(qty*price*USDT_TO_INR):,.0f}"
            _send_alert(_build_alert("Entry", sym, status, side="BUY", quantity=qty, **kw))
            _log_trade({"ts": ts_str, "symbol": sym, "side": "BUY",
                        "entry_px": price, "exit_px": None, "qty": qty,
                        "gross_pnl": None, "net_pnl": None, "reason": "ENTRY",
                        "x_val": lv.get("x_mult", 0), "asset_class": "commodity"})
            log.info("BUY %s @ %.2f qty=%d", sym, price, qty)

        elif price <= lv["sell_below"]:
            _COMM_POSITIONS[sym] = {
                "side": "SELL", "entry_px": price, "qty": qty,
                "buy_sl": lv["sell_sl"], "sell_sl": lv["sell_sl"],
                "targets": [lv[f"ST{i}"] for i in range(1, 6)],
                "retreat_peak_reached": False, "ts": ts_str,
            }
            status = f"SELL TRIGGERED | Capital ₹{(qty*price*USDT_TO_INR):,.0f}"
            _send_alert(_build_alert("Entry", sym, status, side="SELL", quantity=qty, **kw))
            _log_trade({"ts": ts_str, "symbol": sym, "side": "SELL",
                        "entry_px": price, "exit_px": None, "qty": qty,
                        "gross_pnl": None, "net_pnl": None, "reason": "ENTRY",
                        "x_val": lv.get("x_mult", 0), "asset_class": "commodity"})
            log.info("SELL %s @ %.2f qty=%d", sym, price, qty)
        return

    if pos is None:
        return

    # ── Exit logic ────────────────────────────────────────────────────────────
    side     = pos["side"]
    entry_px = pos["entry_px"]
    qty      = pos["qty"]
    brokerage_usdt = BROKERAGE_FLAT / USDT_TO_INR if USDT_TO_INR else BROKERAGE_FLAT

    def _close(exit_px: float, reason: str, arm_reentry: bool = False,
               tgt_label: str = "", tgt_price: float = 0) -> None:
        gross_usdt = (exit_px - entry_px) * qty if side == "BUY" else (entry_px - exit_px) * qty
        net_usdt   = gross_usdt - brokerage_usdt
        gross_inr  = gross_usdt * USDT_TO_INR
        net_inr    = net_usdt * USDT_TO_INR
        pct         = net_inr / 100_000 * 100
        _COMM_POSITIONS[sym] = None
        _COMM_EXITED[sym]    = True
        _COMM_REENTRY[sym]   = None  # clear any stale re-entry watch

        reason_emoji = {"T1_HIT":"✅","T2_HIT":"✅","T3_HIT":"✅","T4_HIT":"✅","T5_HIT":"✅",
                        "ST1_HIT":"✅","ST2_HIT":"✅","ST3_HIT":"✅","ST4_HIT":"✅","ST5_HIT":"✅",
                        "SL_HIT":"🛑","RETREAT":"↩️","EOD_2330":"🌙"}.get(reason, "📤")
        chg_pct = (exit_px - entry_px) / entry_px * 100 if side == "BUY" else (entry_px - exit_px) / entry_px * 100
        status = (f"{reason_emoji} {reason} | Side: {side}\n"
                  f"Gross: ₹{gross_inr:+,.2f}  Brokerage: ₹{BROKERAGE_FLAT:.2f}  Net: ₹{net_inr:+,.2f}  ({pct:+.3f}%)")
        _send_alert(_build_alert("Exit", sym, status, side=side, quantity=qty,
                                  **{**kw, "current_price": exit_px}))
        _log_trade({"ts": ts_str, "symbol": sym, "side": side,
                    "entry_px": entry_px, "exit_px": exit_px, "qty": qty,
                    "gross_pnl": round(gross_inr, 2), "brokerage": BROKERAGE_FLAT,
                    "net_pnl": round(net_inr, 2), "reason": reason,
                    "x_val": lv.get("x_mult", 0), "asset_class": "commodity"})
        log.info("%s %s @ %.2f net_inr=%.2f", reason, sym, exit_px, net_inr)

        # Arm re-entry watch after target exits (not SL/EOD)
        if arm_reentry and tgt_label and tgt_price > 0:
            x_val  = lv.get("x_val", step)
            buffer = 0.75 * x_val
            if side == "BUY":
                rw = {"threshold_direction": "ABOVE",
                      "threshold_price":      tgt_price + buffer,
                      "threshold_entry_side": "BUY",
                      "threshold_entry_price": tgt_price + step,
                      "threshold_entry_sl":    tgt_price,
                      "threshold_entry_target": tgt_price + step * 2,
                      "retouch_entry_side": "SELL",
                      "retouch_entry_price": tgt_price,
                      "retouch_entry_sl":    tgt_price + step,
                      "retouch_entry_target": tgt_price - step,
                      "target_price": tgt_price}
            else:
                rw = {"threshold_direction": "BELOW",
                      "threshold_price":      tgt_price - buffer,
                      "threshold_entry_side": "SELL",
                      "threshold_entry_price": tgt_price - step,
                      "threshold_entry_sl":    tgt_price,
                      "threshold_entry_target": tgt_price - step * 2,
                      "retouch_entry_side": "BUY",
                      "retouch_entry_price": tgt_price,
                      "retouch_entry_sl":    tgt_price - step,
                      "retouch_entry_target": tgt_price + step,
                      "target_price": tgt_price}
            _COMM_REENTRY[sym] = rw
            # Re-entry armed alert
            thresh_dir = "≥" if side == "BUY" else "≤"
            retouch_dir = "≤" if side == "BUY" else "≥"
            armed_msg = (
                f"🔁 {sym} [🏅 MCX Commodity] — Re-entry monitoring armed after {tgt_label}\n\n"
                f"LSP={tgt_price:,.2f} | "
                f"New BA={rw['threshold_entry_price']:,.2f} | "
                f"New SB={rw['retouch_entry_price']:,.2f}\n"
                f"▶ Threshold: price {thresh_dir} {rw['threshold_price']:,.2f} "
                f"→ {rw['threshold_entry_side']} @ {rw['threshold_entry_price']:,.2f}  "
                f"SL {rw['threshold_entry_sl']:,.2f}  Tgt {rw['threshold_entry_target']:,.2f}\n"
                f"▶ Retouch: price {retouch_dir} {tgt_price:,.2f} "
                f"→ {rw['retouch_entry_side']} @ {rw['retouch_entry_price']:,.2f}  "
                f"SL {rw['retouch_entry_sl']:,.2f}  Tgt {rw['retouch_entry_target']:,.2f}"
            )
            _send_alert(armed_msg)

    if side == "BUY":
        # Target exits T1-T5 (full exit per Algofinal — single position, take at first hit)
        for i, tgt in enumerate(pos["targets"], 1):
            if price >= tgt:
                _close(tgt, f"T{i}_HIT", arm_reentry=True, tgt_label=f"T{i}", tgt_price=tgt)
                return
        # Stop Loss — NO re-entry armed after SL
        if price <= lv["buy_sl"]:
            gross_usdt = (price - entry_px) * qty
            net_usdt   = gross_usdt - brokerage_usdt
            gross_inr  = gross_usdt * USDT_TO_INR
            net_inr    = net_usdt * USDT_TO_INR
            status_sl = f"BUY STOP LOSS HIT | Gross ₹{gross_inr:.2f} | Net ₹{net_inr:.2f}"
            _COMM_POSITIONS[sym] = None
            _COMM_EXITED[sym]    = True
            _COMM_REENTRY[sym]   = None
            _send_alert(_build_alert("Exit", sym, status_sl, side="BUY", quantity=qty,
                                      **{**kw, "current_price": price}))
            _log_trade({"ts": ts_str, "symbol": sym, "side": "BUY",
                        "entry_px": entry_px, "exit_px": price, "qty": qty,
                        "gross_pnl": round(gross_inr, 2), "brokerage": BROKERAGE_FLAT,
                        "net_pnl": round(net_inr, 2), "reason": "SL_HIT",
                        "x_val": lv.get("x_mult", 0), "asset_class": "commodity"})
            log.info("SL_HIT %s @ %.2f", sym, price)
            return
        # Retreat 65/45/25 (only WARNING = retreat exit — no other warnings)
        ba = lv["buy_above"]
        if price >= ba + 0.65 * step:
            pos["retreat_peak_reached"] = True
        if pos["retreat_peak_reached"] and price <= ba + 0.25 * step:
            gross_usdt = step * 0.25 * qty
            net_usdt   = gross_usdt - brokerage_usdt
            gross_inr  = gross_usdt * USDT_TO_INR
            net_inr    = net_usdt * USDT_TO_INR
            pct         = net_inr / 100_000 * 100
            status_r = (f"RETREAT 25% EXIT BUY | Price={price:,.2f} "
                        f"| Gross ₹{gross_inr:,.2f} | Net ₹{net_inr:,.2f} ({pct:+.3f}%)")
            _COMM_POSITIONS[sym] = None
            _COMM_EXITED[sym]    = True
            _COMM_REENTRY[sym]   = None
            _send_alert(_build_alert("Exit", sym, status_r, side="BUY", quantity=qty,
                                      **{**kw, "current_price": price}))
            _log_trade({"ts": ts_str, "symbol": sym, "side": "BUY",
                        "entry_px": entry_px, "exit_px": price, "qty": qty,
                        "gross_pnl": round(gross_inr, 2), "brokerage": BROKERAGE_FLAT,
                        "net_pnl": round(net_inr, 2), "reason": "RETREAT",
                        "x_val": lv.get("x_mult", 0), "asset_class": "commodity"})
            log.info("RETREAT %s @ %.2f net_inr=%.2f", sym, price, net_inr)
            return
    else:  # SELL
        for i, tgt in enumerate(pos["targets"], 1):
            if price <= tgt:
                _close(tgt, f"ST{i}_HIT", arm_reentry=True, tgt_label=f"ST{i}", tgt_price=tgt)
                return
        if price >= lv["sell_sl"]:
            gross_usdt = (entry_px - price) * qty
            net_usdt   = gross_usdt - brokerage_usdt
            gross_inr  = gross_usdt * USDT_TO_INR
            net_inr    = net_usdt * USDT_TO_INR
            status_sl = f"SELL STOP LOSS HIT | Gross ₹{gross_inr:.2f} | Net ₹{net_inr:.2f}"
            _COMM_POSITIONS[sym] = None
            _COMM_EXITED[sym]    = True
            _COMM_REENTRY[sym]   = None
            _send_alert(_build_alert("Exit", sym, status_sl, side="SELL", quantity=qty,
                                      **{**kw, "current_price": price}))
            _log_trade({"ts": ts_str, "symbol": sym, "side": "SELL",
                        "entry_px": entry_px, "exit_px": price, "qty": qty,
                        "gross_pnl": round(gross_inr, 2), "brokerage": BROKERAGE_FLAT,
                        "net_pnl": round(net_inr, 2), "reason": "SL_HIT",
                        "x_val": lv.get("x_mult", 0), "asset_class": "commodity"})
            log.info("SL_HIT %s @ %.2f", sym, price)
            return
        sb = lv["sell_below"]
        if price <= sb - 0.65 * step:
            pos["retreat_peak_reached"] = True
        if pos["retreat_peak_reached"] and price >= sb - 0.25 * step:
            gross_usdt = step * 0.25 * qty
            net_usdt   = gross_usdt - brokerage_usdt
            gross_inr  = gross_usdt * USDT_TO_INR
            net_inr    = net_usdt * USDT_TO_INR
            pct         = net_inr / 100_000 * 100
            status_r = (f"RETREAT 25% EXIT SELL | Price={price:,.2f} "
                        f"| Gross ₹{gross_inr:,.2f} | Net ₹{net_inr:,.2f} ({pct:+.3f}%)")
            _COMM_POSITIONS[sym] = None
            _COMM_EXITED[sym]    = True
            _COMM_REENTRY[sym]   = None
            _send_alert(_build_alert("Exit", sym, status_r, side="SELL", quantity=qty,
                                      **{**kw, "current_price": price}))
            _log_trade({"ts": ts_str, "symbol": sym, "side": "SELL",
                        "entry_px": entry_px, "exit_px": price, "qty": qty,
                        "gross_pnl": round(gross_inr, 2), "brokerage": BROKERAGE_FLAT,
                        "net_pnl": round(net_inr, 2), "reason": "RETREAT",
                        "x_val": lv.get("x_mult", 0), "asset_class": "commodity"})
            log.info("RETREAT %s @ %.2f net_inr=%.2f", sym, price, net_inr)
            return
# ════════════════════════════════════════════════════════════════════════════
# PRICE REFRESH LOOP (every 2 seconds)
# ════════════════════════════════════════════════════════════════════════════

def _price_loop() -> None:
    """Main price processing loop — runs every 2s."""
    log.info("Commodity price loop started")
    eod_done     = False
    anchor_930   = False
    while True:
        try:
            now = _now_ist()
            in_sess = _in_session(now)

            # 09:30 re-anchor (only during session)
            if in_sess:
                if (now.hour > 9 or (now.hour == 9 and now.minute >= 30)):
                    if not anchor_930:
                        anchor_930 = True
                        log.info("09:30 re-anchor: recalculating commodity levels")
                        _recalc_levels_930()

                # EOD at 23:30
                if _is_eod(now) and not eod_done:
                    log.info("MCX EOD 23:30 — squaring off all positions")
                    _eod_square_off(now)
                    eod_done = True

            # v10.7 FIX: always publish prices from REST poll cache
            # (REST poller fills _COMM_PRICES continuously regardless of session)
            # This ensures dashboard shows commodity prices even on weekends
            prices_snap: Dict[str, float] = {}
            with _PRICE_LOCK:
                prices_snap = dict(_COMM_PRICES)

            if prices_snap:
                # Always publish to live_prices.json for dashboard display
                _publish_comm_prices(prices_snap)
                # ZMQ only during session (scanners only need prices when MCX is open)
                if _pub and in_sess:
                    _pub.publish(prices_snap, now, topic="commodity")

            # Process each symbol only during active session
            if in_sess:
                for sym in SYMBOLS:
                    px = prices_snap.get(sym)
                    if px and px > 0:
                        _process_price(sym, px, now)

            # Adaptive sleep
            if in_sess:
                time.sleep(1)    # 1s during MCX session
            else:
                time.sleep(10)   # 10s off-hours (REST poller still refreshes every 3s)
        except Exception as exc:
            log.error("Price loop error: %s", exc)
            time.sleep(5)


def _recalc_levels_930() -> None:
    """Re-anchor commodity levels at 09:30 using latest prices."""
    for sym in SYMBOLS:
        with _PRICE_LOCK:
            price = _COMM_PRICES.get(sym)
        if price and price > 0 and _COMM_PREV_CLOSE.get(sym, 0) > 0:
            if abs(price - _COMM_PREV_CLOSE[sym]) / _COMM_PREV_CLOSE[sym] < 0.10:
                _COMM_LEVELS[sym] = _calc_levels(sym, price)
                _COMM_ANCHOR[sym] = price
                COMMODITY_ANCHOR[sym] = price
                log.info("930 re-anchor: %s @ %.2f", sym, price)


def _eod_square_off(ts: datetime) -> None:
    """Close all open positions at EOD."""
    now_s = ts.strftime("%H:%M:%S")
    for sym in SYMBOLS:
        pos = _COMM_POSITIONS[sym]
        if pos is None:
            continue
        with _PRICE_LOCK:
            price = _COMM_PRICES.get(sym, pos["entry_px"])
        side     = pos["side"]
        entry_px = pos["entry_px"]
        qty      = pos["qty"]
        gross_usdt = (price - entry_px) * qty if side == "BUY" else (entry_px - price) * qty
        brokerage_usdt = BROKERAGE / USDT_TO_INR if USDT_TO_INR else BROKERAGE
        net_usdt = gross_usdt - brokerage_usdt
        gross_inr = gross_usdt * USDT_TO_INR
        net_inr   = net_usdt * USDT_TO_INR
        _COMM_POSITIONS[sym] = None
        _COMM_EXITED[sym]    = True
        _log_trade({"ts": now_s, "symbol": sym, "side": side,
                    "entry_px": entry_px, "exit_px": price, "qty": qty,
                    "gross_pnl": round(gross_inr, 2), "brokerage": BROKERAGE,
                    "net_pnl": round(net_inr, 2), "reason": "EOD_2330",
                    "x_val": _COMM_LEVELS.get(sym, {}).get("x_mult", 0),
                    "asset_class": "commodity"})
        log.info("EOD 23:30 square-off: %s %s @ %.2f net_inr=%.2f", side, sym, price, net_inr)
    _send_eod_summary(ts)


def _send_eod_summary(ts: datetime) -> None:
    """Send EOD summary text + Excel to commodity Telegram bot."""
    from tg_async import send_alert, send_document_alert, get_dashboard_url
    date_str = ts.strftime("%Y%m%d")
    now_s    = ts.strftime("%H:%M")

    closed = [t for t in _COMM_TRADES if t.get("exit_px") is not None]
    if not closed:
        return

    gross_total = sum(t.get("gross_pnl", 0) or 0 for t in closed)
    brok_total  = len(closed) * BROKERAGE
    net_total   = sum(t.get("net_pnl",  0) or 0 for t in closed)
    n           = len(closed)
    capital     = n * 100_000
    pct         = (net_total / capital * 100) if capital > 0 else 0.0
    by_pnl      = sorted(closed, key=lambda t: t.get("net_pnl") or 0, reverse=True)
    best        = by_pnl[0]  if by_pnl else {}
    worst       = by_pnl[-1] if by_pnl else {}

    url  = get_dashboard_url()
    msg  = (
        f"EOD Summary - MCX - {ts.strftime('%d %b %Y')} {now_s} IST\n"
        f"─────────────────────────────────\n"
        f"Closed Trades: {n}\n"
        f"Gross P&L:  Rs{gross_total:+,.2f}\n"
        f"Brokerage:  Rs{brok_total:,.2f}\n"
        f"Net P&L:    Rs{net_total:+,.2f}\n"
        f"─────────────────────────────────\n"
        f"Best:  {best.get('symbol','-')} {best.get('side','-')} "
        f"{best.get('reason','-')} Rs{best.get('net_pnl',0):+,.2f}\n"
        f"Worst: {worst.get('symbol','-')} {worst.get('side','-')} "
        f"{worst.get('reason','-')} Rs{worst.get('net_pnl',0):+,.2f}\n"
        f"─────────────────────────────────\n"
        f"Return: {pct:+.3f}%  (target: 0.30%)\n"
        f"Capital: Rs{capital:,.0f}\n"
        f"─────────────────────────────────\n"
        f"Dashboard: {url}"
    )
    send_alert(msg, asset_class="commodity")

    # Write and send EOD Excel
    xl_path = os.path.join("trade_analysis",
                           f"commodity_trade_analysis_{date_str}.xlsx")
    _write_eod_xlsx(date_str, xl_path)
    if os.path.exists(xl_path):
        send_document_alert(xl_path,
                            f"MCX trade analysis (23:30) ready for {ts.strftime('%d %b %Y')}",
                            asset_class="commodity")


# ════════════════════════════════════════════════════════════════════════════
# STARTUP
# ════════════════════════════════════════════════════════════════════════════

def _try_load_cached_levels(date_str: str) -> None:
    """v10.6: Load levels from most recent JSON file for dashboard display on non-trading days."""
    import glob
    try:
        # Try today's file first, then any recent file
        patterns = [
            os.path.join("levels", f"commodity_initial_levels_{date_str}.json"),
            os.path.join("levels", "commodity_initial_levels_*.json"),
        ]
        candidates = []
        for pat in patterns:
            candidates.extend(glob.glob(pat))
        if not candidates:
            log.info("No cached commodity levels file found — prices will show from REST only")
            return
        candidates.sort(key=os.path.getmtime, reverse=True)
        best = candidates[0]
        d = json.load(open(best, encoding="utf-8"))
        lvs = d.get("levels", {})
        for sym, lv in lvs.items():
            if sym in SYMBOLS and isinstance(lv, dict):
                _COMM_LEVELS[sym] = lv
                pc = lv.get("prev_close", 0)
                if pc:
                    # Convert cached INR-ish levels to USDT when needed.
                    lv2 = {}
                    for k, v in lv.items():
                        if k == "x_mult":
                            lv2[k] = v
                            continue
                        if isinstance(v, (int, float)) and v:
                            lv2[k] = _normalise_price_to_usdt(v)
                        else:
                            lv2[k] = v
                    _COMM_LEVELS[sym] = lv2
                    _COMM_PREV_CLOSE[sym] = float(lv2.get("prev_close", 0) or 0)
                    _COMM_ANCHOR[sym] = float(lv2.get("prev_close", 0) or 0)
                    COMMODITY_ANCHOR[sym] = float(lv2.get("prev_close", 0) or 0)
        log.info("Loaded cached MCX levels from %s (%d symbols)", best, len(lvs))
    except Exception as exc:
        log.debug("_try_load_cached_levels: %s", exc)


def startup() -> None:
    """Initialise commodity engine — call once at process start.
    v10.7 FIX: REST poller + price publishing always starts regardless of trading day,
    so dashboard always shows commodity prices (even on weekends).
    """
    from market_calendar import MarketCalendar
    now      = _now_ist()
    date_str = now.strftime("%Y%m%d")

    log.info("═" * 60)
    log.info("AlgoStack v10.7 — Commodity Engine (MCX)")
    log.info("Author: Ridhaant Ajoy Thackur")
    log.info("Symbols: %s", ", ".join(SYMBOLS))
    log.info("Session: 09:00 – 23:30 IST (Mon–Fri)")
    log.info("═" * 60)

    # v10.7 FIX: ALWAYS start REST poller first — provides prices for dashboard
    # display even on weekends/non-trading days. Runs as daemon thread.
    log.info("Starting commodity REST price poller (always-on for dashboard)…")
    _start_commodity_rest_poll()

    if not MarketCalendar.is_trading_day(now):
        log.info("Not a trading day — commodity engine in dashboard-only mode")
        log.info("REST poller will fetch prices from yfinance every 3s for display")
        # Try to load prev-close levels from most recent file for display
        _try_load_cached_levels(date_str)
        return

    # Load prev closes
    log.info("Loading MCX previous closes...")
    pcs = _load_prev_closes()
    for sym in SYMBOLS:
        pc = pcs.get(sym, 0.0)
        _COMM_PREV_CLOSE[sym] = pc
        if pc > 0:
            _COMM_LEVELS[sym] = _calc_levels(sym, pc)
            _COMM_ANCHOR[sym] = pc
            COMMODITY_ANCHOR[sym] = pc
            log.info("  %-14s prev_close=%.2f  x=%.6f",
                     sym, pc, cfg.COMM_X.get(sym, 0))
        else:
            log.warning("  %-14s prev_close unavailable — skipped", sym)

    # Write initial levels
    os.makedirs("levels", exist_ok=True)
    lv_path = os.path.join("levels", f"commodity_initial_levels_{date_str}.json")
    try:
        tmp = lv_path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump({
                "date": date_str,
                "author": "Ridhaant Ajoy Thackur",
                "levels": _COMM_LEVELS,
            }, f, indent=2)
        os.replace(tmp, lv_path)
        log.info("Commodity levels JSON → %s", lv_path)
    except Exception as exc:
        log.warning("Levels JSON write failed: %s", exc)

    # Write initial levels XLSX and send to Telegram
    xl_path = os.path.join("levels",
                           f"commodity_initial_levels_{date_str}.xlsx")
    _write_initial_levels_xlsx(date_str, xl_path)
    try:
        from tg_async import send_document_alert
        send_document_alert(xl_path,
                            f"MCX initial levels for {now.strftime('%d %b %Y')}",
                            asset_class="commodity")
    except Exception:
        pass

    # Start TradingView WS (primary price source during session)
    _start_tradingview_ws()

    # Start ZMQ publisher
    _init_zmq()

    log.info("Commodity engine ready. Starting price loop...")


def main() -> None:
    startup()
    _price_loop()


if __name__ == "__main__":
    main()
