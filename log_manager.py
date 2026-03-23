# ═══════════════════════════════════════════════════════════════════════
# AlgoStack v8.0 | Author: Ridhaant Ajoy Thackur
# log_manager.py — Centralised Logging & Storage Manager
# ═══════════════════════════════════════════════════════════════════════
"""
log_manager.py — AlgoStack Logging & Storage Manager
=====================================================
Provides a unified folder tree for all logs, reports, and data files.
Every file produced by AlgoStack is routed through this module.

Usage:
    from log_manager import LogManager, get_log_path, setup_logging
    path = LogManager.path("trades", f"trade_events_{ds}.jsonl")
    setup_logging("scanner1")
"""
from __future__ import annotations

import logging
import logging.handlers
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import pytz

IST = pytz.timezone("Asia/Kolkata")

# ── Folder structure ──────────────────────────────────────────────────────────
LOG_ROOT = "algostack_logs"

STRUCTURE: Dict[str, str] = {
    "trades":    os.path.join(LOG_ROOT, "trades"),       # trade_events_YYYYMMDD.jsonl
    "levels":    os.path.join(LOG_ROOT, "levels"),       # initial_levels_*.xlsx
    "scanner1":  os.path.join(LOG_ROOT, "scanners", "s1"),
    "scanner2":  os.path.join(LOG_ROOT, "scanners", "s2"),
    "scanner3":  os.path.join(LOG_ROOT, "scanners", "s3"),
    "optimizer": os.path.join(LOG_ROOT, "optimizer"),    # xopt_*.csv
    "bestx":     os.path.join(LOG_ROOT, "bestx"),        # best_x_trades_*.xlsx
    "errors":    os.path.join(LOG_ROOT, "errors"),       # crash logs per process
    "watchdog":  os.path.join(LOG_ROOT, "watchdog"),
    "prices":    os.path.join(LOG_ROOT, "prices"),       # live_prices history
    "reports":   os.path.join(LOG_ROOT, "reports"),      # daily summary PDFs/XLSX
    "sweep":     os.path.join(LOG_ROOT, "sweep"),        # sweep_results copies
}


class LogManager:
    """
    Central routing for all AlgoStack file I/O.

    Ensures:
      - All directories exist before use
      - Author header in every text/CSV file
      - Atomic writes (.tmp -> os.replace) to prevent corruption
      - Crash logs written on process exit
    """

    _initialized = False

    @classmethod
    def init(cls) -> None:
        """Create all log directories. Call once at startup."""
        for category, path in STRUCTURE.items():
            os.makedirs(path, exist_ok=True)
        cls._initialized = True

    @classmethod
    def path(cls, category: str, filename: str) -> str:
        """
        Return the full path for a file in the given log category.
        Creates the directory if it doesn't exist.

        Example:
            path = LogManager.path("trades", "trade_events_20260320.jsonl")
            # -> algostack_logs/trades/trade_events_20260320.jsonl
        """
        dir_path = STRUCTURE.get(category, os.path.join(LOG_ROOT, category))
        os.makedirs(dir_path, exist_ok=True)
        return os.path.join(dir_path, filename)

    @classmethod
    def error_path(cls, process_name: str, ts: Optional[str] = None) -> str:
        """Return path for a crash log file for the given process."""
        ts = ts or datetime.now(IST).strftime("%Y%m%d_%H%M%S")
        err_dir = os.path.join(STRUCTURE["errors"], process_name)
        os.makedirs(err_dir, exist_ok=True)
        return os.path.join(err_dir, f"crash_{ts}.log")

    @classmethod
    def write_atomic(cls, path: str, content: str, encoding: str = "utf-8") -> None:
        """
        Write text content atomically using .tmp -> os.replace pattern.
        Prevents file corruption on process crash during write.
        """
        tmp = path + ".tmp"
        with open(tmp, "w", encoding=encoding) as f:
            f.write(content)
        os.replace(tmp, path)

    @classmethod
    def write_crash_log(cls, process_name: str, tail_lines: list[str],
                        exc: Optional[Exception] = None) -> str:
        """Write a crash log and return the path."""
        ts = datetime.now(IST).strftime("%Y%m%d_%H%M%S")
        path = cls.error_path(process_name, ts)
        header = (
            f"# AlgoStack v8.0 | Author: Ridhaant Ajoy Thackur\n"
            f"# Crash log: {process_name}\n"
            f"# Timestamp: {datetime.now(IST).isoformat()}\n"
            f"# Exception: {exc}\n"
            f"{'─' * 60}\n"
        )
        body = "\n".join(tail_lines[-200:])
        cls.write_atomic(path, header + body)
        return path

    @classmethod
    def xlsx_header_row(cls, report_name: str) -> str:
        """Standard header string for row 1 of every XLSX report."""
        ds = datetime.now(IST).strftime("%d %b %Y")
        return f"AlgoStack v8.0 — {report_name} | Author: Ridhaant Ajoy Thackur | {ds}"

    @classmethod
    def apply_xlsx_header(cls, wb, report_name: str) -> None:
        """
        Apply branded header to openpyxl workbook first sheet.
        Merges A1 across all used columns and styles it.
        """
        try:
            from openpyxl.styles import Alignment, Font, PatternFill
            ws = wb.active
            header = cls.xlsx_header_row(report_name)
            max_col = ws.max_column or 26
            ws.insert_rows(1)
            ws["A1"] = header
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(max_col, 6))
            ws["A1"].font = Font(bold=True, color="FFFFFF", size=11,
                                 name="Segoe UI")
            ws["A1"].fill = PatternFill("solid", fgColor="0D1117")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22
        except Exception as exc:
            logging.getLogger("log_manager").debug("xlsx header apply failed: %s", exc)


def setup_logging(process_name: str, level: str = "INFO") -> logging.Logger:
    """
    Configure logging for a process:
      - Console handler (Rich-compatible)
      - Rotating file handler → algostack_logs/errors/{process_name}/
    Returns the root logger.
    """
    log_level = getattr(logging, level.upper(), logging.INFO)
    log_dir = os.path.join(STRUCTURE["errors"], process_name)
    os.makedirs(log_dir, exist_ok=True)

    today = datetime.now(IST).strftime("%Y-%m-%d")
    log_file = os.path.join(log_dir, f"{today}.log")

    fmt = logging.Formatter(
        f"%(asctime)s [{process_name}] %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    # Console
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    ch.setLevel(log_level)

    # Rotating file (10MB, keep 7 days)
    fh = logging.handlers.RotatingFileHandler(
        log_file, maxBytes=10 * 1024 * 1024, backupCount=7,
        encoding="utf-8",
    )
    fh.setFormatter(fmt)
    fh.setLevel(log_level)

    root = logging.getLogger()
    root.setLevel(log_level)
    if not root.handlers:
        root.addHandler(ch)
        root.addHandler(fh)

    return logging.getLogger(process_name)


# ── Convenience shortcut ──────────────────────────────────────────────────────
def get_log_path(category: str, filename: str) -> str:
    """Shortcut for LogManager.path()."""
    return LogManager.path(category, filename)


# ── Auto-init on import ───────────────────────────────────────────────────────
LogManager.init()
