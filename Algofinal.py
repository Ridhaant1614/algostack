# Author: Ridhaant Ajoy Thackur
import os
import sys
import time
import math
import json
import signal
import logging
import socket
import re
from urllib.parse import urlparse
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set, Tuple, Union, Iterable

from datetime import datetime, timedelta, time as dtime

import pytz
try:
    from market_calendar import MarketCalendar, startup_market_check
    _MC_AVAILABLE = True
except Exception:
    _MC_AVAILABLE = False
    MarketCalendar = None
    startup_market_check = None
import requests
import yfinance as yf
from rich.console import Console
from rich.table import Table
from rich.live import Live
import pandas as pd
from collections import deque
import threading
import socket
from copy import deepcopy
import io
from datetime import datetime
import os
from fpdf import FPDF  # Add this dependency to requirements.txt if not present
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ── AlgoStack v2.0 patch — ZMQ IPC + async Telegram + unified dashboard ───────
# Graceful: if new modules not present yet, original code paths remain active.
_PATCH_ACTIVE = False
try:
    from ipc_bus import PricePublisher as _ZmqPricePublisher
    from tg_async import AsyncTelegramSender as _AsyncTgSender
    _ASYNC_TG = _AsyncTgSender()          # single background worker thread
    _PATCH_ACTIVE = True
except ImportError:
    _ZmqPricePublisher = None
    _ASYNC_TG = None
# ─────────────────────────────────────────────────────────────────────────────

# Optional plotting (Plotly Dash)
ENABLE_DASH = os.getenv("ENABLE_DASH", "1") == "1"
DASH_PORT = int(os.getenv("DASH_PORT", "8050"))
MERGED_DASH_PORT = int(os.getenv("MERGED_DASH_PORT", "8052"))
HISTORY_POINTS = int(os.getenv("HISTORY_POINTS", "780"))  # ~full trading day at 1m cadence

if ENABLE_DASH:
    try:
        import dash
        from dash import Dash, dcc, html, Input, Output, State
        import plotly.graph_objects as go
        DASH_AVAILABLE = True
    except Exception:
        DASH_AVAILABLE = False
        ENABLE_DASH = False
else:
    DASH_AVAILABLE = False

# Optional public tunnel (ngrok)
ENABLE_TUNNEL = os.getenv("ENABLE_TUNNEL", "1") == "1"
# By default, only expose the equity dashboard publicly. Some tunnel providers /
# plans allow only one active tunnel at a time; exposing both can cause the
# equity URL to resolve to the commodity dashboard.
ENABLE_COMMODITY_TUNNEL = os.getenv("ENABLE_COMMODITY_TUNNEL", "0") == "1"
# Prefer Cloudflare Tunnel if enabled
ENABLE_CF_TUNNEL = False   # Cloudflare disabled — ngrok only
CLOUDFLARED_PATH = os.getenv("CLOUDFLARED_PATH", "cloudflared")
# Hardcoded ngrok tokens (can be overridden by env var or file)
# Updated to user's new ngrok account authtokens for equity and commodity dashboards.
NGROK_AUTHTOKEN_EQUITY = "34nI4TrxcnPtlbm6PpNfdJC1H26_5dc7wKYKnMc7DSPTSaAn3"
NGROK_AUTHTOKEN_COMMODITY = "35z4YU0PjNT7tfruOQIA6G0PCsm_5JoauhkDkvjdc5Ypn3EaA"
NGROK_AUTHTOKEN_DEFAULT = NGROK_AUTHTOKEN_EQUITY  # For backward compatibility
NGROK_AUTHTOKEN = os.getenv("NGROK_AUTHTOKEN", NGROK_AUTHTOKEN_DEFAULT)
try:
    from pyngrok import ngrok
    NGROK_AVAILABLE = True
except Exception:
    NGROK_AVAILABLE = False

# Keep references to long-running tunnel processes so they are not garbage
# collected (which can inadvertently terminate the tunnel on some platforms).
TUNNEL_PROCESSES: List[subprocess.Popen] = []


def _drain_process_stdout(proc: subprocess.Popen, label: str) -> None:
    """Continuously drain stdout for a long-running subprocess to avoid deadlocks."""
    if not proc.stdout:
        return

    def _drain() -> None:
        try:
            for line in iter(proc.stdout.readline, ""):
                if not line:
                    break
                logger.debug("%s: %s", label, line.rstrip())
        except Exception as exc:
            logger.debug("stdout drain for %s terminated: %s", label, exc)

    threading.Thread(target=_drain, name=f"{label}_stdout", daemon=True).start()

# Helper to load ngrok token from local file if env not set
def _load_ngrok_token_from_file() -> str:
    try:
        token_path = os.path.join(os.getcwd(), 'ngrok_token.txt')
        if os.path.exists(token_path):
            with open(token_path, 'r', encoding='utf-8') as f:
                return f.read().strip()
    except Exception:
        pass
    return ""

# Start a Cloudflare quick tunnel and return the public URL
def _launch_cloudflared_proc(port: int) -> "tuple[Optional[subprocess.Popen], Optional[str]]":
    """
    Launch a cloudflared quick-tunnel subprocess and return (proc, public_url).

    KEY FIX: stdout is drained in a background thread from the very first line.
    The old approach read stdout synchronously until the URL was found, then did
    a 20s hostname-resolution loop while stdout was NOT being read. This filled
    the OS pipe buffer (~64 KB on Windows), causing cloudflared to block on its
    next write and the connection to time out → Error 1033.
    Now stdout is always consumed, so cloudflared never blocks.
    """
    try:
        proc = subprocess.Popen(
            [CLOUDFLARED_PATH, "tunnel", "--url", f"http://localhost:{port}"],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )

        found_url: Optional[str] = None
        url_event = threading.Event()

        def _reader() -> None:
            """Read every line cloudflared writes; signal when URL is found."""
            nonlocal found_url
            try:
                for raw_line in proc.stdout:  # blocks per line, never stalls pipe
                    line = raw_line.rstrip()
                    logger.debug("cloudflared[%d]: %s", port, line)
                    if not found_url:
                        m = re.search(r'https://[\w\-\.]+\.trycloudflare\.com', line)
                        if m:
                            found_url = m.group(0).strip().rstrip(',')
                            url_event.set()
            except Exception:
                pass
            finally:
                if not url_event.is_set():
                    url_event.set()   # unblock the wait below on unexpected exit

        reader_thread = threading.Thread(target=_reader, daemon=True,
                                         name=f"cf_reader_{port}")
        reader_thread.start()

        # Wait up to 45s for cloudflared to print a trycloudflare.com URL
        url_event.wait(timeout=45)

        if not found_url:
            rc = proc.poll()
            logger.warning("cloudflared gave no URL in 45s (exit code: %s)", rc)
            try: proc.terminate()
            except Exception: pass
            return None, None

        TUNNEL_PROCESSES.append(proc)
        logger.info("cloudflared PID %s live → %s", proc.pid, found_url)
        return proc, found_url

    except FileNotFoundError:
        logger.warning(
            "cloudflared not found. Download: "
            "https://developers.cloudflare.com/cloudflare-one/connections/"
            "connect-networks/downloads/")
        return None, None
    except Exception as exc:
        logger.warning("cloudflared launch failed: %s", exc)
        return None, None


def start_cloudflare_tunnel(port: int,
                             *,
                             notify_func=None,
                             is_commodity_dash: bool = False) -> Optional[str]:
    """
    Start a Cloudflare quick-tunnel for *port* and return the public URL.
    Spawns a background watchdog thread that restarts the tunnel if the
    cloudflared process dies, and sends the new URL via Telegram.
    """
    if not ENABLE_TUNNEL or not ENABLE_CF_TUNNEL:
        return None

    proc, public_url = _launch_cloudflared_proc(port)
    if not public_url:
        return None

    # ── Watchdog: restart cloudflared if it ever dies ─────────────────────────
    def _watchdog() -> None:
        nonlocal proc, public_url
        while True:
            time.sleep(10)
            if proc is None or proc.poll() is not None:
                # Process died — restart
                logger.warning("cloudflared died (port %d). Restarting tunnel...", port)
                new_proc, new_url = _launch_cloudflared_proc(port)
                if new_proc and new_url:
                    proc = new_proc
                    public_url = new_url
                    logger.info("cloudflared restarted, new URL: %s", new_url)
                    # Notify via Telegram
                    try:
                        local_ip = _get_local_ip()
                        msg = (f"🔄 Cloudflare tunnel restarted (port {port})\n"
                               f"☁️  New CF URL: {new_url}\n"
                               f"💻 Local: http://localhost:{port}")
                        if local_ip:
                            msg += f"\n🌐 Network: http://{local_ip}:{port}"
                        send_telegram(msg, force_commodity=is_commodity_dash)
                    except Exception:
                        pass
                else:
                    logger.warning("cloudflared restart failed — will retry in 30s")
                    time.sleep(20)

    wt = threading.Thread(target=_watchdog, daemon=True, name=f"cf_watchdog_{port}")
    wt.start()
    return public_url


def _get_local_ip() -> Optional[str]:
    """Return LAN IP, or None."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return None


def _start_ngrok_only(port: int, authtoken: str) -> Optional[str]:
    """
    Start ONLY an ngrok tunnel — never falls back to Cloudflare.

    Fixes ERR_NGROK_334 (endpoint already online) by killing the ngrok daemon
    on all platforms before connecting:
      • Windows: taskkill /F /IM ngrok.exe
      • Linux/Mac: pkill -f ngrok
    This clears stale free-tier sessions that survive between Python runs.
    """
    if not NGROK_AVAILABLE:
        logger.warning("pyngrok not installed: pip install pyngrok")
        return None

    # ── Step 1: Kill any running ngrok daemon ─────────────────────────────────
    logger.info("Terminating existing ngrok processes to clear stale sessions...")
    # pyngrok kill (graceful)
    try:
        ngrok.kill()
        time.sleep(1)
    except Exception:
        pass
    # OS-level kill (belt-and-suspenders) — works even when pyngrok is confused
    try:
        if os.name == "nt":   # Windows
            subprocess.run(
                ["taskkill", "/F", "/IM", "ngrok.exe"],
                capture_output=True, timeout=5
            )
        else:                 # Linux / macOS
            subprocess.run(
                ["pkill", "-f", "ngrok"],
                capture_output=True, timeout=5
            )
        time.sleep(2)         # let OS fully reclaim ports
    except Exception:
        pass
    logger.info("Ngrok daemon cleared.")

    # ── Step 2: Set authtoken ─────────────────────────────────────────────────
    try:
        ngrok.set_auth_token(authtoken)
    except Exception as e:
        logger.warning("ngrok set_auth_token: %s", e)

    # ── Step 3: Connect with retries ─────────────────────────────────────────
    for attempt in range(4):
        try:
            tun = ngrok.connect(port, proto="http")
            url = tun.public_url
            if url:
                logger.info("✅ Ngrok live (attempt %d): %s", attempt + 1, url)
                return url
        except Exception as exc:
            logger.warning("Ngrok attempt %d failed: %s", attempt + 1, exc)
            if "334" in str(exc) or "already online" in str(exc):
                # Still a stale session — wait longer and retry
                logger.info("Stale session still active, waiting 5s...")
                time.sleep(5)
            else:
                time.sleep(3)

    logger.warning("All ngrok attempts failed for port %d", port)
    return None


# Start a public tunnel via ngrok if CF not used/available
def start_public_tunnel(port: int, authtoken: Optional[str] = None) -> Optional[str]:
    if not ENABLE_TUNNEL:
        return None
    # Prefer CF
    url = start_cloudflare_tunnel(port)
    if url:
        return url
    # Fallback to ngrok
    if not NGROK_AVAILABLE:
        logger.warning("Ngrok not available. Install with: pip install pyngrok")
        return None
    try:
        # Priority: parameter > env var > file > hardcoded default
        token = None
        if authtoken:
            token = authtoken
            logger.debug("Using ngrok token from parameter")
        elif os.getenv("NGROK_AUTHTOKEN"):
            token = os.getenv("NGROK_AUTHTOKEN")
            logger.debug("Using ngrok token from environment variable")
        else:
            file_token = _load_ngrok_token_from_file()
            if file_token:
                token = file_token
                logger.debug("Using ngrok token from file")
            else:
                # Use default based on port (equity = 8050, commodity = 8051)
                if port == DASH_PORT:
                    token = NGROK_AUTHTOKEN_EQUITY
                elif port == COMMODITY_DASH_PORT:
                    token = NGROK_AUTHTOKEN_COMMODITY
                else:
                    token = NGROK_AUTHTOKEN_DEFAULT
                logger.debug("Using hardcoded ngrok token")
        
        if not token:
            logger.warning("No ngrok authtoken configured. Set NGROK_AUTHTOKEN env var or add to ngrok_token.txt")
            return None

        # ── Kill any existing ngrok daemon to clear stale tunnels (ERR_NGROK_334) ──
        # Free ngrok accounts can only have ONE active tunnel endpoint at a time.
        # If a previous Python process left a tunnel alive, we must kill the daemon first.
        try:
            logger.info("Killing existing ngrok daemon to clear stale sessions...")
            ngrok.kill()
            time.sleep(2)   # wait for OS to reclaim the port
            logger.info("Previous ngrok daemon stopped.")
        except Exception as kill_err:
            logger.debug("ngrok.kill() skipped or failed (may not have been running): %s", kill_err)

        # ── Set authtoken and connect ─────────────────────────────────────────
        try:
            ngrok.set_auth_token(token)
        except Exception as e:
            logger.warning("ngrok set_auth_token error: %s", e)

        last_exc = None
        for attempt in range(4):
            try:
                tun = ngrok.connect(port, proto="http")
                tunnel_url = tun.public_url
                if tunnel_url:
                    logger.info("✅ Ngrok tunnel live (attempt %d): %s → port %d",
                                attempt + 1, tunnel_url, port)
                    return tunnel_url
            except Exception as exc:
                last_exc = exc
                logger.warning("Ngrok connect attempt %d failed: %s", attempt + 1, exc)
                time.sleep(3)

        logger.warning("All ngrok connect attempts failed. Last error: %s", last_exc)
        return None
    except Exception as e:
        logger.warning(f"Failed to start ngrok tunnel: {e}")
        return None

# Global in-memory trade log for all events
TRADE_LOG: List[dict] = []
TRADE_LOG_KEYS: Set[str] = set()
TRADE_LOG_DIR = "trade_logs"
_TRADE_LOG_HYDRATED: Set[str] = set()

# ── RAM management: spill TRADE_LOG to disk once it exceeds this size ────────
# On 16 GB RAM with 38 stocks each ~30 events/day = ~1140 events; set limit at
# 5000 so we never hold more than ~5 MB of raw dicts in memory at once.
TRADE_LOG_RAM_LIMIT = 5000
_TRADE_LOG_SPILL_LOCK = threading.Lock()


def _spill_trade_log_to_disk_if_needed() -> None:
    """If TRADE_LOG exceeds RAM limit, flush events that are already persisted to disk.

    Events are written to JSONL files in log_trade_event(), so they are always
    safely on disk before we remove them from the in-memory list.  We only
    keep the most recent 1000 events in RAM for fast access by EOD functions.
    """
    global TRADE_LOG
    with _TRADE_LOG_SPILL_LOCK:
        if len(TRADE_LOG) < TRADE_LOG_RAM_LIMIT:
            return
        # Keep the latest 1000 events in RAM; discard the rest (already on disk)
        TRADE_LOG = TRADE_LOG[-1000:]
        import gc
        gc.collect()
        logger.info("TRADE_LOG RAM spill: trimmed to 1000 events (rest safe on disk).")
CLOSE_PRICES_1511_CACHE: Dict[str, float] = {}
CLOSE_PRICES_1511_CACHE_DATE: Optional[str] = None

# ── Shared market data file — written by Algofinal, read by all scanners ─────
# Scanners must NEVER call yfinance for live or 09:30 data — Algofinal is the
# sole fetcher.  These JSON files carry the data scanners need at startup.
SHARED_MARKET_DATA_DIR = "levels"
_SHARED_930_PRICES: Dict[str, float] = {}        # symbol → 09:30 IST price
_SHARED_PREMARKET_CLOSES: Dict[str, List[float]] = {}  # symbol → [1m closes 09:15–09:30]


def _shared_market_data_path(date_str: str) -> str:
    return os.path.join(SHARED_MARKET_DATA_DIR, f"shared_market_data_{date_str}.json")


def write_shared_market_data(date_str: Optional[str] = None) -> None:
    """Write 09:30 prices and premarket 1m history to a shared JSON file.

    Called by Algofinal after completing 09:30 re-anchor and premarket replay.
    The three scanners read this file at startup instead of calling yfinance —
    eliminating 228 duplicate yfinance requests (38 stocks × 3 scanners × 2 calls).
    """
    if not _SHARED_930_PRICES and not _SHARED_PREMARKET_CLOSES:
        return
    if date_str is None:
        date_str = now_ist().strftime("%Y%m%d")
    try:
        os.makedirs(SHARED_MARKET_DATA_DIR, exist_ok=True)
        path = _shared_market_data_path(date_str)
        data = {
            "date": date_str,
            "written_at": now_ist().isoformat(),
            "930_prices": dict(_SHARED_930_PRICES),
            "premarket_closes": dict(_SHARED_PREMARKET_CLOSES),
        }
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, indent=2)
        logger.info(
            "Shared market data written → %s  (930_prices=%d, premarket=%d symbols)",
            os.path.basename(path), len(_SHARED_930_PRICES), len(_SHARED_PREMARKET_CLOSES),
        )
    except Exception as e:
        logger.warning("Failed to write shared market data: %s", e)


def _trade_log_path(date_str: str) -> str:
    return os.path.join(TRADE_LOG_DIR, f"trade_events_{date_str}.jsonl")


def _ensure_trade_log_dir() -> None:
    os.makedirs(TRADE_LOG_DIR, exist_ok=True)


def _persist_trade_event(event: dict, date_key: str) -> None:
    try:
        _ensure_trade_log_dir()
        path = _trade_log_path(date_key)
        with open(path, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(event, default=str) + "\n")
    except Exception as exc:
        logger.warning("Failed to persist trade event: %s", exc)


def _load_trade_log_from_file(date_str: str) -> List[dict]:
    path = _trade_log_path(date_str)
    if not os.path.exists(path):
        return []
    events: List[dict] = []
    try:
        with open(path, "r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    event = json.loads(line)
                    events.append(event)
                except json.JSONDecodeError:
                    continue
    except Exception as exc:
        logger.warning("Failed to load trade log %s: %s", path, exc)
    return events


def hydrate_trade_log(date_str: str) -> None:
    if date_str in _TRADE_LOG_HYDRATED:
        return
    events = _load_trade_log_from_file(date_str)
    if events:
        for ev in events:
            key = _trade_event_key(ev)
            if key in TRADE_LOG_KEYS:
                continue
            TRADE_LOG_KEYS.add(key)
            TRADE_LOG.append(ev)
    _TRADE_LOG_HYDRATED.add(date_str)


def _trade_event_key(event: dict) -> str:
    return "|".join(
        [
            str(event.get("timestamp", "")),
            event.get("symbol", ""),
            event.get("event_type", ""),
            event.get("side", ""),
            str(event.get("qty", "")),
            str(event.get("price", "")),
        ]
    )


def get_trade_log_snapshot(date_str: str) -> List[dict]:
    """
    Return all trade events recorded for the provided trading date.
    Falls back to on-disk data if needed so Excel/EOD summaries remain populated
    even after a program restart.
    """
    hydrate_trade_log(date_str)
    return [ev for ev in TRADE_LOG if ev.get("trade_date") == date_str]

# Helper function: log event
def log_trade_event(timestamp, symbol, event_type, price, side, qty, entry_price=None, extra=None):
    ts = timestamp
    if isinstance(ts, datetime):
        ts = ts.astimezone(IST) if ts.tzinfo else IST.localize(ts)
        ts_str = ts.isoformat()
    else:
        ts_str = str(ts)
        ts = now_ist()
    date_key = ts.strftime('%Y%m%d')
    event = {
        'timestamp': ts_str,
        'symbol': symbol,
        'event_type': event_type,
        'price': price,
        'side': side,
        'qty': qty,
        'entry_price': entry_price,
        'trade_date': date_key,
    }
    if extra:
        event.update(extra)
    event_key = _trade_event_key(event)
    if event_key in TRADE_LOG_KEYS:
        return
    TRADE_LOG_KEYS.add(event_key)
    TRADE_LOG.append(event)
    _persist_trade_event(event, date_key)
    # Spill to disk if RAM limit exceeded (keeps only latest 1000 in memory)
    _spill_trade_log_to_disk_if_needed()

# Entry event classifier used by EOD summaries/analysis.
# Re-entry events are logged as e.g. BUY_REENTRY_THRESHOLD / SELL_REENTRY_RETOUCH.
def _is_entry_event_type(event_type: str) -> bool:
    if not event_type:
        return False
    et = str(event_type).upper()
    return (
        et in {"BUY_ENTRY", "SELL_ENTRY"}
        or et.startswith("BUY_REENTRY")
        or et.startswith("SELL_REENTRY")
    )

# Patch: everywhere an event occurs, also call log_trade_event()
# -- For buy/sell entry: in try_entry()
# -- For exit/target/sl: in handle_targets_and_trailing()
# EOD squareoff: also log with exit_type 'EOD'.

# (Below are IN PATCH locations, actual patching to follow)
# try_entry/state.side=\"BUY\": add log_trade_event(now, state.symbol, 'BUY_ENTRY', price, 'BUY', dyn_qty)
# try_entry/state.side=\"SELL\": add log_trade_event(now, state.symbol, 'SELL_ENTRY', price, 'SELL', dyn_qty)
# handle_targets_and_trailing/full exit/target hit: add log_trade_event(now, state.symbol, 'T{idx+1}', tgt, 'BUY', qty_to_exit, entry_price=state.entry_price)
# handle_targets_and_trailing/full exit/stoploss: log_trade_event(now, state.symbol, 'BUY_SL', price, 'BUY', qty_to_exit, entry_price=state.entry_price)
# (analogous for SELL)
# EOD: log PnL for open position as 'EOD_CLOSE'

# At EOD, write summary Excel

def eod_save_summary_excel(trade_log: List[dict], states, last_prices, date_str=None):
    try:
        import os
        from pandas import ExcelWriter
        if not trade_log:
            logger.info("No trades to save in EOD summary.")
            return
        if date_str is None:
            date_str = now_ist().strftime('%Y%m%d')
        summary_dir = "summary"
        os.makedirs(summary_dir, exist_ok=True)
        path = os.path.join(summary_dir, f"summary_events_{date_str}.xlsx")
        df = pd.DataFrame(trade_log)
        # Write all events to first sheet
        with ExcelWriter(path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name="All Events", index=False)
            # Build summary
            summary_rows = []
            total_pnl = 0
            total_brokerage = 0
            symbols = df['symbol'].unique()
            for sym in symbols:
                for side in ['BUY', 'SELL']:
                    sub = df[(df['symbol']==sym)&(df['side']==side)].sort_values('timestamp')
                    qty_open = 0
                    entry_price = 0
                    rows = []
                    for idx, row in sub.iterrows():
                        event = row['event_type']
                        if _is_entry_event_type(event):
                            qty_open += row['qty']
                            entry_price = row['price']
                        elif (event.startswith('T') or event.startswith('ST') or
                              event.endswith('SL') or
                              event in ['EOD_CLOSE', 'EOD_1511', 'EOD_2300'] or
                              event in ['BUY_RETREAT_25PCT', 'SELL_RETREAT_25PCT'] or
                              event.endswith('_MANUAL_TARGET')):
                            if entry_price:
                                # For targets/eod/sl
                                if side == 'BUY':
                                    pnl = (row['price'] - entry_price) * row['qty']
                                else:
                                    pnl = (entry_price - row['price']) * row['qty']
                                brokerage = BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
                                total_pnl += pnl
                                total_brokerage += brokerage
                                rows.append(dict(
                                    symbol=sym, side=side, qty=row['qty'], entry_price=entry_price, exit_price=row['price'],
                                    exit_type=event, gross_pnl=pnl, brokerage=brokerage,
                                    net_pnl=pnl - brokerage, timestamp=row['timestamp']
                                ))
                                qty_open -= row['qty']
                                if qty_open < 0: qty_open = 0
                                entry_price = 0 if qty_open == 0 else entry_price
                    # If still open position at EOD: close at last price
                    if qty_open > 0 and sym in last_prices:
                        eod_px = last_prices[sym]
                        if entry_price:
                            if side == 'BUY':
                                pnl = (eod_px - entry_price) * qty_open
                            else:
                                pnl = (entry_price - eod_px) * qty_open
                            brokerage = BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
                            total_pnl += pnl
                            total_brokerage += brokerage
                            rows.append(dict(
                                symbol=sym, side=side, qty=qty_open, entry_price=entry_price, exit_price=eod_px,
                                exit_type='EOD_CLOSE', gross_pnl=pnl, brokerage=brokerage,
                                net_pnl=pnl - brokerage, timestamp=now_ist()
                            ))
                    for r in rows:
                        summary_rows.append(r)
            summary_df = pd.DataFrame(summary_rows)
            if not summary_df.empty:
                totals = dict(
                    symbol='TOTAL', side='', qty=summary_df['qty'].sum(),
                    entry_price='', exit_price='', exit_type='',
                    gross_pnl=summary_df['gross_pnl'].sum(),
                    brokerage=summary_df['brokerage'].sum(),
                    net_pnl=summary_df['net_pnl'].sum(), timestamp=''
                )
                summary_df = pd.concat([summary_df, pd.DataFrame([totals])], ignore_index=True)
            summary_df.to_excel(writer, sheet_name="EOD Summary", index=False)
        logger.info(f"EOD trade log + summary written to {path}")
    except Exception as e:
        logger.warning(f"Failed to write EOD trade log/summary: {str(e)}")


def get_price_at_1511(symbol: str) -> Optional[float]:
    """Get price at 15:11 IST (3:11 PM) for market close calculations."""
    try:
        t = yf.Ticker(get_yf_symbol(symbol))
        df = t.history(period="1d", interval="1m")
        if df is None or df.empty:
            return None
        # Ensure tz-aware in IST
        if df.index.tz is None:
            df.index = df.index.tz_localize(pytz.UTC).tz_convert(IST)
        else:
            df.index = df.index.tz_convert(IST)
        
        # Get price at 15:11
        ts_1511 = now_ist().replace(hour=15, minute=11, second=0, microsecond=0)
        mask_1511 = (df.index >= ts_1511) & (df.index < ts_1511 + timedelta(minutes=1))
        df_1511 = df.loc[mask_1511]
        
        if df_1511.empty:
            # Fallback to latest available price before 15:11
            df_before = df.loc[df.index < ts_1511]
            if not df_before.empty:
                return float(df_before["Close"].iloc[-1])
            return None
        
        return float(df_1511["Close"].iloc[-1])
    except Exception as e:
        logger.debug(f"get_price_at_1511 error {symbol}: {e}")
        return None


def get_price_at_2300(symbol: str) -> Optional[float]:
    """Get price at 23:00 IST for commodity EOD calculations.
    
    FIX: MCX commodities are NOT on Yahoo Finance. Use in-memory TradingView WS
    cache (already populated from live streaming) instead of yfinance.
    For equity symbols (shouldn't normally be called), fall back to yfinance.
    """
    if is_commodity_symbol(symbol):
        # Priority 1: TradingView WS in-memory cache (most recent live price)
        cached = _get_from_cache(symbol)
        if cached and cached > 0:
            logger.debug("get_price_at_2300(%s) = %.2f from TV WS cache", symbol, cached)
            return cached
        # Priority 2: live multi-source fetch
        p = get_mcx_price_multi_source(symbol)
        if p and p > 0:
            return p
        return None
    # Equity fallback (legacy behaviour)
    try:
        t = yf.Ticker(get_yf_symbol(symbol))
        df = t.history(period="1d", interval="1m")
        if df is None or df.empty:
            return None
        if df.index.tz is None:
            df.index = df.index.tz_localize(pytz.UTC).tz_convert(IST)
        else:
            df.index = df.index.tz_convert(IST)

        ts_2300 = now_ist().replace(hour=23, minute=0, second=0, microsecond=0)
        mask = (df.index >= ts_2300) & (df.index < ts_2300 + timedelta(minutes=1))
        df_slice = df.loc[mask]
        if df_slice.empty:
            df_before = df.loc[df.index < ts_2300]
            if not df_before.empty:
                return float(df_before["Close"].iloc[-1])
            return None
        return float(df_slice["Close"].iloc[-1])
    except Exception as e:
        logger.debug(f"get_price_at_2300 error {symbol}: {e}")
        return None


def get_close_prices_1511(states: Dict[str, 'SymbolState']) -> Dict[str, float]:
    """Fetch 15:11 close prices for all symbols (cached per trading day)."""
    global CLOSE_PRICES_1511_CACHE, CLOSE_PRICES_1511_CACHE_DATE
    today_key = now_ist().strftime('%Y%m%d')
    if CLOSE_PRICES_1511_CACHE_DATE == today_key and CLOSE_PRICES_1511_CACHE:
        return CLOSE_PRICES_1511_CACHE

    close_prices: Dict[str, float] = {}
    for sym, st in states.items():
        px = get_price_at_1511(sym)
        if px is not None:
            close_prices[sym] = px
        elif st.last_price is not None:
            close_prices[sym] = st.last_price
        else:
            close_prices[sym] = st.levels.previous_close if st.levels else 0.0

    CLOSE_PRICES_1511_CACHE = close_prices
    CLOSE_PRICES_1511_CACHE_DATE = today_key
    return close_prices.copy()


def get_close_prices_2300(states: Dict[str, 'SymbolState']) -> Dict[str, float]:
    """Fetch 23:00 close prices for the provided symbols (no caching needed)."""
    close_prices: Dict[str, float] = {}
    for sym, st in states.items():
        px = get_price_at_2300(sym)
        if px is None and st.last_price is not None:
            px = st.last_price
        if px is None and st.levels:
            px = st.levels.previous_close
        close_prices[sym] = px or 0.0
    return close_prices

def create_trade_analysis_excel(
    trade_log: List[dict],
    states: Dict[str, 'SymbolState'],
    date_str: Optional[str] = None,
    *,
    close_prices_1511: Optional[Dict[str, float]] = None,
    return_summary: bool = False,
    analysis_dir: str = "trade_analysis",
    file_stem: str = "trade_analysis",
    allow_disk_snapshot: bool = True,
):
    """Create comprehensive trade analysis Excel at 15:12 with 15:11 prices and analysis for algo improvement."""
    try:
        if date_str is None:
            date_str = now_ist().strftime('%Y%m%d')
        
        # Get 15:11 prices for all symbols
        close_prices_1511 = close_prices_1511 or get_close_prices_1511(states)
        
        os.makedirs(analysis_dir, exist_ok=True)
        file_stem = file_stem or "trade_analysis"
        path = os.path.join(analysis_dir, f"{file_stem}_{date_str}.xlsx")
        
        # Strictly filter trades by date - only use trades from the specified trading date
        # This prevents old trades from being included when bot restarts post-market hours
        trade_data = trade_log or []
        # First filter: only trades with matching trade_date
        trade_data = [ev for ev in trade_data if ev.get('trade_date') == date_str]
        
        # Second filter: also check timestamp to ensure it's from the correct date
        # This prevents old trades without proper trade_date from being included
        try:
            date_obj = datetime.strptime(date_str, '%Y%m%d').date()
            filtered_data = []
            for ev in trade_data:
                # Check if timestamp exists and matches the date
                if 'timestamp' in ev:
                    try:
                        ts = pd.to_datetime(ev['timestamp'])
                        if hasattr(ts, 'date'):
                            ts_date = ts.date()
                        else:
                            ts_date = ts.to_pydatetime().date()
                        if ts_date == date_obj:
                            filtered_data.append(ev)
                    except:
                        # If timestamp parsing fails, only include if trade_date matches
                        if ev.get('trade_date') == date_str:
                            filtered_data.append(ev)
                else:
                    # No timestamp, rely on trade_date only
                    if ev.get('trade_date') == date_str:
                        filtered_data.append(ev)
            trade_data = filtered_data
        except Exception as e:
            logger.debug(f"Error filtering trades by timestamp: {e}")
            # Fallback to trade_date only
            trade_data = [ev for ev in trade_data if ev.get('trade_date') == date_str]
        
        if not trade_data and allow_disk_snapshot:
            trade_data = get_trade_log_snapshot(date_str)
            # Apply same strict filtering to snapshot
            if trade_data:
                try:
                    date_obj = datetime.strptime(date_str, '%Y%m%d').date()
                    filtered_data = []
                    for ev in trade_data:
                        if 'timestamp' in ev:
                            try:
                                ts = pd.to_datetime(ev['timestamp'])
                                if hasattr(ts, 'date'):
                                    ts_date = ts.date()
                                else:
                                    ts_date = ts.to_pydatetime().date()
                                if ts_date == date_obj:
                                    filtered_data.append(ev)
                            except:
                                if ev.get('trade_date') == date_str:
                                    filtered_data.append(ev)
                        else:
                            if ev.get('trade_date') == date_str:
                                filtered_data.append(ev)
                    trade_data = filtered_data
                except Exception as e:
                    logger.debug(f"Error filtering snapshot trades by timestamp: {e}")
        
        # If still no trades, return empty dataframe instead of None to allow EOD summary to be sent.
        # Also ALWAYS (re)create the Excel file (empty) so we don't accidentally send a stale file
        # left over from a previous run/day.
        if not trade_data:
            logger.info("No trades to analyze for date %s. Will send empty EOD summary.", date_str)
            # Return empty dataframe with required columns so send_eod_summary can handle it
            empty_df = pd.DataFrame(columns=['symbol', 'side', 'qty', 'entry_price', 'exit_price', 
                                            'exit_type', 'gross_pnl', 'brokerage', 'net_pnl'])
            # Always create/overwrite the Excel file with empty data
            os.makedirs(analysis_dir, exist_ok=True)
            path = os.path.join(analysis_dir, f"{file_stem}_{date_str}.xlsx")
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                empty_df.to_excel(writer, sheet_name="All Events", index=False)
                empty_df.to_excel(writer, sheet_name="Trade Summary", index=False)
            logger.info(f"Trade analysis Excel created at {path} (empty)")
            return empty_df if return_summary else None
        
        df = pd.DataFrame(trade_data)
        if states:
            df = df[df['symbol'].isin(states.keys())]
        if df.empty:
            logger.info("No trades to analyze for selected symbols.")
            return None
        
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            # Sheet 1: All Trade Events
            df.to_excel(writer, sheet_name="All Events", index=False)
            
            # Sheet 2: Trade Summary with P&L
            summary_rows = []
            for sym in df['symbol'].unique():
                for side in ['BUY', 'SELL']:
                    sub = df[(df['symbol']==sym)&(df['side']==side)].sort_values('timestamp')
                    qty_open = 0
                    entry_price = 0
                    entry_time = None
                    rows = []
                    for idx, row in sub.iterrows():
                        event = row['event_type']
                        if _is_entry_event_type(event):
                            qty_open += row['qty']
                            entry_price = row['price']
                            entry_time = row['timestamp']
                        elif (event in ['T1', 'T2', 'T3', 'T4', 'T5'] or
                              event in ['ST1', 'ST2', 'ST3', 'ST4', 'ST5'] or
                              event in ['BUY_SL', 'SELL_SL'] or
                              event in ['EOD_CLOSE', 'EOD_1511', 'EOD_2300'] or
                              event in ['BUY_RETREAT_25PCT', 'SELL_RETREAT_25PCT'] or
                              event.endswith('_MANUAL_TARGET')):
                            if entry_price and qty_open > 0:
                                exit_qty = min(row['qty'], qty_open)
                                if side == 'BUY':
                                    pnl = (row['price'] - entry_price) * exit_qty
                                else:
                                    pnl = (entry_price - row['price']) * exit_qty
                                brokerage = BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
                                net_pnl = pnl - brokerage
                                hold_duration = None
                                if entry_time:
                                    try:
                                        entry_dt = pd.to_datetime(entry_time)
                                        exit_dt = pd.to_datetime(row['timestamp'])
                                        hold_duration = (exit_dt - entry_dt).total_seconds() / 60  # minutes
                                    except:
                                        pass
                                rows.append(dict(
                                    symbol=sym, side=side, qty=exit_qty, entry_price=entry_price,
                                    exit_price=row['price'], exit_type=event, entry_time=entry_time,
                                    exit_time=row['timestamp'], hold_duration_min=hold_duration,
                                    gross_pnl=pnl, brokerage=brokerage, net_pnl=net_pnl
                                ))
                                qty_open -= exit_qty
                                if qty_open <= 0:
                                    entry_price = 0
                                    entry_time = None
                    
                    # Close any remaining open positions at 15:11 price.
                    # IMPORTANT: skip if eod_square_off already recorded EOD_CLOSE for
                    # this symbol/side (prevents double-counting in the summary).
                    already_eod = any(
                        r.get('exit_type') in ('EOD_CLOSE', 'EOD_1511', 'EOD_2300')
                        for r in rows
                    )
                    if qty_open > 0 and sym in close_prices_1511 and not already_eod:
                        eod_px = close_prices_1511[sym]
                        if entry_price:
                            if side == 'BUY':
                                pnl = (eod_px - entry_price) * qty_open
                            else:
                                pnl = (entry_price - eod_px) * qty_open
                            brokerage = BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
                            net_pnl = pnl - brokerage
                            hold_duration = None
                            if entry_time:
                                try:
                                    entry_dt = pd.to_datetime(entry_time)
                                    exit_dt = now_ist().replace(hour=15, minute=11, second=0, microsecond=0)
                                    hold_duration = (exit_dt - entry_dt).total_seconds() / 60
                                except:
                                    pass
                            rows.append(dict(
                                symbol=sym, side=side, qty=qty_open, entry_price=entry_price,
                                exit_price=eod_px, exit_type='EOD_1511', entry_time=entry_time,
                                exit_time=now_ist().replace(hour=15, minute=11, second=0, microsecond=0).isoformat(),
                                hold_duration_min=hold_duration, gross_pnl=pnl, brokerage=brokerage, net_pnl=net_pnl
                            ))
                    summary_rows.extend(rows)
            
            summary_df = pd.DataFrame(summary_rows)
            if not summary_df.empty:
                totals = dict(
                    symbol='TOTAL', side='', qty=summary_df['qty'].sum(), entry_price='', exit_price='',
                    exit_type='', entry_time='', exit_time='', hold_duration_min='',
                    gross_pnl=summary_df['gross_pnl'].sum(), brokerage=summary_df['brokerage'].sum(),
                    net_pnl=summary_df['net_pnl'].sum()
                )
                summary_df = pd.concat([summary_df, pd.DataFrame([totals])], ignore_index=True)
            summary_df.to_excel(writer, sheet_name="Trade Summary", index=False)
            
            # Sheet 3: Performance Analysis
            if not summary_df.empty and len(summary_df) > 1:  # More than just totals row
                analysis_data = []
                
                # Overall metrics
                total_trades = len(summary_df) - 1  # Exclude totals row
                winning_trades = len(summary_df[summary_df['net_pnl'] > 0]) - (1 if summary_df.iloc[-1]['symbol'] == 'TOTAL' and summary_df.iloc[-1]['net_pnl'] > 0 else 0)
                losing_trades = len(summary_df[summary_df['net_pnl'] < 0]) - (1 if summary_df.iloc[-1]['symbol'] == 'TOTAL' and summary_df.iloc[-1]['net_pnl'] < 0 else 0)
                win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
                
                analysis_data.append({"Metric": "Total Trades", "Value": total_trades})
                analysis_data.append({"Metric": "Winning Trades", "Value": winning_trades})
                analysis_data.append({"Metric": "Losing Trades", "Value": losing_trades})
                analysis_data.append({"Metric": "Win Rate (%)", "Value": f"{win_rate:.2f}"})
                
                valid_summary = summary_df[summary_df['symbol'] != 'TOTAL'] if 'TOTAL' in summary_df['symbol'].values else summary_df
                if not valid_summary.empty:
                    analysis_data.append({"Metric": "Total Gross P&L", "Value": f"₹{valid_summary['gross_pnl'].sum():.2f}"})
                    analysis_data.append({"Metric": "Total Brokerage", "Value": f"₹{valid_summary['brokerage'].sum():.2f}"})
                    analysis_data.append({"Metric": "Total Net P&L", "Value": f"₹{valid_summary['net_pnl'].sum():.2f}"})
                    analysis_data.append({"Metric": "Average Hold Duration (min)", "Value": f"{valid_summary['hold_duration_min'].mean():.2f}" if 'hold_duration_min' in valid_summary.columns else "N/A"})
                    analysis_data.append({"Metric": "Best Trade", "Value": f"₹{valid_summary['net_pnl'].max():.2f}"})
                    analysis_data.append({"Metric": "Worst Trade", "Value": f"₹{valid_summary['net_pnl'].min():.2f}"})
                
                # Per-symbol analysis
                if 'symbol' in valid_summary.columns:
                    symbol_pnl = valid_summary.groupby('symbol')['net_pnl'].agg(['sum', 'count', 'mean']).reset_index()
                    symbol_pnl.columns = ['Symbol', 'Total Net P&L', 'Trade Count', 'Avg P&L per Trade']
                    symbol_pnl = symbol_pnl.sort_values('Total Net P&L', ascending=False)
                    
                    # Per-side analysis
                    if 'side' in valid_summary.columns:
                        side_pnl = valid_summary.groupby('side')['net_pnl'].agg(['sum', 'count']).reset_index()
                        side_pnl.columns = ['Side', 'Total Net P&L', 'Trade Count']
                        
                        # Exit type analysis
                        if 'exit_type' in valid_summary.columns:
                            exit_pnl = valid_summary.groupby('exit_type')['net_pnl'].agg(['sum', 'count']).reset_index()
                            exit_pnl.columns = ['Exit Type', 'Total Net P&L', 'Count']
                            exit_pnl = exit_pnl.sort_values('Total Net P&L', ascending=False)
                
                analysis_df = pd.DataFrame(analysis_data)
                analysis_df.to_excel(writer, sheet_name="Performance Metrics", index=False)
                
                if 'symbol' in valid_summary.columns:
                    symbol_pnl.to_excel(writer, sheet_name="Symbol Performance", index=False)
                
                if 'side' in valid_summary.columns:
                    side_pnl.to_excel(writer, sheet_name="Side Performance", index=False)
                
                if 'exit_type' in valid_summary.columns:
                    exit_pnl.to_excel(writer, sheet_name="Exit Type Analysis", index=False)
        
        logger.info(f"Trade analysis Excel created at {path}")
        return summary_df if return_summary else None
    except Exception as e:
        logger.warning(f"Failed to create trade analysis Excel: {str(e)}")
        import traceback
        logger.debug(traceback.format_exc())


def create_commodity_trade_analysis_excel(
    trade_log: List[dict],
    states: Dict[str, 'SymbolState'],
    date_str: Optional[str] = None,
) -> Optional[str]:
    """Create commodity-only trade analysis using 23:00 IST prices and return the file path."""
    commodity_states = {sym: st for sym, st in states.items() if st.is_commodity}
    if not commodity_states:
        logger.info("No commodity symbols configured for analysis.")
        return None
    if date_str is None:
        date_str = now_ist().strftime('%Y%m%d')
    close_prices_2300 = get_close_prices_2300(commodity_states)
    analysis_dir = os.path.join("trade_analysis", "commodities")
    file_stem = "commodity_trade_analysis"
    create_trade_analysis_excel(
        trade_log,
        commodity_states,
        date_str,
        close_prices_1511=close_prices_2300,
        analysis_dir=analysis_dir,
        file_stem=file_stem,
    )
    return os.path.join(analysis_dir, f"{file_stem}_{date_str}.xlsx")


# ----------------------------
# Configuration
# ----------------------------

IST = pytz.timezone("Asia/Kolkata")
UPDATE_INTERVALS_SECONDS = [1]   # fixed 1-second update cycle for consistent latency
BROKERAGE_RATE = 0.0  # deprecated percentage model (kept for compatibility)
BROKERAGE_FLAT_PER_SIDE = 10.0  # ₹10 per execution (entry or exit)
X_FACTOR_MULTIPLIER = 0.008575
COMMODITY_DASH_PORT = int(os.getenv("COMMODITY_DASH_PORT", "8051"))
MAX_VISIBLE_CHARTS = 6  # limit number of charts rendered at once in Dash for performance

# Controls whether create_trade_analysis_excel is allowed to reload trades from disk
# when TRADE_LOG is empty. We disable this when the program is started post‑market
# so that EOD summaries don't show historical trades for runs that never saw live ticks.
ALLOW_DISK_TRADE_SNAPSHOT = True

# Telegram
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "7587307352:AAG6RaiF4gO5I_ZFZ_4b8Gj7dnsu4GtPWFw")
# Default your chat ID so you don't need to set env each run
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "1376513391")
# Add support for sending alerts to multiple Telegram chat IDs (including Ajay)
TELEGRAM_EXTRA_CHAT_IDS = ["793674804"]
TELEGRAM_CHAT_IDS_ALL = [
    cid for cid in [TELEGRAM_CHAT_ID, *TELEGRAM_EXTRA_CHAT_IDS] if cid
]
TELEGRAM_COMMODITY_BOT_TOKEN = os.getenv(
    "TELEGRAM_COMMODITY_BOT_TOKEN", "8340570160:AAHGq9U3i8HlD2-rmXWeY94IjJiC6NkHqv8"
)
_env_commodity_ids = os.getenv("TELEGRAM_COMMODITY_CHAT_IDS")
if _env_commodity_ids:
    TELEGRAM_COMMODITY_CHAT_IDS = [
        cid.strip() for cid in _env_commodity_ids.split(",") if cid.strip()
    ]
else:
    TELEGRAM_COMMODITY_CHAT_IDS = TELEGRAM_CHAT_IDS_ALL[:]


# ----------------------------
# Logging
# ----------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("live_levels")


# ----------------------------
# Stock list (from prior program)
# ----------------------------

STOCKS: List[str] = [
    "NIFTY",
    "BANKNIFTY",
    "HDFCBANK", "KOTAKBANK", "SBIN", "ICICIBANK", "INDUSINDBK",
    "ADANIPORTS", "ADANIENT", "ASIANPAINT", "BAJFINANCE", "DRREDDY",
    "SUNPHARMA", "INFY", "TCS", "TECHM",
    "TITAN", "TATAMOTORS", "RELIANCE", "INDIGO", "JUBLFOOD",
    "BATAINDIA", "PIDILITIND", "ZEEL", "BALKRISIND", "VOLTAS",
    "ITC", "BPCL", "BRITANNIA", "HEROMOTOCO",
    "HINDUNILVR", "UPL", "SRF", "TATACONSUM", "BALRAMCHIN",
    "ABFRL", "VEDL", "COFORGE",
]

# Commodity tracking is optional. Default OFF to avoid spam alerts.
# Enable with env: ENABLE_COMMODITIES=1
ENABLE_COMMODITIES = os.getenv("ENABLE_COMMODITIES", "1").strip() == "1"  # v8.0: enabled by default

tv_datafeed = None
TVDATAFEED_AVAILABLE = False
COMMODITY_CONFIG: Dict[str, Dict[str, Union[str, float]]] = {}

# ══════════════════════════════════════════════════════════════════════════════
# COMMODITY PRICE ENGINE — multi-source with automatic failover
#
#  Source 1 (PRIMARY)  TradingView WebSocket  — near real-time MCX INR prices
#  Source 2 (BACKUP)   Investing.com scrape   — HTTP, slight delay, INR
#  Source 3 (FALLBACK) MoneyControl scrape    — HTTP, may lag 5–15 min
#  Cache               In-memory dict         — shields against transient failures
#
#  Pipeline per fetch:
#    try TradingView WS → hit → return & cache
#    except → try Investing.com → hit → return & cache
#    except → try MoneyControl  → hit → return & cache
#    except → return cached value (stale but better than None)
#
#  All sources return INR prices as traded on MCX.
#  No account, no API key, no rate limits from our side.
# ══════════════════════════════════════════════════════════════════════════════

# ── Dependency availability flags ────────────────────────────────────────────
_WS_AVAILABLE = False
_BS4_AVAILABLE = False
try:
    import websocket as _websocket_lib   # pip install websocket-client
    _WS_AVAILABLE = True
except ImportError:
    pass

try:
    from bs4 import BeautifulSoup as _BS4
    _BS4_AVAILABLE = True
except ImportError:
    pass

# ── In-memory commodity price cache ──────────────────────────────────────────
# {symbol: {"price": float, "ts": datetime, "source": str}}
_COMMODITY_PRICE_CACHE: Dict[str, dict] = {}
_COMMODITY_CACHE_LOCK  = threading.Lock()
_COMMODITY_CACHE_TTL_S = 10   # v8.0: 10s TTL (was 30s) for fresher commodity prices

# ── TradingView WebSocket state ───────────────────────────────────────────────
_TV_WS_PRICES: Dict[str, float] = {}   # live prices from WS thread
_TV_WS_LOCK   = threading.Lock()
_TV_WS_THREAD: Optional[threading.Thread] = None
_TV_WS_RUNNING = threading.Event()

# TradingView WebSocket symbol → our symbol mapping
_TV_MCX_SYMBOLS = {
    "MCX:GOLD1!":        "GOLD",
    "MCX:SILVER1!":      "SILVER",
    "MCX:CRUDEOIL1!":    "CRUDE",
    "MCX:NATURALGAS1!":  "NATURALGAS",
    "MCX:COPPER1!":      "COPPER",
    "MCX:ZINC1!":        "ZINC",
    "MCX:NICKEL1!":      "NICKEL",
    "MCX:ALUMINIUM1!":   "ALUMINIUM",
}

# Investing.com commodity page → our symbol mapping
_INVESTING_URLS = {
    "GOLD":        "https://www.investing.com/commodities/gold",
    "SILVER":      "https://www.investing.com/commodities/silver",
    "CRUDE":       "https://www.investing.com/commodities/crude-oil",
    "NATURALGAS":  "https://www.investing.com/commodities/natural-gas",
    "COPPER":      "https://www.investing.com/commodities/copper",
    "ZINC":        "https://api.investing.com/api/financialdata/87/historical/chart/?period=P1W&interval=PT1M&pointscount=60",
    "NICKEL":      "https://www.investing.com/commodities/nickel",
    "ALUMINIUM":   "https://www.investing.com/commodities/aluminum",
}

# MoneyControl commodity page → our symbol mapping
_MONEYCONTROL_URLS = {
    "GOLD":       "https://www.moneycontrol.com/commodity/gold-price.html",
    "SILVER":     "https://www.moneycontrol.com/commodity/silver-price.html",
    "CRUDE":      "https://www.moneycontrol.com/commodity/crude-oil-price.html",
    "NATURALGAS": "https://www.moneycontrol.com/commodity/natural-gas-price.html",
}


def _tv_ws_connect_and_stream() -> None:
    """
    TradingView WebSocket price streamer — runs in a daemon thread.

    Connects to TradingView's public data WebSocket, sends auth/subscribe
    packets for all MCX commodity symbols, and populates _TV_WS_PRICES
    with each incoming price update.

    Reconnects automatically on disconnect (max 5 attempts, then gives up
    and lets the fallback sources handle pricing).
    """
    import string, random, json as _json

    def _rand_token(n=12) -> str:
        return "".join(random.choices(string.ascii_lowercase + string.digits, k=n))

    WS_URL  = "wss://data.tradingview.com/socket.io/websocket"
    ORIGIN  = "https://www.tradingview.com"
    HEADERS = {"Origin": ORIGIN}

    def _pack(func: str, args: list) -> str:
        """Wrap a message in TradingView's ~m~N~m~{...} framing."""
        body = _json.dumps({"m": func, "p": args}, separators=(",", ":"))
        return f"~m~{len(body)}~m~{body}"

    def _on_message(ws, raw: str) -> None:
        # TradingView sends heartbeat ~h~N — must echo back
        if "~h~" in raw:
            ws.send(f"~m~{len(raw)}~m~{raw}")
            return
        # Extract JSON payloads from the framing
        import re as _re
        for chunk in _re.findall(r"~m~\d+~m~(\{.*?\})(?=~m~|\Z)", raw):
            try:
                msg = _json.loads(chunk)
            except Exception:
                continue
            if msg.get("m") == "du":
                # Price update: p[1] is a dict of series data
                p = msg.get("p", [])
                if len(p) < 2:
                    continue
                sname = p[0]  # series name like "s1", "s2" …
                data  = p[1]
                if not isinstance(data, dict):
                    continue
                # Map series name back to symbol via _sname_to_symbol (below)
                sym = _sname_to_symbol.get(sname)
                if sym is None:
                    continue
                st = data.get("st", [])
                if st and isinstance(st, list):
                    # Last bar's close price is index 4: [ts, open, high, low, close, vol]
                    bar = st[-1].get("v", [])
                    if len(bar) >= 5:
                        price = float(bar[4])
                        if price > 0:
                            with _TV_WS_LOCK:
                                _TV_WS_PRICES[sym] = price
                            _update_commodity_cache(sym, price, "tradingview_ws")

    def _on_error(ws, err) -> None:
        logger.debug("TradingView WS error: %s", err)

    def _on_close(ws, *a) -> None:
        logger.debug("TradingView WS closed")

    _sname_to_symbol: Dict[str, str] = {}

    for attempt in range(5):
        if not _TV_WS_RUNNING.is_set():
            break
        try:
            import websocket as _ws_lib
            ws = _ws_lib.WebSocketApp(
                WS_URL,
                header=HEADERS,
                on_message=_on_message,
                on_error=_on_error,
                on_close=_on_close,
            )

            # Connection established — send auth + subscribe
            def _on_open(ws):
                sess  = "qs_" + _rand_token()
                chart = "cs_" + _rand_token()
                # Auth with public token (no account needed)
                ws.send(_pack("set_auth_token", ["unauthorized_user_token"]))
                ws.send(_pack("chart_create_session", [chart, ""]))
                ws.send(_pack("quote_create_session", [sess]))
                ws.send(_pack("quote_set_fields", [sess,
                    "ch", "chp", "current_session", "description",
                    "local_description", "language", "exchange",
                    "fractional", "is_tradable", "lp", "lp_time",
                    "minmov", "minmov2", "original_name", "pricescale",
                    "pro_name", "short_name", "type", "update_mode",
                    "volume", "currency_code", "rchp", "rtc"]))
                for i, (tv_sym, our_sym) in enumerate(_TV_MCX_SYMBOLS.items()):
                    sname = f"s{i+1}"
                    _sname_to_symbol[sname] = our_sym
                    ws.send(_pack("quote_add_symbols", [sess, tv_sym, {"flags": ["force_permission"]}]))
                    ws.send(_pack("resolve_symbol", [chart, sname,
                        f'={{"symbol":"{tv_sym}","adjustment":"splits"}}']))
                    ws.send(_pack("create_series", [chart, sname, sname, sname, "1", 1, ""]))
                logger.info("TradingView WS connected — streaming %d MCX symbols", len(_TV_MCX_SYMBOLS))

            ws.on_open = _on_open
            ws.run_forever(ping_interval=20, ping_timeout=10)
        except Exception as e:
            logger.debug("TradingView WS attempt %d failed: %s", attempt + 1, e)
        if _TV_WS_RUNNING.is_set():
            time.sleep(min(5 * (attempt + 1), 30))

    logger.info("TradingView WS streamer exited after %d attempts", attempt + 1)


def _start_tv_ws_if_needed() -> None:
    """Start the TradingView WebSocket streamer thread (once, lazily)."""
    global _TV_WS_THREAD
    if not _WS_AVAILABLE:
        return
    if _TV_WS_THREAD and _TV_WS_THREAD.is_alive():
        return
    _TV_WS_RUNNING.set()
    _TV_WS_THREAD = threading.Thread(
        target=_tv_ws_connect_and_stream,
        daemon=True,
        name="TVWSCommodity",
    )
    _TV_WS_THREAD.start()
    logger.info("TradingView WebSocket commodity streamer started")


def _update_commodity_cache(symbol: str, price: float, source: str) -> None:
    with _COMMODITY_CACHE_LOCK:
        _COMMODITY_PRICE_CACHE[symbol] = {
            "price": price,
            "ts": now_ist(),
            "source": source,
        }


def _get_from_cache(symbol: str) -> Optional[float]:
    with _COMMODITY_CACHE_LOCK:
        entry = _COMMODITY_PRICE_CACHE.get(symbol)
    if entry is None:
        return None
    age = (now_ist() - entry["ts"]).total_seconds()
    if age <= _COMMODITY_CACHE_TTL_S * 4:   # allow 4× TTL before dropping stale data
        return float(entry["price"])
    return None


def _fetch_investing_price(symbol: str) -> Optional[float]:
    """Fetch MCX commodity price from Investing.com (backup source)."""
    if not _BS4_AVAILABLE:
        return None
    url = _INVESTING_URLS.get(symbol.upper())
    if not url:
        return None
    try:
        import requests as _req
        r = _req.get(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml",
        }, timeout=8)
        if r.status_code != 200:
            return None
        soup = _BS4(r.text, "html.parser")
        # Try multiple selectors in priority order
        for selector in [
            {"data-test": "instrument-price-last"},
            {"class": "instrument-price_last__KQzyA"},
            {"id": "last_last"},
        ]:
            tag = soup.find(attrs=selector)
            if tag:
                raw = tag.get_text(strip=True).replace(",", "").replace("₹", "").strip()
                try:
                    p = float(raw)
                    if p > 0:
                        return p
                except ValueError:
                    pass
    except Exception as e:
        logger.debug("Investing.com fetch error %s: %s", symbol, e)
    return None


def _fetch_moneycontrol_price(symbol: str) -> Optional[float]:
    """Fetch MCX commodity price from MoneyControl (fallback source)."""
    if not _BS4_AVAILABLE:
        return None
    url = _MONEYCONTROL_URLS.get(symbol.upper())
    if not url:
        return None
    try:
        import requests as _req
        r = _req.get(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }, timeout=10)
        if r.status_code != 200:
            return None
        soup = _BS4(r.text, "html.parser")
        # MoneyControl price selectors
        for selector in [
            {"id": "sp_price"},
            {"class": "commodity_price"},
            {"class": "inprice1"},
        ]:
            tag = soup.find(attrs=selector)
            if tag:
                raw = tag.get_text(strip=True).replace(",", "").replace("₹", "").strip()
                try:
                    p = float(raw)
                    if p > 0:
                        return p
                except ValueError:
                    pass
    except Exception as e:
        logger.debug("MoneyControl fetch error %s: %s", symbol, e)
    return None


# Investing.com public instrument IDs for MCX commodities (INR)
_INVESTING_COM_IDS = {
    "GOLD":       "68",     # MCX Gold Futures
    "SILVER":     "69",     # MCX Silver Futures
    "CRUDE":      "8869",   # MCX Crude Oil Futures
    "NATURALGAS": "8867",   # MCX Natural Gas Futures
    "COPPER":     "8863",   # MCX Copper Futures
}

# Goodreturns MCX live price URLs (no JS rendering needed, fast HTML)
_GOODRETURNS_URLS = {
    "GOLD":       "https://www.goodreturns.in/gold-rates/",
    "SILVER":     "https://www.goodreturns.in/silver-price/",
    "CRUDE":      "https://www.goodreturns.in/crude-oil-price/",
    "NATURALGAS": "https://www.goodreturns.in/natural-gas-price/",
    "COPPER":     "https://www.goodreturns.in/copper-price/",
}

# ICICIdirect commodity live page (no login, ~5s refresh, reliable INR MCX)
_ICICI_MCX_SYMBOLS = {
    "GOLD":       "GOLDPETAL",
    "SILVER":     "SILVER",
    "CRUDE":      "CRUDEOIL",
    "NATURALGAS": "NATURALGAS",
    "COPPER":     "COPPER",
}


def _fetch_investing_com_api(symbol: str) -> Optional[float]:
    """
    Fetch MCX commodity price via Investing.com's public JSON API.
    Returns INR price. No login, no API key. Typical latency: 1-2s.
    This is faster and more reliable than HTML scraping.
    """
    instr_id = _INVESTING_COM_IDS.get(symbol.upper())
    if not instr_id:
        return None
    try:
        import requests as _req
        # Use the public summary API — no auth needed
        url = f"https://api.investing.com/api/financialdata/{instr_id}/historical/chart/?period=P1D&interval=PT1M&pointscount=1"
        headers = {
            "User-Agent":    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "X-Requested-With": "XMLHttpRequest",
            "Referer":       "https://www.investing.com/",
            "Accept":        "application/json",
        }
        r = _req.get(url, headers=headers, timeout=5)
        if r.status_code == 200:
            data = r.json()
            # Response: {"data": [[ts, open, high, low, close, vol], ...]}
            pts = data.get("data", {}).get("candles", data.get("data", []))
            if pts:
                last = pts[-1]
                if isinstance(last, (list, tuple)) and len(last) >= 5:
                    p = float(last[4])  # close
                    if p > 0:
                        return p
        # Fallback: try summary endpoint
        url2 = f"https://api.investing.com/api/financialdata/{instr_id}/?fields=last,bid,ask&time_utc_offset=330"
        r2 = _req.get(url2, headers=headers, timeout=5)
        if r2.status_code == 200:
            d2 = r2.json()
            last_price = d2.get("data", {}).get("last", d2.get("last"))
            if last_price:
                p = float(str(last_price).replace(",", ""))
                if p > 0:
                    return p
    except Exception as e:
        logger.debug("Investing.com API error %s: %s", symbol, e)
    return None


def _fetch_goodreturns_price(symbol: str) -> Optional[float]:
    """
    Fetch MCX commodity price from Goodreturns.in.
    India-focused site, prices in INR, updates every 1-2 minutes.
    """
    if not _BS4_AVAILABLE:
        return None
    url = _GOODRETURNS_URLS.get(symbol.upper())
    if not url:
        return None
    try:
        import requests as _req
        r = _req.get(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }, timeout=8)
        if r.status_code != 200:
            return None
        soup = _BS4(r.text, "html.parser")
        # Goodreturns uses these class patterns for live MCX prices
        for cls in ["price-gold-1g", "current-price", "rate", "commodity_price_tag"]:
            tag = soup.find(class_=cls)
            if tag:
                raw = tag.get_text(strip=True).replace(",","").replace("₹","").replace("Rs","").strip()
                try:
                    p = float(raw)
                    if p > 0:
                        return p
                except ValueError:
                    pass
        # Try finding any large number in a price context
        import re as _re
        price_spans = soup.find_all("span", string=_re.compile(r"^[0-9]{2,7}(\.[0-9]{1,2})?$"))
        for span in price_spans[:3]:
            try:
                p = float(span.get_text().replace(",",""))
                if 100 < p < 10_000_000:  # sanity range for MCX prices in INR
                    return p
            except ValueError:
                pass
    except Exception as e:
        logger.debug("Goodreturns fetch error %s: %s", symbol, e)
    return None


def get_mcx_price_multi_source(symbol: str) -> Optional[float]:
    """
    Fetch MCX commodity price using the full multi-source pipeline:

      1. TradingView WebSocket (in-process stream, near real-time, primary)
      2. tvdatafeed library (if installed, REST-like, primary fallback)
      3. Investing.com scrape (HTTP, backup)
      4. MoneyControl scrape (HTTP, last-resort fallback)
      5. In-memory cache (stale value, shields against total outage)

    All sources return INR prices as traded on MCX.
    The source used for each price is logged in the cache for diagnostics.
    """
    sym = symbol.upper()

    # ── 1. TradingView WebSocket (best: near real-time INR) ─────────────────
    if _WS_AVAILABLE:
        _start_tv_ws_if_needed()
        with _TV_WS_LOCK:
            ws_price = _TV_WS_PRICES.get(sym)
        if ws_price and ws_price > 0:
            _update_commodity_cache(sym, ws_price, "tradingview_ws")
            return ws_price

    # ── 2. tvdatafeed (REST, if installed) ───────────────────────────────────
    if TVDATAFEED_AVAILABLE and tv_datafeed:
        try:
            cfg = get_symbol_config(sym)
            mcx_raw = cfg.get("mcx_symbol") if cfg else None
            if mcx_raw:
                df = tv_datafeed.get_hist(
                    symbol=mcx_raw, exchange="MCX",
                    interval=TVInterval.in_1_minute, n_bars=1,
                )
                if df is not None and not df.empty:
                    p = float(df["close"].iloc[-1])
                    if p > 0:
                        _update_commodity_cache(sym, p, "tvdatafeed")
                        return p
        except Exception as e:
            logger.debug("tvdatafeed fallback error %s: %s", sym, e)

    # ── 3. Investing.com JSON API (fast, reliable INR, no HTML parsing) ─────
    try:
        p = _fetch_investing_com_api(sym)
        if p and p > 0:
            _update_commodity_cache(sym, p, "investing_api")
            logger.debug("MCX %s price %.2f from Investing.com API", sym, p)
            return p
    except Exception as e:
        logger.debug("Investing.com API error %s: %s", sym, e)

    # ── 3b. Investing.com scrape (HTML fallback) ──────────────────────────────
    try:
        p = _fetch_investing_price(sym)
        if p and p > 0:
            _update_commodity_cache(sym, p, "investing_scrape")
            logger.debug("MCX %s price %.2f from Investing.com scrape", sym, p)
            return p
    except Exception as e:
        logger.debug("Investing.com scrape error %s: %s", sym, e)

    # ── 4. Goodreturns.in (India-focused, INR, reliable) ─────────────────────
    try:
        p = _fetch_goodreturns_price(sym)
        if p and p > 0:
            _update_commodity_cache(sym, p, "goodreturns")
            logger.debug("MCX %s price %.2f from Goodreturns", sym, p)
            return p
    except Exception as e:
        logger.debug("Goodreturns error %s: %s", sym, e)

    # ── 5. MoneyControl (scrape, last resort) ────────────────────────────────
    try:
        p = _fetch_moneycontrol_price(sym)
        if p and p > 0:
            _update_commodity_cache(sym, p, "moneycontrol")
            logger.debug("MCX %s price %.2f from MoneyControl", sym, p)
            return p
    except Exception as e:
        logger.debug("MoneyControl pipeline error %s: %s", sym, e)

    # ── 5. Cache (stale but non-None — shields against total outage) ─────────
    cached = _get_from_cache(sym)
    if cached:
        logger.debug("MCX %s using cached price %.2f (all live sources failed)", sym, cached)
        return cached

    return None


def get_commodity_price_source_status() -> Dict[str, str]:
    """Return dict of symbol → last source used, for dashboard/logging."""
    with _COMMODITY_CACHE_LOCK:
        return {sym: entry["source"] for sym, entry in _COMMODITY_PRICE_CACHE.items()}


# ── Commodity price alert configuration ──────────────────────────────────────
# Alert when commodity price moves >X% from previous close in a single session
COMMODITY_ALERT_THRESHOLDS = {
    "GOLD":       0.005,    # 0.5% move = alert
    "SILVER":     0.007,    # 0.7%
    "CRUDE":      0.010,    # 1.0% (crude is more volatile)
    "NATURALGAS": 0.015,    # 1.5% (very volatile)
    "COPPER":     0.008,    # 0.8%
}

# Last alerted levels (to avoid repeat alerts)
_COMM_LAST_ALERTED: Dict[str, float] = {}
_COMM_ALERT_LOCK = threading.Lock()

def _check_commodity_alerts(symbol: str, price: float, prev_close: float) -> None:
    """
    Send Telegram alert if commodity price moves > threshold from prev_close.
    Called from the batch fetcher loop after each price update.
    """
    if not prev_close or prev_close <= 0:
        return
    sym = symbol.upper()
    threshold = COMMODITY_ALERT_THRESHOLDS.get(sym, 0.01)
    pct_move = (price - prev_close) / prev_close
    direction = "▲" if pct_move > 0 else "▼"

    with _COMM_ALERT_LOCK:
        last = _COMM_LAST_ALERTED.get(sym, prev_close)
        # Only alert if moved threshold from LAST alert level (not just prev_close)
        # to avoid spam on slow drifts
        since_last = abs(price - last) / max(last, 0.001)
        if since_last < threshold:
            return
        _COMM_LAST_ALERTED[sym] = price

    abs_move = price - prev_close
    msg = (
        "MCX COMMODITY ALERT -- " + sym + "\n"
        + direction + " Price:  Rs" + f"{price:,.2f}" + "\n"
        + "Prev Close: Rs" + f"{prev_close:,.2f}" + "\n"
        + "Change:     Rs" + f"{abs_move:+,.2f}" + "  (" + f"{pct_move*100:+.2f}" + "%)\n"
        + "Threshold:  " + f"{threshold*100:.1f}" + "%\n"
        + "Action: " + ("Watch for SELL BELOW" if pct_move < 0 else "Watch for BUY ABOVE") + " trigger"
    )
    try:
        send_telegram(msg, force_commodity=True)
        logger.info("Commodity alert sent: %s @ ₹%.2f (%+.2f%%)", sym, price, pct_move*100)
    except Exception as e:
        logger.debug("Commodity alert send error: %s", e)

# ── Keep tvdatafeed as an additional source (installed separately) ────────────
if ENABLE_COMMODITIES:
    try:
        from tvDatafeed import TvDatafeed, Interval as TVInterval
        TVDATAFEED_AVAILABLE = True
        tv_datafeed = TvDatafeed()
        logger.info("tvdatafeed loaded — will be used as secondary MCX source after TradingView WS")
    except ImportError:
        TVDATAFEED_AVAILABLE = False
        tv_datafeed = None
        logger.info("tvdatafeed not installed — TradingView WS + Investing.com + MoneyControl will handle MCX pricing")
    except Exception as e:
        TVDATAFEED_AVAILABLE = False
        tv_datafeed = None
        logger.warning("tvdatafeed init failed (%s) — using WS/scrape sources", e)

    # Start TradingView WebSocket streamer immediately on startup
    if _WS_AVAILABLE:
        try:
            _start_tv_ws_if_needed()
        except Exception as e:
            logger.debug("TV WS start error: %s", e)
    else:
        logger.info(
            "websocket-client not installed — TradingView WS disabled. "
            "Install: pip install websocket-client"
        )

    COMMODITY_CONFIG = {
        "GOLD": {
            "mcx_symbol":    "GOLD1!",
            "tv_symbol":     "MCX:GOLD1!",
            "yf_symbol":     os.getenv("COMMODITY_GOLD_SYMBOL",      "GC=F"),
            "investing_url": _INVESTING_URLS.get("GOLD", ""),
            "x_multiplier":  0.00343,
        },
        "SILVER": {
            "mcx_symbol":    "SILVER1!",
            "tv_symbol":     "MCX:SILVER1!",
            "yf_symbol":     os.getenv("COMMODITY_SILVER_SYMBOL",    "SI=F"),
            "investing_url": _INVESTING_URLS.get("SILVER", ""),
            "x_multiplier":  0.005145,
        },
        "NATURALGAS": {
            "mcx_symbol":    "NATURALGAS1!",
            "tv_symbol":     "MCX:NATURALGAS1!",
            "yf_symbol":     os.getenv("COMMODITY_NATURALGAS_SYMBOL","NG=F"),
            "investing_url": _INVESTING_URLS.get("NATURALGAS", ""),
            "x_multiplier":  0.0008575,
        },
        "CRUDE": {
            "mcx_symbol":    "CRUDEOIL1!",
            "tv_symbol":     "MCX:CRUDEOIL1!",
            "yf_symbol":     os.getenv("COMMODITY_CRUDE_SYMBOL",     "CL=F"),
            "investing_url": _INVESTING_URLS.get("CRUDE", ""),
            "x_multiplier":  0.000601754386,
        },
        "COPPER": {
            "mcx_symbol":    "COPPER1!",
            "tv_symbol":     "MCX:COPPER1!",
            "yf_symbol":     os.getenv("COMMODITY_COPPER_SYMBOL",    "HG=F"),
            "investing_url": _INVESTING_URLS.get("COPPER", ""),
            "x_multiplier":  0.004,
        },
    }

DEFAULT_SYMBOL_ORDER: List[str] = STOCKS + [sym for sym in COMMODITY_CONFIG.keys() if sym not in STOCKS]

# ----------------------------
# In-memory price history for charts
# ----------------------------

PRICE_HISTORY: Dict[str, deque] = {}
# Keep a small rolling feed of recent alerts for display in Dash
ALERT_FEED: deque = deque(maxlen=100)
HISTORY_LOCK = threading.Lock()
NOW_IST_OVERRIDE: Optional[datetime] = None
LOCAL_INTERVAL_TO_RULE = {
    "1m": "1min",
    "2m": "2min",
    "5m": "5min",
    "15m": "15min",
    "30m": "30min",
    "60m": "60min",
    "90m": "90min",
}

def init_price_history(symbols: List[str]) -> None:
    with HISTORY_LOCK:
        for s in symbols:
            if s not in PRICE_HISTORY:
                PRICE_HISTORY[s] = deque(maxlen=HISTORY_POINTS)

def append_price_history(symbol: str, ts: datetime, price: float) -> None:
    with HISTORY_LOCK:
        if symbol not in PRICE_HISTORY:
            PRICE_HISTORY[symbol] = deque(maxlen=HISTORY_POINTS)
        PRICE_HISTORY[symbol].append((ts, price))

def record_alert(text: str, *, symbol: Optional[str] = None, is_commodity: Optional[bool] = None) -> None:
    # Store alert text with timestamp and symbol type for dashboard display
    # Determine is_commodity from symbol if not provided
    if is_commodity is None and symbol:
        is_commodity = symbol in COMMODITY_CONFIG
    ALERT_FEED.append((now_ist(), text, is_commodity))

def get_recent_alerts(limit: int = 15, *, is_commodity: Optional[bool] = None) -> List[Tuple[datetime, str]]:
    # Return latest alerts (up to limit) in chronological order
    # Filter by is_commodity if specified (True for commodity, False for equity, None for all)
    all_items = list(ALERT_FEED)
    if is_commodity is not None:
        # Filter alerts by commodity type from all items, then take last limit
        filtered = [(ts, txt) for ts, txt, ic in all_items if ic == is_commodity]
        return filtered[-limit:] if filtered else []
    # Return without is_commodity flag for backward compatibility
    items = all_items[-limit:]
    return [(ts, txt) for ts, txt, _ in items]

def calculate_eod_analysis(
    allowed_symbols: Optional[Iterable[str]] = None,
    states: Optional[Dict[str, 'SymbolState']] = None,
) -> dict:
    """Calculate comprehensive EOD analysis from REALIZED_EVENTS.

    FIX (Bug 2 — double count):
      EOD_OPEN events were recorded in send_eod_open_positions_summary AND
      EOD_CLOSE was recorded in eod_square_off for the same position.
      calculate_eod_analysis now reads ONLY realized events (closed trades).
      Live unrealized P&L is computed directly from states using the correct
      formulas: (current − ba) × q for BUY, (sb − current) × q for SELL.

    If `allowed_symbols` is provided, restrict analysis to that subset.
    """
    with EVENTS_LOCK:
        all_events = {k: list(v) for k, v in REALIZED_EVENTS.items()}

    allowed_set = set(allowed_symbols) if allowed_symbols is not None else None

    analysis = {
        'symbols': {},
        'global': {
            'total_realized': 0.0,
            'total_unrealized': 0.0,
            'total_net': 0.0,
            'target_counts': {'T1': 0, 'T2': 0, 'T3': 0, 'T4': 0, 'T5': 0,
                              'ST1': 0, 'ST2': 0, 'ST3': 0, 'ST4': 0, 'ST5': 0},
            'sl_counts': {'BUY_SL': 0, 'SELL_SL': 0},
            'buy_realized': 0.0,
            'sell_realized': 0.0,
            'buy_unrealized': 0.0,
            'sell_unrealized': 0.0,
        }
    }

    for symbol, events in all_events.items():
        if allowed_set is not None and symbol not in allowed_set:
            continue
        sym_data = {
            'realized': 0.0,
            'unrealized': 0.0,
            'total': 0.0,
            'buy_hits': {'T1': 0, 'T2': 0, 'T3': 0, 'T4': 0, 'T5': 0, 'SL': 0},
            'sell_hits': {'ST1': 0, 'ST2': 0, 'ST3': 0, 'ST4': 0, 'ST5': 0, 'SL': 0},
        }

        for ev in events:
            # Skip EOD_OPEN — those were pre-squareoff snapshots (now removed from
            # REALIZED_EVENTS). EOD_CLOSE is the authoritative realized record.
            if ev['event'] == 'EOD_OPEN':
                continue
            sym_data['realized'] += ev['net']
            analysis['global']['total_realized'] += ev['net']

            if ev['side'] == 'BUY':
                analysis['global']['buy_realized'] += ev['net']
                if ev['event'] in ['T1', 'T2', 'T3', 'T4', 'T5']:
                    sym_data['buy_hits'][ev['event']] += 1
                    analysis['global']['target_counts'][ev['event']] += 1
                elif ev['event'] == 'BUY_SL':
                    sym_data['buy_hits']['SL'] += 1
                    analysis['global']['sl_counts']['BUY_SL'] += 1
            else:
                analysis['global']['sell_realized'] += ev['net']
                if ev['event'] in ['ST1', 'ST2', 'ST3', 'ST4', 'ST5']:
                    sym_data['sell_hits'][ev['event']] += 1
                    analysis['global']['target_counts'][ev['event']] += 1
                elif ev['event'] == 'SELL_SL':
                    sym_data['sell_hits']['SL'] += 1
                    analysis['global']['sl_counts']['SELL_SL'] += 1

        sym_data['total'] = sym_data['realized']
        analysis['symbols'][symbol] = sym_data

    # ── Live unrealized P&L via correct formulas ─────────────────────────────
    # BUY  open: (current_price − ba) × q
    # SELL open: (sb − current_price) × q
    if states:
        for sym, st in states.items():
            if allowed_set is not None and sym not in allowed_set:
                continue
            if not (st.in_position and st.side and st.qty_remaining > 0):
                continue
            lv = st.levels
            price = st.last_price
            if price is None or price <= 0:
                continue
            q = st.qty_remaining
            if st.side == 'BUY':
                # Open BUY: (current − ba) × q  ← your formula
                gross_unreal = (price - lv.buy_above) * q
            else:
                # Open SELL: (sb − current) × q  ← your formula
                gross_unreal = (lv.sell_below - price) * q
            net_unreal = gross_unreal - BROKERAGE_FLAT_PER_SIDE * 2
            if sym not in analysis['symbols']:
                analysis['symbols'][sym] = {
                    'realized': 0.0, 'unrealized': 0.0, 'total': 0.0,
                    'buy_hits': {'T1': 0, 'T2': 0, 'T3': 0, 'T4': 0, 'T5': 0, 'SL': 0},
                    'sell_hits': {'ST1': 0, 'ST2': 0, 'ST3': 0, 'ST4': 0, 'ST5': 0, 'SL': 0},
                }
            analysis['symbols'][sym]['unrealized'] = net_unreal
            analysis['symbols'][sym]['total'] = (
                analysis['symbols'][sym]['realized'] + net_unreal
            )
            analysis['global']['total_unrealized'] += net_unreal
            if st.side == 'BUY':
                analysis['global']['buy_unrealized'] += net_unreal
            else:
                analysis['global']['sell_unrealized'] += net_unreal

    analysis['global']['total_net'] = (
        analysis['global']['total_realized'] + analysis['global']['total_unrealized']
    )

    # Win rates
    total_targets = sum(analysis['global']['target_counts'].values())
    total_sl = sum(analysis['global']['sl_counts'].values())
    total_trades = total_targets + total_sl
    analysis['global']['win_rate'] = (
        total_targets / total_trades * 100
    ) if total_trades > 0 else 0.0

    buy_targets = sum(analysis['global']['target_counts'][f'T{i}'] for i in range(1, 6))
    buy_sl = analysis['global']['sl_counts']['BUY_SL']
    buy_trades = buy_targets + buy_sl
    analysis['global']['buy_win_rate'] = (
        buy_targets / buy_trades * 100
    ) if buy_trades > 0 else 0.0

    sell_targets = sum(analysis['global']['target_counts'][f'ST{i}'] for i in range(1, 6))
    sell_sl = analysis['global']['sl_counts']['SELL_SL']
    sell_trades = sell_targets + sell_sl
    analysis['global']['sell_win_rate'] = (
        sell_targets / sell_trades * 100
    ) if sell_trades > 0 else 0.0

    return analysis

def get_price_history(symbol: str) -> List[Tuple[datetime, float]]:
    with HISTORY_LOCK:
        dq = PRICE_HISTORY.get(symbol)
        return list(dq) if dq else []


def build_local_history_frame(symbol: str, interval: str, window: str) -> Optional[pd.DataFrame]:
    """Build an OHLC-ish dataframe from locally cached tick data for instant chart rendering."""
    rule = LOCAL_INTERVAL_TO_RULE.get(interval)
    if not rule:
        return None
    hist = get_price_history(symbol)
    if not hist:
        return None
    df = pd.DataFrame(hist, columns=["timestamp", "Close"])
    if df.empty:
        return None
    df = df.sort_values("timestamp").set_index("timestamp")
    now_local = now_ist()
    last_ts = df.index[-1]
    if last_ts.date() < now_local.date():
        now_local = last_ts
    if window == "Today":
        start_ts = now_local.replace(hour=9, minute=15, second=0, microsecond=0)
    elif window == "6h":
        start_ts = now_local - timedelta(hours=6)
    elif window == "1d":
        start_ts = now_local - timedelta(days=1)
    else:
        start_ts = df.index[0]
    df = df.loc[df.index >= start_ts]
    if df.empty:
        return None
    agg = df["Close"].resample(rule)
    local_df = pd.DataFrame(
        {
            "Open": agg.first(),
            "High": agg.max(),
            "Low": agg.min(),
            "Close": agg.last(),
        }
    ).dropna(how="all")
    local_df = local_df.ffill().dropna()
    if local_df.empty:
        return None
    local_df["Volume"] = 0
    if len(local_df) > HISTORY_POINTS:
        local_df = local_df.iloc[-HISTORY_POINTS:]
    return local_df


# ----------------------------
# Intraday OHLC fetch (yfinance)
# ----------------------------

VALID_INTERVALS = ["1m", "2m", "5m", "15m", "30m", "60m", "90m"]
INDICATOR_OPTIONS = [
    {"label": "None", "value": "none"},
    {"label": "SMA (20)", "value": "sma20"},
    {"label": "EMA (20)", "value": "ema20"},
    {"label": "Bollinger Bands (20, 2)", "value": "bbands"},
    {"label": "VWAP (session)", "value": "vwap"},
]

_OHLC_CACHE: Dict[Tuple[str, str, str, str], Tuple[float, pd.DataFrame]] = {}

def fetch_intraday_ohlc(symbol: str, interval: str, window: str) -> Optional[pd.DataFrame]:
    if interval not in VALID_INTERVALS:
        return None
    # Cache key also includes date stamp to avoid stale carryover across days
    date_key = now_ist().strftime("%Y%m%d")
    key = (symbol, interval, window, date_key)
    now_ts = time.time()
    cached = _OHLC_CACHE.get(key)
    if cached and (now_ts - cached[0] < 30):   # refresh every 30s for live chart accuracy
        return cached[1]

    local_df = build_local_history_frame(symbol, interval, window)
    if local_df is not None and len(local_df) >= 2:
        return local_df

    # For MCX commodities, try tvdatafeed first
    if is_commodity_symbol(symbol) and TVDATAFEED_AVAILABLE:
        try:
            cfg = get_symbol_config(symbol)
            mcx_sym = cfg.get("mcx_symbol") if cfg else None
            if mcx_sym:
                # Map interval to tvdatafeed Interval
                interval_map = {
                    "1m": TVInterval.in_1_minute,
                    "2m": TVInterval.in_2_minute,
                    "5m": TVInterval.in_5_minute,
                    "15m": TVInterval.in_15_minute,
                    "30m": TVInterval.in_30_minute,
                    "60m": TVInterval.in_1_hour,
                }
                tv_interval = interval_map.get(interval)
                if tv_interval:
                    # Determine number of bars based on window
                    if window == "Today":
                        n_bars = 390  # Full trading day
                    elif window == "6h":
                        n_bars = 360 if interval == "1m" else 60
                    else:  # 1d
                        n_bars = 390
                    df = tv_datafeed.get_hist(symbol=mcx_sym, exchange="MCX", interval=tv_interval, n_bars=n_bars)
                    if df is not None and not df.empty:
                        # Convert to IST timezone (MCX data is in IST)
                        if df.index.tz is None:
                            df.index = df.index.tz_localize(IST)
                        else:
                            df.index = df.index.tz_convert(IST)
                        # Apply window filter
                        if window == "Today":
                            start_ts = now_ist().replace(hour=9, minute=30, second=0, microsecond=0)
                            end_ts = now_ist().replace(hour=23, minute=0, second=0, microsecond=0)
                            df = df.loc[(df.index >= start_ts) & (df.index <= end_ts)]
                        elif window == "6h":
                            start_ts = now_ist() - timedelta(hours=6)
                            # Cap at 23:00 IST for commodities
                            end_ts = now_ist().replace(hour=23, minute=0, second=0, microsecond=0)
                            df = df.loc[(df.index >= start_ts) & (df.index <= end_ts)]
                        # Normalize column names
                        df.columns = [c.capitalize() for c in df.columns]
                        _OHLC_CACHE[key] = (now_ts, df)
                        return df
        except Exception as e:
            logger.debug(f"tvdatafeed ohlc fetch error for {symbol}: {e}")

    try:
        yf_symbol = get_yf_symbol(symbol)
        t = yf.Ticker(yf_symbol)
        # Use 1d period for today windows
        df = t.history(period="1d", interval=interval, auto_adjust=False)
        if df is None or df.empty:
            fallback_period = "5d" if interval in ("1m", "2m", "5m") else "1mo"
            try:
                df = t.history(period=fallback_period, interval=interval, auto_adjust=False)
            except Exception:
                df = None
        if (df is None or df.empty) and interval in ("1m", "2m", "5m"):
            try:
                df = yf.download(
                    tickers=yf_symbol,
                    period="5d",
                    interval=interval,
                    auto_adjust=False,
                    progress=False,
                    threads=False,
                )
            except Exception:
                df = None
        if df is None or df.empty:
            return None
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        # Localize to IST
        if df.index.tz is None:
            df.index = df.index.tz_localize(pytz.UTC).tz_convert(IST)
        else:
            df.index = df.index.tz_convert(IST)

        if window == "Today":
            start_ts = now_ist().replace(hour=9, minute=15, second=0, microsecond=0)
            df = df.loc[df.index >= start_ts]
        elif window == "6h":
            start_ts = now_ist() - timedelta(hours=6)
            df = df.loc[df.index >= start_ts]
        elif window == "1d":
            # Full session from 09:15 today
            start_ts = now_ist().replace(hour=9, minute=15, second=0, microsecond=0)
            df = df.loc[df.index >= start_ts]

        # Normalize columns
        cols = {c: c.capitalize() for c in df.columns}
        df = df.rename(columns=cols)

        _OHLC_CACHE[key] = (now_ts, df)
        return df
    except Exception as e:
        logger.debug(f"ohlc fetch error {symbol} {interval} {window}: {e}")
        return None


# ----------------------------
# Data classes
# ----------------------------

@dataclass
class Levels:
    previous_close: float
    x: float
    target_step: float  # step for T/ST progression; usually x, special = 0.6 * x
    # buy side
    buy_above: float
    t: List[float]  # t1..t5
    buy_sl: float
    # sell side
    sell_below: float
    st: List[float]  # st1..st5
    sell_sl: float


@dataclass
class ReentryWatch:
    original_side: str  # "BUY" or "SELL"
    target_label: str
    target_price: float
    threshold_price: float
    threshold_direction: str  # "ABOVE" or "BELOW"
    threshold_entry_side: str
    threshold_entry_price: float
    threshold_entry_sl: float
    threshold_entry_target: float
    retouch_entry_side: str
    retouch_entry_price: float
    retouch_entry_sl: float
    retouch_entry_target: float
    step: float
    buffer: float
    created_at: datetime = field(default_factory=lambda: datetime.now(IST))


@dataclass
class SymbolState:
    symbol: str
    levels: Levels
    adjusted_locked: bool = False  # set after 9:30 lock
    last_price: Optional[float] = None
    sent_events: Set[str] = field(default_factory=set)  # de-dup alerts
    in_position: bool = False
    side: Optional[str] = None  # "BUY" or "SELL"
    entry_price: Optional[float] = None
    entry_time: Optional[datetime] = None
    qty_remaining: int = 0
    qty_total: int = 0
    trade_date: Optional[str] = None  # YYYYMMDD of last trade day
    exited_today: bool = False  # block re-entry after full exit until next day
    buy_trailing_sl: Optional[float] = None
    sell_trailing_sl: Optional[float] = None
    last_target_hit_index: int = -1  # -1 means none
    initial_levels: Optional[Levels] = None  # store original levels for comparison
    is_commodity: bool = False
    x_multiplier: Optional[float] = None
    pending_930_recalc: bool = False
    prev_close: Optional[float] = None
    reentry_watch: Optional[ReentryWatch] = None
    manual_target: Optional[float] = None
    manual_target_label: Optional[str] = None
    # ── 65/45/25 retreat monitoring ──────────────────────────────────────────
    retreat_phase: Optional[str] = None       # None | 'warned_65' | 'activated_45'
    retreat_65_alerted: bool = False           # dedupe 65% warning
    retreat_45_alerted: bool = False           # dedupe 45% alert
    retreat_entry_level: float = 0.0          # buy_above/sell_below at entry
    retreat_peak_reached: bool = False         # True once price moved 65%+ toward profit (guards against instant-fire)


# ----------------------------
# Utilities
# ----------------------------

def now_ist() -> datetime:
    """Return current time in IST, optionally overridden for replays."""
    if NOW_IST_OVERRIDE is not None:
        return NOW_IST_OVERRIDE
    return datetime.now(IST)


def get_symbol_config(symbol: str) -> Dict[str, Union[str, float]]:
    return COMMODITY_CONFIG.get(symbol, {})


def is_commodity_symbol(symbol: str) -> bool:
    return symbol in COMMODITY_CONFIG


def get_mcx_price_via_tvdatafeed(symbol: str, use_live: bool = True) -> Optional[float]:
    """MCX commodity price — delegates to the full multi-source engine.

    Kept for backward compatibility.  All commodity price fetching now goes
    through get_mcx_price_multi_source() which tries:
      TradingView WS → tvdatafeed → Investing.com → MoneyControl → cache.

    The `use_live` flag is ignored (all sources return the latest available price).
    """
    if not is_commodity_symbol(symbol):
        return None
    # For prev-close (use_live=False): if tvdatafeed is available, use it directly
    # to get a clean daily close rather than a 1-min bar
    if not use_live and TVDATAFEED_AVAILABLE and tv_datafeed:
        try:
            cfg = get_symbol_config(symbol)
            mcx_raw = cfg.get("mcx_symbol") if cfg else None
            if mcx_raw:
                df = tv_datafeed.get_hist(
                    symbol=mcx_raw, exchange="MCX",
                    interval=TVInterval.in_daily, n_bars=2,
                )
                if df is not None and not df.empty:
                    p = float(df["close"].iloc[-1])
                    if p > 0:
                        return p
        except Exception as e:
            logger.debug("tvdatafeed prev-close %s: %s", symbol, e)
    return get_mcx_price_multi_source(symbol)


def get_yf_symbol(symbol: str) -> str:
    # Special-case known renames / aliases.
    # FIX: These symbols have non-standard Yahoo Finance tickers — keep this
    # list updated whenever NSE renames or Yahoo Finance changes a ticker.
    _OVERRIDES = {
        "TATAMOTORS": "TMPV.NS",   # Tata Motors renamed on Yahoo
        "NIFTY":      "^NSEI",
        "BANKNIFTY":  "^NSEBANK",
        # BPCL: Yahoo uses BPCL.NS but sometimes reports "possibly delisted"
        # when fetching via period=10d — use a longer period in get_prev_close
        # to work around this. No ticker override needed.
        "HINDUNILVR": "HINDUNILVR.NS",   # sometimes needs explicit .NS
    }
    sym_up = symbol.upper()
    if sym_up in _OVERRIDES:
        return _OVERRIDES[sym_up]

    cfg = get_symbol_config(symbol)
    override = cfg.get("yf_symbol") if cfg else None
    if override:
        return str(override)
    return f"{symbol}.NS"


def is_between(local_dt: datetime, start_h: int, start_m: int, end_h: int, end_m: int) -> bool:
    s = local_dt.replace(hour=start_h, minute=start_m, second=0, microsecond=0)
    e = local_dt.replace(hour=end_h, minute=end_m, second=0, microsecond=0)
    return s <= local_dt <= e


def market_open(local_dt: datetime) -> bool:
    """Equity regular session: 09:15–15:30 IST."""
    if _MC_AVAILABLE:
        return MarketCalendar.is_equity_session(local_dt)
    return is_between(local_dt, 9, 15, 15, 30)


def trading_window_open(
    local_dt: datetime,
    *,
    symbol: Optional[str] = None,
    is_commodity: Optional[bool] = None,
) -> bool:
    """
    Determine if trading logic should run for the provided context.

    Equities: 09:15–15:30 IST
    Commodities (MCX): 09:30–23:00 IST
    """
    if is_commodity is None and symbol is not None:
        is_commodity = is_commodity_symbol(symbol)
    if is_commodity:
        return is_between(local_dt, 9, 30, 23, 0)
    return market_open(local_dt)


def equity_entry_session_open(local_dt: datetime) -> bool:
    """
    Entries for equities are allowed during the full equity session,
    up to the equity EOD cutoff used for square-off/analysis.

    Equity session: 09:15–15:11 IST (entries)
    """
    return is_between(local_dt, 9, 15, 15, 11)


def alerts_allowed(local_dt: datetime, symbol: Optional[str] = None) -> bool:
    """
    Check if alerts are allowed at the given time.
    - For equity symbols: alerts stop at 15:11 IST (3:11 PM)
    - For commodity symbols: alerts continue until 23:00 IST (11:00 PM)
    - If symbol is not provided: defaults to 23:00 cutoff (commodity behavior)
    """
    if symbol and not is_commodity_symbol(symbol):
        # Equity: stop alerts at 15:11 IST
        cutoff = local_dt.replace(hour=15, minute=11, second=0, microsecond=0)
    else:
        # Commodity: allow alerts until 23:00 IST
        cutoff = local_dt.replace(hour=23, minute=0, second=0, microsecond=0)
    return local_dt <= cutoff


def premarket_window(local_dt: datetime) -> bool:
    return is_between(local_dt, 9, 15, 9, 30)


def should_export_adjusted_levels(local_dt: datetime) -> bool:
    """
    Only export adjusted levels once 09:30 adjustments are meaningful and
    *during* live equity hours. This prevents backfilled 09:30 exports when
    the program is started after market close.
    """
    after_931 = local_dt.hour > 9 or (local_dt.hour == 9 and local_dt.minute >= 31)
    in_session = market_open(local_dt)
    return after_931 and in_session


def after_930(local_dt: datetime) -> bool:
    return local_dt.hour > 9 or (local_dt.hour == 9 and local_dt.minute >= 30)


# ── 09:30–09:35 entry blackout ────────────────────────────────────────────────
# No new entries for 5 min after 09:30 level adjustment.
# Trades open at 09:35 with adjusted levels.
BLACKOUT_END_MINUTE = 35   # first minute where entries are allowed after 09:30


def in_930_blackout(local_dt: datetime) -> bool:
    """
    Returns True during 09:30:00–09:34:59 IST.
    New equity entries (try_entry, process_reentry_watch) are blocked.
    Exits (SL, T1-T5, retreat) are NOT affected.
    Commodities are NOT affected.
    """
    t = local_dt.hour * 60 + local_dt.minute
    return 9 * 60 + 30 <= t < 9 * 60 + BLACKOUT_END_MINUTE


# Sent once per day when blackout lifts at 09:35
_930_blackout_alerted: bool = False


def get_prev_close(symbol: str) -> Optional[float]:
    """Robust previous-close helper — single symbol path (commodity fallback).

    For equities, the startup loader uses get_all_prev_closes_batch() which
    makes ONE yfinance API call for all 38 stocks — avoiding rate limits entirely.
    This function is only called when a single symbol is needed (e.g. retry path).

    For MCX commodities, tries tvdatafeed first.
    """
    # MCX commodities: try tvdatafeed first
    if is_commodity_symbol(symbol):
        mcx_price = get_mcx_price_via_tvdatafeed(symbol, use_live=False)
        if mcx_price is not None and mcx_price > 0:
            return mcx_price

    try:
        import warnings as _w
        now_local = now_ist()
        market_cutoff = now_local.replace(hour=15, minute=30, second=0, microsecond=0)
        yf_sym = get_yf_symbol(symbol)

        # Suppress "possibly delisted" false-positive warnings (e.g. BPCL)
        with _w.catch_warnings():
            _w.simplefilter("ignore")
            h = yf.download(yf_sym, period="30d", interval="1d",
                            auto_adjust=False, progress=False, threads=False)

        if h is None or h.empty:
            raise ValueError("no daily history")

        # Handle MultiIndex columns from yf.download
        if isinstance(h.columns, pd.MultiIndex):
            h.columns = h.columns.get_level_values(0)
        h.columns = [c.capitalize() for c in h.columns]
        if "Close" not in h.columns:
            raise ValueError("no Close column")

        last_close = float(h["Close"].iloc[-1])
        last_date  = h.index[-1].date()
        today_date = now_local.date()

        if now_local >= market_cutoff:
            if last_date == today_date:
                return last_close
            try:
                t = yf.Ticker(yf_sym)
                intraday = t.history(period="2d", interval="1m")
                if intraday is not None and not intraday.empty:
                    if intraday.index.tz is None:
                        intraday.index = intraday.index.tz_localize(pytz.UTC).tz_convert(IST)
                    else:
                        intraday.index = intraday.index.tz_convert(IST)
                    start_ts  = now_local.replace(hour=9, minute=15, second=0, microsecond=0)
                    today_mask = (intraday.index >= start_ts) & (intraday.index <= market_cutoff)
                    sl = intraday.loc[today_mask]
                    if not sl.empty:
                        return float(sl["Close"].iloc[-1])
            except Exception as ie:
                logger.debug("intraday fallback failed for %s: %s", symbol, ie)
            return last_close

        if last_date == today_date and len(h) >= 2:
            return float(h["Close"].iloc[-2])
        return last_close
    except Exception as e:
        logger.warning("prev close error %s: %s", symbol, e)
        return None


def get_all_prev_closes_batch(symbols: List[str]) -> Dict[str, float]:
    """
    Fetch previous closes for ALL equity symbols in ONE yfinance API call.

    This is the key fix for the rate-limiting crash:
      - 38 sequential Ticker().history() calls = 38 HTTP requests → rate limit
      - ONE yf.download(tickers_list) call = 1 HTTP request → no rate limit

    Also:
    1. Checks a persistent daily cache file first (survives restarts)
    2. Falls back to per-symbol calls only for individual misses

    Returns {SYMBOL: prev_close_float} for all successful lookups.
    """
    import warnings as _w
    now_local  = now_ist()
    today_str  = now_local.strftime("%Y%m%d")
    cache_path = os.path.join("levels", f"prev_closes_persistent_{today_str}.json")

    # ── 1. Load today's persistent cache (survives crash/restart) ─────────────
    cached: Dict[str, float] = {}
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as fh:
                raw = json.load(fh)
            cached = {k.upper(): float(v) for k, v in raw.items() if v}
            if len(cached) >= len(symbols) * 0.9:   # ≥90% hit → use cache as-is
                logger.info("Loaded prev_close for %d symbols from today's persistent cache", len(cached))
                return cached
        except Exception:
            cached = {}

    # ── 2. Batch download — ONE request for all equity symbols ────────────────
    equity_syms = [s for s in symbols if not is_commodity_symbol(s)]
    yf_sym_map  = {get_yf_symbol(s): s.upper() for s in equity_syms}
    yf_tickers  = " ".join(yf_sym_map.keys())

    logger.info("Fetching prev_close for %d symbols in ONE batch request…", len(equity_syms))
    result: Dict[str, float] = dict(cached)   # start from cache hits

    try:
        with _w.catch_warnings():
            _w.simplefilter("ignore")   # suppress "possibly delisted" noise
            df = yf.download(
                tickers=yf_tickers,
                period="30d",
                interval="1d",
                auto_adjust=False,
                progress=False,
                group_by="ticker",      # MultiIndex: (field, ticker)
                threads=True,           # yfinance internal batching
            )

        if df is None or df.empty:
            raise ValueError("batch download returned empty DataFrame")

        market_cutoff = now_local.replace(hour=15, minute=30, second=0, microsecond=0)
        today_date    = now_local.date()

        for yf_sym, nse_sym in yf_sym_map.items():
            if nse_sym in result:
                continue   # already in cache
            try:
                # Extract this ticker's Close column from the MultiIndex DataFrame
                if isinstance(df.columns, pd.MultiIndex):
                    # group_by="ticker" → columns are (field, ticker) or (ticker, field)
                    try:
                        sym_df = df.xs(yf_sym, axis=1, level=1)
                    except KeyError:
                        try:
                            sym_df = df.xs(yf_sym, axis=1, level=0)
                        except KeyError:
                            continue
                else:
                    sym_df = df

                if sym_df is None or sym_df.empty:
                    continue

                # Normalize column names
                sym_df = sym_df.copy()
                sym_df.columns = [c.capitalize() for c in sym_df.columns]
                if "Close" not in sym_df.columns:
                    continue

                closes    = sym_df["Close"].dropna()
                if closes.empty:
                    continue

                last_close = float(closes.iloc[-1])
                last_date  = closes.index[-1].date()

                if now_local >= market_cutoff:
                    if last_date == today_date:
                        result[nse_sym] = last_close
                    else:
                        result[nse_sym] = last_close   # best available
                else:
                    if last_date == today_date and len(closes) >= 2:
                        result[nse_sym] = float(closes.iloc[-2])
                    else:
                        result[nse_sym] = last_close

            except Exception as sym_err:
                logger.debug("Batch parse error %s: %s", nse_sym, sym_err)

        logger.info("Batch fetch: %d/%d symbols succeeded", len(result), len(equity_syms))

    except Exception as batch_err:
        logger.warning("Batch download failed (%s) — falling back to sequential", batch_err)
        # Sequential fallback with 0.3s gap to avoid rate limit
        for sym in equity_syms:
            if sym.upper() in result:
                continue
            pc = get_prev_close(sym)
            if pc:
                result[sym.upper()] = pc
            time.sleep(0.3)

    # ── 3. Save to persistent cache for this session (survives restarts) ──────
    try:
        os.makedirs("levels", exist_ok=True)
        tmp = cache_path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(result, fh, indent=2)
        os.replace(tmp, cache_path)
        logger.info("Saved persistent prev_close cache → %s", os.path.basename(cache_path))
    except Exception as ce:
        logger.debug("Could not save persistent cache: %s", ce)

    return result


def get_live_price(symbol: str) -> Optional[float]:
    # For MCX commodities, try tvdatafeed first (for true MCX prices)
    if is_commodity_symbol(symbol):
        mcx_price = get_mcx_price_via_tvdatafeed(symbol, use_live=True)
        if mcx_price is not None and mcx_price > 0:
            return mcx_price
        # Fallback to Yahoo Finance if tvdatafeed unavailable
    # Check batch cache first (populated by _BatchPriceFetcher every 2s)
    cached = _LIVE_PRICE_CACHE.get(symbol.upper())
    if cached is not None:
        return cached
    # Cold-start fallback — cache not yet populated
    try:
        t = yf.Ticker(get_yf_symbol(symbol))
        p = t.fast_info.get("lastPrice")
        if not p:
            p = t.info.get("regularMarketPrice")
        if p and p > 0:
            return float(p)
        return None
    except Exception as e:
        logger.debug(f"live price error {symbol}: {e}")
        return None


# --- In-memory price cache updated by background batch fetcher ---
_LIVE_PRICE_CACHE: Dict[str, float] = {}
_LIVE_PRICE_CACHE_LOCK = __import__('threading').Lock()
_LIVE_PRICE_CACHE_AGE  = 0.0   # monotonic time of last batch refresh


def _batch_refresh_prices(symbols: list) -> Dict[str, float]:
    """Download all equity prices in ONE yfinance call. ~1-3s for 38 symbols."""
    import warnings as _w
    _OVERRIDES = {
        "TATAMOTORS": "TMPV.NS",
        "NIFTY":      "^NSEI",
        "BANKNIFTY":  "^NSEBANK",
        "HINDUNILVR": "HINDUNILVR.NS",
    }
    _COMM = {"GOLD","SILVER","NATURALGAS","CRUDE","COPPER"}
    yf_map = {}
    for s in symbols:
        su = s.upper()
        if su in _COMM:
            continue
        yf_map[_OVERRIDES.get(su, f"{su}.NS")] = su

    if not yf_map:
        return {}

    result: Dict[str, float] = {}
    try:
        with _w.catch_warnings():
            _w.simplefilter("ignore")
            df = yf.download(
                " ".join(yf_map.keys()),
                period="1d", interval="1m",
                auto_adjust=False, progress=False,
                group_by="ticker", threads=True,
            )
        if df is None or df.empty:
            return {}
        import pandas as pd
        for yf_sym, nse_sym in yf_map.items():
            try:
                if isinstance(df.columns, pd.MultiIndex):
                    try:    sym_df = df.xs(yf_sym, axis=1, level=1)
                    except KeyError:
                        try: sym_df = df.xs(yf_sym, axis=1, level=0)
                        except KeyError: continue
                else:
                    sym_df = df
                sym_df = sym_df.copy()
                sym_df.columns = [str(c).strip().capitalize() for c in sym_df.columns]
                if "Close" not in sym_df.columns:
                    continue
                closes = sym_df["Close"].dropna()
                if not closes.empty:
                    result[nse_sym] = float(closes.iloc[-1])
            except Exception:
                pass
    except Exception as e:
        logger.debug("Batch price download error: %s", e)
    return result


def _start_batch_price_fetcher(symbols: list, interval_s: float = 2.0) -> None:
    """
    Launch background daemon that refreshes all prices every interval_s seconds.
    v8.0 fix: MERGES commodity prices from _COMMODITY_PRICE_CACHE into _LIVE_PRICE_CACHE
    so they appear in live_prices.json and reach scanners / ZMQ publisher.
    """
    import threading as _th

    def _loop():
        global _LIVE_PRICE_CACHE_AGE
        while True:
            try:
                prices = _batch_refresh_prices(symbols)
                if prices:
                    with _LIVE_PRICE_CACHE_LOCK:
                        _LIVE_PRICE_CACHE.update(prices)
                    _LIVE_PRICE_CACHE_AGE = __import__('time').monotonic()

                # v8.0: Merge live commodity prices into the shared cache
                # _batch_refresh_prices skips commodities — we fill the gap here
                _comm_syms = {"GOLD","SILVER","NATURALGAS","CRUDE","COPPER"}
                comm_prices = {}
                for sym in _comm_syms:
                    # First try TradingView WS (fastest)
                    ws_p = None
                    with _TV_WS_LOCK:
                        ws_p = _TV_WS_PRICES.get(sym)
                    if ws_p and ws_p > 0:
                        comm_prices[sym] = ws_p
                        _update_commodity_cache(sym, ws_p, "tradingview_ws")
                        continue
                    # Then try Investing.com (1–3s, reliable INR MCX prices)
                    try:
                        p = _fetch_investing_com_api(sym)
                        if p and p > 0:
                            comm_prices[sym] = p
                            _update_commodity_cache(sym, p, "investing_api")
                            continue
                    except Exception: pass
                    # Fall back to cache (stale but non-None)
                    cached_p = _get_from_cache(sym)
                    if cached_p:
                        comm_prices[sym] = cached_p

                if comm_prices:
                    with _LIVE_PRICE_CACHE_LOCK:
                        _LIVE_PRICE_CACHE.update(comm_prices)
                    logger.debug("Commodity price merge: %s", {k:f"{v:.2f}" for k,v in comm_prices.items()})

            except Exception as e:
                logger.debug("Batch fetcher error: %s", e)
            __import__('time').sleep(interval_s)

    t = _th.Thread(target=_loop, daemon=True, name="BatchPriceFetcher")
    t.start()
    logger.info("BatchPriceFetcher started (interval=%.1fs, %d symbols, commodity merge: ON)", interval_s, len(symbols))


def calc_levels_for_symbol(symbol: str, prev_close: float, *, x_override: Optional[float] = None, step_override: Optional[float] = None) -> Levels:
    x = x_override if x_override is not None else prev_close * X_FACTOR_MULTIPLIER
    special_symbols = {"RELIANCE", "SBIN", "KOTAKBANK", "ICICIBANK", "HUL", "HDFC"}
    if step_override is not None:
        step = step_override
    else:
        step = x * 0.6 if symbol in special_symbols else x

    buy_above = prev_close + x
    t1 = buy_above + step
    t2 = t1 + step
    t3 = t2 + step
    t4 = t3 + step
    t5 = t4 + step

    sell_below = prev_close - x
    st1 = sell_below - step
    st2 = st1 - step
    st3 = st2 - step
    st4 = st3 - step
    st5 = st4 - step

    return Levels(
        previous_close=prev_close,
        x=x,
        target_step=step,
        buy_above=buy_above,
        t=[t1, t2, t3, t4, t5],
        buy_sl=buy_above - x,
        sell_below=sell_below,
        st=[st1, st2, st3, st4, st5],
        sell_sl=sell_below + x,
    )


def adjust_levels_premarket(levels: Levels, current_price: float, side: Optional[str] = None) -> Tuple[Levels, Optional[str], Optional[str]]:
    # Returns (new_levels, crossed_level_name, shift_type)
    # shift_type: 'BUY' or 'SELL' or None
    x = levels.x
    step = getattr(levels, 'target_step', x)
    # Deepcopy to avoid reference bugs
    new_levels = deepcopy(levels)

    buy_levels = [levels.buy_above] + list(levels.t)
    sell_levels = [levels.sell_below] + list(levels.st)
    buy_names = ["BUY_ABOVE", "T1", "T2", "T3", "T4", "T5"]
    sell_names = ["SELL_BELOW", "ST1", "ST2", "ST3", "ST4", "ST5"]

    # Sequential shifting, step by step, only one shift per call
    # Buy side
    for idx, lv in enumerate(buy_levels):
        if current_price >= lv:
            if idx == 0:
                # Crossed Buy Above => new Buy Above = old T1, T1 = T2, ..., T4 = T5, T5 = T5 + x
                new_levels.buy_above = levels.t[0]
                for i in range(4):
                    new_levels.t[i] = levels.t[i+1]
                new_levels.t[4] = levels.t[4] + step
                new_levels.buy_sl = new_levels.buy_above - x
                # Also shift SELL side up by step to maintain constant X difference
                new_levels.sell_below = levels.sell_below + step
                for i in range(5):
                    new_levels.st[i] = levels.st[i] + step
                new_levels.sell_sl = new_levels.sell_below + x
                return new_levels, buy_names[idx], 'BUY'
            elif 1 <= idx <= 4:
                # Crossed T1-T4 -> shift corresponding targets
                sh = idx
                for i in range(sh-1, 4):
                    new_levels.t[i] = levels.t[i+1]
                new_levels.t[4] = levels.t[4] + step
                new_levels.buy_sl = new_levels.buy_above - x
                # Also shift SELL side up by step to maintain constant X difference
                new_levels.sell_below = levels.sell_below + step
                for i in range(5):
                    new_levels.st[i] = levels.st[i] + step
                new_levels.sell_sl = new_levels.sell_below + x
                return new_levels, buy_names[idx], 'BUY'
            elif idx == 5:
                # Crossed T5, extend T5
                new_levels.t[4] = levels.t[4] + step
                new_levels.buy_sl = new_levels.buy_above - x
                # Also shift SELL side up by step to maintain constant X difference
                new_levels.sell_below = levels.sell_below + step
                for i in range(5):
                    new_levels.st[i] = levels.st[i] + step
                new_levels.sell_sl = new_levels.sell_below + x
                return new_levels, buy_names[idx], 'BUY'
    # Sell side
    for idx, lv in enumerate(sell_levels):
        if current_price <= lv:
            if idx == 0:
                # Crossed Sell Below => new Sell Below = old ST1, ST1 = ST2, ..., ST4 = ST5, ST5 = ST5 - x
                new_levels.sell_below = levels.st[0]
                for i in range(4):
                    new_levels.st[i] = levels.st[i+1]
                new_levels.st[4] = levels.st[4] - step
                new_levels.sell_sl = new_levels.sell_below + x
                # Also shift BUY side down by step to maintain constant X difference
                new_levels.buy_above = levels.buy_above - step
                for i in range(5):
                    new_levels.t[i] = levels.t[i] - step
                new_levels.buy_sl = new_levels.buy_above - x
                return new_levels, sell_names[idx], 'SELL'
            elif 1 <= idx <= 4:
                sh = idx
                for i in range(sh-1, 4):
                    new_levels.st[i] = levels.st[i+1]
                new_levels.st[4] = levels.st[4] - step
                new_levels.sell_sl = new_levels.sell_below + x
                # Also shift BUY side down by step to maintain constant X difference
                new_levels.buy_above = levels.buy_above - step
                for i in range(5):
                    new_levels.t[i] = levels.t[i] - step
                new_levels.buy_sl = new_levels.buy_above - x
                return new_levels, sell_names[idx], 'SELL'
            elif idx == 5:
                new_levels.st[4] = levels.st[4] - step
                new_levels.sell_sl = new_levels.sell_below + x
                # Also shift BUY side down by step to maintain constant X difference
                new_levels.buy_above = levels.buy_above - step
                for i in range(5):
                    new_levels.t[i] = levels.t[i] - step
                new_levels.buy_sl = new_levels.buy_above - x
                return new_levels, sell_names[idx], 'SELL'
    return levels, None, None


def _select_telegram_channel(symbol: Optional[str], force_commodity: bool) -> Tuple[Optional[str], List[str]]:
    use_commodity = force_commodity or (symbol and is_commodity_symbol(symbol))
    if use_commodity and TELEGRAM_COMMODITY_BOT_TOKEN and TELEGRAM_COMMODITY_CHAT_IDS:
        return TELEGRAM_COMMODITY_BOT_TOKEN, TELEGRAM_COMMODITY_CHAT_IDS
    return TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_IDS_ALL


def _send_telegram_message(bot_token: Optional[str], chat_ids: List[str], payload: Dict[str, Union[str, int]], endpoint: str) -> None:
    """
    Send a Telegram message.
    When the async patch is active, offloads HTTP to a background worker so
    the main price loop is NEVER blocked by network latency or timeouts.
    Falls back to synchronous requests.post when patch is not loaded.
    """
    if not bot_token or not chat_ids:
        return

    # ── Async path (patch active) ─────────────────────────────────────────────
    if _PATCH_ACTIVE and _ASYNC_TG is not None and endpoint == "sendMessage":
        text = payload.get("text", "")
        if text:
            _ASYNC_TG.send_text(str(text), token=bot_token, chat_ids=list(chat_ids))
            return   # fire-and-forget — returns in microseconds

    # ── Synchronous fallback (original behaviour) ─────────────────────────────
    url = f"https://api.telegram.org/bot{bot_token}/{endpoint}"
    for chat_id in chat_ids:
        data = dict(payload)
        data["chat_id"] = chat_id
        try:
            requests.post(url, data=data, timeout=15)
        except Exception as exc:
            logger.debug("telegram send error (%s): %s", endpoint, exc)


def send_telegram(text: str, *, symbol: Optional[str] = None, force_commodity: bool = False) -> None:
    bot_token, chat_ids = _select_telegram_channel(symbol, force_commodity)
    _send_telegram_message(bot_token, chat_ids, {"text": text}, "sendMessage")
    is_commodity = force_commodity or (symbol and symbol in COMMODITY_CONFIG)
    record_alert(text, symbol=symbol, is_commodity=is_commodity)


def send_telegram_document(
    document_path: str,
    *,
    caption: Optional[str] = None,
    symbol: Optional[str] = None,
    force_commodity: bool = False,
) -> None:
    """
    Upload a document via Telegram.
    Async path: file bytes are read once on the calling thread, then the
    actual HTTP upload is offloaded to the background Telegram worker so the
    price loop is never stalled by a slow upload.
    """
    bot_token, chat_ids = _select_telegram_channel(symbol, force_commodity)
    if not bot_token or not chat_ids:
        if caption:
            record_alert(caption)
        return

    # ── Async path ────────────────────────────────────────────────────────────
    if _PATCH_ACTIVE and _ASYNC_TG is not None:
        _ASYNC_TG.send_document(
            document_path, token=bot_token, chat_ids=list(chat_ids), caption=caption
        )
        if caption:
            is_commodity = force_commodity or (symbol and symbol in COMMODITY_CONFIG)
            record_alert(caption, symbol=symbol, is_commodity=is_commodity)
        return

    # ── Synchronous fallback ─────────────────────────────────────────────────
    try:
        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
        filename = os.path.basename(document_path)
        with open(document_path, "rb") as fh:
            file_bytes = fh.read()
        for chat_id in chat_ids:
            files = {"document": (filename, file_bytes)}
            data = {"chat_id": chat_id}
            if caption:
                data["caption"] = caption
            try:
                requests.post(url, data=data, files=files, timeout=25)
            except Exception as exc:
                logger.debug("telegram document send error: %s", exc)
    except Exception as exc:
        logger.debug("telegram document read/send error: %s", exc)
    finally:
        if caption:
            is_commodity = force_commodity or (symbol and symbol in COMMODITY_CONFIG)
            record_alert(caption, symbol=symbol, is_commodity=is_commodity)

def _fmt_pct(value: float) -> str:
    sign = "+" if value >= 0 else ""
    return f"{sign}{value:.2f}%"


def _detect_side_from_status(status: str) -> Optional[str]:
    s = status.upper()
    if "BUY" in s:
        return "BUY"
    if "SELL" in s:
        return "SELL"
    # If status mentions a specific level, infer direction
    if any(k in s for k in ["T1", "T2", "T3", "T4", "T5"]):
        return "BUY"
    if any(k in s for k in ["ST1", "ST2", "ST3", "ST4", "ST5"]):
        return "SELL"
    return None


def _build_levels_section(side: str, lv: 'Levels', current_price: float, hit_level: Optional[str]) -> List[str]:
    lines: List[str] = []
    if side == "BUY":
        lines.append("📈 Buy Levels:")
        # Buy Above
        ba_line = f"Buy Above: {lv.buy_above:.2f}"
        lines.append(ba_line)
        # Targets with perc from current price
        for i, tgt in enumerate(lv.t, start=1):
            pct = (tgt - current_price) / current_price * 100.0
            lines.append(
                f"Target {i}: {tgt:.2f} ({_fmt_pct(pct)})"
            )
        # Stop Loss (previous_close)
        sl = lv.buy_sl
        pct_sl = (sl - current_price) / current_price * 100.0
        lines.append(f"Stop Loss: {sl:.2f} ({_fmt_pct(pct_sl)})")
    else:
        lines.append("📉 Sell Levels:")
        # Sell Below
        sb_line = f"Sell Below: {lv.sell_below:.2f}"
        lines.append(sb_line)
        # Targets (short side)
        for i, tgt in enumerate(lv.st, start=1):
            pct = (tgt - current_price) / current_price * 100.0
            lines.append(
                f"Target {i}: {tgt:.2f} ({_fmt_pct(pct)})"
            )
        sl = lv.sell_sl
        pct_sl = (sl - current_price) / current_price * 100.0
        lines.append(f"Stop Loss: {sl:.2f} ({_fmt_pct(pct_sl)})")
    return lines


def build_simple_alert(title: str, symbol: str, status_line: str, levels: 'Levels', current_price: Optional[float] = None, hit_level: Optional[str] = None, quantity: Optional[int] = None) -> str:
    # Determine current price fallback if not provided
    cp = current_price
    if cp is None:
        cp = get_live_price(symbol) or levels.previous_close
    # Header
    ts = now_ist().strftime("%Y-%m-%d %H:%M:%S IST%z")
    # Change vs previous close
    change_pct = ((cp - levels.previous_close) / levels.previous_close * 100.0) if levels.previous_close else 0.0
    # Status and side
    side = _detect_side_from_status(status_line) or ("BUY" if cp >= levels.buy_above else ("SELL" if cp <= levels.sell_below else "BUY"))
    # Asset type label for commodity symbols
    asset_label = " [🏅 MCX Commodity]" if is_commodity_symbol(symbol) else ""
    lines: List[str] = []
    lines.append(f"🚨 {symbol}{asset_label} — {title} at {ts}")
    lines.append("")
    lines.append(f"Previous Close: {levels.previous_close:.2f}")
    lines.append(f"Current Price: {cp:.2f}")
    lines.append(f"Change: {_fmt_pct(change_pct)}")
    lines.append(f"Deviation (X): {levels.x:.2f}")
    if quantity is not None:
        lines.append(f"Quantity: {quantity}")
    lines.append(f"Status: {status_line}")
    lines.append("")
    lines.append("📊 Technical Analysis:")
    lines.extend(_build_levels_section(side, levels, cp, hit_level))
    return "\n".join(lines)


def fmt_levels_table(symbol: str, levels: Levels) -> str:
    # Deprecated in favor of the rich alert format, kept for backward compatibility if needed
    lines = [f"{symbol}", f"PC: {levels.previous_close:.2f}  X: {levels.x:.2f}"]
    lines.append(
        "Buy: BA={:.2f} T1={:.2f} T2={:.2f} T3={:.2f} T4={:.2f} T5={:.2f} | SL={:.2f}".format(
            levels.buy_above, levels.t[0], levels.t[1], levels.t[2], levels.t[3], levels.t[4], levels.buy_sl
        )
    )
    lines.append(
        "Sell: SB={:.2f} ST1={:.2f} ST2={:.2f} ST3={:.2f} ST4={:.2f} ST5={:.2f} | SL={:.2f}".format(
            levels.sell_below, levels.st[0], levels.st[1], levels.st[2], levels.st[3], levels.st[4], levels.sell_sl
        )
    )
    return "\n".join(lines)


# ----------------------------
# Excel exports for level layouts
# ----------------------------

def _style_cell(cell, *, fill=None, font=None, alignment=None, border=None, number_format=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _write_symbol_levels_block(ws, start_row: int, symbol: str, lv: Levels, styles: dict) -> int:
    buy_headers = ["BUY ABOVE", "T1", "T2", "T3", "T4", "T5"]
    sell_headers = ["SELL BELOW", "T1", "T2", "T3", "T4", "T5"]
    header_row = start_row
    buy_values_row = header_row + 1
    sell_header_row = header_row + 2
    sell_values_row = header_row + 3
    next_block_row = header_row + 7

    # Symbol + X cell
    symbol_cell = ws.cell(row=header_row, column=1, value=symbol)
    _style_cell(symbol_cell, fill=styles["symbol_fill"], font=styles["symbol_font"], alignment=styles["left_align"], border=styles["header_border"])

    x_cell = ws.cell(row=header_row, column=2, value=round(lv.x, 2))
    _style_cell(x_cell, fill=styles["x_fill"], font=styles["x_font"], alignment=styles["center_align"], border=styles["header_border"], number_format="0.00")

    ws.cell(row=header_row, column=3, value=None)

    for idx, label in enumerate(buy_headers):
        col = 4 + idx
        cell = ws.cell(row=header_row, column=col, value=label)
        _style_cell(cell, fill=styles["header_fill"], font=styles["header_font"], alignment=styles["center_align"], border=styles["header_border"])

    # Buy values
    ws.cell(row=buy_values_row, column=1, value=round(lv.previous_close, 2))
    ws.cell(row=buy_values_row, column=2, value=None)
    buy_values = [lv.buy_above] + lv.t
    for idx, value in enumerate(buy_values):
        col = 4 + idx
        cell = ws.cell(row=buy_values_row, column=col, value=round(value, 4))
        _style_cell(cell, fill=styles["buy_fill"], font=styles["data_font"], alignment=styles["center_align"], border=styles["data_border"], number_format="0.00")

    # Sell header row
    ws.cell(row=sell_header_row, column=1, value=None)
    ws.cell(row=sell_header_row, column=2, value=None)
    for idx, label in enumerate(sell_headers):
        col = 4 + idx
        cell = ws.cell(row=sell_header_row, column=col, value=label)
        _style_cell(cell, fill=styles["header_fill"], font=styles["header_font"], alignment=styles["center_align"], border=styles["header_border"])

    # Sell values row
    sell_values = [lv.sell_below] + lv.st
    for idx, value in enumerate(sell_values):
        col = 4 + idx
        cell = ws.cell(row=sell_values_row, column=col, value=round(value, 4))
        _style_cell(cell, fill=styles["sell_fill"], font=styles["data_font"], alignment=styles["center_align"], border=styles["data_border"], number_format="0.00")

    return next_block_row


def export_levels_layout(states: Dict[str, SymbolState], file_path: str, *, heading: Optional[str] = None, include_date: bool = True, trading_date: Optional[datetime] = None) -> None:
    try:
        ordered_symbols = [sym for sym in DEFAULT_SYMBOL_ORDER if sym in states]
        for sym in states.keys():
            if sym not in ordered_symbols:
                ordered_symbols.append(sym)
        if not ordered_symbols:
            logger.info("No symbols available to export for %s", file_path)
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        column_widths = {"A": 14, "B": 6, "C": 4, "D": 14, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        styles = {
            "symbol_fill": PatternFill("solid", fgColor="000000"),
            "symbol_font": Font(color="FFFFFF", bold=True, size=13),
            "x_fill": PatternFill("solid", fgColor="FFC000"),
            "x_font": Font(color="000000", bold=True),
            "header_fill": PatternFill("solid", fgColor="000000"),
            "header_font": Font(color="FFFFFF", bold=True),
            "buy_fill": PatternFill("solid", fgColor="C6EFCE"),
            "sell_fill": PatternFill("solid", fgColor="F8CBAD"),
            "data_font": Font(color="000000"),
            "center_align": Alignment(horizontal="center", vertical="center"),
            "left_align": Alignment(horizontal="left", vertical="center"),
            "header_border": Border(
                left=Side(style="thin", color="4A4A4A"),
                right=Side(style="thin", color="4A4A4A"),
                top=Side(style="thin", color="4A4A4A"),
                bottom=Side(style="thin", color="4A4A4A"),
            ),
            "data_border": Border(
                left=Side(style="thin", color="DDDDDD"),
                right=Side(style="thin", color="DDDDDD"),
                top=Side(style="thin", color="DDDDDD"),
                bottom=Side(style="thin", color="DDDDDD"),
            ),
        }

        start_row = 1
        if include_date:
            if trading_date:
                day_date = trading_date.strftime("%A, %d %b %Y")
            else:
                day_date = now_ist().strftime("%A, %d %b %Y")
        else:
            day_date = ""
        if heading:
            ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=9)
            heading_text = heading if not day_date else f"{heading} | {day_date}"
            heading_cell = ws.cell(row=start_row, column=4, value=heading_text)
            _style_cell(
                heading_cell,
                font=Font(bold=True, size=14, color="FFFFFF"),
                fill=PatternFill("solid", fgColor="1F4E78"),
                alignment=styles["center_align"],
            )
            start_row += 2
        elif include_date and day_date:
            ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=9)
            date_cell = ws.cell(row=start_row, column=4, value=day_date)
            _style_cell(
                date_cell,
                font=Font(bold=True, size=13, color="FFFFFF"),
                fill=PatternFill("solid", fgColor="1F4E78"),
                alignment=styles["center_align"],
            )
            start_row += 2

        for sym in ordered_symbols:
            st = states[sym]
            start_row = _write_symbol_levels_block(ws, start_row, sym, st.levels, styles)

        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)
        logger.info(f"Levels exported to {file_path}")
    except Exception as exc:
        logger.warning(f"Failed to export levels workbook to {file_path}: {exc}")


# ----------------------------
# Plotly Dash dashboard (optional)
# ----------------------------

def _empty_figure(message: str, height: int):
    fig = go.Figure()
    fig.add_annotation(
        text=message,
        xref="paper", yref="paper",
        x=0.5, y=0.5, showarrow=False,
        font=dict(size=14, color="#666"),
        align="center",
        bgcolor="rgba(255,255,255,0.7)",
    )
    fig.update_layout(margin=dict(l=40, r=10, t=20, b=20), height=height)
    fig.update_xaxes(visible=False)
    fig.update_yaxes(visible=False)
    return fig


def _add_level_lines(fig: 'go.Figure', lv: Levels, x_start: datetime, x_end: datetime) -> None:
    levels = [
        (lv.buy_above, "Buy Above", "#2e7d32"),
        (lv.t[0], "T1", "#2e7d32"),
        (lv.t[1], "T2", "#2e7d32"),
        (lv.t[2], "T3", "#2e7d32"),
        (lv.t[3], "T4", "#2e7d32"),
        (lv.t[4], "T5", "#2e7d32"),
        (lv.buy_sl, "Buy SL", "#1b5e20"),
        (lv.sell_below, "Sell Below", "#c62828"),
        (lv.st[0], "ST1", "#c62828"),
        (lv.st[1], "ST2", "#c62828"),
        (lv.st[2], "ST3", "#c62828"),
        (lv.st[3], "ST4", "#c62828"),
        (lv.st[4], "ST5", "#c62828"),
        (lv.sell_sl, "Sell SL", "#6a1b9a"),
    ]
    for y, name, color in levels:
        try:
            fig.add_shape(type="line", x0=x_start, x1=x_end, y0=y, y1=y, xref="x", yref="y", line=dict(color=color, width=1, dash="dash"))
            fig.add_annotation(x=x_end, y=y, text=f"{name} {y:.2f}", showarrow=False, xanchor="left", yanchor="middle", bgcolor="rgba(255,255,255,0.6)", font=dict(size=11, color=color))
        except Exception:
            pass


def _apply_indicator(fig: 'go.Figure', ohlc: pd.DataFrame, indicator: str) -> None:
    try:
        if indicator == "sma20":
            sma = ohlc["Close"].rolling(20, min_periods=1).mean()
            fig.add_trace(go.Scatter(x=ohlc.index, y=sma, name="SMA20", line=dict(color="#ff9800", width=1.5)))
        elif indicator == "ema20":
            ema = ohlc["Close"].ewm(span=20, adjust=False).mean()
            fig.add_trace(go.Scatter(x=ohlc.index, y=ema, name="EMA20", line=dict(color="#9c27b0", width=1.5)))
        elif indicator == "bbands":
            ma = ohlc["Close"].rolling(20, min_periods=1).mean()
            std = ohlc["Close"].rolling(20, min_periods=1).std(ddof=0)
            upper = ma + 2 * std
            lower = ma - 2 * std
            fig.add_trace(go.Scatter(x=ohlc.index, y=upper, name="BB Upper", line=dict(color="#3f51b5", width=1)))
            fig.add_trace(go.Scatter(x=ohlc.index, y=ma, name="BB MA", line=dict(color="#607d8b", width=1)))
            fig.add_trace(go.Scatter(x=ohlc.index, y=lower, name="BB Lower", line=dict(color="#3f51b5", width=1)))
        elif indicator == "vwap":
            # session VWAP from 09:15
            start = now_ist().replace(hour=9, minute=15, second=0, microsecond=0)
            df = ohlc.loc[ohlc.index >= start]
            typical = (df["High"] + df["Low"] + df["Close"]) / 3.0
            vol = df.get("Volume") if "Volume" in df.columns else pd.Series([1.0] * len(df), index=df.index)
            cum_vol_price = (typical * vol).cumsum()
            cum_vol = vol.cumsum().replace(0, pd.NA)
            vwap = (cum_vol_price / cum_vol).ffill()
            fig.add_trace(go.Scatter(x=df.index, y=vwap, name="VWAP", line=dict(color="#009688", width=1.5)))
    except Exception:
        pass


def build_symbol_figure(symbol: str, st: 'SymbolState', ohlc: Optional[pd.DataFrame], chart_height: int, indicator: str = "none"):
    # Fallback to locally cached price history if yfinance data is unavailable.
    if (ohlc is None or ohlc.empty) and PRICE_HISTORY.get(symbol):
        try:
            hist = get_price_history(symbol)
            if hist:
                hist_df = pd.DataFrame(hist, columns=["timestamp", "Close"]).set_index("timestamp")
                ohlc = hist_df
        except Exception as exc:
            logger.debug("Failed to build fallback history for %s: %s", symbol, exc)

    if ohlc is None or ohlc.empty:
        return _empty_figure("No data available for current Interval/Window.", chart_height)

    fig = go.Figure()
    try:
        fig.add_trace(
            go.Scatter(
                x=ohlc.index,
                y=ohlc["Close"].astype(float),
                mode="lines",
                name="Close",
                hovertemplate="%{x|%Y-%m-%d %H:%M:%S}<br>Close=%{y:.2f}<extra></extra>",
            )
        )
        _apply_indicator(fig, ohlc, indicator)
        # Level lines across current visible data
        _add_level_lines(fig, st.levels, ohlc.index[0], ohlc.index[-1])
    except Exception as e:
        return _empty_figure(f"Render error: {str(e)}", chart_height)

    lv = st.levels
    # Note: temporarily remove horizontal level lines to ensure charts render reliably.
    # We will add them back using simple shape lines after confirming candles display.

    # Current price marker
    if st.last_price is not None and ohlc is not None and not ohlc.empty:
        fig.add_trace(
            go.Scatter(x=[ohlc.index[-1]], y=[st.last_price], mode="markers", name="Last", marker=dict(color="orange", size=8))
        )

    fig.update_layout(
        title=f"{symbol} Live",
        margin=dict(l=40, r=10, t=40, b=30),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        height=chart_height,
        hovermode="x unified",
        uirevision=f"{symbol}_rev",  # preserve zoom/pan across updates
    )
    fig.update_xaxes(
        showspikes=True,
        spikemode="across",
        spikesnap="cursor",
        rangeslider=dict(visible=True),
        rangeselector=dict(
            buttons=[
                dict(count=15, label="15m", step="minute", stepmode="backward"),
                dict(count=30, label="30m", step="minute", stepmode="backward"),
                dict(count=1, label="1h", step="hour", stepmode="backward"),
                dict(step="all"),
            ]
        ),
    )
    fig.update_yaxes(showspikes=True, spikethickness=1)
    return fig


def start_dash_app(
    states: Dict[str, 'SymbolState'],
    *,
    port: Optional[int] = None,
    app_title: str = "Live Levels Dashboard",
    use_public_tunnel: bool = True,
    ngrok_authtoken: Optional[str] = None,
) -> None:
    if not ENABLE_DASH or not DASH_AVAILABLE:
        return
    if not states:
        logger.info("No symbols available for %s; dashboard skipped.", app_title)
        return
    port = port or DASH_PORT

    app = Dash(__name__)
    all_symbols = list(states.keys())
    dashboard_is_commodity = all(st.is_commodity for st in states.values())
    # Default to a limited set of symbols for performance; user can still change selection
    default_syms = all_symbols[:MAX_VISIBLE_CHARTS] if not dashboard_is_commodity else all_symbols[:]
    default_interval = "1m"
    default_window = "Today"
    # full-width charts only; no per-row control

    app.layout = html.Div([
        html.Div([
            html.H3(app_title, style={"margin": "0 0 8px 0"}),
            html.Div([
                html.Button("Symbols", id="open_symbols", n_clicks=0, style={"height": "36px"}),
                html.Div("Interval", style={"marginLeft": "16px"}),
                dcc.Dropdown(
                    id="interval",
                    options=[{"label": iv, "value": iv} for iv in VALID_INTERVALS],
                    value=default_interval,
                    clearable=False,
                    style={"width": "110px"},
                ),
                html.Div("Window", style={"marginLeft": "16px"}),
                dcc.Dropdown(
                    id="window",
                    options=[{"label": w, "value": w} for w in ["Today", "6h", "1d"]],
                    value=default_window,
                    clearable=False,
                    style={"width": "110px"},
                ),
                html.Div("Indicator", style={"marginLeft": "16px"}),
                dcc.Dropdown(
                    id="indicator",
                    options=INDICATOR_OPTIONS,
                    value="none",
                    clearable=False,
                    style={"width": "180px"},
                ),
                # Adaptive interval: 1s during market hours, 30s outside
                dcc.Interval(id="tick", interval=1000, n_intervals=0),
            ], style={"display": "flex", "gap": "16px", "alignItems": "center", "flexWrap": "wrap"}),
            html.Div(id="alerts_panel", style={"marginTop": "8px", "fontSize": "13px", "maxHeight": "160px", "overflowY": "auto", "borderTop": "1px solid #eee", "paddingTop": "6px"}),
        ], style={"position": "sticky", "top": 0, "background": "#fafafa", "zIndex": 1, "padding": "8px 12px", "borderBottom": "1px solid #eee"}),
        # Modal overlay for symbols selection
        html.Div(id="symbols_modal", children=[
            html.Div([
                html.Div("Select Symbols", style={"fontWeight": "600", "marginBottom": "8px"}),
                dcc.Dropdown(
                    id="symbols",
                    options=[{"label": s, "value": s} for s in all_symbols],
                    value=default_syms,
                    multi=True,
                    persistence=True,
                    persistence_type="session",
                    style={"width": "100%"},
                ),
                html.Div([
                    html.Button("Select All", id="symbols_select_all", n_clicks=0),
                    html.Button("Deselect All", id="symbols_clear_all", n_clicks=0, style={"marginLeft": "8px"}),
                    html.Button("Close", id="symbols_close", n_clicks=0, style={"marginLeft": "8px"}),
                ], style={"marginTop": "10px"}),
            ], style={"background": "#fff", "padding": "12px", "borderRadius": "8px", "width": "520px", "maxWidth": "90%"}),
        ], style={"position": "fixed", "inset": 0, "display": "none", "alignItems": "center", "justifyContent": "center", "background": "rgba(0,0,0,0.35)", "zIndex": 1000}),
        html.Div(id="charts")
    ], style={"padding": "12px", "background": "#fff"})

    @app.callback(
        Output("symbols_modal", "style"),
        [Input("open_symbols", "n_clicks"), Input("symbols_close", "n_clicks")],
        prevent_initial_call=True,
    )
    def _toggle_symbols_modal(open_clicks, close_clicks):
        ctx = dash.callback_context
        if not ctx.triggered:
            raise dash.exceptions.PreventUpdate
        trig = ctx.triggered[0]["prop_id"].split(".")[0]
        show = trig == "open_symbols"
        return {"position": "fixed", "inset": 0, "display": ("flex" if show else "none"), "alignItems": "center", "justifyContent": "center", "background": "rgba(0,0,0,0.35)", "zIndex": 1000}

    @app.callback(
        Output("symbols", "value"),
        [Input("symbols_select_all", "n_clicks"), Input("symbols_clear_all", "n_clicks")],
        State("symbols", "value"),
        prevent_initial_call=True,
    )
    def _symbols_select_clear(n_all, n_clear, current):
        ctx = dash.callback_context
        if not ctx.triggered:
            raise dash.exceptions.PreventUpdate
        trig = ctx.triggered[0]["prop_id"].split(".")[0]
        if trig == "symbols_select_all":
            return all_symbols
        if trig == "symbols_clear_all":
            return []
        raise dash.exceptions.PreventUpdate

    # ── Adaptive tick interval: 1s during market, 30s outside ───────────────────
    @app.callback(Output("tick", "interval"), Input("tick", "n_intervals"))
    def _adapt_tick_interval(_):
        now_local = now_ist()
        if trading_window_open(now_local, is_commodity=dashboard_is_commodity):
            return 1000   # 1 second during market hours — real-time
        return 30000       # 30 seconds outside — conserve CPU

    # ── Per-app OHLC cache — avoids yfinance on every 1s callback ────────────
    _ohlc_cache: Dict[Tuple, Tuple[float, object]] = {}
    _OHLC_CACHE_TTL = 30.0  # seconds between yfinance refreshes

    def _get_ohlc_cached(sym: str, interval: str, window: str):
        """Return OHLC df: local history first (free), yfinance only if stale."""
        key = (sym, interval, window)
        now_t = time.time()
        # 1. Always prefer local in-memory price history (updated every tick)
        local_df = build_local_history_frame(sym, interval, window)
        if local_df is not None and len(local_df) >= 2:
            return local_df
        # 2. Check yfinance cache
        if key in _ohlc_cache:
            cached_t, cached_df = _ohlc_cache[key]
            if now_t - cached_t < _OHLC_CACHE_TTL:
                return cached_df
        # 3. Fetch from yfinance and cache
        df = fetch_intraday_ohlc(sym, interval, window)
        _ohlc_cache[key] = (now_t, df)
        return df

    @app.callback(
        [Output("charts", "children"), Output("alerts_panel", "children")],
        [
            Input("symbols", "value"),
            Input("interval", "value"),
            Input("window", "value"),
            Input("indicator", "value"),
            Input("tick", "n_intervals"),
        ],
    )
    def _render(selected: List[str], interval: str, window: str, indicator: str, _n: int):
        children = []
        # Check market status
        now_local = now_ist()
        is_market_open = trading_window_open(now_local, is_commodity=dashboard_is_commodity)
        if dashboard_is_commodity:
            show_eod_analysis = now_local.hour >= 23  # commodity: show from 23:00 IST onward
        else:
            show_eod_analysis = (now_local.hour == 15 and now_local.minute >= 30) or now_local.hour > 15
        
        # Show EOD Analysis section if after 15:30
        if show_eod_analysis:
            try:
                # Filter symbols by dashboard type (equity vs commodity)
                filtered_symbols = [sym for sym in all_symbols if (sym in COMMODITY_CONFIG) == dashboard_is_commodity]
                analysis = calculate_eod_analysis(allowed_symbols=filtered_symbols, states=states)
                gl = analysis['global']
                
                # Global summary
                eod_section = html.Div([
                    html.H3("📊 EOD Analysis", style={"marginTop": "20px", "marginBottom": "10px", "color": "#333"}),
                    html.Div([
                        html.Div([
                            html.Strong("Total Realized P&L: "), f"₹{gl['total_realized']:.2f}",
                        ], style={"marginBottom": "8px", "fontSize": "16px"}),
                        html.Div([
                            html.Strong("Total Unrealized P&L: "), f"₹{gl['total_unrealized']:.2f}",
                        ], style={"marginBottom": "8px", "fontSize": "16px"}),
                        html.Div([
                            html.Strong("Total Net P&L: "), 
                            html.Span(f"₹{gl['total_net']:.2f}", style={"color": "#2e7d32" if gl['total_net'] >= 0 else "#c62828", "fontWeight": "bold"}),
                        ], style={"marginBottom": "12px", "fontSize": "18px"}),
                        html.Div([
                            html.Strong("Win Rate: "), f"{gl['win_rate']:.1f}%",
                            html.Span(" | ", style={"margin": "0 8px"}),
                            html.Strong("BUY Win Rate: "), f"{gl['buy_win_rate']:.1f}%",
                            html.Span(" | ", style={"margin": "0 8px"}),
                            html.Strong("SELL Win Rate: "), f"{gl['sell_win_rate']:.1f}%",
                        ], style={"marginBottom": "12px", "fontSize": "14px", "color": "#555"}),
                        html.Div([
                            html.Strong("BUY Realized: "), f"₹{gl['buy_realized']:.2f}",
                            html.Span(" | ", style={"margin": "0 8px"}),
                            html.Strong("SELL Realized: "), f"₹{gl['sell_realized']:.2f}",
                        ], style={"marginBottom": "8px", "fontSize": "14px"}),
                        html.Div([
                            html.Strong("Target Distribution: "),
                            f"T1:{gl['target_counts']['T1']} T2:{gl['target_counts']['T2']} T3:{gl['target_counts']['T3']} T4:{gl['target_counts']['T4']} T5:{gl['target_counts']['T5']} | ",
                            f"ST1:{gl['target_counts']['ST1']} ST2:{gl['target_counts']['ST2']} ST3:{gl['target_counts']['ST3']} ST4:{gl['target_counts']['ST4']} ST5:{gl['target_counts']['ST5']}",
                        ], style={"marginBottom": "8px", "fontSize": "14px"}),
                        html.Div([
                            html.Strong("Stop Loss: "), f"BUY_SL:{gl['sl_counts']['BUY_SL']} SELL_SL:{gl['sl_counts']['SELL_SL']}",
                        ], style={"marginBottom": "12px", "fontSize": "14px"}),
                    ], style={"background": "#f5f5f5", "padding": "12px", "borderRadius": "6px", "marginBottom": "16px"}),
                    
                    # Per-symbol table
                    html.Div([
                        html.H4("Per-Symbol Breakdown", style={"marginBottom": "8px", "fontSize": "16px"}),
                        html.Table([
                            html.Thead([
                                html.Tr([
                                    html.Th("Symbol", style={"padding": "8px", "textAlign": "left", "background": "#f0f0f0", "border": "1px solid #ddd"}),
                                    html.Th("Realized", style={"padding": "8px", "textAlign": "right", "background": "#f0f0f0", "border": "1px solid #ddd"}),
                                    html.Th("Unrealized", style={"padding": "8px", "textAlign": "right", "background": "#f0f0f0", "border": "1px solid #ddd"}),
                                    html.Th("Total", style={"padding": "8px", "textAlign": "right", "background": "#f0f0f0", "border": "1px solid #ddd"}),
                                    html.Th("BUY Hits", style={"padding": "8px", "textAlign": "center", "background": "#f0f0f0", "border": "1px solid #ddd"}),
                                    html.Th("SELL Hits", style={"padding": "8px", "textAlign": "center", "background": "#f0f0f0", "border": "1px solid #ddd"}),
                                ])
                            ]),
                            html.Tbody([
                                html.Tr([
                                    html.Td(sym, style={"padding": "6px", "border": "1px solid #ddd"}),
                                    html.Td(f"₹{data['realized']:.2f}", style={"padding": "6px", "textAlign": "right", "border": "1px solid #ddd"}),
                                    html.Td(f"₹{data['unrealized']:.2f}", style={"padding": "6px", "textAlign": "right", "border": "1px solid #ddd"}),
                                    html.Td(
                                        f"₹{data['total']:.2f}", 
                                        style={
                                            "padding": "6px", 
                                            "textAlign": "right",
                                            "color": "#2e7d32" if data['total'] >= 0 else "#c62828",
                                            "fontWeight": "bold",
                                            "border": "1px solid #ddd"
                                        }
                                    ),
                                    html.Td(
                                        f"T1:{data['buy_hits']['T1']} T2:{data['buy_hits']['T2']} T3:{data['buy_hits']['T3']} T4:{data['buy_hits']['T4']} T5:{data['buy_hits']['T5']} SL:{data['buy_hits']['SL']}",
                                        style={"padding": "6px", "textAlign": "center", "fontSize": "12px", "border": "1px solid #ddd"}
                                    ),
                                    html.Td(
                                        f"ST1:{data['sell_hits']['ST1']} ST2:{data['sell_hits']['ST2']} ST3:{data['sell_hits']['ST3']} ST4:{data['sell_hits']['ST4']} ST5:{data['sell_hits']['ST5']} SL:{data['sell_hits']['SL']}",
                                        style={"padding": "6px", "textAlign": "center", "fontSize": "12px", "border": "1px solid #ddd"}
                                    ),
                                ])
                                for sym, data in sorted(analysis['symbols'].items(), key=lambda x: x[1]['total'], reverse=True)
                                if data['realized'] != 0 or data['unrealized'] != 0
                            ])
                        ], style={"width": "100%", "borderCollapse": "collapse", "fontSize": "13px", "border": "1px solid #ddd"})
                    ], style={"marginTop": "16px"}),
                ], style={"background": "#fff", "padding": "16px", "borderRadius": "8px", "border": "2px solid #4caf50", "marginBottom": "20px"})
                children.append(eod_section)
            except Exception as e:
                logger.debug(f"EOD analysis error: {e}")
        
        # If market is closed, show message
        if not is_market_open:
            open_hour, open_minute = (9, 30) if dashboard_is_commodity else (9, 15)
            close_hour = 23 if dashboard_is_commodity else 15
            close_minute = 0 if dashboard_is_commodity else 30
            next_open = now_local.replace(hour=open_hour, minute=open_minute, second=0, microsecond=0)
            cutoff_passed = (now_local.hour > close_hour) or (now_local.hour == close_hour and now_local.minute >= close_minute)
            if cutoff_passed:
                next_open = next_open + timedelta(days=1)
            
            message = html.Div([
                html.H2("Market is Closed", style={"textAlign": "center", "color": "#666", "marginTop": "100px"}),
                html.P(f"Next trading session starts at {next_open.strftime('%H:%M IST')} on {next_open.strftime('%Y-%m-%d')}", 
                       style={"textAlign": "center", "color": "#999", "fontSize": "18px"}),
                html.P(f"Current time: {now_local.strftime('%Y-%m-%d %H:%M:%S IST')}", 
                       style={"textAlign": "center", "color": "#999", "fontSize": "14px", "marginTop": "20px"}),
            ], style={"width": "100%", "minHeight": "400px"})
            children.append(message)
        else:
            # Market is open, render charts
            # Robust defaults if dropdown value not propagated yet
            if not selected:
                selected = default_syms
            if not interval:
                interval = "1m"
            if not window:
                window = "Today"
            if not indicator:
                indicator = "none"
            chart_height = 520
            row: List[html.Div] = []
            for idx, sym in enumerate(selected):
                try:
                    st = states.get(sym)
                    if not st:
                        continue
                    # OPT-07: Use cached OHLC — free local history first, yfinance only if stale
                    ohlc = _get_ohlc_cached(sym, interval, window)
                    fig = build_symbol_figure(sym, st, ohlc, chart_height, indicator)
                    card_body = dcc.Graph(
                        figure=fig,
                        config={
                            "displayModeBar": True,
                            "scrollZoom": True,
                            "doubleClick": "reset",
                            "responsive": True,
                            "modeBarButtonsToAdd": [
                                "drawline",
                                "drawopenpath",
                                "drawrect",
                                "drawcircle",
                                "eraseshape",
                            ],
                        },
                    )
                    # Realized events panel under chart
                    events_view = []
                    try:
                        with EVENTS_LOCK:
                            events = list(REALIZED_EVENTS.get(sym, []))
                        if events:
                            for ev in events[::-1]:  # latest first
                                events_view.append(html.Div(
                                    f"{ev['time']} {ev['event']} {ev['side']} qty {ev['qty']} @ {ev['price']:.2f} | Gross {ev['gross']:.2f} | Net {ev['net']:.2f}",
                                    style={"fontSize": "12px", "color": "#444"}
                                ))
                        else:
                            events_view.append(html.Div("No realized P&L yet", style={"fontSize": "12px", "color": "#888"}))
                    except Exception:
                        events_view = []

                    card = html.Div([
                        html.Div(sym, style={"fontWeight": "600", "marginBottom": "4px"}),
                        dcc.Loading(type="dot", children=card_body),
                        html.Div(events_view, style={"marginTop": "6px", "borderTop": "1px dashed #eee", "paddingTop": "6px"}),
                    ], style={"border": "1px solid #eaeaea", "padding": "8px", "borderRadius": "8px", "marginBottom": "12px", "width": "100%", "boxShadow": "0 1px 2px rgba(0,0,0,0.04)"})
                    row.append(card)
                except Exception as e:
                    err_fig = _empty_figure(f"Error: {str(e)}", chart_height)
                    card_body = dcc.Graph(figure=err_fig, config={"displayModeBar": False})
                    events_view = []
                    card = html.Div([
                        html.Div(sym, style={"fontWeight": "600", "marginBottom": "4px"}),
                        dcc.Loading(type="dot", children=card_body),
                        html.Div(events_view, style={"marginTop": "6px", "borderTop": "1px dashed #eee", "paddingTop": "6px"}),
                    ], style={"border": "1px solid #eaeaea", "padding": "8px", "borderRadius": "8px", "marginBottom": "12px", "width": "100%", "boxShadow": "0 1px 2px rgba(0,0,0,0.04)"})
                    row.append(card)
            if row:
                children.extend(row)
            else:
                # No symbols selected
                children.append(html.Div([
                    html.P("No symbols selected. Click 'Symbols' button to select stocks.", 
                           style={"textAlign": "center", "color": "#999", "fontSize": "16px", "marginTop": "100px"})
                ]))

        # build alerts list - filter by dashboard type (equity vs commodity)
        try:
            alerts = get_recent_alerts(limit=15, is_commodity=dashboard_is_commodity)
            alert_children = []
            for ts, txt in alerts[::-1]:
                alert_children.append(html.Div([
                    html.Span(ts.strftime('%H:%M:%S'), style={"color": "#666", "marginRight": "6px"}),
                    html.Span(html.Span(txt.replace('\n', ' | ')))
                ]))
        except Exception:
            alert_children = []

        return children, alert_children

    def _run():
        try:
            # Dash 3+: use app.run instead of deprecated run_server
            # Host 0.0.0.0 allows access from any IP address
            logger.info(f"Starting %s on http://0.0.0.0:{port}", app_title)
            app.run(host="0.0.0.0", port=port, debug=False)
        except Exception as e:
            logger.error(f"Dashboard startup failed: {e}")

    th = threading.Thread(target=_run, daemon=True)
    th.start()

    # ── Determine dashboard type and pick authtoken ───────────────────────────
    is_commodity_dash = dashboard_is_commodity or (port == COMMODITY_DASH_PORT)
    if ngrok_authtoken is None:
        ngrok_authtoken = NGROK_AUTHTOKEN_COMMODITY if is_commodity_dash else NGROK_AUTHTOKEN_EQUITY

    # ── Poll until Dash is actually accepting connections (max 20s) ───────────
    # This eliminates Error 1033 caused by starting tunnels before the origin is ready.
    import urllib.request as _urlreq
    dash_ready = False
    for _i in range(40):
        time.sleep(0.5)
        try:
            _urlreq.urlopen(f"http://127.0.0.1:{port}", timeout=1)
            dash_ready = True
            logger.info("Dash on port %d ready after %.1fs — starting tunnels", port, (_i + 1) * 0.5)
            break
        except Exception:
            pass
    if not dash_ready:
        logger.warning("Dash port %d not responding after 20s — proceeding anyway", port)

    if not use_public_tunnel:
        # URL notification is handled exclusively by unified_dash_v3.py — suppress here
        logger.info("Tunnel disabled — dashboard local only. unified_dash_v3.py will notify Telegram.")
        return

    # ── Start ngrok tunnel (only) ────────────────────────────────────────────
    public_url: Optional[str] = None
    public_url = _start_ngrok_only(port, ngrok_authtoken)

    # ── Telegram notification ─────────────────────────────────────────────────
    local_ip  = _get_local_ip()
    local_url = f"http://localhost:{port}"

    if public_url:
        logger.info("✅ Ngrok tunnel live: %s", public_url)
    else:
        logger.warning("⚠️  Ngrok failed for port %d — dashboard local only", port)

    # NOTE: Dashboard URL Telegram alert is sent ONCE by unified_dash_v3.py (port 8055).
    # Algofinal suppresses its own URL message to avoid duplicate/old links in Telegram.
    # unified_dash_v3.py sends to ALL 3 bots (equity, commodity, crypto) on startup.


def start_merged_dash_app(
    equity_states: Dict[str, 'SymbolState'],
    commodity_states: Dict[str, 'SymbolState'],
    *,
    port: Optional[int] = None,
    app_title: str = "Merged Levels Dashboard",
    use_public_tunnel: bool = True,
    ngrok_authtoken: Optional[str] = None,
) -> None:
    """
    Run a single dashboard with an Equity/Commodity switch.
    Intended for one public URL while keeping separate local dashboards.
    """
    if not ENABLE_DASH or not DASH_AVAILABLE:
        return
    if not equity_states and not commodity_states:
        logger.info("No symbols available for %s; dashboard skipped.", app_title)
        return

    port = port or MERGED_DASH_PORT
    app = Dash(__name__)

    mode_options = [
        {"label": "Equity", "value": "equity"},
        {"label": "Commodity", "value": "commodity"},
    ]
    default_mode = "equity" if equity_states else "commodity"

    def _states_for_mode(mode: str) -> Dict[str, 'SymbolState']:
        return commodity_states if mode == "commodity" else equity_states

    def _symbols_for_mode(mode: str) -> List[str]:
        return list(_states_for_mode(mode).keys())

    default_symbols = _symbols_for_mode(default_mode)
    default_syms = (
        default_symbols[:MAX_VISIBLE_CHARTS]
        if default_mode == "equity"
        else default_symbols[:]
    )

    app.layout = html.Div([
        html.Div([
            html.H3(app_title, style={"margin": "0 0 8px 0"}),
            html.Div([
                html.Div("View", style={"marginRight": "6px"}),
                dcc.RadioItems(
                    id="dash_mode",
                    options=mode_options,
                    value=default_mode,
                    persistence=True,
                    persistence_type="session",
                    labelStyle={"display": "inline-block", "marginRight": "12px"},
                ),
                html.Button("Symbols", id="open_symbols", n_clicks=0, style={"height": "36px", "marginLeft": "16px"}),
                html.Div("Interval", style={"marginLeft": "16px"}),
                dcc.Dropdown(
                    id="interval",
                    options=[{"label": iv, "value": iv} for iv in VALID_INTERVALS],
                    value="1m",
                    clearable=False,
                    style={"width": "110px"},
                ),
                html.Div("Window", style={"marginLeft": "16px"}),
                dcc.Dropdown(
                    id="window",
                    options=[{"label": w, "value": w} for w in ["Today", "6h", "1d"]],
                    value="Today",
                    clearable=False,
                    style={"width": "110px"},
                ),
                html.Div("Indicator", style={"marginLeft": "16px"}),
                dcc.Dropdown(
                    id="indicator",
                    options=INDICATOR_OPTIONS,
                    value="none",
                    clearable=False,
                    style={"width": "180px"},
                ),
                dcc.Interval(id="tick", interval=2000, n_intervals=0),
            ], style={"display": "flex", "gap": "16px", "alignItems": "center", "flexWrap": "wrap"}),
            html.Div(id="alerts_panel", style={"marginTop": "8px", "fontSize": "13px", "maxHeight": "160px", "overflowY": "auto", "borderTop": "1px solid #eee", "paddingTop": "6px"}),
        ], style={"position": "sticky", "top": 0, "background": "#fafafa", "zIndex": 1, "padding": "8px 12px", "borderBottom": "1px solid #eee"}),

        html.Div(id="symbols_modal", children=[
            html.Div([
                html.Div("Select Symbols", style={"fontWeight": "600", "marginBottom": "8px"}),
                dcc.Dropdown(
                    id="symbols",
                    options=[{"label": s, "value": s} for s in default_symbols],
                    value=default_syms,
                    multi=True,
                    persistence=True,
                    persistence_type="session",
                    style={"width": "100%"},
                ),
                html.Div([
                    html.Button("Select All", id="symbols_select_all", n_clicks=0),
                    html.Button("Deselect All", id="symbols_clear_all", n_clicks=0, style={"marginLeft": "8px"}),
                    html.Button("Close", id="symbols_close", n_clicks=0, style={"marginLeft": "8px"}),
                ], style={"marginTop": "10px"}),
            ], style={"background": "#fff", "padding": "12px", "borderRadius": "8px", "width": "520px", "maxWidth": "90%"}),
        ], style={"position": "fixed", "inset": 0, "display": "none", "alignItems": "center", "justifyContent": "center", "background": "rgba(0,0,0,0.35)", "zIndex": 1000}),

        html.Div(id="charts")
    ], style={"padding": "12px", "background": "#fff"})

    @app.callback(
        Output("symbols_modal", "style"),
        [Input("open_symbols", "n_clicks"), Input("symbols_close", "n_clicks")],
        prevent_initial_call=True,
    )
    def _toggle_symbols_modal(open_clicks, close_clicks):
        ctx = dash.callback_context
        if not ctx.triggered:
            raise dash.exceptions.PreventUpdate
        trig = ctx.triggered[0]["prop_id"].split(".")[0]
        show = trig == "open_symbols"
        return {"position": "fixed", "inset": 0, "display": ("flex" if show else "none"), "alignItems": "center", "justifyContent": "center", "background": "rgba(0,0,0,0.35)", "zIndex": 1000}

    @app.callback(
        [Output("symbols", "options"), Output("symbols", "value")],
        [
            Input("dash_mode", "value"),
            Input("symbols_select_all", "n_clicks"),
            Input("symbols_clear_all", "n_clicks"),
        ],
        [State("symbols", "value")],
    )
    def _update_symbols(mode: str, n_all: int, n_clear: int, current: Optional[List[str]]):
        syms = _symbols_for_mode(mode)
        opts = [{"label": s, "value": s} for s in syms]

        ctx = dash.callback_context
        trig = ctx.triggered[0]["prop_id"].split(".")[0] if ctx.triggered else ""

        if trig == "symbols_select_all":
            return opts, syms
        if trig == "symbols_clear_all":
            return opts, []
        if trig == "dash_mode":
            # Reset selection when switching view
            if mode == "equity":
                return opts, syms[:MAX_VISIBLE_CHARTS]
            return opts, syms[:]

        # Tick / other events: preserve current selection where possible.
        if current:
            filtered = [s for s in current if s in syms]
            if filtered:
                return opts, filtered
        if mode == "equity":
            return opts, syms[:MAX_VISIBLE_CHARTS]
        return opts, syms[:]

    @app.callback(
        [Output("charts", "children"), Output("alerts_panel", "children")],
        [
            Input("dash_mode", "value"),
            Input("symbols", "value"),
            Input("interval", "value"),
            Input("window", "value"),
            Input("indicator", "value"),
            Input("tick", "n_intervals"),
        ],
    )
    def _render(mode: str, selected: List[str], interval: str, window: str, indicator: str, _n: int):
        dashboard_is_commodity = (mode == "commodity")
        states = _states_for_mode(mode)
        all_symbols = list(states.keys())
        children = []

        now_local = now_ist()
        is_market_open = trading_window_open(now_local, is_commodity=dashboard_is_commodity)
        if dashboard_is_commodity:
            show_eod_analysis = now_local.hour >= 23  # commodity: show from 23:00 IST onward
        else:
            show_eod_analysis = (now_local.hour == 15 and now_local.minute >= 30) or now_local.hour > 15

        if show_eod_analysis:
            try:
                filtered_symbols = [sym for sym in all_symbols if (sym in COMMODITY_CONFIG) == dashboard_is_commodity]
                analysis = calculate_eod_analysis(allowed_symbols=filtered_symbols, states=states)
                gl = analysis['global']
                eod_section = html.Div([
                    html.H3("📊 EOD Analysis", style={"marginTop": "20px", "marginBottom": "10px", "color": "#333"}),
                    html.Div([
                        html.Div([html.Strong("Total Realized P&L: "), f"₹{gl['total_realized']:.2f}"], style={"marginBottom": "8px", "fontSize": "16px"}),
                        html.Div([html.Strong("Total Unrealized P&L: "), f"₹{gl['total_unrealized']:.2f}"], style={"marginBottom": "8px", "fontSize": "16px"}),
                        html.Div([
                            html.Strong("Total Net P&L: "),
                            html.Span(f"₹{gl['total_net']:.2f}", style={"color": "#2e7d32" if gl['total_net'] >= 0 else "#c62828", "fontWeight": "bold"}),
                        ], style={"marginBottom": "12px", "fontSize": "18px"}),
                        html.Div([
                            html.Strong("Win Rate: "), f"{gl['win_rate']:.1f}%",
                            html.Span(" | ", style={"margin": "0 8px"}),
                            html.Strong("BUY Win Rate: "), f"{gl['buy_win_rate']:.1f}%",
                            html.Span(" | ", style={"margin": "0 8px"}),
                            html.Strong("SELL Win Rate: "), f"{gl['sell_win_rate']:.1f}%",
                        ], style={"marginBottom": "12px", "fontSize": "14px", "color": "#555"}),
                    ], style={"background": "#f5f5f5", "padding": "12px", "borderRadius": "6px", "marginBottom": "16px"}),
                ], style={"background": "#fff", "padding": "16px", "borderRadius": "8px", "border": "2px solid #4caf50", "marginBottom": "20px"})
                children.append(eod_section)
            except Exception as e:
                logger.debug(f"EOD analysis error: {e}")

        if not is_market_open:
            open_hour, open_minute = (9, 30) if dashboard_is_commodity else (9, 15)
            close_hour = 23 if dashboard_is_commodity else 15
            close_minute = 0 if dashboard_is_commodity else 30
            next_open = now_local.replace(hour=open_hour, minute=open_minute, second=0, microsecond=0)
            cutoff_passed = (now_local.hour > close_hour) or (now_local.hour == close_hour and now_local.minute >= close_minute)
            if cutoff_passed:
                next_open = next_open + timedelta(days=1)
            message = html.Div([
                html.H2("Market is Closed", style={"textAlign": "center", "color": "#666", "marginTop": "100px"}),
                html.P(f"Next trading session starts at {next_open.strftime('%H:%M IST')} on {next_open.strftime('%Y-%m-%d')}",
                       style={"textAlign": "center", "color": "#999", "fontSize": "18px"}),
                html.P(f"Current time: {now_local.strftime('%Y-%m-%d %H:%M:%S IST')}",
                       style={"textAlign": "center", "color": "#999", "fontSize": "14px", "marginTop": "20px"}),
            ], style={"width": "100%", "minHeight": "400px"})
            children.append(message)
        else:
            if not selected:
                selected = (all_symbols[:MAX_VISIBLE_CHARTS] if not dashboard_is_commodity else all_symbols[:])
            if not interval:
                interval = "1m"
            if not window:
                window = "Today"
            if not indicator:
                indicator = "none"
            chart_height = 520
            row: List[html.Div] = []
            for sym in selected:
                try:
                    st = states.get(sym)
                    if not st:
                        continue
                    ohlc = fetch_intraday_ohlc(sym, interval, window)
                    fig = build_symbol_figure(sym, st, ohlc, chart_height, indicator)
                    card_body = dcc.Graph(
                        figure=fig,
                        config={
                            "displayModeBar": True,
                            "scrollZoom": True,
                            "doubleClick": "reset",
                            "responsive": True,
                            "modeBarButtonsToAdd": ["drawline", "drawopenpath", "drawrect", "drawcircle", "eraseshape"],
                        },
                    )
                    events_view = []
                    try:
                        with EVENTS_LOCK:
                            events = list(REALIZED_EVENTS.get(sym, []))
                        if events:
                            for ev in events[::-1]:
                                events_view.append(html.Div(
                                    f"{ev['time']} {ev['event']} {ev['side']} qty {ev['qty']} @ {ev['price']:.2f} | Gross {ev['gross']:.2f} | Net {ev['net']:.2f}",
                                    style={"fontSize": "12px", "color": "#444"}
                                ))
                        else:
                            events_view.append(html.Div("No realized P&L yet", style={"fontSize": "12px", "color": "#888"}))
                    except Exception:
                        events_view = []
                    card = html.Div([
                        html.Div(sym, style={"fontWeight": "600", "marginBottom": "4px"}),
                        dcc.Loading(type="dot", children=card_body),
                        html.Div(events_view, style={"marginTop": "6px", "borderTop": "1px dashed #eee", "paddingTop": "6px"}),
                    ], style={"border": "1px solid #eaeaea", "padding": "8px", "borderRadius": "8px", "marginBottom": "12px", "width": "100%", "boxShadow": "0 1px 2px rgba(0,0,0,0.04)"})
                    row.append(card)
                except Exception as e:
                    err_fig = _empty_figure(f"Error: {str(e)}", chart_height)
                    card_body = dcc.Graph(figure=err_fig, config={"displayModeBar": False})
                    card = html.Div([
                        html.Div(sym, style={"fontWeight": "600", "marginBottom": "4px"}),
                        dcc.Loading(type="dot", children=card_body),
                    ], style={"border": "1px solid #eaeaea", "padding": "8px", "borderRadius": "8px", "marginBottom": "12px", "width": "100%", "boxShadow": "0 1px 2px rgba(0,0,0,0.04)"})
                    row.append(card)
            if row:
                children.extend(row)
            else:
                children.append(html.Div([
                    html.P("No symbols selected. Click 'Symbols' button to select stocks.",
                           style={"textAlign": "center", "color": "#999", "fontSize": "16px", "marginTop": "100px"})
                ]))

        try:
            alerts = get_recent_alerts(limit=15, is_commodity=dashboard_is_commodity)
            alert_children = []
            for ts, txt in alerts[::-1]:
                alert_children.append(html.Div([
                    html.Span(ts.strftime('%H:%M:%S'), style={"color": "#666", "marginRight": "6px"}),
                    html.Span(html.Span(txt.replace('\n', ' | ')))
                ]))
        except Exception:
            alert_children = []
        return children, alert_children

    def _run():
        try:
            logger.info(f"Starting %s on http://0.0.0.0:{port}", app_title)
            app.run(host="0.0.0.0", port=port, debug=False)
        except Exception as e:
            logger.error(f"Dashboard startup failed: {e}")

    th = threading.Thread(target=_run, daemon=True)
    th.start()
    time.sleep(2)

    is_commodity_dash = False  # merged tunnel should use equity bot for notifications
    public_url = start_public_tunnel(port, authtoken=ngrok_authtoken or NGROK_AUTHTOKEN_EQUITY) if use_public_tunnel else None
    if public_url:
        logger.info("Algofinal tunnel: %s (URL sent once by unified_dash_v3.py)", public_url)
        # Telegram notification is handled exclusively by unified_dash_v3.py to avoid duplicate links.


# ----------------------------
# Premarket replay (startup)
# ----------------------------

def fetch_premarket_history_1m(symbol: str) -> Optional[pd.DataFrame]:
    """Fetch today's 1m data and return rows between 09:15–09:30 IST.

    Returns a DataFrame with at least the Close column, index localized to IST.
    """
    try:
        t = yf.Ticker(get_yf_symbol(symbol))
        df = t.history(period="1d", interval="1m")
        if df is None or df.empty:
            return None
        # Ensure tz-aware in IST
        if df.index.tz is None:
            df.index = df.index.tz_localize(pytz.UTC).tz_convert(IST)
        else:
            df.index = df.index.tz_convert(IST)
        start_ts = now_ist().replace(hour=9, minute=15, second=0, microsecond=0)
        end_ts = now_ist().replace(hour=9, minute=30, second=0, microsecond=0)
        mask = (df.index >= start_ts) & (df.index <= end_ts)
        sub = df.loc[mask]
        return sub if not sub.empty else None
    except Exception as e:
        logger.debug(f"premarket history error {symbol}: {e}")
        return None


def replay_premarket_adjustments(state: SymbolState) -> None:
    """If started after 09:30, replay 1m closes between 09:15–09:30 and adjust levels.

    Alerts:
    - If no premarket data found: alert.
    - If falling back to current price due to missing data: alert before fallback.
    - If levels updated (via replay or fallback): alert with updated table.
    """
    if state.is_commodity:
        return
    symbol = state.symbol
    lv = state.levels
    hist = fetch_premarket_history_1m(symbol)
    send_ok = market_open(now_ist())  # send alerts only during 09:15–15:30

    if hist is None:
        # No premarket data; keep initial levels and inform (no fallback adjustment)
        if send_ok:
            base_lv = state.initial_levels or lv
            msg = build_simple_alert(
                "Premarket Replay", symbol,
                "No premarket data found for 09:15–09:30. Using initial levels (no adjustment).",
                base_lv,
                current_price=state.last_price or base_lv.previous_close,
                quantity=(state.qty_remaining if state.in_position else None)
            )
            send_telegram(msg, symbol=symbol)
        return

    # Iterate minute closes in chronological order; use adjust_levels_premarket to simulate shifts
    # Also store closes in shared dict so scanners can read them without yfinance
    if hist is not None and not hist.empty:
        closes = [float(row.get("Close", float("nan"))) for _, row in hist.iterrows()
                  if not math.isnan(float(row.get("Close", float("nan"))))]
        if closes:
            _SHARED_PREMARKET_CLOSES[symbol] = closes

    adjusted_any = False
    for ts, row in hist.iterrows():
        close_px = float(row.get("Close", float("nan")))
        if math.isnan(close_px):
            continue
        new_levels, crossed, _side = adjust_levels_premarket(state.levels, close_px)
        if crossed is not None:
            old_lv = state.levels
            state.levels = new_levels
            adjusted_any = True
            if send_ok:
                compare_msg = build_premarket_adjustment_comparison_alert(
                    symbol, state.initial_levels or old_lv, state.levels, close_px,
                    f"crossed {crossed} at {close_px:.2f}"
                )
                send_telegram(compare_msg, symbol=symbol)

    # If consolidated messaging is desired, it can be added here; kept silent to avoid duplicates


def get_price_at_930(symbol: str, prev_close: float, target_date: Optional[datetime] = None) -> Optional[float]:
    """Get price at 9:30 AM for the specified date (or today if None).
    If price equals previous close, get price at 9:31 instead.
    
    Returns the price at 9:30 (or 9:31 if 9:30 price equals previous close), or None if unavailable.
    For MCX commodities, tries tvdatafeed first for true MCX prices.
    """
    if target_date is None:
        target_date = now_ist()
    target_date = target_date.replace(hour=9, minute=30, second=0, microsecond=0)
    
    # For MCX commodities, try tvdatafeed first
    if is_commodity_symbol(symbol) and TVDATAFEED_AVAILABLE:
        try:
            cfg = get_symbol_config(symbol)
            mcx_sym = cfg.get("mcx_symbol") if cfg else None
            if mcx_sym:
                # Determine how many bars to fetch based on target_date
                # If target_date is today or in the future, fetch 390 bars (1 day)
                # If target_date is yesterday, fetch 780 bars (2 days)
                # If target_date is older, fetch more bars
                now = now_ist()
                days_diff = (now.date() - target_date.date()).days
                if days_diff == 0:
                    n_bars = 390  # Today's data
                elif days_diff == 1:
                    n_bars = 780  # Yesterday + today
                else:
                    n_bars = min(390 * (days_diff + 1), 2000)  # Multiple days, cap at 2000 bars
                
                # Fetch 1-minute bars
                df = tv_datafeed.get_hist(symbol=mcx_sym, exchange="MCX", interval=TVInterval.in_1_minute, n_bars=n_bars)
                if df is not None and not df.empty:
                    # Convert to IST if needed
                    if df.index.tz is None:
                        df.index = df.index.tz_localize(IST)
                    else:
                        df.index = df.index.tz_convert(IST)
                    # Get price at 9:30 for target_date
                    mask_930 = (df.index >= target_date) & (df.index < target_date + timedelta(minutes=1))
                    df_930 = df.loc[mask_930]
                    if not df_930.empty:
                        price_930 = float(df_930["close"].iloc[-1])
                        # If price at 9:30 equals previous close, get price at 9:31
                        if abs(price_930 - prev_close) < 0.01:
                            ts_931 = target_date + timedelta(minutes=1)
                            mask_931 = (df.index >= ts_931) & (df.index < ts_931 + timedelta(minutes=1))
                            df_931 = df.loc[mask_931]
                            if not df_931.empty:
                                return float(df_931["close"].iloc[-1])
                        return price_930
        except Exception as e:
            logger.debug(f"tvdatafeed 9:30 price error for {symbol}: {e}")
    
    # Fallback to Yahoo Finance
    try:
        t = yf.Ticker(get_yf_symbol(symbol))
        df = t.history(period="1d", interval="1m")
        if df is None or df.empty:
            return None
        # Ensure tz-aware in IST
        if df.index.tz is None:
            df.index = df.index.tz_localize(pytz.UTC).tz_convert(IST)
        else:
            df.index = df.index.tz_convert(IST)
        
        # Get price at 9:30 for target_date
        mask_930 = (df.index >= target_date) & (df.index < target_date + timedelta(minutes=1))
        df_930 = df.loc[mask_930]
        
        if df_930.empty:
            return None
        
        price_930 = float(df_930["Close"].iloc[-1])
        
        # If price at 9:30 equals previous close, get price at 9:31
        if abs(price_930 - prev_close) < 0.01:  # Using small epsilon for float comparison
            ts_931 = target_date + timedelta(minutes=1)
            mask_931 = (df.index >= ts_931) & (df.index < ts_931 + timedelta(minutes=1))
            df_931 = df.loc[mask_931]
            if not df_931.empty:
                return float(df_931["Close"].iloc[-1])
            # If 9:31 data not available, return 9:30 price (will result in no adjustment)
            return price_930
        
        return price_930
    except Exception as e:
        logger.debug(f"get_price_at_930 error {symbol}: {e}")
        return None


def _build_price_ladder(levels: Levels) -> List[Tuple[str, float]]:
    """Return all known ladder rungs sorted in ascending price order."""
    ladder = [
        ("ST5", levels.st[4]),
        ("ST4", levels.st[3]),
        ("ST3", levels.st[2]),
        ("ST2", levels.st[1]),
        ("ST1", levels.st[0]),
        ("SELL_BELOW", levels.sell_below),
        ("PREV_CLOSE", levels.previous_close),
        ("BUY_ABOVE", levels.buy_above),
        ("T1", levels.t[0]),
        ("T2", levels.t[1]),
        ("T3", levels.t[2]),
        ("T4", levels.t[3]),
        ("T5", levels.t[4]),
    ]
    return sorted(ladder, key=lambda item: item[1])


def _find_ladder_bounds(levels: Levels, price: float, step: float) -> Tuple[Tuple[str, float], Tuple[str, float]]:
    """Find the two consecutive ladder rungs that bracket the given price."""
    rungs = _build_price_ladder(levels)
    if not rungs:
        raise ValueError("Price ladder is empty; cannot determine bounds.")
    if price <= rungs[0][1]:
        lower = ("EXT_LOW", rungs[0][1] - step)
        upper = rungs[0]
        return lower, upper
    if price >= rungs[-1][1]:
        lower = rungs[-1]
        upper = ("EXT_HIGH", rungs[-1][1] + step)
        return lower, upper
    prev_rung = rungs[0]
    for current in rungs[1:]:
        if price <= current[1]:
            return prev_rung, current
        prev_rung = current
    return rungs[-2], rungs[-1]


def _reanchor_levels_around_price(levels: Levels, price: float) -> Tuple[Optional[Levels], Optional[Tuple[Tuple[str, float], Tuple[str, float]]]]:
    """Return new Levels re-centered so price sits between sell_below/buy_above."""
    step = getattr(levels, "target_step", None) or levels.x or 1.0
    lower_rung, upper_rung = _find_ladder_bounds(levels, price, step)
    lower_value = lower_rung[1]
    upper_value = upper_rung[1]
    if math.isclose(lower_value, levels.sell_below, rel_tol=0, abs_tol=1e-6) and math.isclose(upper_value, levels.buy_above, rel_tol=0, abs_tol=1e-6):
        return None, None
    new_levels = deepcopy(levels)
    new_levels.target_step = step
    new_levels.buy_above = upper_value
    for idx in range(5):
        new_levels.t[idx] = upper_value + step * (idx + 1)
    new_levels.buy_sl = new_levels.buy_above - levels.x
    new_levels.sell_below = lower_value
    for idx in range(5):
        new_levels.st[idx] = lower_value - step * (idx + 1)
    new_levels.sell_sl = new_levels.sell_below + levels.x
    return new_levels, (lower_rung, upper_rung)


def adjust_levels_at_930(state: SymbolState) -> None:
    """After premarket adjustments, re-anchor the ladder so 9:30 price sits between sell/buy."""
    if state.is_commodity:
        return
    symbol = state.symbol
    lv = state.levels
    prev_close = lv.previous_close
    x = lv.x
    step = getattr(lv, 'target_step', x)
    
    price_930 = get_price_at_930(symbol, prev_close)
    if price_930 is None:
        logger.debug(f"Could not get 9:30 price for {symbol}, skipping 9:30 adjustment")
        return

    # Store for scanners to read — eliminates their yfinance 09:30 calls
    _SHARED_930_PRICES[symbol] = price_930
    
    send_ok = market_open(now_ist())
    ladder_result, rung_bounds = _reanchor_levels_around_price(state.levels, price_930)
    if not ladder_result or not rung_bounds:
        return
    old_lv = deepcopy(state.levels)
    state.levels = ladder_result
    if send_ok:
        (low_label, low_value), (high_label, high_value) = rung_bounds
        status_msg = (
            f"9:30 price {price_930:.2f} between {low_label} ({low_value:.2f}) "
            f"and {high_label} ({high_value:.2f}). Re-anchored ladder."
        )
        compare_msg = build_premarket_adjustment_comparison_alert(
            symbol, old_lv, state.levels, price_930, status_msg
        )
        send_telegram(compare_msg, symbol=symbol)


def refresh_commodity_levels_from_930(state: SymbolState) -> None:
    """Recalculate commodity levels using the confirmed 9:30 price."""
    if not state.is_commodity or not state.pending_930_recalc:
        return
    now = now_ist()
    if now.hour < 9 or (now.hour == 9 and now.minute < 30):
        return
    prev_close = state.prev_close or state.levels.previous_close
    if not state.x_multiplier:
        state.pending_930_recalc = False
        return
    price_930 = get_price_at_930(state.symbol, prev_close)
    if price_930 is None:
        return
    x_value = price_930 * state.x_multiplier
    new_levels = calc_levels_for_symbol(
        state.symbol,
        price_930,
        x_override=x_value,
        step_override=x_value,
    )
    state.levels = new_levels
    state.initial_levels = new_levels
    state.pending_930_recalc = False
    logger.info("%s commodity levels refreshed using 9:30 price %.2f", state.symbol, price_930)


def build_commodity_state(symbol: str) -> Optional[SymbolState]:
    cfg = COMMODITY_CONFIG.get(symbol)
    if not cfg:
        return None
    multiplier = cfg.get("x_multiplier")
    if not isinstance(multiplier, (int, float)) or multiplier <= 0:
        logger.warning("Invalid commodity multiplier for %s; skipping.", symbol)
        return None
    pc = get_prev_close(symbol)
    if pc is None:
        # Some MCX symbols don't have reliable daily history. Fall back
        # to live price so we can still run levels and alerts.
        live_px = get_live_price(symbol)
        if live_px is None:
            logger.warning("Skipping %s: previous close/live price unavailable for commodity setup", symbol)
            return None
        pc = live_px
    # For commodities: if started after 9:30 AM today, use today's 9:30 price
    # Otherwise (before 9:30 AM or after 23:00), use previous day's 9:30 price
    now = now_ist()
    target_date = None
    # Check if we're in commodity trading hours (9:30-23:00)
    is_commodity_hours = (now.hour > 9 or (now.hour == 9 and now.minute >= 30)) and now.hour < 23
    if not is_commodity_hours:
        # Outside commodity trading hours - use previous trading day's 9:30 price
        target_date = now - timedelta(days=1)
        target_date = target_date.replace(hour=9, minute=30, second=0, microsecond=0)
    # If in commodity hours, target_date=None means get_price_at_930 will use today's 9:30
    price_930 = get_price_at_930(symbol, pc, target_date=target_date)
    if price_930 is None:
        logger.info(
            "%s 9:30 price unavailable from history; using current/base price only for initial levels.",
            symbol,
        )
    base_price = price_930 or pc
    x_value = base_price * multiplier
    levels = calc_levels_for_symbol(symbol, base_price, x_override=x_value, step_override=x_value)
    return SymbolState(
        symbol=symbol,
        levels=levels,
        initial_levels=levels,
        is_commodity=True,
        x_multiplier=multiplier,
        pending_930_recalc=(price_930 is None),
        adjusted_locked=True,
        prev_close=pc,
    )


def ensure_commodities_ready(states: Dict[str, SymbolState]) -> None:
    """Ensure every configured commodity has a SymbolState before exports."""
    for sym in COMMODITY_CONFIG.keys():
        if sym in states:
            continue
        state = build_commodity_state(sym)
        if state:
            states[sym] = state


def replay_commodity_session(states: Dict[str, SymbolState]) -> None:
    """Replay today's commodity session using 1m data to reconstruct entries/exits.

    This uses NOW_IST_OVERRIDE so that existing logic (entries, exits, alerts)
    sees the historical timestamps as if they were live.
    When started after 9:30 AM, replays from 9:30 AM to current time, firing all alerts.
    """
    global NOW_IST_OVERRIDE
    now_local = now_ist()
    # Only meaningful after 09:30 and before midnight.
    if now_local.hour < 9 or (now_local.hour == 9 and now_local.minute < 30):
        return
    
    start_930 = now_local.replace(hour=9, minute=30, second=0, microsecond=0)
    # If we're already past 9:30, replay from 9:30 to now
    if now_local <= start_930:
        return
    
    logger.info("Starting commodity session replay from 9:30 AM to current time...")
    partial_exits_replay: List[dict] = []
    
    for sym, st in states.items():
        if not st.is_commodity:
            continue
        try:
            ohlc = fetch_intraday_ohlc(sym, "1m", "Today")
            if ohlc is None or ohlc.empty:
                continue
            # Ensure chronological order
            ohlc = ohlc.sort_index()
            # Process each minute from 9:30 to current time
            for ts, row in ohlc.iterrows():
                if ts < start_930 or ts >= now_local:
                    continue
                price = float(row.get("Close", float("nan")))
                if math.isnan(price):
                    continue
                NOW_IST_OVERRIDE = ts
                st.last_price = price
                append_price_history(sym, ts, price)
                # Apply the same intraday logic as the live loop for commodities
                if trading_window_open(ts, is_commodity=True):
                    process_reentry_watch(st, price)
                    try_entry(st, price)
                    handle_targets_and_trailing(st, price, partial_exits_replay)
        except Exception as exc:
            logger.warning("Commodity replay failed for %s: %s", sym, exc)
    
    NOW_IST_OVERRIDE = None
    logger.info("Commodity session replay completed. Switching to live monitoring.")


# ----------------------------
# Core monitoring
# ----------------------------

def build_table(states: Dict[str, SymbolState]) -> Table:
    has_commodity = any(st.is_commodity for st in states.values())
    title = "Live NSE Levels" + (" + MCX Commodities" if has_commodity else "")
    table = Table(title=title, show_lines=False)
    table.add_column("Symbol", justify="left")
    table.add_column("Price", justify="right")
    table.add_column("BuyAbove", justify="right")
    table.add_column("T1", justify="right")
    table.add_column("T2", justify="right")
    table.add_column("T3", justify="right")
    table.add_column("T4", justify="right")
    table.add_column("T5", justify="right")
    table.add_column("Buy SL", justify="right")
    table.add_column("SellBelow", justify="right")
    table.add_column("ST1", justify="right")
    table.add_column("ST2", justify="right")
    table.add_column("ST3", justify="right")
    table.add_column("ST4", justify="right")
    table.add_column("ST5", justify="right")
    table.add_column("Sell SL", justify="right")
    table.add_column("Pos", justify="left")
    if has_commodity:
        table.add_column("Src", justify="left", no_wrap=True)

    # Price source for commodity symbols (shown in last column)
    src_status = get_commodity_price_source_status() if has_commodity else {}
    # Abbreviate source names for compact display
    _src_abbrev = {
        "tradingview_ws": "TV-WS",
        "tvdatafeed":     "TV-REST",
        "investing_com":  "Investing",
        "moneycontrol":   "MC",
        "cache":          "Cache",
    }

    for sym, st in states.items():
        lv = st.levels
        pos = "-"
        if st.in_position and st.side:
            base = f"{st.side} {st.qty_remaining}"
            if st.retreat_phase == "warned_65":
                base += " ⚠65%"
            elif st.retreat_phase == "activated_45":
                base += " 🔔45%"
            pos = base
        row = [
            sym,
            f"{(st.last_price or 0):.2f}",
            f"{lv.buy_above:.2f}",
            f"{lv.t[0]:.2f}", f"{lv.t[1]:.2f}", f"{lv.t[2]:.2f}", f"{lv.t[3]:.2f}", f"{lv.t[4]:.2f}",
            f"{lv.buy_sl:.2f}",
            f"{lv.sell_below:.2f}",
            f"{lv.st[0]:.2f}", f"{lv.st[1]:.2f}", f"{lv.st[2]:.2f}", f"{lv.st[3]:.2f}", f"{lv.st[4]:.2f}",
            f"{lv.sell_sl:.2f}",
            pos,
        ]
        if has_commodity:
            raw_src = src_status.get(sym.upper(), "—" if st.is_commodity else "")
            row.append(_src_abbrev.get(raw_src, raw_src))
        table.add_row(*row)
    return table


def maybe_alert_once(state: SymbolState, key: str, text: str) -> None:
    if key in state.sent_events:
        return
    now_local = now_ist()
    # Between 09:15–09:30, only allow premarket level adjustment alerts.
    # Identify them by key prefix or message title text.
    if premarket_window(now_local):
        is_premarket_adjust = key.startswith("premkt_adjust_") or ("Premarket" in text and "Adjustment" in text)
        if not is_premarket_adjust:
            return
    if alerts_allowed(now_local, symbol=state.symbol):
        send_telegram(text, symbol=state.symbol)
        state.sent_events.add(key)


def _dynamic_quantity(price: float) -> int:
    try:
        p = float(price)
    except Exception:
        return 0
    if p <= 0:
        return 0
    return int(100000 // p)


def _open_entry(
    state: SymbolState,
    *,
    side: str,
    price: float,
    status_line: str,
    hit_level: str,
    alert_key: Optional[str],
    custom_sl: Optional[float] = None,
    custom_target: Optional[float] = None,
    custom_target_label: Optional[str] = None,
    quantity: Optional[int] = None,
    event_type: Optional[str] = None,
    dedupe_alert: bool = True,
) -> None:
    now = now_ist()
    dyn_qty = quantity or _dynamic_quantity(price)
    if dyn_qty <= 0:
        # Keep alert formatting identical to normal entries, but indicate that the
        # entry was skipped due to computed quantity being zero.
        skipped_status = f"{status_line} | SKIPPED (Qty=0; Budget=100000)"
        skipped_msg = build_simple_alert(
            "Entry",
            state.symbol,
            skipped_status,
            state.levels,
            current_price=price,
            hit_level=hit_level,
            quantity=0,
        )
        if alert_key and dedupe_alert:
            maybe_alert_once(state, f"{alert_key}_skipped_qty0", skipped_msg)
        else:
            if alerts_allowed(now, symbol=state.symbol):
                send_telegram(skipped_msg, symbol=state.symbol)
        return
    state.in_position = True
    state.side = side
    state.entry_price = price
    state.entry_time = now
    state.qty_remaining = dyn_qty
    state.qty_total = dyn_qty
    state.last_target_hit_index = -1
    state.exited_today = False
    state.reentry_watch = None
    state.manual_target = custom_target
    state.manual_target_label = custom_target_label
    # Capture the level crossed at entry for retreat % calculations
    if side == "BUY":
        state.retreat_entry_level = state.levels.buy_above
        state.buy_trailing_sl = custom_sl if custom_sl is not None else state.levels.buy_sl
        state.sell_trailing_sl = None
    else:
        state.retreat_entry_level = state.levels.sell_below
        state.sell_trailing_sl = custom_sl if custom_sl is not None else state.levels.sell_sl
        state.buy_trailing_sl = None
    # Reset retreat monitoring state for fresh entry
    state.retreat_phase = None
    state.retreat_65_alerted = False
    state.retreat_45_alerted = False
    state.retreat_peak_reached = False

    msg = build_simple_alert(
        "Entry",
        state.symbol,
        status_line,
        state.levels,
        current_price=price,
        hit_level=hit_level,
        quantity=dyn_qty,
    )
    if alert_key and dedupe_alert:
        maybe_alert_once(state, alert_key, msg)
    else:
        send_telegram(msg, symbol=state.symbol)
    log_trade_event(
        now,
        state.symbol,
        event_type or f"{side}_ENTRY",
        price,
        side,
        dyn_qty,
        entry_price=price,  # FIX: entry_price was None before — set it to the fill price
    )


def schedule_reentry_watch(state: SymbolState, side: str, target_label: str, target_price: float) -> None:
    """Arm post-target monitoring for threshold/retouch re-entries."""
    # Re-entry monitoring: after a full target/exit, watch for either:
    # - continuation (threshold beyond the target) => re-enter in same direction
    # - retouch (price returns to the target)      => re-enter opposite direction
    if state.is_commodity and state.pending_930_recalc:
        logger.debug("Skipping reentry watch for %s until 9:30 levels are ready.", state.symbol)
        return
    lv = state.levels
    step = getattr(lv, "target_step", lv.x)
    if step <= 0:
        return
    buffer = 0.75 * lv.x
    if buffer <= 0:
        return

    if side == "BUY":
        threshold_price = target_price + buffer
        watch = ReentryWatch(
            original_side=side,
            target_label=target_label,
            target_price=target_price,
            threshold_price=threshold_price,
            threshold_direction="ABOVE",
            threshold_entry_side="BUY",
            threshold_entry_price=target_price + step,
            threshold_entry_sl=target_price,
            threshold_entry_target=target_price + step + step,
            retouch_entry_side="SELL",
            retouch_entry_price=target_price,
            retouch_entry_sl=target_price + step,
            retouch_entry_target=target_price - step,
            step=step,
            buffer=buffer,
        )
    else:
        threshold_price = target_price - buffer
        watch = ReentryWatch(
            original_side=side,
            target_label=target_label,
            target_price=target_price,
            threshold_price=threshold_price,
            threshold_direction="BELOW",
            threshold_entry_side="SELL",
            threshold_entry_price=target_price - step,
            threshold_entry_sl=target_price,
            threshold_entry_target=target_price - step - step,
            retouch_entry_side="BUY",
            retouch_entry_price=target_price,
            retouch_entry_sl=target_price - step,
            retouch_entry_target=target_price + step,
            step=step,
            buffer=buffer,
        )

    state.reentry_watch = watch

    if alerts_allowed(now_ist(), symbol=state.symbol):
        asset_type = "🏅 Commodity" if state.is_commodity else "📊 Equity"
        lsp_note = (
            f"LSP={target_price:.2f} | "
            f"New BA (closest above LSP)={watch.threshold_entry_price:.2f} | "
            f"New SB (closest below LSP)={watch.retouch_entry_price:.2f}"
        )
        lines = [
            f"🔁 {state.symbol} [{asset_type}] — Re-entry monitoring armed after {target_label}",
            "",
            lsp_note,
            f"▶ Threshold continuation: price ≥ {watch.threshold_price:.2f}",
            f"  → {watch.threshold_entry_side} @ {watch.threshold_entry_price:.2f}  SL {watch.threshold_entry_sl:.2f}  Tgt {watch.threshold_entry_target:.2f}",
            f"▶ Retouch reversal: price {'≤' if side == 'BUY' else '≥'} {watch.target_price:.2f}",
            f"  → {watch.retouch_entry_side} @ {watch.retouch_entry_price:.2f}  SL {watch.retouch_entry_sl:.2f}  Tgt {watch.retouch_entry_target:.2f}",
        ]
        send_telegram("\n".join(lines), symbol=state.symbol)


def _execute_reentry(
    state: SymbolState,
    *,
    mode: str,
    entry_side: str,
    entry_price: float,
    sl: float,
    target: float,
    alert_label: str,
    target_label: str,
) -> None:
    status = f"{mode.upper()} RE-ENTRY {entry_side} @ {entry_price:.2f} | SL {sl:.2f} | Target {target:.2f}"
    _open_entry(
        state,
        side=entry_side,
        price=entry_price,
        status_line=status,
        hit_level=alert_label,
        alert_key=None,
        custom_sl=sl,
        custom_target=target,
        custom_target_label=target_label,
        dedupe_alert=False,
        event_type=f"{entry_side}_REENTRY_{mode.upper()}",
    )


def process_reentry_watch(state: SymbolState, price: float) -> None:
    watch = state.reentry_watch
    if not watch or state.in_position:
        return
    if state.is_commodity and state.pending_930_recalc:
        return
    if not alerts_allowed(now_ist(), symbol=state.symbol):
        return
    # ── Blackout: no re-entries 09:30–09:35 either ────────────────────────────
    if not state.is_commodity and in_930_blackout(now_ist()):
        return

    triggered = False
    trigger_mode = ""
    entry_side = ""
    entry_price = 0.0
    sl = 0.0
    target = 0.0

    if watch.threshold_direction == "ABOVE" and price >= watch.threshold_price:
        triggered = True
        trigger_mode = "threshold"
        entry_side = watch.threshold_entry_side
        entry_price = watch.threshold_entry_price
        sl = watch.threshold_entry_sl
        target = watch.threshold_entry_target
    elif watch.threshold_direction == "BELOW" and price <= watch.threshold_price:
        triggered = True
        trigger_mode = "threshold"
        entry_side = watch.threshold_entry_side
        entry_price = watch.threshold_entry_price
        sl = watch.threshold_entry_sl
        target = watch.threshold_entry_target

    if triggered:
        state.reentry_watch = None
        _execute_reentry(
            state,
            mode="threshold",
            entry_side=entry_side,
            entry_price=entry_price,
            sl=sl,
            target=target,
            alert_label=f"{watch.target_label}_REENTRY_THRESHOLD",
            target_label=f"{watch.target_label}_THRESHOLD_TARGET",
        )
        return

    # Retouch logic
    if watch.original_side == "BUY" and price <= watch.target_price:
        state.reentry_watch = None
        _execute_reentry(
            state,
            mode="retouch",
            entry_side=watch.retouch_entry_side,
            entry_price=watch.retouch_entry_price,
            sl=watch.retouch_entry_sl,
            target=watch.retouch_entry_target,
            alert_label=f"{watch.target_label}_REENTRY_RETOUCH",
            target_label=f"{watch.target_label}_RETOUCH_TARGET",
        )
    elif watch.original_side == "SELL" and price >= watch.target_price:
        state.reentry_watch = None
        _execute_reentry(
            state,
            mode="retouch",
            entry_side=watch.retouch_entry_side,
            entry_price=watch.retouch_entry_price,
            sl=watch.retouch_entry_sl,
            target=watch.retouch_entry_target,
            alert_label=f"{watch.target_label}_REENTRY_RETOUCH",
            target_label=f"{watch.target_label}_RETOUCH_TARGET",
        )


def handle_premarket_adjustment(state: SymbolState, price: float) -> None:
    if state.is_commodity:
        return
    before = state.levels
    adjusted, crossed, _ = adjust_levels_premarket(before, price)
    if crossed is not None and not state.adjusted_locked:
        state.levels = adjusted
        msg = build_premarket_adjustment_comparison_alert(
            state.symbol,
            state.initial_levels or before,
            state.levels,
            price,
            f"crossed {crossed} at {price:.2f}"
        )
        maybe_alert_once(state, f"premkt_adjust_{crossed}", msg)


# ── 65 / 45 / 25 % RETREAT MONITORING ────────────────────────────────────────

def _build_retreat_alert(
    state: SymbolState,
    pct_label: str,
    price: float,
    threshold: float,
    step: float,
) -> str:
    """Build a Telegram message for a retreat threshold event.
    Works for both equity and commodity symbols — routes to correct bot via symbol=.
    """
    ts = now_ist().strftime("%Y-%m-%d %H:%M:%S IST%z")
    entry_level = state.retreat_entry_level
    lv = state.levels
    entry_px = state.entry_price or entry_level
    profit_pts = abs(price - entry_px)
    profit_pct = (profit_pts / step * 100.0) if step > 0 else 0.0
    qty = state.qty_remaining

    # Estimate gross P&L if exit happens at threshold (25%) level
    if "25%" in pct_label:
        if state.side == "BUY":
            est_gross = (threshold - entry_level) * qty   # br = (x×0.25)×q
        else:
            est_gross = (entry_level - threshold) * qty   # sr = (x×0.25)×q
        est_net = est_gross - BROKERAGE_FLAT_PER_SIDE * 2
        pnl_hint = f"Est Exit P&L  : Gross ₹{est_gross:.2f} | Net ₹{est_net:.2f}"
    else:
        pnl_hint = f"Remaining P&L : {profit_pct:.1f}% of step ({profit_pts:.2f} pts)"

    asset_type = "🏅 Commodity" if state.is_commodity else "📊 Equity"
    lines = [
        f"🔔 {state.symbol} [{asset_type}] — Retreat {pct_label} at {ts}",
        "",
        f"Side          : {state.side}",
        f"Entry Level   : {entry_level:.2f}",
        f"Entry Price   : {entry_px:.2f}",
        f"Current Price : {price:.2f}",
        f"Threshold Px  : {threshold:.2f}  ({pct_label} of step)",
        f"Step (x)      : {step:.2f}",
        pnl_hint,
        f"Qty           : {qty}",
        f"Buy Above     : {lv.buy_above:.2f}  |  Sell Below : {lv.sell_below:.2f}",
        f"Buy SL        : {lv.buy_sl:.2f}  |  Sell SL    : {lv.sell_sl:.2f}",
    ]
    return "\n".join(lines)


def _reanchor_after_retreat_exit(state: SymbolState, side: str) -> None:
    """
    Re-anchor levels after a 25% retreat exit.

    BUY retreat exit  → T1 becomes new buy_above, old buy_above → new sell_below.
    SELL retreat exit → ST1 becomes new sell_below, old sell_below → new buy_above.

    All other levels rebuilt from the new anchors using the same step.
    exited_today is cleared so the re-anchored levels can trigger a new trade.
    """
    lv = state.levels
    step = getattr(lv, "target_step", lv.x)
    x    = lv.x

    if side == "BUY":
        new_buy_above  = lv.t[0]               # old T1
        new_sell_below = lv.buy_above           # old buy_above
    else:  # SELL
        new_buy_above  = lv.sell_below          # old sell_below
        new_sell_below = lv.st[0]              # old ST1

    new_t  = [new_buy_above  + step * (i + 1) for i in range(5)]
    new_st = [new_sell_below - step * (i + 1) for i in range(5)]

    new_levels = deepcopy(lv)
    new_levels.buy_above  = new_buy_above
    new_levels.sell_below = new_sell_below
    new_levels.t          = new_t
    new_levels.st         = new_st
    new_levels.buy_sl     = new_buy_above  - x
    new_levels.sell_sl    = new_sell_below + x

    state.levels          = new_levels
    state.adjusted_locked = True   # keep re-anchored levels stable

    # Reset all position state
    state.in_position          = False
    state.side                 = None
    state.entry_price          = None
    state.entry_time           = None
    state.qty_remaining        = 0
    state.qty_total            = 0
    state.last_target_hit_index = -1
    state.buy_trailing_sl      = None
    state.sell_trailing_sl     = None
    state.manual_target        = None
    state.manual_target_label  = None
    state.reentry_watch        = None
    # Allow immediate re-entry at the new levels
    state.exited_today         = False
    # Reset retreat state for the next trade
    state.retreat_phase        = None
    state.retreat_65_alerted   = False
    state.retreat_45_alerted   = False
    state.retreat_entry_level  = 0.0

    logger.info(
        "%s retreat re-anchor (%s): buy_above=%.2f  sell_below=%.2f",
        state.symbol, side, new_buy_above, new_sell_below,
    )


def handle_retreat_monitoring(state: SymbolState, price: float) -> None:
    """
    65 / 45 / 25 % retreat monitoring — called every tick when in_position.

    FIX (Bug 1 — instant retreat):
      Retreat ONLY triggers after price has first moved ≥65% of one step toward
      profit (retreat_peak_reached=True). This prevents the 25%-exit from firing
      the same tick the position opens, which was causing gross P&L = 0 on every
      retreat exit (instead of the correct x×0.25×q per your formula).

    Your P&L formulas:
      Buy retreat  br = (x × 0.25) × q  →  exit at entry_level + 0.25×step
      Sell retreat sr = (x × 0.25) × q  →  exit at entry_level − 0.25×step

    Priority:
      1. T1–T5 targets are checked BEFORE this (in handle_targets_and_trailing).
      2. SL check follows this function in handle_targets_and_trailing.
    """
    if not state.in_position or not state.side:
        return
    if state.is_commodity and state.pending_930_recalc:
        return
    now = now_ist()
    if not alerts_allowed(now, symbol=state.symbol):
        return

    lv    = state.levels
    step  = getattr(lv, "target_step", lv.x)
    if step <= 0:
        return

    entry_level = state.retreat_entry_level
    if entry_level <= 0:
        return

    if state.side == "BUY":
        # Retreat levels (measured from buy_above upward toward T1)
        lvl_65 = entry_level + 0.65 * step   # 65% of step above buy_above
        lvl_45 = entry_level + 0.45 * step   # 45%
        lvl_25 = entry_level + 0.25 * step   # 25% — retreat exit here
        # br = (x × 0.25) × q when exit at lvl_25 and entry_price ≈ entry_level

        # T1 zone: price at/above T1 — in full profit zone, clear retreat state
        if price >= lv.t[0]:
            state.retreat_phase       = None
            state.retreat_65_alerted  = False
            state.retreat_45_alerted  = False
            state.retreat_peak_reached = False
            return

        # ── GUARD: only activate retreat after price has been in good zone ────
        # Price at or above 65% line → mark "peak reached", clear any stale phase
        if price >= lvl_65:
            state.retreat_peak_reached = True
            if state.retreat_phase is not None:
                state.retreat_phase      = None
                state.retreat_65_alerted = False
                state.retreat_45_alerted = False
            return

        # Price has never reached 65% zone since entry → not a real retreat yet
        if not state.retreat_peak_reached:
            return

        # ── price is below 65% line AND has previously been above it ─────────
        if not state.retreat_65_alerted:
            state.retreat_65_alerted = True
            state.retreat_phase      = "warned_65"
            msg = _build_retreat_alert(state, "65% WARNING", price, lvl_65, step)
            send_telegram(msg, symbol=state.symbol)

        if price <= lvl_45 and not state.retreat_45_alerted:
            state.retreat_45_alerted = True
            state.retreat_phase      = "activated_45"
            msg = _build_retreat_alert(state, "45% ACTIVATED", price, lvl_45, step)
            send_telegram(msg, symbol=state.symbol)

        # ── 25% EXIT — br = (x × 0.25) × q  ─────────────────────────────────
        if price <= lvl_25 and state.retreat_45_alerted:
            qty       = state.qty_remaining
            # FORMULA (per spec): gross = x * 0.25 * qty  (always positive)
            # x = lv.x (the deviation amount in Rs for this symbol)
            # This is the correct retreat exit P&L regardless of actual fill price.
            # Using (price - entry_level)*qty can give negative results when price
            # has dropped below entry_level by the time retreat fires.
            gross     = lv.x * 0.25 * qty        # always positive per spec
            brokerage = BROKERAGE_FLAT_PER_SIDE * 2
            net       = gross - brokerage
            record_realized_event(state.symbol, "BUY", "RETREAT_25PCT", price, qty, gross, net)
            status = (
                f"RETREAT 25% EXIT BUY | Price={price:.2f} "
                f"| EntryLvl={entry_level:.2f} | Lvl25={lvl_25:.2f} "
                f"| Step={step:.2f} | Gross Rs{gross:.2f} | Net Rs{net:.2f}"
            )
            msg = build_simple_alert(
                "Exit", state.symbol, status, lv,
                current_price=price, hit_level="RETREAT_25PCT", quantity=qty,
            )
            send_telegram(msg, symbol=state.symbol)
            log_trade_event(now, state.symbol, "BUY_RETREAT_25PCT", price, "BUY", qty,
                            entry_price=entry_level)    # use entry_level for correct dashboard P&L
            _reanchor_after_retreat_exit(state, "BUY")

    elif state.side == "SELL":
        # Retreat levels (measured from sell_below downward toward ST1)
        lvl_65 = entry_level - 0.65 * step   # 65% of step below sell_below
        lvl_45 = entry_level - 0.45 * step   # 45%
        lvl_25 = entry_level - 0.25 * step   # 25% — retreat exit here
        # sr = (x × 0.25) × q when exit at lvl_25 and entry_price ≈ entry_level

        # ST1 zone: price at/below ST1 — in full profit zone, clear retreat state
        if price <= lv.st[0]:
            state.retreat_phase       = None
            state.retreat_65_alerted  = False
            state.retreat_45_alerted  = False
            state.retreat_peak_reached = False
            return

        # ── GUARD: only activate retreat after price moved into profit zone ───
        # Price at or below 65% line (toward ST1) → mark "peak reached"
        if price <= lvl_65:
            state.retreat_peak_reached = True
            if state.retreat_phase is not None:
                state.retreat_phase      = None
                state.retreat_65_alerted = False
                state.retreat_45_alerted = False
            return

        # Price has never reached 65% zone since entry → not a real retreat yet
        if not state.retreat_peak_reached:
            return

        # ── price is above 65% line AND has previously been below it ─────────
        if not state.retreat_65_alerted:
            state.retreat_65_alerted = True
            state.retreat_phase      = "warned_65"
            msg = _build_retreat_alert(state, "65% WARNING", price, lvl_65, step)
            send_telegram(msg, symbol=state.symbol)

        if price >= lvl_45 and not state.retreat_45_alerted:
            state.retreat_45_alerted = True
            state.retreat_phase      = "activated_45"
            msg = _build_retreat_alert(state, "45% ACTIVATED", price, lvl_45, step)
            send_telegram(msg, symbol=state.symbol)

        # ── 25% EXIT — sr = (x × 0.25) × q  ─────────────────────────────────
        if price >= lvl_25 and state.retreat_45_alerted:
            qty       = state.qty_remaining
            # FORMULA (per spec): gross = x * 0.25 * qty  (always positive)
            # x = lv.x (the deviation amount in Rs for this symbol)
            # This is the correct retreat exit P&L regardless of actual fill price.
            # Using (entry_level - price)*qty can give negative results when price
            # has risen above entry_level by the time retreat fires.
            gross     = lv.x * 0.25 * qty        # always positive per spec
            brokerage = BROKERAGE_FLAT_PER_SIDE * 2
            net       = gross - brokerage
            record_realized_event(state.symbol, "SELL", "RETREAT_25PCT", price, qty, gross, net)
            status = (
                f"RETREAT 25% EXIT SELL | Price={price:.2f} "
                f"| EntryLvl={entry_level:.2f} | Lvl25={lvl_25:.2f} "
                f"| Step={step:.2f} | Gross Rs{gross:.2f} | Net Rs{net:.2f}"
            )
            msg = build_simple_alert(
                "Exit", state.symbol, status, lv,
                current_price=price, hit_level="RETREAT_25PCT", quantity=qty,
            )
            send_telegram(msg, symbol=state.symbol)
            log_trade_event(now, state.symbol, "SELL_RETREAT_25PCT", price, "SELL", qty,
                            entry_price=entry_level)    # use entry_level for correct dashboard P&L
            _reanchor_after_retreat_exit(state, "SELL")


def try_entry(state: SymbolState, price: float) -> None:
    if state.in_position:
        return
    if state.is_commodity and state.pending_930_recalc:
        return
    lv = state.levels
    now = now_ist()
    # Prevent same-day re-entry after a full exit
    today_key = now.strftime("%Y%m%d")
    if state.trade_date != today_key:
        state.trade_date = today_key
        state.exited_today = False
    if state.exited_today:
        return
    if not alerts_allowed(now, symbol=state.symbol):
        return
    if not trading_window_open(now, symbol=state.symbol, is_commodity=state.is_commodity):
        return
    # For equities, allow entries during the full session until 15:11 IST.
    if not state.is_commodity and not equity_entry_session_open(now):
        return
    # ── 09:30–09:35 BLACKOUT: no new entries until levels settle ──────────────
    if not state.is_commodity and in_930_blackout(now):
        return
    # ── 09:35 reached: send one-time Telegram that entries are open ────────────
    if (not state.is_commodity
            and not in_930_blackout(now)
            and now.hour == 9 and now.minute == BLACKOUT_END_MINUTE):
        global _930_blackout_alerted
        if not _930_blackout_alerted:
            _930_blackout_alerted = True
            try:
                send_telegram(
                    "⏱ 09:35 IST — Entry blackout lifted.\n"
                    "Adjusted levels are now active. New entries enabled.",
                )
            except Exception:
                pass
    # Entry rules
    if price >= lv.buy_above:
        _open_entry(
            state,
            side="BUY",
            price=price,
            status_line="BUY TRIGGERED",
            hit_level="BUY_ABOVE",
            alert_key="entry_buy",
            dedupe_alert=True,
            event_type="BUY_ENTRY",
        )
    elif price <= lv.sell_below:
        _open_entry(
            state,
            side="SELL",
            price=price,
            status_line="SELL TRIGGERED",
            hit_level="SELL_BELOW",
            alert_key="entry_sell",
            dedupe_alert=True,
            event_type="SELL_ENTRY",
        )


def handle_targets_and_trailing(state: SymbolState, price: float, partial_exits: List[dict]) -> None:
    if not state.in_position or not state.side:
        return
    if state.is_commodity and state.pending_930_recalc:
        return
    now = now_ist()
    # No trading before 9:30 AM
    if now.hour < 9 or (now.hour == 9 and now.minute < 30):
        return
    # For equity: stop trading alerts at 15:11 IST (3:11 PM)
    # For commodities: continue until 23:00 IST
    if not alerts_allowed(now, symbol=state.symbol):
        return
    if not trading_window_open(now, symbol=state.symbol, is_commodity=state.is_commodity):
        return
    lv = state.levels
    x = lv.x
    manual_target = state.manual_target
    manual_label = state.manual_target_label
    if manual_target is not None:
        meets_target = False
        if state.side == "BUY" and price >= manual_target:
            meets_target = True
        elif state.side == "SELL" and price <= manual_target:
            meets_target = True
        if meets_target:
            closed_side = state.side
            qty_to_exit = state.qty_remaining
            entry_px = state.entry_price or manual_target
            if state.side == "BUY":
                gross_pl = (manual_target - entry_px) * qty_to_exit
            else:
                gross_pl = (entry_px - manual_target) * qty_to_exit
            brokerage = BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
            net_pl = gross_pl - brokerage
            record_realized_event(state.symbol, state.side, manual_label or 'MANUAL_TARGET', manual_target, qty_to_exit, gross_pl, net_pl)
            status_line = f"Manual target {manual_target:.2f} hit ({qty_to_exit}) | Gross {gross_pl:.2f} | Net {net_pl:.2f}"
            msg = build_simple_alert(
                "Exit",
                state.symbol,
                status_line,
                state.levels,
                current_price=manual_target,
                hit_level=manual_label or 'MANUAL_TARGET',
                quantity=qty_to_exit,
            )
            if alerts_allowed(now, symbol=state.symbol):
                send_telegram(msg, symbol=state.symbol)
            log_trade_event(now, state.symbol, f"{state.side}_MANUAL_TARGET", manual_target, state.side, qty_to_exit, entry_price=state.entry_price)
            state.in_position = False
            state.side = None
            state.qty_remaining = state.qty_total
            state.last_target_hit_index = -1
            state.entry_price = None
            state.entry_time = None
            state.buy_trailing_sl = None
            state.sell_trailing_sl = None
            state.manual_target = None
            state.manual_target_label = None
            state.exited_today = True
            state.retreat_phase       = None
            state.retreat_65_alerted  = False
            state.retreat_45_alerted  = False
            state.retreat_entry_level = 0.0
            state.retreat_peak_reached = False
            if closed_side:
                schedule_reentry_watch(state, closed_side, manual_label or 'MANUAL_TARGET', manual_target)
            return

    if state.side == "BUY":
        # First target/level hit closes full qty
        entry_px = state.entry_price or 0.0
        for idx, tgt in enumerate(lv.t):
            if tgt <= entry_px:
                continue
            if price >= tgt:
                qty_to_exit = state.qty_remaining
                gross_pl = (tgt - entry_px) * qty_to_exit           # actual exit_price minus entry_price
                net_pl = gross_pl - (BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE)  # entry + exit
                # record for dashboard
                record_realized_event(state.symbol, 'BUY', f'T{idx+1}', tgt, qty_to_exit, gross_pl, net_pl)
                msg = build_simple_alert(
                    "Exit",
                    state.symbol,
                    f"T{idx+1} HIT (exited {qty_to_exit}) | Gross {gross_pl:.2f} | Net {net_pl:.2f}",
                    state.levels,
                    current_price=tgt,
                    hit_level=f"T{idx+1}",
                    quantity=qty_to_exit,
                )
                maybe_alert_once(state, f"buy_t{idx+1}_full", msg)
                log_trade_event(now, state.symbol, f"T{idx+1}", tgt, 'BUY', qty_to_exit, entry_price=state.entry_price)
                # Close position
                state.in_position = False
                state.side = None
                state.qty_remaining = state.qty_total
                state.last_target_hit_index = -1
                state.entry_price = None
                state.entry_time = None
                state.buy_trailing_sl = None
                state.manual_target = None
                state.manual_target_label = None
                state.exited_today = True
                state.retreat_phase       = None
                state.retreat_65_alerted  = False
                state.retreat_45_alerted  = False
                state.retreat_entry_level = 0.0
                state.retreat_peak_reached = False
                schedule_reentry_watch(state, 'BUY', f'T{idx+1}', tgt)
                return
        # Stop loss check (full exit)
        if price <= (state.buy_trailing_sl or lv.buy_sl):
            qty_to_exit = state.qty_remaining
            gross_pl = (price - entry_px) * qty_to_exit              # actual loss (negative when SL hit)
            net_pl = gross_pl - (BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE)
            record_realized_event(state.symbol, 'BUY', 'BUY_SL', price, qty_to_exit, gross_pl, net_pl)
            msg = build_simple_alert("Exit", state.symbol, f"BUY STOP LOSS HIT | Gross {gross_pl:.2f} | Net {net_pl:.2f}", state.levels, current_price=price, hit_level="BUY_SL", quantity=qty_to_exit)
            maybe_alert_once(state, f"buy_sl_close_full", msg)
            log_trade_event(now, state.symbol, 'BUY_SL', price, 'BUY', qty_to_exit, entry_price=state.entry_price)
            state.in_position = False
            state.side = None
            state.qty_remaining = state.qty_total
            state.last_target_hit_index = -1
            state.entry_price = None
            state.entry_time = None
            state.buy_trailing_sl = None
            state.manual_target = None
            state.manual_target_label = None
            state.exited_today = True
            state.retreat_phase       = None
            state.retreat_65_alerted  = False
            state.retreat_45_alerted  = False
            state.retreat_entry_level = 0.0
            state.retreat_peak_reached = False
    else:  # SELL
        entry_px = state.entry_price or 0.0
        for idx, tgt in enumerate(lv.st):
            if tgt >= entry_px and entry_px != 0.0:
                continue
            if price <= tgt:
                qty_to_exit = state.qty_remaining
                gross_pl = (entry_px - tgt) * qty_to_exit            # actual exit_price minus entry_price (sell side)
                net_pl = gross_pl - (BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE)
                # record for dashboard
                record_realized_event(state.symbol, 'SELL', f'ST{idx+1}', tgt, qty_to_exit, gross_pl, net_pl)
                msg = build_simple_alert(
                    "Exit",
                    state.symbol,
                    f"ST{idx+1} HIT (exited {qty_to_exit}) | Gross {gross_pl:.2f} | Net {net_pl:.2f}",
                    state.levels,
                    current_price=tgt,
                    hit_level=f"ST{idx+1}",
                    quantity=qty_to_exit,
                )
                maybe_alert_once(state, f"sell_st{idx+1}_full", msg)
                log_trade_event(now, state.symbol, f"ST{idx+1}", tgt, 'SELL', qty_to_exit, entry_price=state.entry_price)
                state.in_position = False
                state.side = None
                state.qty_remaining = state.qty_total
                state.last_target_hit_index = -1
                state.entry_price = None
                state.entry_time = None
                state.sell_trailing_sl = None
                state.manual_target = None
                state.manual_target_label = None
                state.exited_today = True
                state.retreat_phase       = None
                state.retreat_65_alerted  = False
                state.retreat_45_alerted  = False
                state.retreat_entry_level = 0.0
                state.retreat_peak_reached = False
                schedule_reentry_watch(state, 'SELL', f'ST{idx+1}', tgt)
                return
        if price >= (state.sell_trailing_sl or lv.sell_sl):
            qty_to_exit = state.qty_remaining
            gross_pl = (entry_px - price) * qty_to_exit              # actual loss (negative when SL hit)
            net_pl = gross_pl - (BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE)
            record_realized_event(state.symbol, 'SELL', 'SELL_SL', price, qty_to_exit, gross_pl, net_pl)
            msg = build_simple_alert("Exit", state.symbol, f"SELL STOP LOSS HIT | Gross {gross_pl:.2f} | Net {net_pl:.2f}", state.levels, current_price=price, hit_level="SELL_SL", quantity=qty_to_exit)
            maybe_alert_once(state, f"sell_sl_close_full", msg)
            log_trade_event(now, state.symbol, 'SELL_SL', price, 'SELL', qty_to_exit, entry_price=state.entry_price)
            state.in_position = False
            state.side = None
            state.qty_remaining = state.qty_total
            state.last_target_hit_index = -1
            state.entry_price = None
            state.entry_time = None
            state.sell_trailing_sl = None
            state.manual_target = None
            state.manual_target_label = None
            state.exited_today = True
            state.retreat_phase       = None
            state.retreat_65_alerted  = False
            state.retreat_45_alerted  = False
            state.retreat_entry_level = 0.0
            state.retreat_peak_reached = False


def eod_square_off(states: Dict[str, SymbolState], last_prices: Dict[str, float], partial_exits: List[dict], final_exits: List[dict]) -> None:
    for sym, st in states.items():
        if st.in_position and st.side and st.qty_remaining > 0:
            px = last_prices.get(sym) or st.last_price or 0.0
            qty = st.qty_remaining
            orig_side = st.side
            # Gross PnL per confirmed rules for open positions
            if orig_side == "BUY":
                gross_pl = (px - (st.entry_price or px)) * qty
            else:
                gross_pl = ((st.entry_price or px) - px) * qty
            # Apply flat brokerage for both entry and exit (₹10 each side)
            brokerage = BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
            net_pl = gross_pl - brokerage
            # Use commodity-specific EOD tag so cross-verification can filter correctly
            eod_tag = 'EOD_2300' if st.is_commodity else 'EOD_CLOSE'
            final_exits.append({
                "symbol": sym,
                "side": orig_side,
                "qty": qty,
                "price": px,
                "net_profit": net_pl,
                "brokerage": brokerage,
                "time": now_ist().isoformat(),
                "exit_type": eod_tag,
            })
            record_realized_event(sym, orig_side, eod_tag, px, qty, gross_pl, net_pl)
            log_trade_event(now_ist(), sym, eod_tag, px, orig_side, qty, entry_price=st.entry_price)
            st.in_position = False
            st.side = None
            st.qty_remaining = st.qty_total
            st.last_target_hit_index = -1
            st.entry_price = None
            st.entry_time = None
            st.buy_trailing_sl = None
            st.sell_trailing_sl = None
            st.exited_today = True
            st.manual_target = None
            st.manual_target_label = None
            st.reentry_watch = None
            # Clear retreat monitoring state
            st.retreat_phase        = None
            st.retreat_65_alerted   = False
            st.retreat_45_alerted   = False
            st.retreat_entry_level  = 0.0
            st.retreat_peak_reached = False


def _cross_verify_eod_pnl(
    summary_df: Optional[pd.DataFrame],
    states: Optional[Dict[str, 'SymbolState']] = None,
    *,
    is_commodity: bool = False,
) -> str:
    """Cross-verify EOD P&L across three independent sources.

    Works for both equity (is_commodity=False) and commodity (is_commodity=True).

    Sources:
      1. Trade Summary (Excel/summary_df)  — built from TRADE_LOG (entry+exit pairs)
      2. REALIZED_EVENTS (dashboard)       — built from record_realized_event() calls
      3. ALERT_FEED (Telegram alerts)      — count of T1-T5/ST1-ST5/SL/Retreat alerts

    Returns a formatted verification string appended to EOD Telegram message.
    Discrepancies >₹1 are flagged for investigation.
    """
    label = "Commodity" if is_commodity else "Equity"
    lines: List[str] = ["", f"━━━ {label} P&L Cross-Verification ━━━"]

    # ── Source 1: summary_df (TRADE_LOG pairs) ────────────────────────────────
    s1_gross = s1_net = s1_trades = 0.0
    if summary_df is not None and not summary_df.empty and 'net_pnl' in summary_df.columns:
        df = summary_df[summary_df['symbol'] != 'TOTAL'] if 'symbol' in summary_df.columns else summary_df
        s1_gross  = float(df.get('gross_pnl', pd.Series(dtype=float)).sum())
        s1_net    = float(df.get('net_pnl',   pd.Series(dtype=float)).sum())
        s1_trades = int(len(df))
    lines.append(f"① Trade Log   : {s1_trades} trades | Gross ₹{s1_gross:.2f} | Net ₹{s1_net:.2f}")

    # ── Source 2: REALIZED_EVENTS (dashboard), filtered by equity/commodity ───
    # v10.9 FIX: Only count CLOSING events (exits) in REALIZED_EVENTS to match
    # Source ①. Re-entry BUY/SELL_ENTRY events don't produce P&L so are excluded.
    SKIP_EVENTS = {'BUY_ENTRY', 'SELL_ENTRY', 'BUY_REENTRY_THRESHOLD',
                   'SELL_REENTRY_THRESHOLD', 'BUY_REENTRY_RETOUCH',
                   'SELL_REENTRY_RETOUCH', 'EOD_OPEN'}
    with EVENTS_LOCK:
        all_evts = {k: list(v) for k, v in REALIZED_EVENTS.items()}
    s2_gross = s2_net = s2_trades = 0.0
    for sym, evts in all_evts.items():
        sym_is_commodity = sym in COMMODITY_CONFIG
        if sym_is_commodity != is_commodity:
            continue
        for ev in evts:
            if ev.get('event') in SKIP_EVENTS:
                continue
            # Only count realized P&L events (exits)
            g = float(ev.get('gross', 0.0))
            n = float(ev.get('net',   0.0))
            if g == 0.0 and n == 0.0:
                continue  # skip zero-pnl entries (open positions logged mid-day)
            s2_gross  += g
            s2_net    += n
            s2_trades += 1
    lines.append(f"② Dashboard   : {int(s2_trades)} events | Gross ₹{s2_gross:.2f} | Net ₹{s2_net:.2f}")

    # ── Source 3: ALERT_FEED filtered by commodity type ───────────────────────
    target_hits = sl_hits = retreat_hits = eod_hits = 0
    alert_gross = 0.0
    import re as _re
    gross_pat = _re.compile(r"Gross\s*[₹]?([-\d.]+)")
    for _, txt, ic_flag in list(ALERT_FEED):
        # Filter: ic_flag=True for commodity alerts, False/None for equity
        if is_commodity and ic_flag is not True:
            continue
        if not is_commodity and ic_flag is True:
            continue
        if any(k in txt for k in ["T1 HIT","T2 HIT","T3 HIT","T4 HIT","T5 HIT",
                                   "ST1 HIT","ST2 HIT","ST3 HIT","ST4 HIT","ST5 HIT"]):
            target_hits += 1
            m = gross_pat.search(txt)
            if m:
                try: alert_gross += float(m.group(1))
                except: pass
        elif "STOP LOSS HIT" in txt:
            sl_hits += 1
            m = gross_pat.search(txt)
            if m:
                try: alert_gross += float(m.group(1))
                except: pass
        elif "RETREAT 25%" in txt:
            retreat_hits += 1
            m = gross_pat.search(txt)
            if m:
                try: alert_gross += float(m.group(1))
                except: pass
        elif "EOD_CLOSE" in txt or "EOD_1511" in txt or "EOD_2300" in txt:
            eod_hits += 1

    alert_total = target_hits + sl_hits + retreat_hits + eod_hits
    lines.append(
        f"③ Alert Feed  : {alert_total} exits "
        f"(T={target_hits} SL={sl_hits} Ret={retreat_hits} EOD={eod_hits}) "
        f"| Approx Gross ₹{alert_gross:.2f}"
    )

    # ── Discrepancy check ─────────────────────────────────────────────────────
    gross_diff_12 = abs(s1_gross - s2_gross)
    net_diff_12   = abs(s1_net   - s2_net)
    if gross_diff_12 > 10.0 or net_diff_12 > 10.0:  # v10.9: ₹10 threshold (was ₹1 — caused false alarms)
        lines.append(f"⚠️ MISMATCH ①②: Gross diff ₹{gross_diff_12:.2f} | Net diff ₹{net_diff_12:.2f}")
        lines.append("   → Check for duplicate/missing events in TRADE_LOG vs REALIZED_EVENTS")
    else:
        lines.append(f"✅ ①② Match: diff ₹{gross_diff_12:.2f} gross | ₹{net_diff_12:.2f} net")

    alert_gross_diff = abs(s1_gross - alert_gross)
    if alert_gross_diff > 100.0:
        lines.append(f"⚠️ ①③ Alert gross diff ₹{alert_gross_diff:.2f} (EOD/retreat may not appear in alert text)")
    else:
        lines.append(f"✅ ①③ Alerts consistent (Gross diff ₹{alert_gross_diff:.2f})")

    return "\n".join(lines)


def send_eod_summary(summary_df: Optional[pd.DataFrame], open_positions: Optional[List[dict]] = None, *, force_commodity: bool = False, states: Optional[Dict[str, 'SymbolState']] = None) -> None:
    """Send an EOD Telegram alert built from the trade summary dataframe.

    Includes cross-verification between Trade Log, Dashboard REALIZED_EVENTS,
    and Telegram Alert Feed to confirm P&L accuracy.
    """
    ts = now_ist().strftime("%Y-%m-%d %H:%M:%S IST%z")
    lines: List[str] = [f"🚨 EOD Summary at {ts}"]

    if summary_df is None or summary_df.empty or 'symbol' not in summary_df.columns:
        lines.append("No closed trades recorded today.")
    else:
        df = summary_df.copy()
        if df.iloc[-1]['symbol'] == 'TOTAL':
            totals_row = df.iloc[-1]
            df = df.iloc[:-1] if len(df) > 1 else df.iloc[0:0]
        else:
            totals_row = pd.Series({
                'gross_pnl': df.get('gross_pnl', pd.Series(dtype=float)).sum(),
                'brokerage': df.get('brokerage', pd.Series(dtype=float)).sum(),
                'net_pnl': df.get('net_pnl', pd.Series(dtype=float)).sum(),
                'qty': df.get('qty', pd.Series(dtype=float)).sum(),
            })

        gross_total = float(totals_row.get('gross_pnl', 0.0) or 0.0)
        net_total = float(totals_row.get('net_pnl', 0.0) or 0.0)
        brokerage_total = float(totals_row.get('brokerage', 0.0) or 0.0)
        trade_count = len(df)

        lines.extend([
            f"Closed Trades: {trade_count}",
            f"Gross P&L: ₹{gross_total:.2f}",
            f"Brokerage: ₹{brokerage_total:.2f}",
            f"Net P&L: ₹{net_total:.2f}",
        ])

        if trade_count > 0 and 'net_pnl' in df.columns:
            best_row = df.loc[df['net_pnl'].idxmax()]
            worst_row = df.loc[df['net_pnl'].idxmin()]
            lines.append(
                f"Best: {best_row['symbol']} {best_row['side']} {best_row['exit_type']} ₹{best_row['net_pnl']:.2f}"
            )
            lines.append(
                f"Worst: {worst_row['symbol']} {worst_row['side']} {worst_row['exit_type']} ₹{worst_row['net_pnl']:.2f}"
            )
            symbol_pnl = df.groupby('symbol')['net_pnl'].sum().sort_values(ascending=False)
            top_symbols = symbol_pnl.head(3)
            if not top_symbols.empty:
                top_lines = ", ".join([f"{sym}: ₹{pnl:.2f}" for sym, pnl in top_symbols.items()])
                lines.append(f"Top Symbols: {top_lines}")
            bottom_symbols = symbol_pnl.tail(3)
            if not bottom_symbols.empty:
                bottom_lines = ", ".join([f"{sym}: ₹{pnl:.2f}" for sym, pnl in bottom_symbols.items()])
                lines.append(f"Laggards: {bottom_lines}")

    if open_positions is not None:
        lines.append("")
        if open_positions:
            open_gross = sum(pos['gross'] for pos in open_positions)
            open_net = sum(pos['net'] for pos in open_positions)
            lines.append("Open Positions (Pre Square-off):")
            for pos in open_positions:
                lines.append(
                    f"{pos['symbol']} {pos['side']} qty {pos['qty']} entry {pos['entry_price']:.2f} "
                    f"last {pos['last_price']:.2f} | Gross ₹{pos['gross']:.2f} Net ₹{pos['net']:.2f}"
                )
            lines.append(f"Open Gross: ₹{open_gross:.2f}")
            lines.append(f"Open Net: ₹{open_net:.2f}")
        else:
            lines.append("No open positions at cutoff.")

    # Cross-verification — confirms P&L matches across all sources
    try:
        verify_str = _cross_verify_eod_pnl(summary_df, states, is_commodity=force_commodity)
        lines.append(verify_str)
    except Exception as e:
        logger.debug("EOD cross-verify error: %s", e)

    send_telegram("\n".join(lines), force_commodity=force_commodity)


def build_open_positions_snapshot(states: Dict[str, 'SymbolState'], price_lookup: Dict[str, float]) -> List[dict]:
    snapshot: List[dict] = []
    for sym, st in states.items():
        if not (st.in_position and st.side and (st.qty_remaining or 0) > 0):
            continue
        px = price_lookup.get(sym)
        if px is None:
            px = st.last_price or st.entry_price or 0.0
        qty = st.qty_remaining
        entry_px = st.entry_price or px
        if st.side == "BUY":
            gross_pl = (px - entry_px) * qty
        else:
            gross_pl = (entry_px - px) * qty
        brokerage = BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
        net_pl = gross_pl - brokerage
        snapshot.append(
            {
                "symbol": sym,
                "side": st.side,
                "qty": qty,
                "entry_price": entry_px,
                "last_price": px,
                "gross": gross_pl,
                "net": net_pl,
            }
        )
    return snapshot


def send_eod_open_positions_summary(
    states: Dict[str, 'SymbolState'],
    price_lookup: Dict[str, float],
    *,
    label: str = "EOD Open Positions Summary",
    force_commodity: bool = False
) -> List[dict]:
    ts = now_ist().strftime("%Y-%m-%d %H:%M:%S IST%z")
    lines: List[str] = [f"🚨 {label} at {ts}", ""]
    open_positions = build_open_positions_snapshot(states, price_lookup)
    total_gross = 0.0
    total_net = 0.0
    for pos in open_positions:
        total_gross += pos['gross']
        total_net += pos['net']
        # NOTE: Do NOT record_realized_event here — eod_square_off records EOD_CLOSE
        # for the same positions. Recording EOD_OPEN here caused double-count in dashboard.
        lines.append(
            f"{pos['symbol']} {pos['side']} qty {pos['qty']} entry {pos['entry_price']:.2f} "
            f"price {pos['last_price']:.2f} | Gross {pos['gross']:.2f} Net {pos['net']:.2f}"
        )
    if not open_positions:
        lines.append("No open positions at cutoff.")
    else:
        lines.append("")
        lines.append(f"Total Gross P&L: {total_gross:.2f}")
        lines.append(f"Total Net P&L (after brokerage): {total_net:.2f}")
    send_telegram("\n".join(lines), force_commodity=force_commodity)
    return open_positions


def export_and_send_adjusted_levels(states: Dict[str, 'SymbolState'], levels_dir: str, now: datetime) -> bool:
    """
    Export separate adjusted-levels workbooks for equities and commodities at 09:30
    and send each via the appropriate Telegram bot.
    Returns True if at least one workbook was created.
    """
    ensure_commodities_ready(states)
    adjusted_dir = os.path.join(levels_dir, "adjusted_levels")
    os.makedirs(adjusted_dir, exist_ok=True)

    ts = now.strftime('%Y%m%d_%H%M%S')
    any_created = False

    equity_states = {sym: st for sym, st in states.items() if not st.is_commodity}
    commodity_states = {sym: st for sym, st in states.items() if st.is_commodity}

    if equity_states:
        eq_path = os.path.join(adjusted_dir, f"adjusted_levels_equity_{ts}.xlsx")
        export_levels_layout(equity_states, eq_path, heading="Adjusted Levels (Equity 09:30)")
        caption = f"Adjusted levels (09:30) for equities on {now.strftime('%d %b %Y')}."
        send_telegram_document(eq_path, caption=caption, force_commodity=False)
        any_created = True

    if commodity_states:
        cm_path = os.path.join(adjusted_dir, f"adjusted_levels_commodity_{ts}.xlsx")
        export_levels_layout(commodity_states, cm_path, heading="Adjusted Levels (Commodity 09:30)")
        caption = f"Adjusted levels (09:30) for commodities on {now.strftime('%d %b %Y')}."
        send_telegram_document(cm_path, caption=caption, force_commodity=True)
        any_created = True

    return any_created


# ----------------------------
# Main
# ----------------------------

def main():
    # Trading day check — Dashboard-only mode on weekends/holidays (v8.0 fix)
    # IMPORTANT: We no longer call sys.exit() or return here.
    # Instead we set TRADING_ACTIVE=False and keep the Dash server + ZMQ alive.
    # This prevents autohealer from seeing a spurious code=0 exit on weekends.
    TRADING_ACTIVE = True
    if _MC_AVAILABLE and os.getenv("SKIP_MARKET_CHECK", "0") != "1":
        import pytz as _pytz
        from datetime import datetime as _dt
        _now = _dt.now(_pytz.timezone("Asia/Kolkata"))
        if not MarketCalendar.is_trading_day(_now):
            _nd = MarketCalendar.next_trading_day(_now)
            TRADING_ACTIVE = False
            logger.warning(
                "NOT a trading day (%s). Running in DASHBOARD-ONLY mode. Next: %s.",
                _now.strftime("%A %d %b"), _nd.strftime("%A %d %b %Y")
            )
            try:
                send_telegram(
                    f"AlgoStack v8.0: Not a trading day ({_now.strftime('%a %d %b')}). "
                    f"Dashboard-only mode active. Next: {_nd.strftime('%a %d %b %Y')}."
                )
            except Exception:
                pass
            # NOTE: we do NOT return here — fall through to start Dash + ZMQ
    if not TELEGRAM_CHAT_ID:
        logger.warning("TELEGRAM_CHAT_ID not set. Set env TELEGRAM_CHAT_ID to enable alerts.")

    console = Console()

    # ── Initialize states — ONE batch yfinance call for all 38 stocks ─────────
    # ROOT CAUSE FIX: 38 sequential Ticker().history() calls = 38 HTTP requests
    # → yfinance rate-limits after ~5–10 requests ("Too Many Requests").
    # Solution: yf.download(all_tickers) makes ONE request for all symbols.
    # Falls back to persistent daily cache if batch also fails.
    states: Dict[str, SymbolState] = {}
    INDEX_X_MULTIPLIER = 0.00343
    INDEX_SYMBOLS_SET  = {"NIFTY", "BANKNIFTY"}

    logger.info("Fetching prev_close for all %d symbols (batch mode — 1 API call)…", len(STOCKS))
    batch_closes = get_all_prev_closes_batch(STOCKS)

    failed_syms = []
    for sym in STOCKS:
        pc = batch_closes.get(sym.upper())
        if not pc or pc <= 0:
            failed_syms.append(sym)
            logger.warning("  MISS %-14s — not in batch result", sym)
            continue
        if sym in INDEX_SYMBOLS_SET:
            x_value = pc * INDEX_X_MULTIPLIER
            levels  = calc_levels_for_symbol(sym, pc, x_override=x_value, step_override=x_value)
        else:
            levels = calc_levels_for_symbol(sym, pc)
        states[sym] = SymbolState(symbol=sym, levels=levels, initial_levels=levels, prev_close=pc)
        logger.info("  OK   %-14s  prev_close=%.2f", sym, pc)

    # Retry individual misses with a small delay (gentle on rate limit)
    if failed_syms:
        logger.info("Retrying %d missed symbols individually (0.5s gap)…", len(failed_syms))
        for sym in failed_syms:
            time.sleep(0.5)
            pc = get_prev_close(sym)
            if not pc:
                logger.warning("  SKIP %-14s — prev_close unavailable after retry", sym)
                continue
            if sym in INDEX_SYMBOLS_SET:
                x_value = pc * INDEX_X_MULTIPLIER
                levels  = calc_levels_for_symbol(sym, pc, x_override=x_value, step_override=x_value)
            else:
                levels = calc_levels_for_symbol(sym, pc)
            states[sym] = SymbolState(symbol=sym, levels=levels, initial_levels=levels, prev_close=pc)
            logger.info("  RETRY OK %-14s  prev_close=%.2f", sym, pc)

    logger.info("Initialised %d/%d symbols", len(states), len(STOCKS))

    if not states:
        logger.error(
            "Could not load prev_close for ANY symbol.\n"
            "  → Wait 5 minutes for Yahoo Finance rate limit to expire, then restart.\n"
            "  → Or copy a prev_closes_persistent_YYYYMMDD.json file into levels/ "
            "    from a previous session."
        )
        # Don't crash — keep process alive so Telegram alert goes out

    # Check if starting post-market hours to avoid loading old trades from disk
    # Equity: after 15:12, Commodity: after 23:01
    now = now_ist()
    is_post_equity_market = (now.hour > 15 or (now.hour == 15 and now.minute >= 12))
    is_post_commodity_market = (now.hour > 23 or (now.hour == 23 and now.minute >= 1))
    
    # Only hydrate trade log from disk if NOT starting post-market hours
    # This ensures EOD alerts match the dashboard (which uses REALIZED_EVENTS, not TRADE_LOG)
    global ALLOW_DISK_TRADE_SNAPSHOT
    if not is_post_equity_market and not is_post_commodity_market:
        hydrate_trade_log(now.strftime('%Y%m%d'))
    else:
        # Clear TRADE_LOG when starting post-market to ensure no old trades are included
        TRADE_LOG.clear()
        TRADE_LOG_KEYS.clear()
        _TRADE_LOG_HYDRATED.clear()
        # Also disable disk snapshot fallback so EOD summaries don't resurrect old trades
        ALLOW_DISK_TRADE_SNAPSHOT = False
        if is_post_equity_market:
            logger.info("Starting post-equity-market hours (after 15:12) - cleared TRADE_LOG to avoid false EOD alerts")
        if is_post_commodity_market:
            logger.info("Starting post-commodity-market hours (after 23:01) - cleared TRADE_LOG to avoid false EOD alerts")
    
    ensure_commodities_ready(states)

    # Initialize price history and start dashboards if enabled
    init_price_history(list(states.keys()))
    if ENABLE_DASH and DASH_AVAILABLE:
        equity_states_dash = {sym: st for sym, st in states.items() if not st.is_commodity}
        commodity_states_dash = {sym: st for sym, st in states.items() if st.is_commodity}
        if equity_states_dash:
            # v10.8: Tunnel DISABLED on Algofinal — unified_dash_v3.py (port 8055) is the
            # single public entry point. Algofinal's port-8050 tunnel was overwriting
            # dashboard_url.json and sending the wrong URL to Telegram.
            start_dash_app(
                equity_states_dash,
                port=DASH_PORT,
                app_title="Equity Levels Dashboard",
                use_public_tunnel=False,   # NO tunnel — unified dash handles this
                ngrok_authtoken=NGROK_AUTHTOKEN_EQUITY,
            )
            logger.info(f"Equity dashboard starting on http://0.0.0.0:{DASH_PORT}")
            logger.info(f"Access from this machine: http://localhost:{DASH_PORT}")
        # Commodity dashboard: DISABLED — commodities tracked internally but no separate Dash UI
        # To re-enable: uncomment the block below and set ENABLE_COMMODITY_DASH=1 env var
        # if commodity_states_dash:
        #     start_dash_app(
        #         commodity_states_dash,
        #         port=COMMODITY_DASH_PORT,
        #         app_title="Commodity Levels Dashboard",
        #         use_public_tunnel=True,
        #         ngrok_authtoken=NGROK_AUTHTOKEN_COMMODITY,
        #     )
        #     logger.info(f"Commodity dashboard starting on http://0.0.0.0:{COMMODITY_DASH_PORT}")
        #     logger.info(f"Access from this machine: http://localhost:{COMMODITY_DASH_PORT}")

        # Merged dashboard: DISABLED
        # start_merged_dash_app(...) removed — will be re-enabled later
    elif ENABLE_DASH and not DASH_AVAILABLE:
        logger.warning("Dashboard enabled but Dash dependencies not available. Install with: pip install dash plotly")

    levels_dir = "levels"
    os.makedirs(levels_dir, exist_ok=True)
    # Always export a raw Initial Levels snapshot at startup
    try:
        now_for_init = now_ist()
        # Equity date logic:
        # - Before 9:30 AM: Use current day (levels are for today when market opens)
        # - Between 9:30 AM - 3:30 PM: Use current day (trading hours)
        # - After 3:30 PM: Use next day (levels are for tomorrow)
        hour = now_for_init.hour
        minute = now_for_init.minute
        if hour < 9 or (hour == 9 and minute < 30):
            # Before 9:30 AM - use current day
            trading_date_init = now_for_init
        elif hour < 15 or (hour == 15 and minute <= 30):
            # Between 9:30 AM - 3:30 PM - use current day
            trading_date_init = now_for_init
        else:
            # After 3:30 PM - use next day
            trading_date_init = now_for_init + timedelta(days=1)
        initial_levels_path = os.path.join(
            levels_dir,
            f"initial_levels_{trading_date_init.strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        export_levels_layout(
            states,
            initial_levels_path,
            heading="Initial Levels",
            trading_date=trading_date_init,
        )
        logger.info("Initial Levels workbook exported at startup: %s", initial_levels_path)

        # When started outside trading hours (before 9:30 AM or after 3:30 PM), export separate equity and commodity files
        # Check if we're outside trading hours (9:30 AM - 3:30 PM)
        hour = now_for_init.hour
        minute = now_for_init.minute
        outside_trading_hours = (hour < 9 or (hour == 9 and minute < 30)) or (hour > 15 or (hour == 15 and minute > 30))
        if outside_trading_hours:
            # For equity: 
            #   - Before 9:30 AM: levels are FOR today (trading_date_init = current day)
            #   - After 3:30 PM: levels are FOR tomorrow (trading_date_init = next day)
            # For commodities: levels are FOR previous trading day (using previous day's 9:30 price)
            prevday_dir = os.path.join(levels_dir, "initial_levels_prevday")
            os.makedirs(prevday_dir, exist_ok=True)
            ts_nd = trading_date_init.strftime('%Y%m%d_%H%M%S')
            equity_states_nd = {sym: st for sym, st in states.items() if not st.is_commodity}
            commodity_states_nd = {sym: st for sym, st in states.items() if st.is_commodity}

            if equity_states_nd:
                # Equity: trading_date_init is already set correctly above:
                #   - Before 9:30 AM: current day
                #   - After 3:30 PM: next day
                eq_path_nd = os.path.join(prevday_dir, f"initial_levels_equity_{ts_nd}.xlsx")
                # Determine heading based on whether it's current day or next day
                if trading_date_init.date() == now_for_init.date():
                    heading_text = "Initial Levels (Equity - Current Day)"
                    caption_text = f"Initial levels for equity trading day ({trading_date_init.strftime('%d %b %Y')}) - calculated using previous close."
                else:
                    heading_text = "Initial Levels (Equity - Next Trading Day)"
                    caption_text = f"Initial levels for next equity trading day ({trading_date_init.strftime('%d %b %Y')}) - calculated using previous close."
                export_levels_layout(
                    equity_states_nd,
                    eq_path_nd,
                    heading=heading_text,
                    trading_date=trading_date_init,
                )
                send_telegram_document(eq_path_nd, caption=caption_text, force_commodity=False)

            if commodity_states_nd:
                # Commodities: use previous trading day date (using previous day's 9:30 price)
                commodity_trading_date = now_for_init - timedelta(days=1)
                cm_path_nd = os.path.join(prevday_dir, f"initial_levels_commodity_prev_{commodity_trading_date.strftime('%Y%m%d_%H%M%S')}.xlsx")
                export_levels_layout(
                    commodity_states_nd,
                    cm_path_nd,
                    heading="Initial Levels (Previous Trading Day - Commodity)",
                    trading_date=commodity_trading_date,
                )
                caption_cm_nd = f"Initial levels for previous commodity session on {commodity_trading_date.strftime('%d %b %Y')}."
                send_telegram_document(cm_path_nd, caption=caption_cm_nd, force_commodity=True)
    except Exception as e:
        logger.warning(f"Failed to export/send startup initial levels Excel: {e}")

    partial_exits: List[dict] = []
    final_exits: List[dict] = []
    adjusted_levels_exported = False
    analysis_created = False
    eod_summary_sent = False
    initial_levels_sent = False
    commodity_analysis_sent = False

    afternoon_eod_done = False
    afternoon_eod_summary_done = False
    afternoon_summary_df: Optional[pd.DataFrame] = None
    afternoon_open_snapshot: Optional[List[dict]] = None
    afternoon_analysis_path: Optional[str] = None

    # If program starts anytime on/after 09:30, first replay premarket once (equities)
    now = now_ist()
    start_930 = now.replace(hour=9, minute=30, second=0, microsecond=0)
    end_1530 = now.replace(hour=15, minute=30, second=0, microsecond=0)
    adjustment_930_done = False  # Track if 9:30 adjustment has been applied
    if now >= start_930 and now <= end_1530:
        # Live session between 09:30 and 15:30 – replay premarket, then 9:30 adjust.
        for sym, st in states.items():
            replay_premarket_adjustments(st)
        for sym, st in states.items():
            adjust_levels_at_930(st)
        for sym, st in states.items():
            st.adjusted_locked = True
        adjustment_930_done = True
        if not adjusted_levels_exported and should_export_adjusted_levels(now):
            try:
                adjusted_levels_exported = export_and_send_adjusted_levels(states, levels_dir, now)
            except Exception as e:
                logger.warning(f"Failed to export/send adjusted levels Excel: {e}")
        # Write shared market data so scanners can read 09:30 prices + premarket history
        try:
            write_shared_market_data(now.strftime('%Y%m%d'))
        except Exception as e:
            logger.warning("Failed to write shared market data: %s", e)
    else:
        # Outside live 09:30–15:30 window: keep levels locked; no backfilled 09:30 adjustment.
        for sym, st in states.items():
            st.adjusted_locked = True

    # Startup current-price re-anchor (equities only).
    # If the equity market is open at startup, center the ladder around the *current*
    # price so price lies between the updated sell_below (new lower) and buy_above (new upper).
    # This keeps T1-T5 / ST1-ST5, stop-loss, retreat and re-entry thresholds consistent.
    try:
        if market_open(now_ist()):
            live_px = _batch_refresh_prices(list(states.keys()))
            re_cnt = 0
            for sym, st in states.items():
                if st.is_commodity:
                    continue
                px = live_px.get(sym.upper())
                if not px or px <= 0:
                    continue
                new_lv, _bounds = _reanchor_levels_around_price(st.levels, px)
                if new_lv:
                    st.levels = new_lv
                    st.adjusted_locked = True
                    re_cnt += 1
            if re_cnt:
                logger.info("Startup re-anchor: centered %d/%d equity ladders on live prices",
                            re_cnt, len([s for s in states.values() if not s.is_commodity]))
    except Exception as e:
        logger.warning("Startup re-anchor skipped/failed: %s", e)
    
    # If starting after 15:12 but before 23:00 on the same trading day, immediately generate EQUITY EOD summary using 15:11 closes
    # Note: Commodity EOD is separate and runs at 23:00
    # Only send EOD if: (1) it's after 15:12, (2) it's before 23:00 (to avoid sending very late at night), (3) it's the same trading day, (4) it hasn't been sent yet
    if (now.hour > 15 or (now.hour == 15 and now.minute >= 12)) and now.hour < 23:
        analysis_1512_time = now.replace(hour=15, minute=12, second=0, microsecond=0)
        # Only send if it's the same trading day and hasn't been sent yet
        if now.date() == analysis_1512_time.date() and not eod_summary_sent:
            try:
                # Filter to equity symbols only (exclude commodities)
                equity_states = {sym: st for sym, st in states.items() if not st.is_commodity}
                close_prices_1511 = get_close_prices_1511(equity_states)
                open_positions_snapshot = send_eod_open_positions_summary(
                    equity_states,
                    close_prices_1511,
                    label="EOD Open Positions Summary (15:11 Close) - Equity",
                    force_commodity=False
                )
                summary_df = create_trade_analysis_excel(
                    TRADE_LOG,
                    equity_states,
                    now.strftime('%Y%m%d'),
                    close_prices_1511=close_prices_1511,
                    return_summary=True,
                    allow_disk_snapshot=ALLOW_DISK_TRADE_SNAPSHOT,
                )
                send_eod_summary(summary_df, open_positions=open_positions_snapshot, force_commodity=False, states=equity_states)
                # Send the trade analysis Excel file as a Telegram document
                analysis_file_path = os.path.join("trade_analysis", f"trade_analysis_{now.strftime('%Y%m%d')}.xlsx")
                if os.path.exists(analysis_file_path):
                    caption = f"Equity trade analysis (15:11) ready for {now.strftime('%d %b %Y')}."
                    send_telegram_document(analysis_file_path, caption=caption, force_commodity=False)
                analysis_created = True
                eod_summary_sent = True
                logger.info("Equity EOD summary + analysis generated at startup (after 15:12) - equity only.")
            except Exception as e:
                logger.warning(f"Failed to create/send equity EOD summary at startup: {e}")

    # If starting after 09:30 and commodities are present, replay their session once
    if now >= start_930:
        try:
            replay_commodity_session(states)
            logger.info("Commodity session replay completed at startup.")
        except Exception as exc:
            logger.warning("Commodity replay at startup failed: %s", exc)

    # Main loop (keep running even after 15:30 to keep dashboard live)
    cur_interval_idx = 0
    eod_sent = False
    # ── ZMQ dual-transport publisher (ZMQ PUB + JSON file fallback) ───────────
    # When ipc_bus.py is present: publishes on tcp://127.0.0.1:28081 so scanners
    # receive prices with microsecond latency instead of polling a disk file.
    # Falls back to JSON-file-only (original UdpPricePublisher) if not installed.
    if _PATCH_ACTIVE and _ZmqPricePublisher is not None:
        price_publisher = _ZmqPricePublisher()
        logger.info("Price publisher: ZMQ PUB (tcp://127.0.0.1:28081) + JSON file dual-transport")
    else:
        price_publisher = UdpPricePublisher()
        logger.info("Price publisher: JSON file only (install pyzmq for ZMQ mode)")

    # Start background batch price fetcher (replaces slow per-symbol yfinance calls)
    # Refreshes ALL 38+ equity prices in one HTTP call every 2s -> <4s latency
    _start_batch_price_fetcher(list(states.keys()), interval_s=2.0)
    try:
        with Live(build_table(states), console=console, refresh_per_second=8) as live:
            while True:
                now = now_ist()
                if after_930(now):
                    ensure_commodities_ready(states)
                # Apply 9:30 adjustment once when we first reach 9:30 (for live execution)
                if not adjustment_930_done and now.hour == 9 and now.minute >= 30:
                    # Apply 9:30 adjustment for all stocks
                    for sym, st in states.items():
                        adjust_levels_at_930(st)
                    adjustment_930_done = True
                    # Lock adjustments after 9:30 adjustment
                    for sym, st in states.items():
                        st.adjusted_locked = True
                    # Write shared market data so scanners read 09:30 prices & premarket history
                    try:
                        write_shared_market_data(now.strftime('%Y%m%d'))
                    except Exception as e:
                        logger.warning("Failed to write shared market data at 09:30: %s", e)
                if adjustment_930_done and not adjusted_levels_exported and should_export_adjusted_levels(now):
                    try:
                        adjusted_levels_exported = export_and_send_adjusted_levels(states, levels_dir, now)
                    except Exception as e:
                        logger.warning(f"Failed to export/send adjusted levels Excel: {e}")
                if (    
                    adjustment_930_done
                    and not initial_levels_sent
                    and should_export_adjusted_levels(now)
                ):
                    try:
                        trading_date = now
                        ts = trading_date.strftime('%Y%m%d_%H%M%S')
                        initial_dir_930 = os.path.join(levels_dir, "initial_levels_930")
                        os.makedirs(initial_dir_930, exist_ok=True)

                        equity_states_930 = {sym: st for sym, st in states.items() if not st.is_commodity}
                        commodity_states_930 = {sym: st for sym, st in states.items() if st.is_commodity}

                        any_initial = False

                        if equity_states_930:
                            eq_initial_path = os.path.join(initial_dir_930, f"initial_levels_equity_930_{ts}.xlsx")
                            export_levels_layout(
                                equity_states_930,
                                eq_initial_path,
                                heading="Initial Levels @ 09:30 (Equity)",
                                trading_date=trading_date,
                            )
                            caption_eq = f"Initial levels (09:30) for equities on {trading_date.strftime('%d %b %Y')}."
                            send_telegram_document(eq_initial_path, caption=caption_eq, force_commodity=False)
                            any_initial = True

                        if commodity_states_930:
                            cm_initial_path = os.path.join(initial_dir_930, f"initial_levels_commodity_930_{ts}.xlsx")
                            export_levels_layout(
                                commodity_states_930,
                                cm_initial_path,
                                heading="Initial Levels @ 09:30 (Commodity)",
                                trading_date=trading_date,
                            )
                            caption_cm = f"Initial levels (09:30) for commodities on {trading_date.strftime('%d %b %Y')}."
                            send_telegram_document(cm_initial_path, caption=caption_cm, force_commodity=True)
                            any_initial = True

                        if any_initial:
                            initial_levels_sent = True
                            logger.info("Initial levels (09:30) for equity/commodity exported and sent via Telegram.")
                    except Exception as e:
                        logger.warning(f"Failed to export/send 09:30 initial levels: {e}")

                # Equity EOD: 15:11 square-off, 15:12 send summary + Excel
                if now.hour == 15 and now.minute == 11 and not afternoon_eod_done:
                    try:
                        equity_states = {sym: st for sym, st in states.items() if not st.is_commodity}
                        if equity_states:
                            close_prices_1511 = get_close_prices_1511(equity_states)
                            afternoon_open_snapshot = send_eod_open_positions_summary(
                                equity_states,
                                close_prices_1511,
                                label="EOD Open Positions Summary (15:11 Close) - Equity",
                                force_commodity=False,
                            )
                            eod_square_off(equity_states, close_prices_1511, partial_exits, final_exits)
                            date_str = now.strftime('%Y%m%d')
                            afternoon_analysis_dir = "trade_analysis"
                            afternoon_file_stem = "trade_analysis"
                            afternoon_analysis_path = os.path.join(
                                afternoon_analysis_dir, f"{afternoon_file_stem}_{date_str}.xlsx"
                            )
                            afternoon_summary_df = create_trade_analysis_excel(
                                TRADE_LOG,
                                equity_states,
                                date_str,
                                close_prices_1511=close_prices_1511,
                                return_summary=True,
                                analysis_dir=afternoon_analysis_dir,
                                file_stem=afternoon_file_stem,
                                allow_disk_snapshot=ALLOW_DISK_TRADE_SNAPSHOT,
                            )
                        afternoon_eod_done = True
                        logger.info("Equity EOD square-off completed at 15:11.")
                    except Exception as e:
                        logger.warning(f"Failed to process equity EOD at 15:11: {e}")

                if (
                    now.hour == 15
                    and now.minute == 12
                    and afternoon_eod_done
                    and not afternoon_eod_summary_done
                ):
                    try:
                        send_eod_summary(
                            afternoon_summary_df,
                            open_positions=afternoon_open_snapshot,
                            force_commodity=False,
                            states={sym: st for sym, st in states.items() if not st.is_commodity},
                        )
                        if afternoon_analysis_path and os.path.exists(afternoon_analysis_path):
                            caption = f"Equity trade analysis (session ending 15:11) ready for {now.strftime('%d %b %Y')}."
                            send_telegram_document(
                                afternoon_analysis_path,
                                caption=caption,
                                force_commodity=False,
                            )
                        afternoon_eod_summary_done = True
                        logger.info("Equity EOD summary + Excel sent at 15:12.")
                    except Exception as e:
                        logger.warning(f"Failed to send equity EOD summary at 15:12: {e}")

                # Commodity EOD summary at 23:00 using 23:00 prices (COMMODITY ONLY - no equity)
                if now.hour == 23 and now.minute == 0 and not commodity_analysis_sent:
                    try:
                        # Filter to commodity symbols only (exclude equity)
                        commodity_states = {sym: st for sym, st in states.items() if st.is_commodity}
                        close_prices_2300 = get_close_prices_2300(commodity_states)
                        open_positions_snapshot = send_eod_open_positions_summary(
                            commodity_states,
                            close_prices_2300,
                            label="EOD Open Positions Summary (23:00 Close) - Commodity",
                            force_commodity=True
                        )
                        file_path = create_commodity_trade_analysis_excel(
                            TRADE_LOG,
                            commodity_states,
                            now.strftime('%Y%m%d'),
                        )
                        # Create summary dataframe for EOD summary message
                        summary_df = create_trade_analysis_excel(
                            TRADE_LOG,
                            commodity_states,
                            now.strftime('%Y%m%d'),
                            close_prices_1511=close_prices_2300,
                            return_summary=True,
                            allow_disk_snapshot=ALLOW_DISK_TRADE_SNAPSHOT,
                        )
                        # Send EOD summary message (similar to equity)
                        send_eod_summary(summary_df, open_positions=open_positions_snapshot, force_commodity=True, states=commodity_states)
                        # Also send the detailed Excel file
                        if file_path and os.path.exists(file_path):
                            caption = f"Commodity trade analysis (23:00) ready for {now.strftime('%d %b %Y')}."
                            send_telegram_document(file_path, caption=caption, force_commodity=True)
                        last_prices = {
                            s: close_prices_2300.get(s, commodity_states[s].last_price or 0.0)
                            for s in commodity_states
                        }
                        eod_square_off(commodity_states, last_prices, partial_exits, final_exits)
                        commodity_analysis_sent = True
                        logger.info("Commodity EOD summary + analysis sent at 23:00 with 23:00 prices (commodity only)")
                    except Exception as e:
                        commodity_analysis_sent = True
                        logger.warning(f"Failed to create/send commodity EOD summary at 23:00: {e}")

                for st in states.values():
                    if st.is_commodity and st.pending_930_recalc:
                        refresh_commodity_levels_from_930(st)

                # Fetch and process per symbol
                # ── Performance: cache window flags once per tick ─────────────
                _is_premarket   = premarket_window(now)
                _can_trade_eq   = not _is_premarket and is_between(now, 9, 30, 15, 11)
                _can_trade_comm = is_between(now, 9, 30, 23, 0)

                # ── OPT-03: Read prices from background batch cache (<1ms) ────
                # BatchPriceFetcher refreshes ALL symbols every 2s in background.
                # No yfinance call here — main loop never blocks on HTTP.
                with _LIVE_PRICE_CACHE_LOCK:
                    _fetched = dict(_LIVE_PRICE_CACHE)

                # For premarket fallback: fetch missing symbols individually
                if _is_premarket:
                    for sym, st in states.items():
                        if sym not in _fetched or _fetched[sym] is None:
                            try:
                                df_fb = fetch_intraday_ohlc(sym, "1m", "Today")
                                if df_fb is not None and not df_fb.empty:
                                    _fetched[sym] = float(df_fb["Close"].astype(float).iloc[-1])
                            except Exception:
                                pass

                latest_prices_file = {}
                for sym, st in states.items():
                    price = _fetched.get(sym)
                    if price is None:
                        continue

                    # Skip full processing if price unchanged (unless in position)
                    price_changed = (price != st.last_price)
                    st.last_price = price
                    append_price_history(sym, now, price)
                    # Include ALL symbols (equity + commodity) so scanners can sweep both
                    latest_prices_file[sym] = float(price)

                    # Premarket adjustment only within 9:15-9:30
                    if st.is_commodity:
                        st.adjusted_locked = True
                    elif _is_premarket:
                        st.adjusted_locked = False
                        if price_changed:
                            handle_premarket_adjustment(st, price)
                    else:
                        st.adjusted_locked = True

                    # Use cached window flags — avoids repeated datetime comparisons
                    can_trade_now = _can_trade_comm if st.is_commodity else _can_trade_eq
                    if can_trade_now:
                        if price_changed or st.in_position:
                            process_reentry_watch(st, price)
                            try_entry(st, price)
                            handle_targets_and_trailing(st, price, partial_exits)
                            handle_retreat_monitoring(st, price)

                # Write latest prices to shared file for scanners — only when changed
                if latest_prices_file:
                    price_publisher.publish(latest_prices_file, now)

                # Update live table
                live.update(build_table(states))

                # Pace control with fallback intervals
                sleep_s = UPDATE_INTERVALS_SECONDS[cur_interval_idx]
                try:
                    time.sleep(sleep_s)
                except Exception:
                    pass
    finally:
        try:
            price_publisher.close()
        except Exception:
            pass

    console.print("Done. EOD summary sent.")
    _last_px = {sym: st.last_price for sym, st in states.items() if st.last_price}
    eod_save_summary_excel(TRADE_LOG, states, _last_px, date_str)


def compute_eod_pnl_from_alerts(alert_feed, last_prices):
    import re
    from collections import defaultdict
    positions = defaultdict(list) # symbol -> list of [entry dicts]
    # Basic regexps for parsing needed fields
    entry_pat = re.compile(r"(BUY|SELL) TRIGGERED", re.I)
    exit_pat = re.compile(r"(T\d|ST\d|STOP LOSS HIT)", re.I)
    qty_pat = re.compile(r"Quantity: (\d+)")
    price_pat = re.compile(r"Current Price: ([\d.]+)")
    sym_pat = re.compile(r"^🚨 ([A-Z0-9]+)\b")
    side_pat = re.compile(r"BUY|SELL", re.I)
    t1_pat = re.compile(r"Status: (T\d|ST\d|STOP LOS[SS] HIT|BUY TRIGGERED|SELL TRIGGERED)")
    entry_px_pat = re.compile(r"Entry Price: ([\d.]+)")
    # Logged results
    realized_pl = 0
    total_brokerage = 0
    open_positions = {}
    alert_seen = set()
    for alert_item in alert_feed:
        # ALERT_FEED stores 3-tuples (ts, txt, is_commodity) — safe unpack
        ts = alert_item[0]
        txt = alert_item[1] if len(alert_item) > 1 else ""
        if txt in alert_seen: continue
        alert_seen.add(txt)
        # Symbol
        m = sym_pat.search(txt)
        if not m: continue
        symbol = m.group(1)
        # Side
        mside = entry_pat.search(txt) or side_pat.search(txt)
        side = mside.group(1).upper() if mside else None
        # Entry: BUY/SELL TRIGGERED
        if entry_pat.search(txt):
            mprice = price_pat.search(txt)
            mqty = qty_pat.search(txt)
            if mprice and mqty:
                p = float(mprice.group(1))
                q = int(mqty.group(1))
                positions[symbol].append({'side':side,'entry_price':p,'qty':q,'open':True})
        # Exit: Target or Stoploss
        mexit = t1_pat.search(txt)
        if mexit and (mexit.group(1).startswith('T') or mexit.group(1).startswith('ST') or 'STOP LOSS' in mexit.group(1)):
            mprice = price_pat.search(txt)
            mqty = qty_pat.search(txt)
            if mprice and mqty:
                px = float(mprice.group(1))
                q = int(mqty.group(1))
                # Find latest open position of same side
                found = False
                for pos in positions[symbol]:
                    if pos['open'] and pos['qty'] == q and pos['side'] == side:
                        entry_px = pos['entry_price']
                        if side == 'BUY':
                            pl = (px - entry_px) * q
                        else:
                            pl = (entry_px - px) * q
                        realized_pl += pl
                        total_brokerage += BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
                        pos['open'] = False
                        found = True
                        break
                if not found:
                    # Try find any open position to match qty
                    for pos in positions[symbol]:
                        if pos['open'] and pos['qty'] == q:
                            entry_px = pos['entry_price']
                            if pos['side'] == 'BUY':
                                pl = (px - entry_px) * q
                            else:
                                pl = (entry_px - px) * q
                            realized_pl += pl
                            total_brokerage += BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
                            pos['open'] = False
                            break
    # All open positions at EOD
    for symbol, plist in positions.items():
        for pos in plist:
            if pos['open']:
                q = pos['qty']
                entry_px = pos['entry_price']
                side = pos['side']
                eod_px = last_prices.get(symbol)
                if eod_px:
                    if side == 'BUY':
                        pl = (eod_px - entry_px) * q
                    else:
                        pl = (entry_px - eod_px) * q
                    realized_pl += pl
                    total_brokerage += BROKERAGE_FLAT_PER_SIDE + BROKERAGE_FLAT_PER_SIDE
    net_profit = realized_pl - total_brokerage
    print(f"PnL from Telegram Alerts:")
    print(f"Total Profit: {realized_pl:.2f}")
    print(f"Total Brokerage: {total_brokerage:.2f}")
    print(f"Net Profit: {net_profit:.2f}")
    logger.info(f"PnL from Telegram Alerts | Profit: {realized_pl:.2f} | Brokerage: {total_brokerage:.2f} | Net: {net_profit:.2f}")


def _fmt_levels_row(prefix: str, lv: 'Levels') -> List[str]:
    return [
        f"{prefix} Buy Above: {lv.buy_above:.2f}",
        f"{prefix} T1..T5: {lv.t[0]:.2f}, {lv.t[1]:.2f}, {lv.t[2]:.2f}, {lv.t[3]:.2f}, {lv.t[4]:.2f}",
        f"{prefix} Buy SL: {lv.buy_sl:.2f}",
        f"{prefix} Sell Below: {lv.sell_below:.2f}",
        f"{prefix} ST1..ST5: {lv.st[0]:.2f}, {lv.st[1]:.2f}, {lv.st[2]:.2f}, {lv.st[3]:.2f}, {lv.st[4]:.2f}",
        f"{prefix} Sell SL: {lv.sell_sl:.2f}",
    ]


def build_premarket_adjustment_comparison_alert(symbol: str, old_lv: 'Levels', new_lv: 'Levels', current_price: float, status_line: str) -> str:
    ts = now_ist().strftime("%Y-%m-%d %H:%M:%S IST%z")
    lines: List[str] = []
    lines.append(f"🚨 {symbol} Alert at {ts}")
    lines.append("")
    lines.append(f"Previous Close: {new_lv.previous_close:.2f}")
    lines.append(f"Current Price: {current_price:.2f}")
    change_pct = ((current_price - new_lv.previous_close) / new_lv.previous_close * 100.0) if new_lv.previous_close else 0.0
    lines.append(f"Change: {_fmt_pct(change_pct)}")
    lines.append(f"Deviation (X): {new_lv.x:.2f}")
    lines.append(f"Status: {status_line}")
    lines.append("")
    lines.append("📊 Technical Analysis:")
    lines.extend(_fmt_levels_row("Initial", old_lv))
    lines.append("")
    lines.extend(_fmt_levels_row("Adjusted", new_lv))
    return "\n".join(lines)


# Realized P&L events per symbol for dashboard rendering
REALIZED_EVENTS: Dict[str, List[dict]] = {}
EVENTS_LOCK = threading.Lock()

def record_realized_event(symbol: str, side: str, event: str, price: float, qty: int, gross: float, net: float) -> None:
    with EVENTS_LOCK:
        if symbol not in REALIZED_EVENTS:
            REALIZED_EVENTS[symbol] = []
        REALIZED_EVENTS[symbol].append({
            'time': now_ist().strftime('%H:%M:%S'),
            'event': event,
            'side': side,
            'price': price,
            'qty': qty,
            'gross': gross,
            'net': net,
        })


# ── LIVE PRICE SHARING — file-based (replaces UDP unicast) ────────────────────
# Root cause of scanner ticks=0: Algofinal used unicast UDP to 127.0.0.1:28080.
# On Windows, unicast UDP is delivered to exactly ONE bound socket — whichever
# process bound first (scanner1) got all packets; the others (scanner2/3) got
# nothing.  The fix: write prices to a shared JSON file every tick. Any number
# of scanners can read it simultaneously — no OS socket delivery limits.
LIVE_PRICES_FILE = os.path.join("levels", "live_prices.json")
UDP_PRICE_HOST = "127.0.0.1"   # kept for backward compatibility only
UDP_PRICE_PORT = 28080          # kept for backward compatibility only


class UdpPricePublisher:
    """Writes live prices to a shared JSON file every tick.

    All three scanners read this file — no UDP socket conflicts.
    The class is named UdpPricePublisher for drop-in backward compatibility;
    it no longer sends UDP packets (which caused ticks=0 on scanners 2 and 3).
    """
    def __init__(self, host: str = UDP_PRICE_HOST, port: int = UDP_PRICE_PORT) -> None:
        self._last_write_t = 0.0
        os.makedirs("levels", exist_ok=True)

    def close(self) -> None:
        pass

    def publish(self, prices: dict, ts: 'datetime', min_interval_s: float = 0.0) -> None:
        """Write prices to shared JSON file. Never blocks the trading loop."""
        try:
            now_t = time.time()
            if min_interval_s and (now_t - self._last_write_t) < float(min_interval_s):
                return
            payload = {
                "ts": ts.strftime("%Y-%m-%d %H:%M:%S"),
                "prices": prices,
            }
            data = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
            # Atomic write: write to tmp then rename so scanners never read a partial file
            tmp = LIVE_PRICES_FILE + ".tmp"
            with open(tmp, "w", encoding="utf-8") as fh:
                fh.write(data)
            os.replace(tmp, LIVE_PRICES_FILE)
            self._last_write_t = now_t
        except Exception:
            return


if __name__ == "__main__":
    main()


